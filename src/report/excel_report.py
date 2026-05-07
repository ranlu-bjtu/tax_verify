from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.models.compare_result import CompareResult, CompareStatus
from src.report.report_generator import BaseReportGenerator

REPORT_COLUMNS = [
    "序号", "企业名称", "纳税人识别号", "税款所属期", "税种",
    "表单编码", "表单名称", "页码", "行名称", "栏次",
    "列名称", "接口字段ID", "接口JSONPath", "数据类型",
    "接口原始值", "网页原始值", "接口归一化值", "网页归一化值",
    "是否一致", "差异类型", "差异值", "容差", "差异原因", "备注",
]

STATUS_DISPLAY = {
    CompareStatus.MATCH: "一致",
    CompareStatus.TOLERANCE_MATCH: "一致(容差内)",
    CompareStatus.MISMATCH: "不一致",
    CompareStatus.API_MISSING: "接口缺失",
    CompareStatus.WEB_MISSING: "网页缺失",
    CompareStatus.BOTH_MISSING: "双方缺失",
    CompareStatus.PARSE_ERROR: "解析错误",
    CompareStatus.MAPPING_ERROR: "映射错误",
    CompareStatus.SKIP: "跳过",
}

MISMATCH_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
MATCH_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


class ExcelReport(BaseReportGenerator):

    def generate(self, result: CompareResult) -> str:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"compare_{result.batch_id}_{result.tax_type}_{result.period}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "比对明细"

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_COLUMNS))
        title_cell = ws.cell(row=1, column=1,
                             value=f"税务申报数据比对报告 - {result.company_name} - {result.period}")
        title_cell.font = Font(bold=True, size=14)

        # Summary row
        summary = result.summary
        summary_text = (
            f"总字段数: {summary.total_fields} | "
            f"一致: {summary.match_count} | "
            f"容差内一致: {summary.tolerance_match_count} | "
            f"不一致: {summary.mismatch_count} | "
            f"一致率: {summary.match_rate}%"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REPORT_COLUMNS))
        ws.cell(row=2, column=1, value=summary_text).font = Font(bold=True, size=11)

        # Header
        for col_idx, col_name in enumerate(REPORT_COLUMNS, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, fr in enumerate(result.field_results, 4):
            status_display = STATUS_DISPLAY.get(fr.status, fr.status.value)
            row_data = [
                row_idx - 3,
                result.company_name,
                result.taxpayer_id,
                result.period,
                result.tax_type,
                result.form_code,
                result.form_name,
                1,  # page
                fr.row_name,
                fr.line_no,
                fr.column_name,
                fr.field_id,
                "",  # json_path
                fr.data_type.value,
                str(fr.api_raw_value) if fr.api_raw_value is not None else "",
                str(fr.web_raw_value) if fr.web_raw_value is not None else "",
                str(fr.api_normalized) if fr.api_normalized is not None else "",
                str(fr.web_normalized) if fr.web_normalized is not None else "",
                status_display,
                fr.diff_type or "",
                fr.diff_value or "",
                str(fr.tolerance) if fr.tolerance is not None else "",
                fr.detail or "",
                "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)

            # Color mismatch rows
            if fr.status in (CompareStatus.MISMATCH, CompareStatus.API_MISSING,
                             CompareStatus.WEB_MISSING, CompareStatus.BOTH_MISSING):
                for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = MISMATCH_FILL
            elif fr.status in (CompareStatus.MATCH, CompareStatus.TOLERANCE_MATCH):
                ws.cell(row=row_idx, column=19).fill = MATCH_FILL

        wb.save(str(filepath))
        return str(filepath)