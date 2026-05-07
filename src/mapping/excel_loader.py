import logging
from pathlib import Path
from typing import Optional

import openpyxl

from src.mapping.mapping_cleaner import MappingCleaner
from src.mapping.mapping_validator import MappingValidator
from src.models.field_mapping import FieldMapping

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Loads and cleans field mappings from an Excel file."""

    def __init__(
        self,
        file_path: str,
        sheet: Optional[str] = None,
        header_row: int = 1,
        data_start_row: int = 2,
    ):
        self.file_path = Path(file_path)
        self.sheet = sheet
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.cleaner = MappingCleaner()
        self.validator = MappingValidator()

    def load(self) -> list[FieldMapping]:
        if not self.file_path.exists():
            logger.error(f"Mapping file not found: {self.file_path}")
            return []

        wb = openpyxl.load_workbook(str(self.file_path), data_only=True)

        sheet_name = self.sheet
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]

        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in {self.file_path}")
            return []

        ws = wb[sheet_name]

        # Detect header row if not explicitly set
        header_idx = self.header_row
        headers = []
        for cell in ws[header_idx]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())

        # Read data rows
        raw_rows = []
        for row_idx in range(self.data_start_row, ws.max_row + 1):
            row_data = {}
            all_empty = True
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                value = cell.value
                if value is not None:
                    all_empty = False
                row_data[header] = value
            if not all_empty:
                raw_rows.append(row_data)

        wb.close()

        logger.info(
            f"Loaded {len(raw_rows)} rows from '{sheet_name}' "
            f"(headers: {len(headers)})"
        )

        cleaned = self.cleaner.clean_rows(raw_rows)
        validated = self.validator.validate(cleaned)

        logger.info(f"Validated {len(validated)} field mappings")
        return validated

    def list_sheets(self) -> list[str]:
        if not self.file_path.exists():
            return []
        wb = openpyxl.load_workbook(str(self.file_path), data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets