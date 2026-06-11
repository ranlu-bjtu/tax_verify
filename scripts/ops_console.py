"""Local operator console for batch collection and verification.

The console is intentionally dependency-free. It wraps the existing
scripts/batch_collect_verify.py entry point and exposes a small local web UI for
operators who should not need to use PowerShell directly.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import signal
import subprocess
import sys
import time
import urllib.parse
import urllib.request
import webbrowser
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.coverage.analyzer import write_coverage_status
from src.coverage.registry import (
    build_coverage_targets,
    declaration_statuses_for_collect_statuses,
    normalize_collect_status_keys,
    normalize_tax_type_keys,
    supported_tax_types,
)


RUNTIME_DIR = PROJECT_ROOT / "runtime" / "ops_console"
JOBS_FILE = RUNTIME_DIR / "jobs.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "batch_runs"
DEFAULT_ACCOUNTSET_OUTPUT_DIR = PROJECT_ROOT / "output" / "accountset_runs"
DEFAULT_BACKEND_LOGIN_OUTPUT_DIR = PROJECT_ROOT / "output" / "backend_login_runs"
DEFAULT_PRIVACY_PHONE_OUTPUT_DIR = PROJECT_ROOT / "output" / "privacy_phone_runs"
DEFAULT_PLUGIN_PATH = Path(r"C:\Users\Administrator\Downloads\EtaxPlugin")
DEFAULT_USER_DATA_DIR = PROJECT_ROOT / "browser_profile" / "etax_compare_forms"
DEFAULT_CHROME_PATH = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")
DEFAULT_PORT = 8765
MANUAL_VERIFICATION_REQUIRED_MARKER = "MANUAL_VERIFICATION_REQUIRED"
ACTIVE_PROCESSES: dict[str, subprocess.Popen[bytes]] = {}


INDEX_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>税务取数验证工作台</title>
  <style>
    :root {
      --bg: #f6f8fb;
      --surface: #ffffff;
      --ink: #17212f;
      --muted: #667085;
      --line: #d6dde8;
      --focus: #2563eb;
      --ok: #0f766e;
      --warn: #9a6700;
      --danger: #b42318;
      --idle: #475467;
      --code-bg: #101828;
      --code-text: #e4e7ec;
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: Arial, "Microsoft YaHei", sans-serif; background: var(--bg); color: var(--ink); }
    header { padding: 14px 22px; border-bottom: 1px solid var(--line); background: var(--surface); display: flex; align-items: center; justify-content: space-between; gap: 12px; }
    h1 { margin: 0; font-size: 20px; line-height: 1.3; letter-spacing: 0; }
    h2 { margin: 0; padding: 11px 13px; font-size: 15px; border-bottom: 1px solid var(--line); background: #f9fbfd; }
    main { display: grid; grid-template-columns: minmax(390px, 520px) minmax(560px, 1fr); gap: 16px; padding: 16px 22px 24px; align-items: start; }
    section { background: var(--surface); border: 1px solid var(--line); border-radius: 8px; overflow: hidden; }
    .stack { display: grid; gap: 12px; align-content: start; min-width: 0; }
    .entry-tabs-section { position: sticky; top: 0; z-index: 5; }
    .entry-tabs { display: flex; gap: 6px; padding: 8px; overflow-x: auto; background: #fff; }
    .entry-tab { white-space: nowrap; min-height: 32px; padding: 6px 10px; border-radius: 6px; color: #344054; background: #f8fafc; }
    .entry-tab.active { background: var(--focus); border-color: var(--focus); color: #fff; font-weight: 700; }
    .entry-panel { display: none; }
    .entry-panel.active { display: block; }
    .form-grid { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; padding: 13px; }
    .full { grid-column: 1 / -1; }
    label { display: block; font-size: 12px; font-weight: 700; color: #344054; margin-bottom: 5px; }
    input, select, textarea { width: 100%; border: 1px solid var(--line); border-radius: 6px; background: #fff; color: var(--ink); padding: 8px 9px; font: inherit; min-height: 36px; }
    textarea { min-height: 120px; resize: vertical; }
    input:focus, select:focus, textarea:focus, button:focus { outline: 2px solid rgba(37, 99, 235, 0.22); outline-offset: 1px; border-color: var(--focus); }
    .checks { display: flex; flex-wrap: wrap; gap: 8px 14px; align-items: center; }
    .checks label { display: flex; align-items: center; gap: 6px; margin: 0; font-weight: 500; color: var(--ink); }
    .coverage-checks { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 7px 12px; padding: 8px 9px; border: 1px solid var(--line); border-radius: 6px; background: #fff; }
    .coverage-checks label { display: flex; align-items: center; gap: 6px; margin: 0; font-weight: 500; color: var(--ink); }
    input[type="checkbox"] { width: 16px; min-height: 16px; }
    .actions { padding: 0 13px 13px; display: flex; flex-wrap: wrap; gap: 8px; }
    button { border: 1px solid var(--line); background: #fff; color: var(--ink); min-height: 36px; border-radius: 6px; padding: 8px 11px; font: inherit; cursor: pointer; }
    button.primary { background: var(--focus); border-color: var(--focus); color: #fff; font-weight: 700; }
    button.danger { color: var(--danger); border-color: #f0b8b3; }
    button:disabled { opacity: 0.55; cursor: not-allowed; }
    .status-list { display: grid; gap: 6px; padding: 10px 13px 13px; }
    .status-item { display: grid; grid-template-columns: 90px 68px 1fr; gap: 8px; align-items: start; border-bottom: 1px solid #eef2f6; padding: 6px 0; font-size: 13px; }
    .status-item:last-child { border-bottom: 0; }
    .badge { display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 12px; line-height: 1.3; text-align: center; }
    .ok { background: #e3fcef; color: var(--ok); }
    .warn { background: #fff4d6; color: var(--warn); }
    .fail { background: #ffe3e3; color: var(--danger); }
    .idle { background: #eef2f6; color: var(--idle); }
    .muted { color: var(--muted); }
    .job { padding: 12px 13px; display: grid; gap: 9px; }
    .job-head { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; justify-content: space-between; }
    .mono, code, pre { font-family: Consolas, "Courier New", monospace; }
    pre { margin: 0; background: var(--code-bg); color: var(--code-text); padding: 10px; border-radius: 6px; overflow: auto; max-height: 260px; font-size: 12px; line-height: 1.45; white-space: pre-wrap; }
    table { border-collapse: collapse; width: 100%; font-size: 12px; }
    th, td { border: 1px solid var(--line); padding: 6px 7px; text-align: left; vertical-align: middle; }
    th { background: #f0f4f8; }
    .runs { padding: 12px 13px; overflow-x: auto; }
    .runs td { white-space: nowrap; }
    .review-table { min-width: 1180px; }
    .review-table td { white-space: nowrap; }
    .review-table .wrap { white-space: normal; min-width: 180px; max-width: 300px; }
    .review-table select, .review-table input { min-height: 28px; padding: 4px 6px; font-size: 12px; }
    .mini-actions { padding: 10px 13px 0; display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
    .run-name { max-width: 260px; overflow: hidden; text-overflow: ellipsis; }
    a { color: var(--focus); text-decoration: none; }
    a:hover { text-decoration: underline; }
    @media (min-width: 981px) {
      .results-stack { position: sticky; top: 12px; max-height: calc(100vh - 24px); overflow: auto; padding-right: 2px; }
    }
    @media (max-width: 980px) {
      main { grid-template-columns: 1fr; padding: 12px; }
      header { padding: 12px; }
      .entry-tabs-section { top: 0; }
    }
    @media (max-width: 720px) {
      .form-grid, .coverage-checks { grid-template-columns: 1fr; }
      .full { grid-column: auto; }
      .actions button { flex: 1 1 140px; }
    }
  </style>
</head>
<body>
  <header>
    <h1>税务取数验证工作台</h1>
    <div class="muted" id="clock"></div>
  </header>
  <main>
    <div class="stack">
      <section class="entry-tabs-section">
        <div class="entry-tabs" role="tablist" aria-label="操作入口">
          <button type="button" class="entry-tab active" data-entry-panel="taskPanel" role="tab" aria-selected="true">取数验证</button>
          <button type="button" class="entry-tab" data-entry-panel="accountsetPanel" role="tab" aria-selected="false">创建账套</button>
          <button type="button" class="entry-tab" data-entry-panel="manualAccountsetPanel" role="tab" aria-selected="false">手工创建</button>
          <button type="button" class="entry-tab" data-entry-panel="privacyPhonePanel" role="tab" aria-selected="false">隐私号同步</button>
          <button type="button" class="entry-tab" data-entry-panel="healthPanel" role="tab" aria-selected="false">环境检查</button>
        </div>
      </section>

      <section id="taskPanel" class="entry-panel active" role="tabpanel">
        <h2>任务参数</h2>
        <form id="taskForm">
          <div class="form-grid">
            <div>
              <label for="mode">任务类型</label>
              <select id="mode" name="mode">
                <option value="full">完整链路：发取数并验证</option>
                <option value="collect_only">只发起取数</option>
                <option value="verify_existing">只验证已有任务</option>
              </select>
            </div>
            <div>
              <label for="period">所属期</label>
              <input id="period" name="period" placeholder="202604" required>
            </div>
            <div>
              <label for="enterprise">企业</label>
              <input id="enterprise" name="enterprise" value="蓝天之爱" required>
            </div>
            <div>
              <label for="ydzUsername">易代账账号</label>
              <input id="ydzUsername" name="ydzUsername" autocomplete="username" placeholder="浏览器已登录可不填">
            </div>
            <div>
              <label for="ydzPassword">易代账密码</label>
              <input id="ydzPassword" name="ydzPassword" type="password" autocomplete="current-password" placeholder="仅本次运行使用">
            </div>
            <div>
              <label for="targets">验证范围</label>
              <input id="targets" name="targets" value="auto">
            </div>
            <div class="full">
              <label>需要覆盖税种</label>
              <div class="coverage-checks">
                <label><input type="checkbox" name="coverageTaxTypes" value="VAT_GENERAL" checked> 增值税（一般纳税人）</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="VAT_SMALL" checked> 增值税（小规模纳税人）</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="CIT_A"> 企业所得税（A类）</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="CULTURE_FEE" checked> 文化事业建设费</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="CONSUMPTION_TAX" checked> 消费税</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="CBJ_PERSONAL" checked> 个税残保金</label>
                <label><input type="checkbox" name="coverageTaxTypes" value="CBJ_ANNUAL" checked> 汇算清缴残保金</label>
              </div>
            </div>
            <div class="full">
              <label>后台补齐取数状态</label>
              <div class="coverage-checks">
                <label><input type="checkbox" name="coverageCollectStatuses" value="collected" checked> 已取数</label>
                <label><input type="checkbox" name="coverageCollectStatuses" value="not_collected" checked> 未取数</label>
              </div>
            </div>
            <div>
              <label for="cdpPort">浏览器端口</label>
              <input id="cdpPort" name="cdpPort" type="number" value="9222">
            </div>
            <div class="full">
              <label for="chromePath">Chrome 路径</label>
              <input id="chromePath" name="chromePath" value="C:\Program Files\Google\Chrome\Application\chrome.exe">
            </div>
            <div class="full">
              <label for="pluginPath">税局插件路径</label>
              <input id="pluginPath" name="pluginPath" value="C:\Users\Administrator\Downloads\EtaxPlugin">
            </div>
            <div class="full">
              <label for="userDataDir">浏览器数据目录</label>
              <input id="userDataDir" name="userDataDir" value="browser_profile\etax_compare_forms">
            </div>
            <div>
              <label for="pollTimeout">取数等待秒数</label>
              <input id="pollTimeout" name="pollTimeout" type="number" value="600">
            </div>
            <div>
              <label for="taxTimeout">税局等待秒数</label>
              <input id="taxTimeout" name="taxTimeout" type="number" value="600">
            </div>
            <div>
              <label for="logLevel">日志级别</label>
              <select id="logLevel" name="logLevel">
                <option value="INFO">INFO</option>
                <option value="DEBUG">DEBUG</option>
                <option value="WARNING">WARNING</option>
              </select>
            </div>
            <div class="full">
              <label for="taxNos">税号清单</label>
              <textarea id="taxNos" name="taxNos" placeholder="每行一个税号，也支持逗号分隔"></textarea>
            </div>
            <div class="full checks">
              <label><input type="checkbox" name="force"> 强制重新发起取数</label>
              <label><input type="checkbox" name="rerunVerified"> 已验证也重跑</label>
              <label><input type="checkbox" name="skipBrowser"> 跳过税局浏览器验证</label>
              <label><input type="checkbox" name="skipPdf"> 不保存 PDF</label>
              <label><input type="checkbox" name="skipCoverageSupplement"> 跳过后台补齐</label>
            </div>
          </div>
          <div class="actions">
            <button class="primary" type="submit">开始任务</button>
            <button type="button" id="refreshHealth">刷新检查</button>
            <button type="button" id="openLatest">打开最新报告</button>
            <button type="button" id="resumeRun">继续未完成</button>
            <button type="button" id="regenerateSummary">重新生成汇总页</button>
            <button type="button" id="refreshRuns">刷新批次</button>
          </div>
        </form>
      </section>

      <section id="accountsetPanel" class="entry-panel" role="tabpanel">
        <h2>创建账套</h2>
        <form id="accountsetForm">
          <div class="form-grid">
            <div>
              <label for="accountsetEnv">环境</label>
              <select id="accountsetEnv" name="accountsetEnv">
                <option value="inte">集测</option>
                <option value="prod">线上</option>
              </select>
            </div>
            <div>
              <label for="accountsetOpeningPeriod">启用期间</label>
              <input id="accountsetOpeningPeriod" name="accountsetOpeningPeriod" value="202501">
            </div>
            <div>
              <label for="accountsetTaxpayerType">纳税性质</label>
              <select id="accountsetTaxpayerType" name="accountsetTaxpayerType">
                <option value="SMALL_TAXPAYER">小规模纳税人</option>
                <option value="NORMAL_TAXPAYER">一般纳税人</option>
              </select>
            </div>
            <div>
              <label for="accountsetIndustryId">行业</label>
              <input id="accountsetIndustryId" name="accountsetIndustryId" value="11079">
            </div>
            <div>
              <label for="accountsetCdpPort">浏览器端口</label>
              <input id="accountsetCdpPort" name="accountsetCdpPort" type="number" value="9222">
            </div>
            <div>
              <label for="accountsetSessionTimeout">登录等待秒数</label>
              <input id="accountsetSessionTimeout" name="accountsetSessionTimeout" type="number" value="120">
            </div>
            <div class="full">
              <label for="accountsetLookbackDays">后台查询范围</label>
              <input id="accountsetLookbackDays" name="accountsetLookbackDays" value="30,180,730,1460">
            </div>
            <div class="full">
              <label for="accountsetYdzWorkUrl">易代账工作台地址</label>
              <input id="accountsetYdzWorkUrl" name="accountsetYdzWorkUrl" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetYdzAuthMode">易代账认证方式</label>
              <select id="accountsetYdzAuthMode" name="accountsetYdzAuthMode">
                <option value="">自动（账号密码优先）</option>
                <option value="browser">浏览器登录态</option>
                <option value="password">账号密码直登</option>
                <option value="token">环境变量 token</option>
              </select>
            </div>
            <div class="full">
              <label for="accountsetEnvFile">登录配置文件</label>
              <input id="accountsetEnvFile" name="accountsetEnvFile" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetYdzUsername">易代账账号</label>
              <input id="accountsetYdzUsername" name="accountsetYdzUsername" autocomplete="username" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetYdzPassword">易代账密码</label>
              <input id="accountsetYdzPassword" name="accountsetYdzPassword" type="password" autocomplete="current-password" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetYdzEnterprise">易代账企业</label>
              <input id="accountsetYdzEnterprise" name="accountsetYdzEnterprise" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetBackendUsername">报税后台账号</label>
              <input id="accountsetBackendUsername" name="accountsetBackendUsername" autocomplete="username" placeholder="可不填">
            </div>
            <div>
              <label for="accountsetBackendPassword">报税后台密码</label>
              <input id="accountsetBackendPassword" name="accountsetBackendPassword" type="password" autocomplete="current-password" placeholder="可不填">
            </div>
            <div class="full">
              <label for="accountsetTaxNos">税号清单</label>
              <textarea id="accountsetTaxNos" name="accountsetTaxNos" placeholder="每行一个税号，也支持逗号分隔"></textarea>
            </div>
            <div class="full checks">
              <label><input type="checkbox" name="accountsetDryRun"> 只预检查不保存</label>
              <label><input type="checkbox" name="accountsetNoLaunchChrome"> 不自动打开浏览器</label>
              <label><input type="checkbox" name="accountsetSkipAutoLogin"> 不自动登录</label>
            </div>
          </div>
          <div class="actions">
            <button class="primary" type="submit">开始创建账套</button>
          </div>
        </form>
      </section>

      <section id="manualAccountsetPanel" class="entry-panel" role="tabpanel">
        <h2>手工创建账套</h2>
        <form id="manualAccountsetForm">
          <div class="form-grid">
            <div>
              <label for="manualAccountsetEnv">环境</label>
              <select id="manualAccountsetEnv" name="manualAccountsetEnv">
                <option value="inte">集测</option>
                <option value="prod">线上</option>
              </select>
            </div>
            <div>
              <label for="manualAccountsetOpeningPeriod">启用期间</label>
              <input id="manualAccountsetOpeningPeriod" name="manualAccountsetOpeningPeriod" value="202501">
            </div>
            <div>
              <label for="manualAccountsetTaxpayerType">纳税性质</label>
              <select id="manualAccountsetTaxpayerType" name="manualAccountsetTaxpayerType">
                <option value="SMALL_TAXPAYER">小规模纳税人</option>
                <option value="NORMAL_TAXPAYER">一般纳税人</option>
              </select>
            </div>
            <div>
              <label for="manualAccountsetIndustryId">行业</label>
              <input id="manualAccountsetIndustryId" name="manualAccountsetIndustryId" value="11079">
            </div>
            <div>
              <label for="manualAccountsetCdpPort">浏览器端口</label>
              <input id="manualAccountsetCdpPort" name="manualAccountsetCdpPort" type="number" value="9222">
            </div>
            <div>
              <label for="manualAccountsetSessionTimeout">登录等待秒数</label>
              <input id="manualAccountsetSessionTimeout" name="manualAccountsetSessionTimeout" type="number" value="120">
            </div>
            <div>
              <label for="manualAccountsetTaxNo">税号</label>
              <input id="manualAccountsetTaxNo" name="manualAccountsetTaxNo">
            </div>
            <div>
              <label for="manualAccountsetCustomerName">客户名称</label>
              <input id="manualAccountsetCustomerName" name="manualAccountsetCustomerName">
            </div>
            <div>
              <label for="manualAccountsetAreaName">地区</label>
              <input id="manualAccountsetAreaName" name="manualAccountsetAreaName" placeholder="如 北京、河北、山东">
            </div>
            <div>
              <label for="manualAccountsetLoginMethod">登录方式</label>
              <select id="manualAccountsetLoginMethod" name="manualAccountsetLoginMethod">
                <option value="YSHDL">税局隐私号登录</option>
                <option value="DLYW-YSHDL">税局隐私号-代理登录</option>
                <option value="SDSRDX">税局手工录入验证码登录</option>
                <option value="DLYW-SDSRDX">税局手工录入验证码-代理登录</option>
                <option value="SBZMDL">申报账密登录</option>
              </select>
            </div>
            <div>
              <label for="manualAccountsetProxyTaxNo">代理公司税号</label>
              <input id="manualAccountsetProxyTaxNo" name="manualAccountsetProxyTaxNo" placeholder="代理登录时必填">
            </div>
            <div>
              <label for="manualAccountsetPrivacyNo">隐私号/手机号</label>
              <input id="manualAccountsetPrivacyNo" name="manualAccountsetPrivacyNo">
            </div>
            <div>
              <label for="manualAccountsetPassword">个人用户密码</label>
              <input id="manualAccountsetPassword" name="manualAccountsetPassword" type="password" autocomplete="current-password">
            </div>
            <div>
              <label for="manualAccountsetYdzWorkUrl">易代账工作台地址</label>
              <input id="manualAccountsetYdzWorkUrl" name="manualAccountsetYdzWorkUrl" placeholder="可不填">
            </div>
            <div>
              <label for="manualAccountsetYdzAuthMode">易代账认证方式</label>
              <select id="manualAccountsetYdzAuthMode" name="manualAccountsetYdzAuthMode">
                <option value="">自动（账号密码优先）</option>
                <option value="browser">浏览器登录态</option>
                <option value="password">账号密码直登</option>
                <option value="token">环境变量 token</option>
              </select>
            </div>
            <div>
              <label for="manualAccountsetEnvFile">登录配置文件</label>
              <input id="manualAccountsetEnvFile" name="manualAccountsetEnvFile" placeholder="可不填">
            </div>
            <div>
              <label for="manualAccountsetYdzUsername">易代账账号</label>
              <input id="manualAccountsetYdzUsername" name="manualAccountsetYdzUsername" autocomplete="username" placeholder="可不填">
            </div>
            <div>
              <label for="manualAccountsetYdzPassword">易代账密码</label>
              <input id="manualAccountsetYdzPassword" name="manualAccountsetYdzPassword" type="password" autocomplete="current-password" placeholder="可不填">
            </div>
            <div>
              <label for="manualAccountsetYdzEnterprise">易代账企业</label>
              <input id="manualAccountsetYdzEnterprise" name="manualAccountsetYdzEnterprise" placeholder="可不填">
            </div>
            <div class="full checks">
              <label><input type="checkbox" name="manualAccountsetDryRun"> 只预检查不保存</label>
              <label><input type="checkbox" name="manualAccountsetNoLaunchChrome"> 不自动打开浏览器</label>
              <label><input type="checkbox" name="manualAccountsetSkipAutoLogin"> 不自动登录</label>
            </div>
          </div>
          <div class="actions">
            <button class="primary" type="submit">手工创建账套</button>
          </div>
        </form>
      </section>

      <section id="privacyPhonePanel" class="entry-panel" role="tabpanel">
        <h2>隐私号同步</h2>
        <form id="privacyPhoneForm">
          <div class="form-grid">
            <div>
              <label for="privacyPhoneCdpPort">浏览器端口</label>
              <input id="privacyPhoneCdpPort" name="privacyPhoneCdpPort" type="number" value="9222">
            </div>
            <div>
              <label for="privacyPhoneSessionTimeout">登录等待秒数</label>
              <input id="privacyPhoneSessionTimeout" name="privacyPhoneSessionTimeout" type="number" value="120">
            </div>
            <div>
              <label for="privacyPhoneBackendUsername">后台账号</label>
              <input id="privacyPhoneBackendUsername" name="privacyPhoneBackendUsername" autocomplete="username">
            </div>
            <div>
              <label for="privacyPhoneBackendPassword">后台密码</label>
              <input id="privacyPhoneBackendPassword" name="privacyPhoneBackendPassword" type="password" autocomplete="current-password">
            </div>
            <div class="full">
              <label for="privacyPhoneEnvFile">登录配置文件</label>
              <input id="privacyPhoneEnvFile" name="privacyPhoneEnvFile" placeholder="可不填">
            </div>
            <div class="full">
              <label for="privacyPhones">隐私号清单</label>
              <textarea id="privacyPhones" name="privacyPhones" placeholder="每行一个隐私号，也支持逗号分隔"></textarea>
            </div>
            <div class="full checks">
              <label><input type="checkbox" name="privacyPhoneDryRun"> 只查询不复制</label>
              <label><input type="checkbox" name="privacyPhoneNoLaunchChrome"> 不自动打开浏览器</label>
              <label><input type="checkbox" name="privacyPhoneSkipAutoLogin"> 不自动登录，仅检查现有登录态</label>
            </div>
          </div>
          <div class="actions">
            <button class="primary" type="submit">开始同步隐私号</button>
          </div>
        </form>
      </section>

      <section id="healthPanel" class="entry-panel" role="tabpanel">
        <h2>环境检查</h2>
        <div id="health" class="status-list"></div>
      </section>
    </div>

    <div class="stack results-stack">
      <section>
        <h2>当前任务</h2>
        <div class="job">
          <div class="job-head">
            <div id="jobTitle" class="muted">暂无运行任务</div>
            <div>
              <button type="button" id="stopJob" class="danger">停止任务</button>
            </div>
          </div>
          <div id="jobMeta" class="muted"></div>
          <div id="jobSummary" class="muted"></div>
          <pre id="jobLog">等待任务启动。</pre>
        </div>
      </section>

      <section>
        <h2>税号进度</h2>
        <div class="runs">
          <table>
            <thead><tr><th>税号</th><th>地区</th><th>企业</th><th>阶段</th><th>状态</th><th>taskId</th><th>原因</th><th>操作</th></tr></thead>
            <tbody id="progress"><tr><td colspan="8" class="muted">暂无进度。</td></tr></tbody>
          </table>
        </div>
      </section>

      <section>
        <h2>覆盖检查</h2>
        <div class="mini-actions">
          <button type="button" id="refreshCoverage">刷新覆盖</button>
          <span class="muted" id="coverageMeta"></span>
        </div>
        <div class="runs">
          <table>
            <thead><tr><th>税种</th><th>状态</th><th>覆盖</th><th>代表税号</th><th>taskId</th></tr></thead>
            <tbody id="coverageRows"><tr><td colspan="5" class="muted">暂无覆盖数据。</td></tr></tbody>
          </table>
        </div>
      </section>

      <section>
        <h2>问题处理</h2>
        <div class="mini-actions">
          <button type="button" id="refreshReview">刷新问题</button>
          <button type="button" id="exportReview">导出问题清单</button>
          <span class="muted" id="reviewMeta"></span>
        </div>
        <div class="runs">
          <table class="review-table">
            <thead><tr><th>税号</th><th>企业</th><th>表单</th><th>字段</th><th>差异</th><th>处理状态</th><th>备注</th><th>证据</th></tr></thead>
            <tbody id="reviewIssues"><tr><td colspan="8" class="muted">暂无问题。</td></tr></tbody>
          </table>
        </div>
      </section>

      <section>
        <h2>最近批次</h2>
        <div class="runs">
          <table>
            <thead><tr><th>批次</th><th>状态</th><th>税号</th><th>需处理</th><th>差异</th><th>已处理</th><th>报告</th></tr></thead>
            <tbody id="runs"></tbody>
          </table>
        </div>
      </section>
    </div>
  </main>
  <script>
    const $ = id => document.getElementById(id);
    let CURRENT_JOB = null;

    function previousMonthPeriod() {
      const d = new Date();
      d.setMonth(d.getMonth() - 1);
      return `${d.getFullYear()}${String(d.getMonth() + 1).padStart(2, '0')}`;
    }

    function formPayload() {
      const form = new FormData($('taskForm'));
      return {
        mode: form.get('mode'),
        taxNos: form.get('taxNos') || '',
        period: form.get('period') || '',
        enterprise: form.get('enterprise') || '',
        ydzUsername: form.get('ydzUsername') || '',
        ydzPassword: form.get('ydzPassword') || '',
        targets: form.get('targets') || 'auto',
        coverageTaxTypes: Array.from(document.querySelectorAll('input[name="coverageTaxTypes"]:checked')).map(item => item.value),
        coverageCollectStatuses: Array.from(document.querySelectorAll('input[name="coverageCollectStatuses"]:checked')).map(item => item.value),
        cdpPort: Number(form.get('cdpPort') || 9222),
        chromePath: form.get('chromePath') || '',
        pluginPath: form.get('pluginPath') || '',
        userDataDir: form.get('userDataDir') || '',
        pollTimeout: Number(form.get('pollTimeout') || 600),
        taxTimeout: Number(form.get('taxTimeout') || 600),
        logLevel: form.get('logLevel') || 'INFO',
        force: form.has('force'),
        rerunVerified: form.has('rerunVerified'),
        skipBrowser: form.has('skipBrowser'),
        skipPdf: form.has('skipPdf'),
        skipCoverageSupplement: form.has('skipCoverageSupplement')
      };
    }

    function accountsetPayload() {
      const form = new FormData($('accountsetForm'));
      return {
        accountsetEnv: form.get('accountsetEnv') || 'inte',
        accountsetTaxNos: form.get('accountsetTaxNos') || '',
        accountsetOpeningPeriod: form.get('accountsetOpeningPeriod') || '202501',
        accountsetTaxpayerType: form.get('accountsetTaxpayerType') || 'SMALL_TAXPAYER',
        accountsetIndustryId: form.get('accountsetIndustryId') || '11079',
        accountsetCdpPort: Number(form.get('accountsetCdpPort') || 9222),
        accountsetSessionTimeout: Number(form.get('accountsetSessionTimeout') || 120),
        accountsetLookbackDays: form.get('accountsetLookbackDays') || '30,180,730,1460',
        accountsetYdzWorkUrl: form.get('accountsetYdzWorkUrl') || '',
        accountsetYdzAuthMode: form.get('accountsetYdzAuthMode') || '',
        accountsetEnvFile: form.get('accountsetEnvFile') || '',
        accountsetYdzUsername: form.get('accountsetYdzUsername') || '',
        accountsetYdzPassword: form.get('accountsetYdzPassword') || '',
        accountsetYdzEnterprise: form.get('accountsetYdzEnterprise') || '',
        accountsetBackendUsername: form.get('accountsetBackendUsername') || '',
        accountsetBackendPassword: form.get('accountsetBackendPassword') || '',
        accountsetChromePath: $('chromePath') ? $('chromePath').value : '',
        accountsetDryRun: form.has('accountsetDryRun'),
        accountsetNoLaunchChrome: form.has('accountsetNoLaunchChrome'),
        accountsetSkipAutoLogin: form.has('accountsetSkipAutoLogin')
      };
    }

    function manualAccountsetPayload() {
      const form = new FormData($('manualAccountsetForm'));
      return {
        accountsetManualSource: true,
        accountsetEnv: form.get('manualAccountsetEnv') || 'inte',
        accountsetManualTaxNo: form.get('manualAccountsetTaxNo') || '',
        accountsetManualCustomerName: form.get('manualAccountsetCustomerName') || '',
        accountsetManualAreaName: form.get('manualAccountsetAreaName') || '',
        accountsetManualLoginMethod: form.get('manualAccountsetLoginMethod') || 'YSHDL',
        accountsetManualProxyTaxNo: form.get('manualAccountsetProxyTaxNo') || '',
        accountsetManualPrivacyNo: form.get('manualAccountsetPrivacyNo') || '',
        accountsetManualPassword: form.get('manualAccountsetPassword') || '',
        accountsetOpeningPeriod: form.get('manualAccountsetOpeningPeriod') || '202501',
        accountsetTaxpayerType: form.get('manualAccountsetTaxpayerType') || 'SMALL_TAXPAYER',
        accountsetIndustryId: form.get('manualAccountsetIndustryId') || '11079',
        accountsetCdpPort: Number(form.get('manualAccountsetCdpPort') || 9222),
        accountsetSessionTimeout: Number(form.get('manualAccountsetSessionTimeout') || 120),
        accountsetYdzWorkUrl: form.get('manualAccountsetYdzWorkUrl') || '',
        accountsetYdzAuthMode: form.get('manualAccountsetYdzAuthMode') || '',
        accountsetEnvFile: form.get('manualAccountsetEnvFile') || '',
        accountsetYdzUsername: form.get('manualAccountsetYdzUsername') || '',
        accountsetYdzPassword: form.get('manualAccountsetYdzPassword') || '',
        accountsetYdzEnterprise: form.get('manualAccountsetYdzEnterprise') || '',
        accountsetChromePath: $('chromePath') ? $('chromePath').value : '',
        accountsetDryRun: form.has('manualAccountsetDryRun'),
        accountsetNoLaunchChrome: form.has('manualAccountsetNoLaunchChrome'),
        accountsetSkipAutoLogin: form.has('manualAccountsetSkipAutoLogin')
      };
    }

    function privacyPhonePayload() {
      const form = new FormData($('privacyPhoneForm'));
      return {
        privacyPhones: form.get('privacyPhones') || '',
        privacyPhoneCdpPort: Number(form.get('privacyPhoneCdpPort') || 9222),
        privacyPhoneSessionTimeout: Number(form.get('privacyPhoneSessionTimeout') || 120),
        privacyPhoneEnvFile: form.get('privacyPhoneEnvFile') || '',
        privacyPhoneChromePath: $('chromePath') ? $('chromePath').value : '',
        backendUsername: form.get('privacyPhoneBackendUsername') || '',
        backendPassword: form.get('privacyPhoneBackendPassword') || '',
        privacyPhoneDryRun: form.has('privacyPhoneDryRun'),
        privacyPhoneNoLaunchChrome: form.has('privacyPhoneNoLaunchChrome'),
        privacyPhoneSkipAutoLogin: form.has('privacyPhoneSkipAutoLogin')
      };
    }

    function escapeHtml(value) {
      return String(value ?? '').replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]));
    }

    function hasMetric(value) {
      return value !== undefined && value !== null && value !== '';
    }

    function activateEntryPanel(panelId) {
      const panels = Array.from(document.querySelectorAll('.entry-panel'));
      const target = panels.some(panel => panel.id === panelId) ? panelId : 'taskPanel';
      panels.forEach(panel => {
        const active = panel.id === target;
        panel.classList.toggle('active', active);
        panel.hidden = !active;
      });
      document.querySelectorAll('.entry-tab').forEach(tab => {
        const active = tab.dataset.entryPanel === target;
        tab.classList.toggle('active', active);
        tab.setAttribute('aria-selected', active ? 'true' : 'false');
      });
      try {
        localStorage.setItem('opsConsoleEntryPanel', target);
      } catch (_err) {}
    }

    function initEntryTabs() {
      document.querySelectorAll('.entry-tab').forEach(tab => {
        tab.addEventListener('click', () => activateEntryPanel(tab.dataset.entryPanel));
      });
      let saved = 'taskPanel';
      try {
        saved = localStorage.getItem('opsConsoleEntryPanel') || saved;
      } catch (_err) {}
      activateEntryPanel(saved);
    }

    async function api(path, options) {
      const res = await fetch(path, options);
      const data = await res.json();
      if (!res.ok || data.ok === false) {
        throw new Error(data.error || data.message || '请求失败');
      }
      return data;
    }

    function badge(status) {
      const cls = status === 'ok' ? 'ok' : status === 'fail' ? 'fail' : status === 'warn' ? 'warn' : 'idle';
      const label = status === 'ok' ? '正常' : status === 'fail' ? '异常' : status === 'warn' ? '注意' : '未知';
      return `<span class="badge ${cls}">${label}</span>`;
    }

    async function refreshHealth() {
      const payload = formPayload();
      const params = new URLSearchParams({
        cdpPort: payload.cdpPort,
        chromePath: payload.chromePath,
        pluginPath: payload.pluginPath,
        userDataDir: payload.userDataDir
      });
      const data = await api('/api/health?' + params.toString());
      $('health').innerHTML = data.checks.map(item => `
        <div class="status-item">
          <div>${item.name}</div>
          <div>${badge(item.status)}</div>
          <div>${item.message || ''}</div>
        </div>
      `).join('');
    }

    async function startTask(event) {
      event.preventDefault();
      const data = await api('/api/start', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(formPayload())
      });
      renderJob(data.job);
      refreshRuns();
    }

    async function startAccountset(event) {
      event.preventDefault();
      const data = await api('/api/create-accountset', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(accountsetPayload())
      });
      renderJob(data.job);
      refreshRuns();
    }

    async function startManualAccountset(event) {
      event.preventDefault();
      const data = await api('/api/create-accountset', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(manualAccountsetPayload())
      });
      renderJob(data.job);
      refreshRuns();
    }

    async function startPrivacyPhoneSync(event) {
      event.preventDefault();
      const data = await api('/api/sync-privacy-phone', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(privacyPhonePayload())
      });
      renderJob(data.job);
      refreshRuns();
    }

    function renderJob(job) {
      CURRENT_JOB = job || null;
      if (!job) {
        $('jobTitle').textContent = '暂无运行任务';
        $('jobMeta').textContent = '';
        $('jobSummary').textContent = '';
        $('jobLog').textContent = '等待任务启动。';
        return;
      }
      const isAccountset = job.jobType === 'accountset';
      const isBackendLogin = job.jobType === 'backend-login';
      const isPrivacyPhone = job.jobType === 'privacy-phone-sync';
      const status = job.status || 'unknown';
      const badgeClass = status === 'running' ? 'warn' : status === 'success' ? 'ok' : status === 'failed' || status === 'missing' ? 'fail' : 'idle';
      const statusLabel = job.statusLabel || status || '未知';
      $('jobTitle').innerHTML = `<span class="badge ${badgeClass}">${escapeHtml(statusLabel)}</span> <span class="mono">${escapeHtml(job.runId || '')}</span>`;
      const report = job.summaryPath
        ? (isAccountset || isBackendLogin || isPrivacyPhone
          ? ` · <a href="#" onclick="openPath(${JSON.stringify(job.summaryPath)}); return false;">打开结果</a>`
          : ` · <a href="#" onclick="openReport('${job.runId}'); return false;">打开报告</a>`)
        : '';
      const metaParts = [
        `pid=${escapeHtml(job.pid || '-')}`,
        job.startedAt ? `开始 ${escapeHtml(job.startedAt)}` : '',
        job.finishedAt ? `结束 ${escapeHtml(job.finishedAt)}` : '',
        job.durationText ? `耗时 ${escapeHtml(job.durationText)}` : '',
        escapeHtml(job.runDir || '')
      ].filter(Boolean);
      $('jobMeta').innerHTML = `${metaParts.join(' · ')}${report}`;
      const summaryParts = [
        job.resultLabel ? `结果：${job.resultLabel}` : '',
        hasMetric(job.totalTaxNos) ? `税号 ${job.totalTaxNos}` : '',
        hasMetric(job.manualRequired) ? `需处理 ${job.manualRequired}` : '',
        hasMetric(job.problemCount) ? `差异 ${job.problemCount}` : '',
        job.operatorAction ? escapeHtml(job.operatorAction) : '',
        job.exitCode !== undefined && job.exitCode !== null && Number(job.exitCode) !== 0 ? `退出码 ${job.exitCode}` : ''
      ].filter(Boolean);
      $('jobSummary').textContent = summaryParts.join(' · ');
      $('jobLog').textContent = job.logTail || job.displayCommand || '暂无日志。';
      if (job.runId) {
        if (isAccountset || isBackendLogin || isPrivacyPhone) {
          if (isAccountset) {
            renderAccountsetProgress(job.accountsetResults || []);
          } else if (isPrivacyPhone) {
            renderPrivacyPhoneProgress(job.privacyPhoneResults || []);
          } else {
            renderBackendLoginProgress(job);
          }
          $('reviewIssues').innerHTML = `<tr><td colspan="8" class="muted">${isPrivacyPhone ? '隐私号同步任务不产生字段差异。' : (isBackendLogin ? '后台登录任务不产生字段差异。' : '创建账套任务不产生字段差异。')}</td></tr>`;
          $('reviewMeta').textContent = '';
          $('coverageRows').innerHTML = `<tr><td colspan="5" class="muted">${isPrivacyPhone ? '隐私号同步任务不产生覆盖数据。' : (isBackendLogin ? '后台登录任务不产生覆盖数据。' : '创建账套任务不产生覆盖数据。')}</td></tr>`;
          $('coverageMeta').textContent = '';
        } else {
          refreshProgress(job.runId).catch(() => {});
          refreshReview(job.runId).catch(() => {});
          refreshCoverage(job.runId).catch(() => {});
        }
      }
    }

    async function refreshJob() {
      const data = await api('/api/jobs');
      renderJob(data.current || null);
    }

    async function stopJob() {
      await api('/api/stop', {method: 'POST'});
      await refreshJob();
      await refreshRuns();
    }

    async function refreshProgress(runId) {
      if (!runId) {
        $('progress').innerHTML = '<tr><td colspan="8" class="muted">暂无进度。</td></tr>';
        return;
      }
      const data = await api('/api/job-status?runId=' + encodeURIComponent(runId));
      const items = (data.status && data.status.items) || [];
      $('progress').innerHTML = items.map(item => `
        <tr>
          <td class="mono">${item.taxNo || ''}</td>
          <td>${item.region || ''}</td>
          <td class="run-name" title="${item.custName || ''}">${item.custName || ''}</td>
          <td>${item.stageName || item.stage || ''}</td>
          <td>${statusBadge(item.status, item.statusLabel)}</td>
          <td class="mono">${item.taskId || '-'}</td>
          <td title="${item.action || ''}">${item.reason || '-'}</td>
          <td>
            <button type="button" onclick="verifyOne('${data.status.runId}', '${item.taxNo}')">继续验证</button>
            <button type="button" onclick="retryCollect('${data.status.runId}', '${item.taxNo}')">重试取数</button>
            <button type="button" onclick="skipOne('${data.status.runId}', '${item.taxNo}')">跳过</button>
          </td>
        </tr>
      `).join('') || '<tr><td colspan="8" class="muted">暂无进度。</td></tr>';
    }

    function renderAccountsetProgress(results) {
      if (!results.length) {
        $('progress').innerHTML = '<tr><td colspan="8" class="muted">等待账套创建结果。</td></tr>';
        return;
      }
      $('progress').innerHTML = results.map(item => {
        const ok = item.status === 'OK' || item.status === 'DRY_RUN';
        const status = ok ? 'success' : (item.status === 'PARTIAL' ? 'manual' : 'failed');
        const errors = (item.errors || []).join('; ');
        const privacy = item.privacyPhoneStatus ? ` · 隐私号 ${item.privacyPhoneStatus}` : '';
        return `
          <tr>
            <td class="mono">${escapeHtml(item.taxNo || '')}</td>
            <td>${escapeHtml(item.areaName || item.areaCode || '')}</td>
            <td class="run-name" title="${escapeHtml(item.name || '')}">${escapeHtml(item.name || '')}</td>
            <td>${escapeHtml(item.action || '')}</td>
            <td>${statusBadge(status, item.status || '')}</td>
            <td class="mono">${escapeHtml(item.custId || '-')}</td>
            <td title="${escapeHtml((item.privacyPhoneMessage || '') + (errors ? ' · ' + errors : ''))}">${escapeHtml(item.loginMethod || '')}${escapeHtml(privacy)}${errors ? ' · ' + escapeHtml(errors) : ''}</td>
            <td>-</td>
          </tr>
        `;
      }).join('');
    }

    function renderBackendLoginProgress(job) {
      const status = job.backendReady ? 'success' : (job.status === 'running' ? 'running' : 'failed');
      const label = job.backendReady ? '已登录' : (job.status === 'running' ? '登录中' : '未登录');
      $('progress').innerHTML = `
        <tr>
          <td class="mono">报税后台</td>
          <td>-</td>
          <td>public-manage</td>
          <td>登录</td>
          <td>${statusBadge(status, label)}</td>
          <td class="mono">-</td>
          <td>${escapeHtml(job.resultLabel || '')}</td>
          <td>-</td>
        </tr>
      `;
    }

    function renderPrivacyPhoneProgress(results) {
      if (!results.length) {
        $('progress').innerHTML = '<tr><td colspan="8" class="muted">等待隐私号同步结果。</td></tr>';
        return;
      }
      $('progress').innerHTML = results.map(item => {
        const okStatuses = new Set(['OK', 'DRY_RUN', 'EXISTS', 'PULLED', 'DRY_RUN_EXISTS', 'DRY_RUN_MISSING', 'SKIPPED']);
        const ok = okStatuses.has(item.status || '');
        const status = ok ? 'success' : 'failed';
        const errors = (item.errors || []).join('; ');
        const orgNames = (item.inteSummaryRows || item.summaryRows || []).map(row => row.orgName || row.orgId || '').filter(Boolean).join(', ');
        const mode = item.inteSummaryCount !== undefined ? '集测可查询' : '复制同步';
        const count = item.inteSummaryCount !== undefined ? item.inteSummaryCount : (item.detailCount || 0);
        const message = item.pullMessage || item.copyMessage || errors || '-';
        return `
          <tr>
            <td class="mono">${escapeHtml(item.privatePhone || '')}</td>
            <td>-</td>
            <td class="run-name" title="${escapeHtml(orgNames)}">${escapeHtml(orgNames || '-')}</td>
            <td>${escapeHtml(mode)}</td>
            <td>${statusBadge(status, item.status || '')}</td>
            <td class="mono">${escapeHtml(String(count || 0))}</td>
            <td title="${escapeHtml(errors)}">${escapeHtml(message)}</td>
            <td>-</td>
          </tr>
        `;
      }).join('');
    }

    async function refreshReview(runId) {
      if (!runId) {
        $('reviewIssues').innerHTML = '<tr><td colspan="8" class="muted">暂无问题。</td></tr>';
        $('reviewMeta').textContent = '';
        return;
      }
      const data = await api('/api/review?runId=' + encodeURIComponent(runId));
      const items = data.items || [];
      $('reviewMeta').textContent = `${items.length} 条问题`;
      $('reviewIssues').innerHTML = items.slice(0, 200).map(item => `
        <tr>
          <td class="mono">${escapeHtml(item.taxNo)}</td>
          <td class="run-name" title="${escapeHtml(item.custName)}">${escapeHtml(item.custName)}</td>
          <td class="wrap" title="${escapeHtml(item.formName)}">${escapeHtml(item.formShortName || item.formName)}</td>
          <td class="mono" title="${escapeHtml(item.position)}">${escapeHtml(item.fieldId)}</td>
          <td class="wrap">${escapeHtml(item.reason)}<br><span class="muted">${escapeHtml(item.apiRawValue)} / ${escapeHtml(item.webRawValue)}</span></td>
          <td>
            <select onchange="updateReview('${data.runId}', '${item.key}', 'reviewStatus', this.value)">
              ${reviewStatusOptions(item.reviewStatus)}
            </select>
          </td>
          <td><input value="${escapeHtml(item.note || '')}" onchange="updateReview('${data.runId}', '${item.key}', 'note', this.value)" placeholder="处理备注"></td>
          <td>${reviewEvidenceLinks(item)}</td>
        </tr>
      `).join('') || '<tr><td colspan="8" class="muted">暂无问题。</td></tr>';
    }

    async function refreshCoverage(runId) {
      if (!runId) {
        $('coverageRows').innerHTML = '<tr><td colspan="5" class="muted">暂无覆盖数据。</td></tr>';
        $('coverageMeta').textContent = '';
        return;
      }
      const data = await api('/api/coverage?runId=' + encodeURIComponent(runId));
      const coverage = data.coverage || {};
      const summary = coverage.summary || {};
      $('coverageMeta').textContent = `${summary.coveredTargets || 0}/${summary.totalTargets || 0} 已覆盖`;
      $('coverageRows').innerHTML = (coverage.targets || []).map(target => {
        const example = ((target.examples || [])[0]) || {};
        return `
          <tr>
            <td>${escapeHtml(target.taxTypeName || target.taxType)}</td>
            <td>${escapeHtml(target.declarationStatusName || target.declarationStatus)}</td>
            <td>${statusBadge(target.covered ? 'success' : 'manual', target.covered ? '已覆盖' : '缺口')}</td>
            <td class="mono">${escapeHtml(example.taxNo || '-')}</td>
            <td class="mono">${escapeHtml(example.taskId || '-')}</td>
          </tr>
        `;
      }).join('') || '<tr><td colspan="5" class="muted">暂无覆盖数据。</td></tr>';
    }

    function reviewStatusOptions(current) {
      const values = ['待处理', '处理中', '已确认接口问题', '已确认网页解析问题', '已确认税局数据问题', '已忽略', '已完成'];
      return values.map(value => `<option value="${escapeHtml(value)}" ${value === current ? 'selected' : ''}>${escapeHtml(value)}</option>`).join('');
    }

    function reviewEvidenceLinks(item) {
      const links = [];
      if (item.summaryPath) links.push(`<a href="#" onclick="openPath(${JSON.stringify(item.summaryPath)}); return false;">报告</a>`);
      if (item.pdfPath) links.push(`<a href="#" onclick="openPath(${JSON.stringify(item.pdfPath)}); return false;">PDF</a>`);
      if (item.excelPath) links.push(`<a href="#" onclick="openPath(${JSON.stringify(item.excelPath)}); return false;">Excel</a>`);
      return links.join(' · ') || '-';
    }

    async function updateReview(runId, key, field, value) {
      await api('/api/review-update', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({runId, key, fields: {[field]: value}})
      });
    }

    async function exportReview() {
      const job = await api('/api/jobs');
      const runId = (job.current && job.current.runId) || '';
      if (!runId) throw new Error('没有可导出的批次');
      const data = await api('/api/export-review', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({runId})
      });
      if (data.fileUrl) window.open(data.fileUrl, '_blank');
    }

    async function openPath(path) {
      await api('/api/open-path', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({path})
      });
    }

    function statusBadge(status, label) {
      const cls = status === 'success' ? 'ok' : status === 'warning' || status === 'running' ? 'warn' : status === 'manual' || status === 'failed' ? 'fail' : 'idle';
      return `<span class="badge ${cls}">${label || status || '未知'}</span>`;
    }

    async function runAction(path, payload) {
      const merged = Object.assign({}, formPayload(), payload || {});
      const data = await api(path, {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(merged)
      });
      if (data.job) renderJob(data.job);
      if (data.status) refreshProgress(data.status.runId);
      if (data.status) refreshReview(data.status.runId);
      if (data.status) refreshCoverage(data.status.runId);
      refreshRuns();
      return data;
    }

    async function resumeRun() {
      const job = await api('/api/jobs');
      const runId = (job.current && job.current.runId) || '';
      if (!runId) throw new Error('没有可续跑的批次');
      await runAction('/api/resume-run', {runId});
    }

    async function regenerateSummary() {
      const job = await api('/api/jobs');
      const runId = (job.current && job.current.runId) || '';
      if (!runId) throw new Error('没有可重新生成的批次');
      const data = await runAction('/api/regenerate-summary', {runId});
      if (data.fileUrl) window.open(data.fileUrl, '_blank');
    }

    async function verifyOne(runId, taxNo) {
      await runAction('/api/verify-one', {runId, taxNo});
    }

    async function retryCollect(runId, taxNo) {
      await runAction('/api/retry-one', {runId, taxNo});
    }

    async function skipOne(runId, taxNo) {
      await runAction('/api/skip-one', {runId, taxNo});
    }

    async function openLatestReport() {
      const data = await api('/api/open-latest-report', {method: 'POST'});
      if (data.fileUrl) window.open(data.fileUrl, '_blank');
    }

    async function openReport(runId) {
      const data = await api('/api/open-report', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({runId})
      });
      if (data.fileUrl) window.open(data.fileUrl, '_blank');
    }

    async function refreshRuns() {
      const data = await api('/api/runs');
      $('runs').innerHTML = data.runs.map(run => `
        <tr>
          <td class="run-name mono" title="${run.runDir || ''}">${run.runId}</td>
          <td>${run.statusLabel}</td>
          <td>${run.totalTaxNos}</td>
          <td>${run.manualRequired}</td>
          <td>${run.problemCount}</td>
          <td>${run.reviewedCount || 0}</td>
          <td>${run.summaryPath ? `<a href="#" onclick="openReport('${run.runId}'); return false;">打开</a>` : '-'}</td>
        </tr>
      `).join('') || '<tr><td colspan="6" class="muted">暂无批次</td></tr>';
    }

    function tickClock() {
      $('clock').textContent = new Date().toLocaleString();
    }

    $('period').value = previousMonthPeriod();
    initEntryTabs();
    $('taskForm').addEventListener('submit', startTask);
    $('accountsetForm').addEventListener('submit', startAccountset);
    $('manualAccountsetForm').addEventListener('submit', startManualAccountset);
    $('privacyPhoneForm').addEventListener('submit', startPrivacyPhoneSync);
    $('refreshHealth').addEventListener('click', refreshHealth);
    $('refreshRuns').addEventListener('click', refreshRuns);
    $('openLatest').addEventListener('click', openLatestReport);
    $('refreshReview').addEventListener('click', async () => {
      const job = await api('/api/jobs');
      const runId = (job.current && job.current.runId) || '';
      refreshReview(runId).catch(err => alert(err.message));
    });
    $('refreshCoverage').addEventListener('click', async () => {
      const job = await api('/api/jobs');
      const runId = (job.current && job.current.runId) || '';
      refreshCoverage(runId).catch(err => alert(err.message));
    });
    $('exportReview').addEventListener('click', () => exportReview().catch(err => alert(err.message)));
    $('resumeRun').addEventListener('click', () => resumeRun().catch(err => alert(err.message)));
    $('regenerateSummary').addEventListener('click', () => regenerateSummary().catch(err => alert(err.message)));
    $('stopJob').addEventListener('click', stopJob);
    setInterval(tickClock, 1000);
    setInterval(refreshJob, 3000);
    setInterval(refreshRuns, 10000);
    tickClock();
    refreshHealth().catch(err => $('health').textContent = err.message);
    refreshJob().catch(() => {});
    refreshRuns().catch(() => {});
  </script>
</body>
</html>
"""


def previous_month_period(today: date | None = None) -> str:
    today = today or date.today()
    year = today.year
    month = today.month - 1
    if month == 0:
        year -= 1
        month = 12
    return f"{year}{month:02d}"


def parse_tax_nos_text(text: str) -> list[str]:
    values = [part.strip() for part in re.split(r"[\s,，;；]+", text or "") if part.strip()]
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def parse_private_phones_text(text: str) -> list[str]:
    values = [part.strip() for part in re.split(r"[\s,，;；]+", text or "") if part.strip()]
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def safe_run_id(value: str | None = None) -> str:
    raw = value or f"ops_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    clean = re.sub(r"[^0-9A-Za-z_.-]+", "_", raw).strip("._-")
    return clean or f"ops_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


def positive_int(value: Any, default: int, minimum: int = 1, maximum: int = 999999) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return min(max(number, minimum), maximum)


def browser_path_args(payload: dict[str, Any]) -> list[str]:
    return [
        "--chrome-path",
        str(payload.get("chromePath") or DEFAULT_CHROME_PATH),
        "--user-data-dir",
        str(payload.get("userDataDir") or DEFAULT_USER_DATA_DIR),
        "--plugin-path",
        str(payload.get("pluginPath") or DEFAULT_PLUGIN_PATH),
    ]


def all_coverage_tax_type_keys() -> list[str]:
    return [definition.tax_type for definition in supported_tax_types()]


def default_workbench_coverage_tax_type_keys() -> list[str]:
    return [key for key in all_coverage_tax_type_keys() if key != "CIT_A"]


def all_coverage_collect_status_keys() -> list[str]:
    return ["collected", "not_collected"]


def selected_coverage_tax_types(payload: dict[str, Any], state: dict[str, Any] | None = None) -> list[str]:
    if "coverageTaxTypes" in payload:
        raw = payload.get("coverageTaxTypes")
        values = raw if isinstance(raw, list) else [str(raw or "")]
        selected = normalize_tax_type_keys(values)
        if not selected:
            raise ValueError("请至少选择一个需要覆盖的税种。")
        return selected
    if state and state.get("coverageTaxTypes"):
        selected = normalize_tax_type_keys(state.get("coverageTaxTypes") or [])
        if selected:
            return selected
    return default_workbench_coverage_tax_type_keys()


def selected_coverage_collect_statuses(payload: dict[str, Any], state: dict[str, Any] | None = None) -> list[str]:
    if "coverageCollectStatuses" in payload:
        raw = payload.get("coverageCollectStatuses")
        values = raw if isinstance(raw, list) else [str(raw or "")]
        selected = normalize_collect_status_keys(values)
        if not selected:
            raise ValueError("请至少选择一个后台补齐取数状态。")
        return selected
    if state and state.get("coverageCollectStatuses"):
        selected = normalize_collect_status_keys(state.get("coverageCollectStatuses") or [])
        if selected:
            return selected
    return all_coverage_collect_status_keys()


def coverage_command_args(coverage_tax_types: list[str], coverage_collect_statuses: list[str]) -> list[str]:
    return [
        "--coverage-tax-types",
        ",".join(coverage_tax_types),
        "--coverage-collect-statuses",
        ",".join(coverage_collect_statuses),
    ]


def coverage_supplement_scan_args(payload: dict[str, Any]) -> list[str]:
    has_work_urls = bool(str(payload.get("ydzSupplementWorkUrls") or "").strip())
    scan_enterprises = bool(payload.get("scanYdzEnterprises"))
    if not scan_enterprises and not has_work_urls:
        return []
    args = ["--coverage-supplement-refresh-cit-from-ydz"]
    if scan_enterprises:
        args.append("--coverage-supplement-scan-ydz-enterprises")
    return args


def build_batch_command(
    payload: dict[str, Any],
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, Any]:
    mode = str(payload.get("mode") or "full")
    if mode not in {"full", "collect_only", "verify_existing"}:
        raise ValueError("任务类型无效。")

    tax_nos = parse_tax_nos_text(str(payload.get("taxNos") or ""))
    if not tax_nos:
        raise ValueError("请至少填写一个税号。")

    period = str(payload.get("period") or previous_month_period()).strip()
    if not re.fullmatch(r"\d{6}", period):
        raise ValueError("所属期必须是 YYYYMM，例如 202604。")

    enterprise = str(payload.get("enterprise") or "蓝天之爱").strip()
    if not enterprise:
        raise ValueError("企业不能为空。")

    run_id = safe_run_id(str(payload.get("runId") or ""))
    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    tax_no_file = run_dir / "tax_nos.txt"
    tax_no_file.write_text("\n".join(tax_nos) + "\n", encoding="utf-8")
    coverage_tax_types = selected_coverage_tax_types(payload)
    coverage_collect_statuses = selected_coverage_collect_statuses(payload)

    command = [
        sys.executable,
        str(PROJECT_ROOT / "scripts" / "batch_collect_verify.py"),
        "--tax-no-file",
        str(tax_no_file),
        "--period",
        period,
        "--enterprise",
        enterprise,
        "--run-id",
        run_id,
        "--output-dir",
        str(output_dir),
        "--targets",
        str(payload.get("targets") or "auto").strip() or "auto",
        "--cdp-port",
        str(positive_int(payload.get("cdpPort"), 9222, 1, 65535)),
        *browser_path_args(payload),
        "--poll-timeout",
        str(positive_int(payload.get("pollTimeout"), 600, 30)),
        "--tax-timeout",
        str(positive_int(payload.get("taxTimeout"), 600, 30)),
        "--log-level",
        str(payload.get("logLevel") or "INFO").upper(),
        *coverage_command_args(coverage_tax_types, coverage_collect_statuses),
        *coverage_supplement_scan_args(payload),
    ]
    if mode in {"full", "verify_existing"}:
        command.append("--verify")
    if mode == "verify_existing":
        command.append("--skip-collect")
    if bool(payload.get("force")):
        command.append("--force")
    if bool(payload.get("rerunVerified")):
        command.append("--rerun-verified")
    if bool(payload.get("skipBrowser")):
        command.append("--skip-browser")
    if bool(payload.get("skipPdf")):
        command.append("--skip-pdf")
    if bool(payload.get("skipCoverageSupplement")):
        command.append("--skip-coverage-supplement")

    return {
        "runId": run_id,
        "taxNos": tax_nos,
        "runDir": run_dir,
        "taxNoFile": tax_no_file,
        "coverageTaxTypes": coverage_tax_types,
        "coverageCollectStatuses": coverage_collect_statuses,
        "command": command,
        "displayCommand": powershell_command(command),
    }


def normalize_accountset_env(value: Any) -> str:
    env = str(value or "inte").strip().lower()
    if env not in {"inte", "prod"}:
        raise ValueError("创建账套环境无效。")
    return env


def normalize_accountset_lookback_days(value: Any) -> str:
    parts = [part for part in re.split(r"[\s,，;；]+", str(value or "")) if part]
    days = [str(positive_int(part, 30, 1, 3650)) for part in parts]
    return ",".join(days or ["30", "180", "730", "1460"])


def normalize_accountset_login_method(value: Any) -> str:
    text = str(value or "").strip()
    upper = text.upper()
    supported = {"YSHDL", "DLYW-YSHDL", "SDSRDX", "DLYW-SDSRDX", "SBZMDL"}
    if upper in supported:
        return upper
    compact = "".join(text.split())
    has_proxy = "代理" in compact or upper.startswith("DLYW-")
    has_privacy = "隐私" in compact
    has_manual_captcha = (
        "SDSRDX" in upper
        or "手工录入验证码" in compact
        or "手动录入验证码" in compact
        or ("验证码" in compact and ("手工" in compact or "手动" in compact))
    )
    if has_manual_captcha:
        return "DLYW-SDSRDX" if has_proxy else "SDSRDX"
    if has_proxy and has_privacy:
        return "DLYW-YSHDL"
    if has_privacy:
        return "YSHDL"
    if "申报账密" in compact or "账号密码" in compact:
        return "SBZMDL"
    raise ValueError("手工登录方式无效。")


def build_create_accountset_command(
    payload: dict[str, Any],
    output_dir: Path = DEFAULT_ACCOUNTSET_OUTPUT_DIR,
) -> dict[str, Any]:
    env = normalize_accountset_env(payload.get("accountsetEnv"))
    manual_source = bool(payload.get("accountsetManualSource"))
    tax_nos_text = (
        str(payload.get("accountsetManualTaxNo") or "")
        if manual_source
        else str(payload.get("accountsetTaxNos") or payload.get("taxNos") or "")
    )
    tax_nos = parse_tax_nos_text(tax_nos_text)
    if not tax_nos:
        raise ValueError("请至少填写一个需要创建账套的税号。")
    if manual_source and len(tax_nos) != 1:
        raise ValueError("手工创建账套一次只能填写一个税号。")

    opening_period = str(payload.get("accountsetOpeningPeriod") or "202501").strip()
    if not re.fullmatch(r"\d{6}", opening_period):
        raise ValueError("启用期间必须是 YYYYMM，例如 202501。")

    taxpayer_type = str(payload.get("accountsetTaxpayerType") or "SMALL_TAXPAYER").strip().upper()
    if taxpayer_type not in {"SMALL_TAXPAYER", "NORMAL_TAXPAYER"}:
        raise ValueError("纳税性质无效。")

    industry_id = str(payload.get("accountsetIndustryId") or "11079").strip()
    if not industry_id:
        raise ValueError("行业不能为空。")

    if manual_source:
        customer_name = str(payload.get("accountsetManualCustomerName") or "").strip()
        login_method = normalize_accountset_login_method(payload.get("accountsetManualLoginMethod"))
        proxy_tax_no = str(payload.get("accountsetManualProxyTaxNo") or "").strip()
        privacy_no = str(payload.get("accountsetManualPrivacyNo") or "").strip()
        manual_password = str(payload.get("accountsetManualPassword") or "")
        if not customer_name:
            raise ValueError("手工创建账套必须填写客户名称。")
        if login_method in {"YSHDL", "DLYW-YSHDL", "SDSRDX", "DLYW-SDSRDX"} and not privacy_no:
            raise ValueError("手工创建账套必须填写隐私号/手机号。")
        if login_method.startswith("DLYW-") and not proxy_tax_no:
            raise ValueError("代理登录必须填写代理公司税号。")
        if login_method != "SBZMDL" and not manual_password:
            raise ValueError("手工创建账套必须填写个人用户密码。")

    prefix = "accountset_manual" if manual_source else "accountset"
    run_id = safe_run_id(str(payload.get("runId") or f"{prefix}_{env}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"))
    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    tax_no_file = run_dir / "tax_nos.txt"
    tax_no_file.write_text("\n".join(tax_nos) + "\n", encoding="utf-8")
    output_json = run_dir / "accountset_summary.json"
    script_path = PROJECT_ROOT / "scripts" / "ydz_create_customers.py"
    if not script_path.exists():
        raise FileNotFoundError(f"未找到账套创建脚本：{script_path}")

    command = [
        sys.executable,
        str(script_path),
        "--env",
        env,
        "--tax-no-file",
        str(tax_no_file),
        "--cdp-port",
        str(positive_int(payload.get("accountsetCdpPort"), 9222, 1, 65535)),
        "--chrome-path",
        str(payload.get("accountsetChromePath") or DEFAULT_CHROME_PATH),
        "--opening-period",
        opening_period,
        "--taxpayer-type",
        taxpayer_type,
        "--industry-id",
        industry_id,
        "--session-timeout",
        str(positive_int(payload.get("accountsetSessionTimeout"), 120, 10, 1800)),
        "--lookback-days",
        normalize_accountset_lookback_days(payload.get("accountsetLookbackDays")),
        "--output-json",
        str(output_json),
        "--log-level",
        str(payload.get("accountsetLogLevel") or payload.get("logLevel") or "INFO").upper(),
    ]
    if manual_source:
        command.extend(["--manual-source-env", "--skip-privacy-phone-sync"])
    ydz_work_url = str(payload.get("accountsetYdzWorkUrl") or "").strip()
    if ydz_work_url:
        command.extend(["--ydz-work-url", ydz_work_url])
    ydz_auth_mode = str(payload.get("accountsetYdzAuthMode") or "").strip().lower()
    if ydz_auth_mode:
        if ydz_auth_mode not in {"browser", "token", "password"}:
            raise ValueError("易代账认证方式无效。")
        command.extend(["--ydz-auth-mode", ydz_auth_mode])
    env_file = str(payload.get("accountsetEnvFile") or "").strip()
    if env_file:
        command.extend(["--env-file", env_file])
    if bool(payload.get("accountsetDryRun")):
        command.append("--dry-run")
    if bool(payload.get("accountsetNoLaunchChrome")):
        command.append("--no-launch-chrome")
    if bool(payload.get("accountsetSkipAutoLogin")):
        command.append("--skip-auto-login")

    return {
        "runId": run_id,
        "taxNos": tax_nos,
        "runDir": run_dir,
        "taxNoFile": tax_no_file,
        "outputJson": output_json,
        "command": command,
        "displayCommand": powershell_command(command),
    }


def cit_a_source_candidate_rows(state: dict[str, Any]) -> list[dict[str, str]]:
    supplement = state.get("coverageSupplement") if isinstance(state, dict) else {}
    if not isinstance(supplement, dict):
        supplement = {}
    target_keys = {
        str(key)
        for key in (supplement.get("missingKeys") or supplement.get("requestedTargetKeys") or [])
        if str(key).startswith("CIT_A:")
    }
    for row in supplement.get("sourceReadiness") or []:
        key = str((row or {}).get("targetKey") or "")
        if key.startswith("CIT_A:"):
            target_keys.add(key)
    matched_task_ids_by_key: dict[str, set[str]] = {key: set() for key in target_keys}
    for row in (supplement.get("sourceReadiness") or []) + (supplement.get("diagnostics") or []):
        if not isinstance(row, dict):
            continue
        key = str(row.get("targetKey") or "")
        if not key.startswith("CIT_A:"):
            continue
        target_keys.add(key)
        matched_task_ids_by_key.setdefault(key, set()).update(
            str(task_id)
            for task_id in (row.get("backendMatchedTaskIds") or row.get("matchedTaskIds") or [])
            if str(task_id or "")
        )

    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    def add_row(
        *,
        target_key: str,
        tax_no: str,
        period: str = "",
        task_id: str = "",
        reason: str = "",
        source: str = "",
    ) -> None:
        if not target_key.startswith("CIT_A:") or not tax_no:
            return
        key = (target_key, tax_no.upper(), period)
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "targetKey": target_key,
                "taxNo": tax_no.upper(),
                "period": period,
                "taskId": task_id,
                "reason": reason,
                "source": source,
            }
        )

    for item in (state.get("items") or {}).values():
        if not isinstance(item, dict):
            continue
        collect = item.get("collect") if isinstance(item.get("collect"), dict) else {}
        verify = item.get("verify") if isinstance(item.get("verify"), dict) else {}
        resolved = collect.get("resolvedTask") if isinstance(collect.get("resolvedTask"), dict) else {}
        target_key = str(resolved.get("coverageTarget") or "")
        if not target_key.startswith("CIT_A:"):
            continue
        tax_no = str(collect.get("taxNo") or item.get("taxNo") or "").strip()
        period = str(collect.get("period") or state.get("period") or "").strip()
        task_ids = {
            str(value)
            for value in (
                collect.get("verifyTaskId"),
                resolved.get("taskId"),
                *((collect.get("verifyTaskIds") or []) if isinstance(collect.get("verifyTaskIds"), list) else []),
            )
            if str(value or "")
        }
        task_ids.update(str(task_id) for task_id in (item.get("verifyTasks") or {}).keys() if str(task_id or ""))
        matched_task_ids = matched_task_ids_by_key.get(target_key) or set()
        if matched_task_ids and not task_ids.intersection(matched_task_ids):
            continue
        add_row(
            target_key=target_key,
            tax_no=tax_no,
            period=period,
            task_id=next((task_id for task_id in task_ids if not matched_task_ids or task_id in matched_task_ids), ""),
            reason=str(verify.get("reason") or ""),
            source="backend_supplement_item",
        )

    for record in supplement.get("freshYdzRefresh") or []:
        if not isinstance(record, dict):
            continue
        target_key = str(record.get("targetKey") or "")
        reason = str(record.get("reason") or "")
        tax_no = str(record.get("taxNo") or "").strip()
        if reason == "no_ydz_account_in_current_enterprise":
            add_row(
                target_key=target_key,
                tax_no=tax_no,
                period=str(record.get("period") or state.get("period") or "").strip(),
                task_id=str(record.get("sourceTaskId") or ""),
                reason=reason,
                source=str(record.get("source") or "fresh_ydz_refresh"),
            )

    order = {"CIT_A:filed": 0, "CIT_A:unfiled": 1}
    return sorted(rows, key=lambda row: (order.get(row["targetKey"], 9), row["taxNo"], row["period"]))


def build_cit_a_accountset_precheck_command(
    payload: dict[str, Any],
    batch_output_dir: Path = DEFAULT_OUTPUT_DIR,
    output_dir: Path = DEFAULT_ACCOUNTSET_OUTPUT_DIR,
) -> dict[str, Any]:
    source_run_id = str(payload.get("sourceRunId") or payload.get("runId") or "").strip()
    if not source_run_id:
        raise ValueError("缺少需要预检查的批次ID。")
    _run_dir, state = load_run_state(source_run_id, output_dir=batch_output_dir)
    candidates = cit_a_source_candidate_rows(state)
    tax_nos = parse_tax_nos_text("\n".join(row["taxNo"] for row in candidates))
    if not tax_nos:
        raise ValueError("当前批次没有可用于企业所得税 A 类账套预检查的候选税号。")

    precheck_payload = dict(payload)
    precheck_payload["accountsetTaxNos"] = "\n".join(tax_nos)
    precheck_payload["accountsetDryRun"] = True
    precheck_payload["runId"] = safe_run_id(
        str(payload.get("precheckRunId") or f"cit_a_accountset_precheck_{source_run_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    )
    spec = build_create_accountset_command(precheck_payload, output_dir=output_dir)
    spec["sourceRunId"] = source_run_id
    spec["sourceCandidates"] = candidates
    return spec


def build_backend_login_command(
    payload: dict[str, Any],
    output_dir: Path = DEFAULT_BACKEND_LOGIN_OUTPUT_DIR,
) -> dict[str, Any]:
    env = normalize_accountset_env(payload.get("backendLoginEnv") or payload.get("accountsetEnv") or "inte")
    run_id = safe_run_id(str(payload.get("runId") or f"backend_login_{env}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"))
    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    output_json = run_dir / "backend_login_status.json"
    script_path = PROJECT_ROOT / "scripts" / "ydz_create_customers.py"
    if not script_path.exists():
        raise FileNotFoundError(f"未找到后台登录脚本：{script_path}")

    command = [
        sys.executable,
        str(script_path),
        "--env",
        env,
        "--login-only",
        "--login-target",
        "backend",
        "--cdp-port",
        str(positive_int(payload.get("backendCdpPort") or payload.get("accountsetCdpPort"), 9222, 1, 65535)),
        "--chrome-path",
        str(payload.get("backendChromePath") or payload.get("accountsetChromePath") or DEFAULT_CHROME_PATH),
        "--session-timeout",
        str(positive_int(payload.get("backendSessionTimeout"), 120, 10, 1800)),
        "--output-json",
        str(output_json),
        "--log-level",
        str(payload.get("backendLogLevel") or payload.get("logLevel") or "INFO").upper(),
    ]
    env_file = str(payload.get("backendEnvFile") or payload.get("accountsetEnvFile") or "").strip()
    if env_file:
        command.extend(["--env-file", env_file])
    if bool(payload.get("backendNoLaunchChrome")):
        command.append("--no-launch-chrome")
    if bool(payload.get("backendSkipAutoLogin")):
        command.append("--skip-auto-login")

    return {
        "runId": run_id,
        "runDir": run_dir,
        "outputJson": output_json,
        "command": command,
        "displayCommand": powershell_command(command),
    }


def build_privacy_phone_sync_command(
    payload: dict[str, Any],
    output_dir: Path = DEFAULT_PRIVACY_PHONE_OUTPUT_DIR,
) -> dict[str, Any]:
    private_phones = parse_private_phones_text(str(payload.get("privacyPhones") or payload.get("privatePhones") or ""))
    if not private_phones:
        raise ValueError("请至少填写一个需要同步的隐私号。")

    run_id = safe_run_id(str(payload.get("runId") or f"privacy_phone_{datetime.now().strftime('%Y%m%d_%H%M%S')}"))
    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    private_phone_file = run_dir / "private_phones.txt"
    private_phone_file.write_text("\n".join(private_phones) + "\n", encoding="utf-8")
    output_json = run_dir / "privacy_phone_summary.json"
    script_path = PROJECT_ROOT / "scripts" / "sync_privacy_phone.py"
    if not script_path.exists():
        raise FileNotFoundError(f"未找到隐私号同步脚本：{script_path}")

    command = [
        sys.executable,
        str(script_path),
        "--private-phone-file",
        str(private_phone_file),
        "--cdp-port",
        str(positive_int(payload.get("privacyPhoneCdpPort"), 9222, 1, 65535)),
        "--chrome-path",
        str(payload.get("privacyPhoneChromePath") or DEFAULT_CHROME_PATH),
        "--session-timeout",
        str(positive_int(payload.get("privacyPhoneSessionTimeout"), 120, 10, 1800)),
        "--output-json",
        str(output_json),
        "--log-level",
        str(payload.get("privacyPhoneLogLevel") or payload.get("logLevel") or "INFO").upper(),
    ]
    env_file = str(payload.get("privacyPhoneEnvFile") or "").strip()
    if env_file:
        command.extend(["--env-file", env_file])
    if bool(payload.get("privacyPhoneDryRun")):
        command.append("--dry-run")
    if bool(payload.get("privacyPhoneNoLaunchChrome")):
        command.append("--no-launch-chrome")
    if bool(payload.get("privacyPhoneSkipAutoLogin")):
        command.append("--skip-auto-login")

    return {
        "runId": run_id,
        "privatePhones": private_phones,
        "runDir": run_dir,
        "privatePhoneFile": private_phone_file,
        "outputJson": output_json,
        "command": command,
        "displayCommand": powershell_command(command),
    }


def load_run_state(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> tuple[Path, dict[str, Any]]:
    safe = safe_run_id(run_id)
    run_dir = output_dir / safe
    state_path = run_dir / "state.json"
    if not state_path.exists():
        raise FileNotFoundError(f"未找到批次状态文件：{state_path}")
    state = json.loads(state_path.read_text(encoding="utf-8"))
    if not isinstance(state, dict):
        raise ValueError("批次状态文件格式无效。")
    return run_dir, state


def build_existing_run_command(
    payload: dict[str, Any],
    tax_nos: list[str],
    skip_collect: bool,
    force: bool = False,
    verify: bool = True,
    rerun_verified: bool = True,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, Any]:
    run_dir, state = load_run_state(str(payload.get("runId") or ""), output_dir=output_dir)
    if not tax_nos:
        raise ValueError("没有可处理的税号。")
    tax_no_file = run_dir / f"ops_selected_{int(time.time())}.txt"
    tax_no_file.write_text("\n".join(tax_nos) + "\n", encoding="utf-8")
    coverage_tax_types = selected_coverage_tax_types(payload, state=state)
    coverage_collect_statuses = selected_coverage_collect_statuses(payload, state=state)
    command = [
        sys.executable,
        str(PROJECT_ROOT / "scripts" / "batch_collect_verify.py"),
        "--tax-no-file",
        str(tax_no_file),
        "--period",
        str(state.get("period") or payload.get("period") or previous_month_period()),
        "--enterprise",
        str(state.get("enterprise") or payload.get("enterprise") or "蓝天之爱"),
        "--run-id",
        str(state.get("runId") or run_dir.name),
        "--output-dir",
        str(output_dir),
        "--targets",
        str(payload.get("targets") or "auto").strip() or "auto",
        "--cdp-port",
        str(positive_int(payload.get("cdpPort"), 9222, 1, 65535)),
        *browser_path_args(payload),
        "--poll-timeout",
        str(positive_int(payload.get("pollTimeout"), 600, 30)),
        "--tax-timeout",
        str(positive_int(payload.get("taxTimeout"), 600, 30)),
        "--log-level",
        str(payload.get("logLevel") or "INFO").upper(),
        *coverage_command_args(coverage_tax_types, coverage_collect_statuses),
        *coverage_supplement_scan_args(payload),
    ]
    if verify:
        command.append("--verify")
    if skip_collect:
        command.append("--skip-collect")
    if force:
        command.append("--force")
    if rerun_verified:
        command.append("--rerun-verified")
    if bool(payload.get("skipBrowser")):
        command.append("--skip-browser")
    if bool(payload.get("skipPdf")):
        command.append("--skip-pdf")
    if bool(payload.get("skipCoverageSupplement")):
        command.append("--skip-coverage-supplement")
    return {
        "runId": str(state.get("runId") or run_dir.name),
        "taxNos": tax_nos,
        "runDir": run_dir,
        "coverageTaxTypes": coverage_tax_types,
        "coverageCollectStatuses": coverage_collect_statuses,
        "command": command,
        "displayCommand": powershell_command(command),
    }


def unfinished_tax_nos(state: dict[str, Any], include_manual: bool = True) -> list[str]:
    values: list[str] = []
    for tax_no, item in (state.get("items") or {}).items():
        collect = item.get("collect") or {}
        verify = item.get("verify") or {}
        verify_status = str(verify.get("status") or "")
        collect_status = str(collect.get("status") or "").upper()
        done = verify_status in {"success", "completed_with_differences"} or collect_status == "NO_NEED_COLLECTED"
        if done:
            continue
        if bool(collect.get("manualRequired")) and not include_manual:
            continue
        values.append(str(tax_no))
    return values


def powershell_command(command: list[str]) -> str:
    return " ".join(quote_powershell_arg(part) for part in command)


def quote_powershell_arg(value: Any) -> str:
    text = str(value)
    if re.fullmatch(r"[0-9A-Za-z_./:\\-]+", text):
        return text
    return "'" + text.replace("'", "''") + "'"


def load_jobs() -> list[dict[str, Any]]:
    if not JOBS_FILE.exists():
        return []
    try:
        data = json.loads(JOBS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []
    if isinstance(data, list):
        return data
    return []


def save_jobs(jobs: list[dict[str, Any]]) -> None:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    JOBS_FILE.write_text(json.dumps(jobs[-50:], ensure_ascii=False, indent=2), encoding="utf-8")


def pid_is_running(pid: Any) -> bool:
    try:
        pid_int = int(pid)
    except Exception:
        return False
    if pid_int <= 0:
        return False
    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            kernel32 = ctypes.windll.kernel32
            process_query_limited_information = 0x1000
            still_active = 259
            handle = kernel32.OpenProcess(process_query_limited_information, False, pid_int)
            if handle:
                try:
                    exit_code = wintypes.DWORD()
                    if kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code)):
                        return int(exit_code.value) == still_active
                    return True
                finally:
                    kernel32.CloseHandle(handle)
        except Exception:
            pass
        try:
            result = subprocess.run(
                ["tasklist", "/FI", f"PID eq {pid_int}", "/NH"],
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
                encoding="utf-8",
                errors="ignore",
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                timeout=5,
            )
            return str(pid_int) in result.stdout
        except Exception:
            return False
    try:
        os.kill(pid_int, 0)
        return True
    except Exception:
        return False


def current_job() -> dict[str, Any] | None:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    save_jobs(jobs)
    running = [job for job in jobs if job.get("status") == "running"]
    if running:
        return running[-1]
    return jobs[-1] if jobs else None


def job_status_label(status: str) -> str:
    labels = {
        "running": "运行中",
        "success": "已完成",
        "failed": "已失败",
        "finished": "已结束",
        "missing": "批次文件缺失",
    }
    return labels.get(status, status or "未知")


def parse_job_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    minutes, second = divmod(seconds, 60)
    hours, minute = divmod(minutes, 60)
    if hours:
        return f"{hours}小时{minute}分{second}秒"
    if minute:
        return f"{minute}分{second}秒"
    return f"{second}秒"


def summarize_accountset_result_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "summaryPath": str(path),
            "accountsetResults": [],
            "resultLabel": "等待结果",
            "totalTaxNos": "",
            "manualRequired": "",
            "problemCount": "",
        }
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = []
    rows = data if isinstance(data, list) else []
    total = len(rows)
    manual = sum(1 for row in rows if str((row or {}).get("status") or "") not in {"OK", "DRY_RUN"})
    dry_run = total > 0 and all(str((row or {}).get("status") or "") == "DRY_RUN" for row in rows)
    if manual:
        result_label = "需处理"
    elif dry_run:
        result_label = "预检查完成"
    elif total:
        result_label = "已完成"
    else:
        result_label = "无结果"
    return {
        "summaryPath": str(path),
        "accountsetResults": rows,
        "resultLabel": result_label,
        "totalTaxNos": total,
        "manualRequired": manual,
        "problemCount": manual,
    }


def accountset_log_requires_manual_verification(log_tail: str) -> bool:
    text = str(log_tail or "")
    return MANUAL_VERIFICATION_REQUIRED_MARKER in text or "requires manual slider verification" in text


def summarize_backend_login_result_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "summaryPath": str(path),
            "resultLabel": "等待登录结果",
            "backendReady": False,
            "totalTaxNos": "",
            "manualRequired": "",
            "problemCount": "",
        }
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    row = data if isinstance(data, dict) else {}
    backend_ready = bool(row.get("backendReady") or (row.get("target") == "backend" and row.get("ready")))
    result_label = "后台已登录" if backend_ready else "后台未登录"
    return {
        "summaryPath": str(path),
        "resultLabel": result_label,
        "backendReady": backend_ready,
        "backendLoginStatus": row,
        "totalTaxNos": "",
        "manualRequired": 0 if backend_ready else 1,
        "problemCount": 0 if backend_ready else 1,
    }


def summarize_privacy_phone_result_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "summaryPath": str(path),
            "privacyPhoneResults": [],
            "resultLabel": "等待同步结果",
            "totalTaxNos": "",
            "manualRequired": "",
            "problemCount": "",
        }
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = []
    rows = data if isinstance(data, list) else []
    total = len(rows)
    success_statuses = {"OK", "DRY_RUN", "EXISTS", "PULLED", "DRY_RUN_EXISTS", "DRY_RUN_MISSING", "SKIPPED"}
    failed = sum(1 for row in rows if str((row or {}).get("status") or "") not in success_statuses)
    dry_run = total > 0 and all(str((row or {}).get("status") or "").startswith("DRY_RUN") for row in rows)
    if failed:
        result_label = "需要处理"
    elif dry_run:
        result_label = "预检查完成"
    elif total:
        result_label = "已完成"
    else:
        result_label = "无结果"
    return {
        "summaryPath": str(path),
        "privacyPhoneResults": rows,
        "resultLabel": result_label,
        "totalTaxNos": total,
        "manualRequired": failed,
        "problemCount": failed,
    }


def refresh_job_status(job: dict[str, Any]) -> dict[str, Any]:
    run_id = str(job.get("runId") or "")
    proc = ACTIVE_PROCESSES.get(run_id)
    if proc is not None:
        code = proc.poll()
        if code is None:
            job["status"] = "running"
        else:
            job["exitCode"] = code
            job["status"] = "success" if code == 0 else "failed"
            job["finishedAt"] = job.get("finishedAt") or datetime.now().isoformat(timespec="seconds")
            ACTIVE_PROCESSES.pop(run_id, None)
    elif job.get("status") == "running" and not pid_is_running(job.get("pid")):
        job["status"] = "finished"
        job["finishedAt"] = job.get("finishedAt") or datetime.now().isoformat(timespec="seconds")

    run_dir = Path(str(job.get("runDir") or ""))
    if str(job.get("jobType") or "") == "accountset":
        summary_path = Path(str(job.get("summaryPath") or run_dir / "accountset_summary.json"))
        job.update(summarize_accountset_result_file(summary_path))
    elif str(job.get("jobType") or "") == "backend-login":
        summary_path = Path(str(job.get("summaryPath") or run_dir / "backend_login_status.json"))
        job.update(summarize_backend_login_result_file(summary_path))
    elif str(job.get("jobType") or "") == "privacy-phone-sync":
        summary_path = Path(str(job.get("summaryPath") or run_dir / "privacy_phone_summary.json"))
        job.update(summarize_privacy_phone_result_file(summary_path))
    else:
        summary_path = run_dir / "batch_summary.html"
        if summary_path.exists():
            job["summaryPath"] = str(summary_path)
            summary = summarize_run(run_dir)
            job["resultLabel"] = summary.get("statusLabel") or ""
            job["totalTaxNos"] = summary.get("totalTaxNos", "")
            job["manualRequired"] = summary.get("manualRequired", "")
            job["problemCount"] = summary.get("problemCount", "")
            job["reviewedCount"] = summary.get("reviewedCount", "")
    log_path = Path(str(job.get("logPath") or ""))
    job["logTail"] = tail_text(log_path, 16000)
    status = str(job.get("status") or "")
    job["statusLabel"] = job_status_label(status)
    if str(job.get("jobType") or "") == "accountset" and accountset_log_requires_manual_verification(job.get("logTail", "")):
        if status == "running":
            job["statusLabel"] = "需人工验证"
            job["resultLabel"] = "等待易代账滑块"
            job["manualRequired"] = 1
            job["problemCount"] = 0
            job["operatorAction"] = "请在已打开的 Chrome 中完成易代账滑块验证，完成后脚本会自动继续。"
        elif status in {"failed", "finished"} and not job.get("accountsetResults"):
            job["resultLabel"] = "滑块验证未完成"
            job["manualRequired"] = 1
            job["problemCount"] = 1
            job["operatorAction"] = "重新运行前请先完成易代账登录，或使用已登录的浏览器会话重跑。"
    started_at = parse_job_datetime(job.get("startedAt"))
    finished_at = parse_job_datetime(job.get("finishedAt")) if job.get("finishedAt") else None
    if started_at:
        end_at = finished_at or datetime.now()
        job["durationSeconds"] = max(0, int((end_at - started_at).total_seconds()))
        job["durationText"] = format_duration(float(job["durationSeconds"]))
    return job


def start_batch_job(payload: dict[str, Any]) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    spec = build_batch_command(payload)
    job = start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=len(spec["taxNos"]),
        payload=payload,
        jobs=jobs,
    )
    return job


def start_create_accountset_job(payload: dict[str, Any]) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    spec = build_create_accountset_command(payload)
    return start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=len(spec["taxNos"]),
        payload=payload,
        jobs=jobs,
        job_type="accountset",
        summary_path=spec["outputJson"],
    )


def start_cit_a_accountset_precheck_job(payload: dict[str, Any]) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    spec = build_cit_a_accountset_precheck_command(payload)
    return start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=len(spec["taxNos"]),
        payload={**payload, "accountsetDryRun": True},
        jobs=jobs,
        job_type="accountset",
        summary_path=spec["outputJson"],
    )


def start_backend_login_job(payload: dict[str, Any]) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    spec = build_backend_login_command(payload)
    return start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=0,
        payload=payload,
        jobs=jobs,
        job_type="backend-login",
        summary_path=spec["outputJson"],
    )


def start_privacy_phone_sync_job(payload: dict[str, Any]) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    spec = build_privacy_phone_sync_command(payload)
    return start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=len(spec["privatePhones"]),
        payload=payload,
        jobs=jobs,
        job_type="privacy-phone-sync",
        summary_path=spec["outputJson"],
    )


def start_existing_run_job(
    payload: dict[str, Any],
    tax_nos: list[str],
    skip_collect: bool,
    force: bool = False,
    verify: bool = True,
    rerun_verified: bool = True,
) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")
    spec = build_existing_run_command(
        payload,
        tax_nos=tax_nos,
        skip_collect=skip_collect,
        force=force,
        verify=verify,
        rerun_verified=rerun_verified,
    )
    return start_process_job(
        run_id=spec["runId"],
        run_dir=spec["runDir"],
        command=spec["command"],
        display_command=spec["displayCommand"],
        tax_no_count=len(spec["taxNos"]),
        payload=payload,
        jobs=jobs,
    )


def resume_run_job(payload: dict[str, Any]) -> dict[str, Any]:
    _run_dir, state = load_run_state(str(payload.get("runId") or ""))
    tax_nos = unfinished_tax_nos(state, include_manual=True)
    return start_existing_run_job(payload, tax_nos, skip_collect=False, force=False, verify=True, rerun_verified=True)


def verify_one_job(payload: dict[str, Any]) -> dict[str, Any]:
    tax_no = str(payload.get("taxNo") or "").strip()
    if not tax_no:
        raise ValueError("缺少税号。")
    return start_existing_run_job(payload, [tax_no], skip_collect=True, force=False, verify=True, rerun_verified=True)


def retry_one_collect_job(payload: dict[str, Any]) -> dict[str, Any]:
    tax_no = str(payload.get("taxNo") or "").strip()
    if not tax_no:
        raise ValueError("缺少税号。")
    return start_existing_run_job(payload, [tax_no], skip_collect=False, force=True, verify=True, rerun_verified=True)


def start_process_job(
    run_id: str,
    run_dir: Path,
    command: list[str],
    display_command: str,
    tax_no_count: int,
    payload: dict[str, Any],
    jobs: list[dict[str, Any]] | None = None,
    job_type: str = "batch",
    summary_path: Path | None = None,
) -> dict[str, Any]:
    jobs = [refresh_job_status(job) for job in (jobs if jobs is not None else load_jobs())]
    active = [job for job in jobs if job.get("status") == "running"]
    if active:
        raise RuntimeError(f"已有任务正在运行：{active[-1].get('runId')}")

    logs_dir = run_dir / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / "ops_console.log"
    log_file = log_path.open("ab")
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    process = subprocess.Popen(
        command,
        cwd=PROJECT_ROOT,
        stdout=log_file,
        stderr=subprocess.STDOUT,
        creationflags=creationflags,
        env=build_subprocess_env(payload),
    )
    log_file.close()
    ACTIVE_PROCESSES[run_id] = process

    job = {
        "runId": run_id,
        "pid": process.pid,
        "status": "running",
        "startedAt": datetime.now().isoformat(timespec="seconds"),
        "runDir": str(run_dir),
        "logPath": str(log_path),
        "summaryPath": str(summary_path or run_dir / "batch_summary.html"),
        "displayCommand": display_command,
        "taxNoCount": tax_no_count,
        "jobType": job_type,
    }
    jobs.append(job)
    save_jobs(jobs)
    return refresh_job_status(job)


def build_subprocess_env(payload: dict[str, Any]) -> dict[str, str]:
    env = os.environ.copy()
    env.setdefault("NODE_NO_WARNINGS", "1")
    username = str(payload.get("ydzUsername") or "").strip()
    password = str(payload.get("ydzPassword") or "")
    if username:
        env["YDZ_USERNAME"] = username
    if password:
        env["YDZ_PASSWORD"] = password
    supplement_work_urls = str(payload.get("ydzSupplementWorkUrls") or "").strip()
    if supplement_work_urls:
        env["YDZ_SUPPLEMENT_WORK_URLS"] = supplement_work_urls
    backend_username = str(payload.get("backendUsername") or "").strip()
    backend_password = str(payload.get("backendPassword") or "")
    backend_url = str(payload.get("backendUrl") or "").strip()
    if backend_username:
        env["TAX_BACKEND_USERNAME"] = backend_username
    if backend_password:
        env["TAX_BACKEND_PASSWORD"] = backend_password
    if backend_url:
        env["TAX_BACKEND_URL"] = backend_url
    if "accountsetEnv" in payload:
        accountset_env = normalize_accountset_env(payload.get("accountsetEnv"))
        prefix = "YDZ_INTE" if accountset_env == "inte" else "YDZ_PROD"
        accountset_ydz_username = str(payload.get("accountsetYdzUsername") or "").strip()
        accountset_ydz_password = str(payload.get("accountsetYdzPassword") or "")
        accountset_enterprise = str(payload.get("accountsetYdzEnterprise") or "").strip()
        accountset_work_url = str(payload.get("accountsetYdzWorkUrl") or "").strip()
        backend_username = str(payload.get("accountsetBackendUsername") or "").strip()
        backend_password = str(payload.get("accountsetBackendPassword") or "")
        backend_url = str(payload.get("accountsetBackendUrl") or "").strip()
        if accountset_ydz_username:
            env[f"{prefix}_USERNAME"] = accountset_ydz_username
        if accountset_ydz_password:
            env[f"{prefix}_PASSWORD"] = accountset_ydz_password
        if accountset_enterprise:
            env[f"{prefix}_ENTERPRISE"] = accountset_enterprise
        if accountset_work_url:
            env[f"{prefix}_WORK_URL"] = accountset_work_url
        if backend_username:
            env["TAX_BACKEND_USERNAME"] = backend_username
        if backend_password:
            env["TAX_BACKEND_PASSWORD"] = backend_password
        if backend_url:
            env["TAX_BACKEND_URL"] = backend_url
        if bool(payload.get("accountsetManualSource")):
            manual_fields = {
                "YDZ_MANUAL_TAX_NO": str(payload.get("accountsetManualTaxNo") or "").strip().upper(),
                "YDZ_MANUAL_CUSTOMER_NAME": str(payload.get("accountsetManualCustomerName") or "").strip(),
                "YDZ_MANUAL_AREA_NAME": str(payload.get("accountsetManualAreaName") or "").strip(),
                "YDZ_MANUAL_AREA_CODE": str(payload.get("accountsetManualAreaCode") or "").strip(),
                "YDZ_MANUAL_LOGIN_METHOD": normalize_accountset_login_method(payload.get("accountsetManualLoginMethod")),
                "YDZ_MANUAL_PROXY_TAX_NO": str(payload.get("accountsetManualProxyTaxNo") or "").strip(),
                "YDZ_MANUAL_PRIVACY_NO": str(payload.get("accountsetManualPrivacyNo") or "").strip(),
                "YDZ_MANUAL_PASSWORD": str(payload.get("accountsetManualPassword") or ""),
            }
            for key, value in manual_fields.items():
                if value:
                    env[key] = value
    return env


def stop_running_job() -> dict[str, Any] | None:
    jobs = [refresh_job_status(job) for job in load_jobs()]
    running = [job for job in jobs if job.get("status") == "running"]
    if not running:
        save_jobs(jobs)
        return None
    job = running[-1]
    run_id = str(job.get("runId") or "")
    proc = ACTIVE_PROCESSES.get(run_id)
    try:
        if proc is not None and proc.poll() is None:
            terminate_process_tree(proc.pid)
        elif job.get("pid"):
            terminate_process_tree(int(job["pid"]))
    except Exception as exc:
        job["stopError"] = str(exc)
    job["status"] = "stopping"
    job["finishedAt"] = datetime.now().isoformat(timespec="seconds")
    save_jobs(jobs)
    return refresh_job_status(job)


def terminate_process_tree(pid: int) -> None:
    if os.name == "nt":
        subprocess.run(["taskkill", "/T", "/F", "/PID", str(pid)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return
    try:
        os.kill(pid, signal.SIGTERM)
    except ProcessLookupError:
        return


def tail_text(path: Path, max_chars: int = 12000) -> str:
    if not path.exists():
        return ""
    try:
        data = path.read_bytes()
    except Exception:
        return ""
    return data[-max_chars:].decode("utf-8", errors="replace")


def environment_checks(params: dict[str, str]) -> list[dict[str, str]]:
    cdp_port = positive_int(params.get("cdpPort"), 9222, 1, 65535)
    chrome_path = Path(params.get("chromePath") or DEFAULT_CHROME_PATH)
    plugin_path = Path(params.get("pluginPath") or DEFAULT_PLUGIN_PATH)
    user_data_dir = Path(params.get("userDataDir") or DEFAULT_USER_DATA_DIR)
    checks = [
        check_item("项目文件", (PROJECT_ROOT / "main.py").exists() and (PROJECT_ROOT / "scripts" / "batch_collect_verify.py").exists(), "项目入口文件完整", "未找到 main.py 或批量脚本"),
        check_item("账套创建脚本", (PROJECT_ROOT / "scripts" / "ydz_create_customers.py").exists(), "账套创建入口文件完整", "未找到 scripts/ydz_create_customers.py"),
        check_item("Python", Path(sys.executable).exists(), sys.executable, "Python 路径不可用"),
        check_item("Chrome", chrome_path.exists(), str(chrome_path), "未找到 Chrome，请检查页面填写的路径"),
        check_item("税局插件", plugin_path.exists(), str(plugin_path), "未找到 EtaxPlugin，请检查路径"),
        check_item("浏览器目录", user_data_dir.exists(), str(user_data_dir), "目录不存在，启动浏览器时会自动创建", warn_when_false=True),
        check_cdp(cdp_port),
        check_credentials(),
        check_proxy(),
        check_output_dir(DEFAULT_OUTPUT_DIR, "批量输出目录"),
        check_output_dir(DEFAULT_ACCOUNTSET_OUTPUT_DIR, "账套输出目录"),
        check_output_dir(DEFAULT_BACKEND_LOGIN_OUTPUT_DIR, "后台登录输出目录"),
        check_output_dir(DEFAULT_PRIVACY_PHONE_OUTPUT_DIR, "隐私号同步输出目录"),
        check_active_job(),
    ]
    return checks


def check_item(name: str, ok: bool, ok_message: str, fail_message: str, warn_when_false: bool = False) -> dict[str, str]:
    return {
        "name": name,
        "status": "ok" if ok else ("warn" if warn_when_false else "fail"),
        "message": ok_message if ok else fail_message,
    }


def check_cdp(port: int) -> dict[str, str]:
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=2) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        browser = data.get("Browser") or "Chrome CDP 已连接"
        return {"name": "浏览器连接", "status": "ok", "message": browser}
    except Exception:
        return {"name": "浏览器连接", "status": "warn", "message": f"未连接 127.0.0.1:{port}，运行完整链路前需启动带 CDP 的 Chrome"}


def check_credentials() -> dict[str, str]:
    has_username = bool(os.environ.get("YDZ_USERNAME"))
    has_password = bool(os.environ.get("YDZ_PASSWORD"))
    if has_username and has_password:
        return {"name": "易代账凭据", "status": "ok", "message": "当前进程已读取到 YDZ_USERNAME/YDZ_PASSWORD"}
    return {
        "name": "易代账凭据",
        "status": "warn",
        "message": "当前进程未读取到环境变量；可在任务参数中临时填写账号密码，或先手动登录易代账",
    }


def check_proxy() -> dict[str, str]:
    proxies = urllib.request.getproxies()
    if not proxies:
        return {"name": "代理", "status": "ok", "message": "未检测到系统代理环境变量"}
    names = ", ".join(sorted(proxies))
    return {"name": "代理", "status": "warn", "message": f"检测到代理配置：{names}；税局异常时优先关闭代理后重试"}


def check_output_dir(path: Path, name: str = "输出目录") -> dict[str, str]:
    try:
        path.mkdir(parents=True, exist_ok=True)
        probe = path / ".ops_console_write_test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return {"name": name, "status": "ok", "message": str(path)}
    except Exception as exc:
        return {"name": name, "status": "fail", "message": f"不可写：{exc}"}


def check_active_job() -> dict[str, str]:
    job = current_job()
    if job and job.get("status") == "running":
        return {"name": "运行任务", "status": "warn", "message": f"{job.get('runId')} 正在运行"}
    return {"name": "运行任务", "status": "ok", "message": "当前无运行中的任务"}


def list_recent_runs(limit: int = 10, output_dir: Path = DEFAULT_OUTPUT_DIR) -> list[dict[str, Any]]:
    if not output_dir.exists():
        return []
    run_dirs = [path for path in output_dir.iterdir() if path.is_dir()]
    run_dirs.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return [summarize_run(path) for path in run_dirs[:limit]]


def summarize_run(run_dir: Path) -> dict[str, Any]:
    state_path = run_dir / "state.json"
    summary_path = run_dir / "batch_summary.html"
    run_id = run_dir.name
    total = manual = problems = 0
    status_label = "未生成结果"
    updated_at = ""
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
            run_id = str(state.get("runId") or run_id)
            updated_at = str(state.get("updatedAt") or "")
            items = state.get("items") or {}
            total = len(items)
        except Exception:
            pass
    csv_path = run_dir / "batch_summary.csv"
    if csv_path.exists():
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            total = total or len(rows)
            manual = sum(1 for row in rows if row.get("manualCategory"))
            for row in rows:
                try:
                    problems += int(row.get("problemCount") or 0)
                except ValueError:
                    pass
        except Exception:
            pass
    if manual:
        status_label = "需处理"
    elif problems:
        status_label = "有差异"
    elif summary_path.exists():
        status_label = "已完成"
    return {
        "runId": run_id,
        "runDir": str(run_dir),
        "updatedAt": updated_at,
        "totalTaxNos": total,
        "manualRequired": manual,
        "problemCount": problems,
        "reviewedCount": reviewed_count(run_dir),
        "summaryPath": str(summary_path) if summary_path.exists() else "",
        "statusLabel": status_label,
    }


def reviewed_count(run_dir: Path) -> int:
    review = load_review_state(run_dir)
    count = 0
    for item in (review.get("items") or {}).values():
        status = str((item or {}).get("reviewStatus") or "")
        if status and status != "待处理":
            count += 1
    return count


def latest_report(output_dir: Path = DEFAULT_OUTPUT_DIR) -> Path | None:
    runs = list_recent_runs(limit=50, output_dir=output_dir)
    for run in runs:
        path = Path(str(run.get("summaryPath") or ""))
        if path.exists():
            return path
    return None


def report_for_run(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> Path | None:
    safe = safe_run_id(run_id)
    path = output_dir / safe / "batch_summary.html"
    return path if path.exists() else None


def missing_run_status(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR, reason: str = "") -> dict[str, Any]:
    safe = safe_run_id(run_id)
    run_dir = output_dir / safe
    return {
        "runId": safe,
        "runDir": str(run_dir),
        "status": "missing",
        "statusLabel": "批次文件缺失",
        "summaryPath": str(run_dir / "batch_summary.html") if (run_dir / "batch_summary.html").exists() else "",
        "items": [],
        "logTail": reason or f"未找到批次状态文件：{run_dir / 'state.json'}",
    }


def job_status_for_run(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> dict[str, Any]:
    try:
        run_dir, state = load_run_state(run_id, output_dir=output_dir)
    except FileNotFoundError as exc:
        return missing_run_status(run_id, output_dir=output_dir, reason=str(exc))
    status_path = run_dir / "ops_status.json"
    if status_path.exists():
        try:
            status = json.loads(status_path.read_text(encoding="utf-8"))
            enrich_status_reasons(status, run_dir)
            return status
        except Exception:
            pass
    status = fallback_ops_status(run_dir, state)
    enrich_status_reasons(status, run_dir)
    return status


def coverage_for_run(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> dict[str, Any]:
    try:
        run_dir, state = load_run_state(run_id, output_dir=output_dir)
    except FileNotFoundError as exc:
        return {
            "runId": safe_run_id(run_id),
            "summary": {"totalTargets": 0, "coveredTargets": 0, "missingTargets": 0, "hitCount": 0},
            "targets": [],
            "missingTargets": [],
            "hits": [],
            "supplement": {"enabled": False, "status": "missing", "message": str(exc)},
        }
    targets = build_coverage_targets(
        declaration_statuses=declaration_statuses_for_collect_statuses(state.get("coverageCollectStatuses") or []),
        tax_types=normalize_tax_type_keys(state.get("coverageTaxTypes") or []),
    )
    return write_coverage_status(run_dir, report_root=PROJECT_ROOT / "output" / "reports", targets=targets)


def fallback_ops_status(run_dir: Path, state: dict[str, Any]) -> dict[str, Any]:
    items = []
    for tax_no, item in sorted((state.get("items") or {}).items()):
        collect = item.get("collect") or {}
        verify = item.get("verify") or {}
        account = collect.get("account") or {}
        items.append(
            {
                "taxNo": tax_no,
                "period": item.get("period") or state.get("period") or "",
                "region": str(account.get("areaCode") or ""),
                "custName": account.get("custName") or "",
                "stage": item.get("stage") or "queued",
                "stageName": item.get("stage") or "排队中",
                "status": fallback_item_status(collect, verify),
                "statusLabel": fallback_item_status_label(collect, verify),
                "taskId": "、".join(str(value) for value in collect.get("verifyTaskIds") or [collect.get("verifyTaskId")] if value),
                "collectStatus": collect.get("status") or "",
                "verifyStatus": verify.get("status") or "",
                "reason": first_non_empty((collect.get("errors") or []) + [verify.get("reason") or ""]),
                "action": "",
                "summaryPath": verify.get("summaryPath") or "",
            }
        )
    return {
        "runId": state.get("runId") or run_dir.name,
        "period": state.get("period") or "",
        "enterprise": state.get("enterprise") or "",
        "status": "running" if any(item["status"] == "running" for item in items) else "finished",
        "updatedAt": state.get("updatedAt") or "",
        "items": items,
    }


def fallback_item_status(collect: dict[str, Any], verify: dict[str, Any]) -> str:
    if collect.get("manualRequired"):
        return "manual"
    status = str(verify.get("status") or "")
    if status == "success":
        return "success"
    if status == "completed_with_differences":
        return "warning"
    if status == "failed":
        return "failed"
    if status == "skipped":
        return "skipped"
    return "running"


def fallback_item_status_label(collect: dict[str, Any], verify: dict[str, Any]) -> str:
    return {
        "manual": "需人工",
        "success": "完成",
        "warning": "有差异",
        "failed": "失败",
        "skipped": "跳过",
        "running": "进行中",
    }.get(fallback_item_status(collect, verify), "未知")


def first_non_empty(values: list[Any]) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def enrich_status_reasons(status: dict[str, Any], run_dir: Path) -> None:
    log_text = tail_text(run_dir / "logs" / "ops_console.log", 20000)
    inferred = infer_blocker_reason(log_text)
    for item in status.get("items") or []:
        if not item.get("reason") and inferred:
            item["reason"] = inferred["reason"]
            item["action"] = inferred["action"]


def infer_blocker_reason(text: str) -> dict[str, str] | None:
    if not text:
        return None
    if "PendingTaxLoginJobError" in text or "已有进税局任务未完成" in text or "之前执行过 进税局" in text:
        match = re.search(r"进税局\((\d{12,})\)", text)
        suffix = f"占用任务：{match.group(1)}。" if match else ""
        return {
            "reason": f"已有进税局任务未完成，新的税局登录被后台拒绝。{suffix}",
            "action": "等待占用任务结束，或在后台处理后点击继续验证。",
        }
    if "DeclarationQueryAuthError" in text or "统一登录页" in text or "数字账户认证" in text:
        return {
            "reason": "税局登录态或数字账户认证已失效。",
            "action": "人工重新进入对应税局/数字账户后点击继续验证。",
        }
    checks = [
        ("YDZ_USERNAME and YDZ_PASSWORD are required", "易代账未登录，需要填写账号密码。", "在任务参数中填写易代账账号密码后重试。"),
        ("Timed out waiting for collection terminal status", "取数任务长时间未完成。", "在易代账任务列表确认取数是否仍在执行，必要时重试取数。"),
        ("Tax bureau login timeout", "税局登录超时。", "完成税局登录后点击继续验证。"),
        ("Could not navigate to declaration query page", "未能进入申报信息查询页。", "重新进入税局后点击继续验证。"),
        ("数字账户登录失效", "数字账户登录失效。", "关闭失效页面，重新进入数字账户后点击继续验证。"),
        ("low web extraction coverage", "网页解析覆盖率低。", "查看对应报告和截图，确认页面是否打开到正确表单。"),
        ("社会保险费", "社会保险费不支持自动取数。", "重新发起取数，确认提交税种不含社会保险费。"),
    ]
    for keyword, reason, action in checks:
        if keyword in text:
            return {"reason": reason, "action": action}
    if "getClientJob" in text and "taskId" in text:
        return {"reason": "后台未返回内部 taskId。", "action": "在后台任务列表确认是否生成取数任务，必要时重新发起取数。"}
    return None


def regenerate_summary(run_id: str) -> Path:
    run_dir, state = load_run_state(run_id)
    from scripts.batch_collect_verify import render_summary, write_state

    write_state(state, run_dir)
    return render_summary(state, run_dir)


def skip_one_tax_no(run_id: str, tax_no: str) -> dict[str, Any]:
    run_dir, state = load_run_state(run_id)
    item = (state.get("items") or {}).get(tax_no)
    if not item:
        raise ValueError(f"批次中没有税号：{tax_no}")
    item["stage"] = "skipped"
    item["stageUpdatedAt"] = datetime.now().isoformat(timespec="seconds")
    item["verify"] = {
        "status": "skipped",
        "reason": "运营手动跳过。",
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
    }
    from scripts.batch_collect_verify import render_summary, write_state

    write_state(state, run_dir)
    render_summary(state, run_dir)
    return job_status_for_run(run_id)


def review_path(run_dir: Path) -> Path:
    return run_dir / "ops_review.json"


def load_review_state(run_dir: Path) -> dict[str, Any]:
    path = review_path(run_dir)
    if not path.exists():
        return {"items": {}, "createdAt": datetime.now().isoformat(timespec="seconds")}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"items": {}, "createdAt": datetime.now().isoformat(timespec="seconds")}
    if not isinstance(data, dict):
        return {"items": {}, "createdAt": datetime.now().isoformat(timespec="seconds")}
    data.setdefault("items", {})
    return data


def save_review_state(run_dir: Path, review: dict[str, Any]) -> None:
    review["updatedAt"] = datetime.now().isoformat(timespec="seconds")
    review_path(run_dir).write_text(json.dumps(review, ensure_ascii=False, indent=2), encoding="utf-8")


def problem_review_key(detail: dict[str, Any]) -> str:
    raw = "|".join(
        str(detail.get(key) or "")
        for key in ("taxNo", "taskId", "formId", "fieldId", "lineNo", "rowName", "columnName", "status")
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def load_problem_details(run_dir: Path) -> list[dict[str, Any]]:
    details_path = run_dir / "batch_problem_details.csv"
    if not details_path.exists():
        return []
    with details_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    unique: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = problem_review_key(row)
        if key not in unique:
            row["key"] = key
            unique[key] = row
    return list(unique.values())


def review_items_for_run(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> dict[str, Any]:
    try:
        run_dir, _state = load_run_state(run_id, output_dir=output_dir)
    except FileNotFoundError as exc:
        safe = safe_run_id(run_id)
        return {"runId": safe, "items": [], "reviewPath": "", "message": str(exc)}
    review = load_review_state(run_dir)
    review_items = review.get("items") or {}
    details = []
    for detail in load_problem_details(run_dir):
        key = detail["key"]
        saved = review_items.get(key) or {}
        evidence = evidence_links_for_detail(detail)
        merged = {
            **detail,
            "key": key,
            "formShortName": short_form_name(str(detail.get("formName") or detail.get("formId") or "")),
            "position": format_problem_position(detail),
            "reason": problem_reason(detail),
            "suggestion": problem_suggestion(detail),
            "reviewStatus": saved.get("reviewStatus") or "待处理",
            "assignee": saved.get("assignee") or "",
            "note": saved.get("note") or "",
            "needDev": bool(saved.get("needDev")),
            "reviewUpdatedAt": saved.get("updatedAt") or "",
            **evidence,
        }
        details.append(merged)
    return {"runId": run_id, "items": details, "reviewPath": str(review_path(run_dir))}


def update_review_item(run_id: str, key: str, fields: dict[str, Any], output_dir: Path = DEFAULT_OUTPUT_DIR) -> dict[str, Any]:
    run_dir, _state = load_run_state(run_id, output_dir=output_dir)
    valid = {"reviewStatus", "assignee", "note", "needDev"}
    review = load_review_state(run_dir)
    item = dict((review.get("items") or {}).get(key) or {})
    for field, value in fields.items():
        if field not in valid:
            continue
        item[field] = bool(value) if field == "needDev" else str(value or "")
    item["updatedAt"] = datetime.now().isoformat(timespec="seconds")
    item["key"] = key
    review.setdefault("items", {})[key] = item
    save_review_state(run_dir, review)
    return item


def export_review(run_id: str, output_dir: Path = DEFAULT_OUTPUT_DIR) -> Path:
    run_dir, _state = load_run_state(run_id, output_dir=output_dir)
    payload = review_items_for_run(run_id, output_dir=output_dir)
    rows = payload["items"]
    export_dir = run_dir / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = export_dir / f"问题处理清单_{run_id}.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "字段差异明细"
        headers = review_export_headers()
        ws.append([label for _key, label in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF1F8")
        for row in rows:
            ws.append([review_export_value(row, key) for key, _label in headers])
        ws.freeze_panes = "A2"
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
            ws.column_dimensions[letter].width = width
        overview = wb.create_sheet("批次总览", 0)
        overview.append(["批次", run_id])
        overview.append(["问题数", len(rows)])
        overview.append(["导出时间", datetime.now().isoformat(timespec="seconds")])
        wb.save(xlsx_path)
        return xlsx_path
    except Exception:
        csv_path = export_dir / f"问题处理清单_{run_id}.csv"
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=[key for key, _label in review_export_headers()])
            writer.writerow({key: label for key, label in review_export_headers()})
            for row in rows:
                writer.writerow({key: review_export_value(row, key) for key, _label in review_export_headers()})
        return csv_path


def review_export_headers() -> list[tuple[str, str]]:
    return [
        ("taxNo", "税号"),
        ("custName", "企业"),
        ("region", "地区"),
        ("taxTypeName", "税种"),
        ("declarationStatus", "已申报/未申报"),
        ("formName", "表单"),
        ("fieldId", "字段"),
        ("position", "Excel位置"),
        ("apiRawValue", "接口值"),
        ("webRawValue", "网页值"),
        ("status", "差异类型"),
        ("reason", "差异说明"),
        ("suggestion", "处理建议"),
        ("reviewStatus", "处理状态"),
        ("assignee", "处理人"),
        ("note", "备注"),
        ("summaryPath", "报告链接"),
        ("pdfPath", "PDF链接"),
        ("excelPath", "接口填充Excel"),
    ]


def review_export_value(row: dict[str, Any], key: str) -> Any:
    return row.get(key) or ""


def format_problem_position(detail: dict[str, Any]) -> str:
    parts = []
    if detail.get("lineNo"):
        parts.append(f"行 {detail.get('lineNo')}")
    if detail.get("rowName"):
        parts.append(str(detail.get("rowName")))
    if detail.get("columnName"):
        parts.append(str(detail.get("columnName")))
    return " / ".join(parts)


def problem_reason(detail: dict[str, Any]) -> str:
    status = str(detail.get("status") or "")
    return {
        "mismatch": "接口值与网页值不一致",
        "api_missing": "接口未返回该字段",
        "web_missing": "网页未解析到该字段",
        "parse_error": "字段解析失败",
        "mapping_error": "字段映射错误",
    }.get(status, status or "字段存在问题")


def problem_suggestion(detail: dict[str, Any]) -> str:
    status = str(detail.get("status") or "")
    if status == "api_missing":
        return "优先确认接口是否应返回该字段。"
    if status == "web_missing":
        return "优先确认税局页面是否打开正确以及网页解析规则。"
    if status == "mismatch":
        return "对照 PDF/网页截图和接口填充 Excel 复核。"
    if status == "mapping_error":
        return "检查字段 ID 与 Excel 位置映射。"
    return "查看报告和证据文件后判定。"


def short_form_name(form_name: str) -> str:
    names = [
        ("附列资料（一）", "附表一"),
        ("附列资料（二）", "附表二"),
        ("附列资料（三）", "附表三"),
        ("附列资料（四）", "附表四"),
        ("附列资料（五）", "附表五"),
        ("一般纳税人适用", "增值税主表"),
        ("文化事业建设费", "文化事业建设费"),
    ]
    for needle, label in names:
        if needle in form_name:
            return label
    return form_name[:18]


def evidence_links_for_detail(detail: dict[str, Any]) -> dict[str, str]:
    task_id = str(detail.get("taskId") or "")
    form_id = str(detail.get("formId") or "")
    report_dir = PROJECT_ROOT / "output" / "reports" / task_id
    result = {
        "summaryPath": normalize_existing_path(detail.get("summaryPath")),
        "pdfPath": "",
        "excelPath": "",
    }
    if not report_dir.exists() or not form_id:
        return result
    result["pdfPath"] = str(latest_matching_file(report_dir / "pdf", form_id, "*.pdf") or "")
    result["excelPath"] = str(latest_matching_file(report_dir / "excel", form_id, "*.xlsx") or "")
    write_report_index(report_dir)
    return result


def normalize_existing_path(value: Any) -> str:
    text = str(value or "")
    if not text:
        return ""
    path = Path(text)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return str(path) if path.exists() else text


def latest_matching_file(folder: Path, token: str, pattern: str) -> Path | None:
    if not folder.exists():
        return None
    matches = [path for path in folder.glob(pattern) if token in path.name]
    if not matches:
        return None
    return max(matches, key=lambda path: path.stat().st_mtime)


def write_report_index(report_dir: Path) -> None:
    forms: dict[str, dict[str, str]] = {}
    for subdir, key, pattern in (("pdf", "pdf", "*.pdf"), ("excel", "filledExcel", "*.xlsx")):
        folder = report_dir / subdir
        if not folder.exists():
            continue
        for path in folder.glob(pattern):
            form_id = extract_form_id_from_name(path.name)
            if not form_id:
                continue
            forms.setdefault(form_id, {})[key] = str(path)
    for path in report_dir.glob("compare_summary_*.html"):
        for form in forms.values():
            form.setdefault("summaryHtml", str(path))
    if forms:
        (report_dir / "report_index.json").write_text(
            json.dumps({"taskId": report_dir.name, "forms": forms}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def extract_form_id_from_name(name: str) -> str:
    known = (
        "vat_general_appendix5",
        "vat_general_appendix4",
        "vat_general_appendix3",
        "vat_general_appendix2",
        "vat_general_appendix1",
        "vat_general_main",
        "vat_small_appendix2",
        "vat_small_appendix1",
        "vat_small_main",
        "culture_fee_deduction",
        "culture_fee_main",
        "cit_a_main",
    )
    for form_id in known:
        if form_id in name:
            return form_id
    return ""


def open_operator_path(value: str) -> Path:
    path = Path(str(value or ""))
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    resolved = path.resolve()
    allowed = [PROJECT_ROOT / "output", PROJECT_ROOT / "runtime"]
    if not any(str(resolved).lower().startswith(str(root.resolve()).lower()) for root in allowed):
        raise ValueError("只允许打开项目输出目录中的文件。")
    if not resolved.exists():
        raise FileNotFoundError(str(resolved))
    return resolved


def open_local_file(path: Path) -> None:
    if hasattr(os, "startfile"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    else:
        webbrowser.open(path.resolve().as_uri())


class OpsConsoleHandler(BaseHTTPRequestHandler):
    server_version = "OpsConsole/1.0"

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/":
            self.send_html(INDEX_HTML)
        elif parsed.path == "/api/health":
            params = {key: values[-1] for key, values in urllib.parse.parse_qs(parsed.query).items()}
            self.send_json({"ok": True, "checks": environment_checks(params)})
        elif parsed.path == "/api/jobs":
            self.send_json({"ok": True, "current": current_job()})
        elif parsed.path == "/api/runs":
            self.send_json({"ok": True, "runs": list_recent_runs()})
        elif parsed.path == "/api/job-status":
            params = {key: values[-1] for key, values in urllib.parse.parse_qs(parsed.query).items()}
            self.send_json({"ok": True, "status": job_status_for_run(str(params.get("runId") or ""))})
        elif parsed.path == "/api/coverage":
            params = {key: values[-1] for key, values in urllib.parse.parse_qs(parsed.query).items()}
            self.send_json({"ok": True, "coverage": coverage_for_run(str(params.get("runId") or ""))})
        elif parsed.path == "/api/review":
            params = {key: values[-1] for key, values in urllib.parse.parse_qs(parsed.query).items()}
            self.send_json({"ok": True, **review_items_for_run(str(params.get("runId") or ""))})
        elif parsed.path == "/api/latest-report":
            path = latest_report()
            self.send_json({"ok": True, "summaryPath": str(path) if path else "", "fileUrl": path.resolve().as_uri() if path else ""})
        else:
            self.send_error(404)

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        try:
            if parsed.path == "/api/start":
                payload = self.read_json()
                job = start_batch_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/create-accountset":
                payload = self.read_json()
                job = start_create_accountset_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/precheck-cit-accountset":
                payload = self.read_json()
                job = start_cit_a_accountset_precheck_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/login-backend":
                payload = self.read_json()
                job = start_backend_login_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/sync-privacy-phone":
                payload = self.read_json()
                job = start_privacy_phone_sync_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/stop":
                job = stop_running_job()
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/open-latest-report":
                path = latest_report()
                if not path:
                    raise RuntimeError("未找到批量汇总报告。")
                open_local_file(path)
                self.send_json({"ok": True, "summaryPath": str(path), "fileUrl": path.resolve().as_uri()})
            elif parsed.path == "/api/open-report":
                payload = self.read_json()
                path = report_for_run(str(payload.get("runId") or ""))
                if not path:
                    raise RuntimeError("未找到该批次报告。")
                open_local_file(path)
                self.send_json({"ok": True, "summaryPath": str(path), "fileUrl": path.resolve().as_uri()})
            elif parsed.path == "/api/regenerate-summary":
                payload = self.read_json()
                path = regenerate_summary(str(payload.get("runId") or ""))
                self.send_json({"ok": True, "summaryPath": str(path), "fileUrl": path.resolve().as_uri(), "status": job_status_for_run(str(payload.get("runId") or ""))})
            elif parsed.path == "/api/analyze-coverage":
                payload = self.read_json()
                coverage = coverage_for_run(str(payload.get("runId") or ""))
                self.send_json({"ok": True, "coverage": coverage})
            elif parsed.path == "/api/resume-run":
                payload = self.read_json()
                job = resume_run_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/verify-one":
                payload = self.read_json()
                job = verify_one_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/retry-one":
                payload = self.read_json()
                job = retry_one_collect_job(payload)
                self.send_json({"ok": True, "job": job})
            elif parsed.path == "/api/skip-one":
                payload = self.read_json()
                status = skip_one_tax_no(str(payload.get("runId") or ""), str(payload.get("taxNo") or ""))
                self.send_json({"ok": True, "status": status})
            elif parsed.path == "/api/review-update":
                payload = self.read_json()
                item = update_review_item(
                    str(payload.get("runId") or ""),
                    str(payload.get("key") or ""),
                    payload.get("fields") if isinstance(payload.get("fields"), dict) else {},
                )
                self.send_json({"ok": True, "item": item})
            elif parsed.path == "/api/export-review":
                payload = self.read_json()
                path = export_review(str(payload.get("runId") or ""))
                self.send_json({"ok": True, "path": str(path), "fileUrl": path.resolve().as_uri()})
            elif parsed.path == "/api/open-path":
                payload = self.read_json()
                path = open_operator_path(str(payload.get("path") or ""))
                open_local_file(path)
                self.send_json({"ok": True, "path": str(path), "fileUrl": path.resolve().as_uri()})
            else:
                self.send_error(404)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        data = json.loads(raw or "{}")
        if not isinstance(data, dict):
            raise ValueError("请求内容格式无效。")
        return data

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, html_text: str) -> None:
        body = html_text.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, _format: str, *_args: Any) -> None:
        return


def bind_server(host: str, port: int) -> ThreadingHTTPServer:
    last_error: Exception | None = None
    for candidate in range(port, port + 20):
        try:
            return ThreadingHTTPServer((host, candidate), OpsConsoleHandler)
        except OSError as exc:
            last_error = exc
    raise RuntimeError(f"无法绑定本地端口：{last_error}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Start local operator console.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--open", action="store_true", help="Open the console in the default browser.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    server = bind_server(args.host, args.port)
    host, port = server.server_address
    url = f"http://{host}:{port}/"
    print(f"Operator console: {url}")
    if args.open:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
