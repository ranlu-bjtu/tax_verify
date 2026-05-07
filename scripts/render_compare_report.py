"""Render a CompareResult JSON file as a readable standalone HTML report."""

from __future__ import annotations

import argparse
import html
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook


STATUS_META = {
    "mismatch": ("不一致", "bad"),
    "api_missing": ("接口缺失", "warn"),
    "web_missing": ("网页缺失", "warn"),
    "parse_error": ("解析失败", "bad"),
    "mapping_error": ("映射错误", "bad"),
    "both_missing": ("双方为空", "muted"),
    "tolerance_match": ("容差通过", "ok"),
    "match": ("一致", "ok"),
    "skip": ("跳过", "muted"),
}
STATUS_ORDER = {
    "mismatch": 0,
    "parse_error": 1,
    "mapping_error": 2,
    "api_missing": 3,
    "web_missing": 4,
    "both_missing": 5,
    "tolerance_match": 6,
    "match": 7,
    "skip": 8,
}
PROBLEM_STATUSES = {"mismatch", "api_missing", "web_missing", "parse_error", "mapping_error"}
WORKBOOK_ROOT = (
    Path.home()
    / "xwechat_files"
    / "wxid_ok3uvjq21ydu22_098f"
    / "msg"
    / "file"
    / "2026-04"
)
LOCAL_WORKBOOK_ROOT = Path("mappings") / "id_workbooks"
VAT_WORKBOOK = "增值税小规模纳税人ID定义-5.16增加附列资料二ok-8.27增加发票归集统计表ok-12.29ok-1.16ok-2.10ok菁菁开发提供主表附加明细ID修改.xlsx"
VAT_GENERAL_WORKBOOK = "增值税一般纳税人.xlsx"
CIT_WORKBOOK = "企业所得税主表.xlsx"
FIELD_ID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render compare JSON to HTML.")
    parser.add_argument("json_report", help="Path to compare result JSON.")
    parser.add_argument("--output", default="", help="HTML output path. Defaults beside the JSON file.")
    parser.add_argument("--workbook", default="", help="ID Excel workbook path for Chinese labels/layout.")
    parser.add_argument("--sheet", default="", help="Workbook sheet name. Defaults from report or first sheet.")
    parser.add_argument("--include-all", action="store_true", help="Also render matched/empty fields.")
    return parser.parse_args()


def h(value: Any) -> str:
    if value is None:
        return "<span class=\"empty\">空</span>"
    text = str(value)
    if text == "":
        return "<span class=\"empty\">空</span>"
    return html.escape(text)


def status_label(status: str) -> str:
    label, cls = STATUS_META.get(status, (status, "muted"))
    return f"<span class=\"pill {cls}\">{html.escape(label)}</span>"


def load_report(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def find_workbook(name: str) -> Path | None:
    matches = [
        p
        for root in (LOCAL_WORKBOOK_ROOT, WORKBOOK_ROOT)
        if root.exists()
        for p in root.rglob(name)
        if not p.name.startswith("~$")
    ]
    return matches[0] if matches else None


def auto_workbook(report: dict[str, Any]) -> Path | None:
    form_code = str(report.get("form_code", ""))
    tax_type = str(report.get("tax_type", ""))
    if "VAT_SMALL_SCALE" in form_code or "VAT_SMALL_SCALE" in tax_type:
        return find_workbook(VAT_WORKBOOK)
    if "VAT_GENERAL" in form_code or "VAT_GENERAL" in tax_type:
        return find_workbook(VAT_GENERAL_WORKBOOK)
    if "CIT" in form_code or "QYSDS" in form_code or "所得税" in str(report.get("form_name", "")):
        return find_workbook(CIT_WORKBOOK)
    return None


def load_excel_metadata(workbook_path: Path | None, sheet_name: str = "") -> dict[str, dict[str, Any]]:
    if not workbook_path or not workbook_path.exists():
        return {}
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    metadata: dict[str, dict[str, Any]] = {}
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            field_id = str(value).strip() if value is not None else ""
            if not FIELD_ID_RE.match(field_id):
                continue
            metadata[field_id] = {
                "field_id": field_id,
                "excel_row": row_idx,
                "excel_col": col_idx,
                "row_name": find_row_name(row, col_idx),
                "line_no": find_row_line_no(row),
                "column_name": build_column_name(rows, row_idx, col_idx),
                "section": find_section(rows, row_idx),
            }
    wb.close()
    return metadata


def find_row_name(row: tuple[Any, ...], col_idx: int) -> str:
    candidates = []
    for value in row[: max(col_idx - 1, 0)]:
        text = str(value).strip() if value is not None else ""
        if text and not FIELD_ID_RE.match(text) and not re.fullmatch(r"\d+(\D.*)?", text):
            candidates.append(text)
    return candidates[-1] if candidates else ""


def find_row_line_no(row: tuple[Any, ...]) -> str:
    for value in row[:4]:
        text = str(value).strip() if value is not None else ""
        if re.fullmatch(r"\d+(\D.*)?", text):
            return text
    return ""


def build_column_name(rows: list[tuple[Any, ...]], row_idx: int, col_idx: int) -> str:
    labels = []
    for header_idx in range(0, row_idx - 1):
        row = rows[header_idx]
        text = header_text_at(row, col_idx)
        if text and not FIELD_ID_RE.match(text) and not skip_header_label(text):
            labels.append(text)
    return " / ".join(dedupe(labels[-3:]))


def skip_header_label(text: str) -> bool:
    return any(token in text for token in ("申报表", "纳税人", "税款所属", "金额单位"))


def header_text_at(row: tuple[Any, ...], col_idx: int) -> str:
    for idx in range(col_idx - 1, -1, -1):
        value = row[idx] if idx < len(row) else None
        text = str(value).strip() if value is not None else ""
        if text:
            return text
    return ""


def find_section(rows: list[tuple[Any, ...]], row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        row = rows[idx]
        first = str(row[0]).strip() if row and row[0] is not None else ""
        if first and not FIELD_ID_RE.match(first):
            return first
    return ""


def dedupe(values: list[str]) -> list[str]:
    result = []
    for value in values:
        if value not in result:
            result.append(value)
    return result


def enrich_fields(fields: list[dict[str, Any]], metadata: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = []
    for item in fields:
        merged = dict(item)
        meta = metadata.get(str(item.get("field_id", "")), {})
        for key in ("row_name", "line_no", "column_name"):
            if meta.get(key) and not merged.get(key):
                merged[key] = meta[key]
        for key in ("excel_row", "excel_col", "section"):
            if meta.get(key):
                merged[key] = meta[key]
        row_name = merged.get("row_name", "")
        column_name = merged.get("column_name", "")
        if row_name and column_name:
            merged["business_name"] = f"{row_name} - {column_name}"
        else:
            merged["business_name"] = row_name or column_name or merged.get("display_name") or merged.get("field_id")
        enriched.append(merged)
    return enriched


def choose_output(input_path: Path, output: str) -> Path:
    if output:
        return Path(output)
    return input_path.with_suffix(".html")


def render_cards(report: dict[str, Any], counts: Counter[str]) -> str:
    summary = report.get("summary", {})
    total = summary.get("total_fields", len(report.get("field_results", [])))
    match_rate = summary.get("match_rate", 0)
    problem_count = sum(counts[s] for s in PROBLEM_STATUSES)
    cards = [
        ("通过率", f"{match_rate}%", "ok" if problem_count == 0 else "warn"),
        ("问题字段", problem_count, "bad" if problem_count else "ok"),
        ("不一致", counts["mismatch"], "bad" if counts["mismatch"] else "ok"),
        ("接口缺失", counts["api_missing"], "warn" if counts["api_missing"] else "ok"),
        ("网页缺失", counts["web_missing"], "warn" if counts["web_missing"] else "ok"),
        ("总字段", total, "muted"),
    ]
    return "\n".join(
        f"""
        <section class="metric {cls}">
          <div class="metric-label">{html.escape(label)}</div>
          <div class="metric-value">{html.escape(str(value))}</div>
        </section>
        """
        for label, value, cls in cards
    )


def render_problem_list(fields: list[dict[str, Any]]) -> str:
    problems = problem_fields(fields)
    if not problems:
        return "<p class=\"quiet\">没有发现需要优先处理的问题字段。</p>"
    rows = []
    for item in problems:
        rows.append(
            f"""
            <tr>
              <td>{status_label(item.get("status", ""))}</td>
              <td>{h(item.get("line_no"))}</td>
              <td>{h(item.get("business_name"))}</td>
              <td><code>{h(item.get("field_id"))}</code></td>
              <td>{h(item.get("api_raw_value"))}</td>
              <td>{h(item.get("web_raw_value"))}</td>
              <td>{h(item.get("diff_value"))}</td>
            </tr>
            """
        )
    return f"""
    <table>
      <thead>
        <tr>
          <th>状态</th><th>行次</th><th>项目</th><th>字段</th><th>接口值</th><th>网页值</th><th>差异</th>
        </tr>
      </thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    """


def problem_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        item for item in sorted(
            fields,
            key=lambda x: (
                STATUS_ORDER.get(x.get("status", ""), 99),
                int(x.get("excel_row") or 999999),
                str(x.get("line_no", "")),
                x.get("field_id", ""),
            ),
        )
        if item.get("status") in PROBLEM_STATUSES
    ]


def render_all_rows(fields: list[dict[str, Any]]) -> str:
    rows = []
    for item in sorted(fields, key=lambda x: (STATUS_ORDER.get(x.get("status", ""), 99), str(x.get("line_no", "")), x.get("field_id", ""))):
        rows.append(
            f"""
            <tr data-status="{html.escape(str(item.get('status', '')))}">
              <td>{status_label(item.get("status", ""))}</td>
              <td>{h(item.get("line_no"))}</td>
              <td>{h(item.get("business_name"))}</td>
              <td><code>{h(item.get("field_id"))}</code></td>
              <td>{h(item.get("data_type"))}</td>
              <td>{h(item.get("api_raw_value"))}</td>
              <td>{h(item.get("web_raw_value"))}</td>
              <td>{h(item.get("api_normalized"))}</td>
              <td>{h(item.get("web_normalized"))}</td>
              <td>{h(item.get("detail"))}</td>
            </tr>
            """
        )
    return "\n".join(rows)


def render_html(report: dict[str, Any], source_path: Path, metadata: dict[str, dict[str, Any]], include_all: bool = False) -> str:
    fields = enrich_fields(report.get("field_results", []), metadata)
    counts = Counter(item.get("status", "") for item in fields)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title = report.get("form_name") or report.get("form_code") or source_path.name
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)} - 对比报告</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #172033;
      --muted: #667085;
      --line: #d8dee8;
      --bad: #b42318;
      --bad-bg: #fee4e2;
      --warn: #b54708;
      --warn-bg: #fef0c7;
      --ok: #027a48;
      --ok-bg: #dcfae6;
      --soft: #eef2f7;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font: 14px/1.5 "Segoe UI", "Microsoft YaHei", Arial, sans-serif;
    }}
    header {{
      padding: 24px 32px 16px;
      background: var(--panel);
      border-bottom: 1px solid var(--line);
    }}
    h1 {{ margin: 0 0 8px; font-size: 22px; font-weight: 650; }}
    h2 {{ margin: 24px 0 12px; font-size: 17px; }}
    .meta {{ color: var(--muted); display: flex; gap: 16px; flex-wrap: wrap; }}
    main {{ padding: 20px 32px 36px; }}
    .metrics {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 12px;
      margin-bottom: 22px;
    }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-left: 5px solid var(--line);
      padding: 14px 16px;
      border-radius: 8px;
    }}
    .metric.ok {{ border-left-color: var(--ok); }}
    .metric.warn {{ border-left-color: var(--warn); }}
    .metric.bad {{ border-left-color: var(--bad); }}
    .metric-label {{ color: var(--muted); font-size: 13px; }}
    .metric-value {{ font-size: 24px; font-weight: 700; margin-top: 2px; }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 18px;
      overflow: auto;
    }}
    .toolbar {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      align-items: center;
      margin-bottom: 12px;
    }}
    input, select {{
      height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 0 10px;
      background: #fff;
      min-width: 180px;
    }}
    table {{ width: 100%; border-collapse: collapse; min-width: 980px; }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 9px 10px;
      text-align: left;
      vertical-align: top;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #f9fafb;
      color: #344054;
      font-weight: 650;
      z-index: 1;
    }}
    tr:hover td {{ background: #fafcff; }}
    code {{ font-family: Consolas, "SFMono-Regular", monospace; font-size: 13px; }}
    .pill {{
      display: inline-flex;
      align-items: center;
      min-height: 22px;
      padding: 2px 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 650;
      white-space: nowrap;
    }}
    .pill.bad {{ color: var(--bad); background: var(--bad-bg); }}
    .pill.warn {{ color: var(--warn); background: var(--warn-bg); }}
    .pill.ok {{ color: var(--ok); background: var(--ok-bg); }}
    .pill.muted {{ color: #475467; background: var(--soft); }}
    .empty, .quiet {{ color: var(--muted); }}
  </style>
</head>
<body>
  <header>
    <h1>{html.escape(title)}</h1>
    <div class="meta">
      <span>批次：{html.escape(str(report.get("batch_id", "")))}</span>
      <span>表单：{html.escape(str(report.get("form_code", "")))}</span>
      <span>报告时间：{html.escape(str(report.get("timestamp", "")))}</span>
      <span>生成时间：{generated_at}</span>
      <span>来源：{html.escape(str(source_path))}</span>
    </div>
  </header>
  <main>
    <div class="metrics">{render_cards(report, counts)}</div>
    <section class="panel">
      <h2>优先处理</h2>
      {render_problem_list(fields)}
    </section>
    {render_all_section(fields) if include_all else ""}
  </main>
  {render_filter_script() if include_all else ""}
</body>
</html>
"""


def render_all_section(fields: list[dict[str, Any]]) -> str:
    return f"""
    <section class="panel">
      <h2>全部字段</h2>
      <div class="toolbar">
        <input id="search" placeholder="搜索字段、行次或取值">
        <select id="status">
          <option value="">全部状态</option>
          {''.join(f'<option value="{html.escape(k)}">{html.escape(v[0])} ({counts[k]})</option>' for k, v in STATUS_META.items() if counts[k])}
        </select>
      </div>
      <table id="all">
        <thead>
          <tr>
            <th>状态</th><th>行次</th><th>项目</th><th>字段</th><th>类型</th><th>接口原值</th>
            <th>网页原值</th><th>接口归一</th><th>网页归一</th><th>说明</th>
          </tr>
        </thead>
        <tbody>{render_all_rows(fields)}</tbody>
      </table>
    </section>
    """


def render_filter_script() -> str:
    return """
  <script>
    const search = document.getElementById('search');
    const status = document.getElementById('status');
    const rows = Array.from(document.querySelectorAll('#all tbody tr'));
    function applyFilters() {{
      const q = search.value.trim().toLowerCase();
      const s = status.value;
      for (const row of rows) {{
        const okStatus = !s || row.dataset.status === s;
        const okText = !q || row.innerText.toLowerCase().includes(q);
        row.style.display = okStatus && okText ? '' : 'none';
      }}
    }}
    search.addEventListener('input', applyFilters);
    status.addEventListener('change', applyFilters);
  </script>
"""


def main() -> int:
    args = parse_args()
    input_path = Path(args.json_report)
    report = load_report(input_path)
    workbook_path = Path(args.workbook) if args.workbook else auto_workbook(report)
    metadata = load_excel_metadata(workbook_path, args.sheet or report.get("sheet_name", ""))
    output_path = choose_output(input_path, args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_html(report, input_path, metadata, include_all=args.include_all), encoding="utf-8")
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
