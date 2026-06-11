"""Compare configured tax declaration forms by outer taskId.

This is the extensible version of the small VAT main-table smoke script.  A
new form comparison should normally be added as one CompareTarget entry: API
tax code/table, ID workbook/sheet, query-result keywords, optional detail form
keywords, and an extraction strategy.
"""

from __future__ import annotations

import argparse
import base64
import calendar
import hashlib
import html
import json
import logging
import re
import sys
import time
import urllib.parse
import urllib.parse
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import requests
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

sys.path.insert(0, ".")

from src.api.api_client import APIClient
from src.chanjet_admin.task_execution_log import (
    current_period_flag_from_logs,
    fetch_current_period_flag as fetch_task_current_period_flag,
    fetch_task_execution_logs,
)
from src.compare.comparator import Comparator
from src.compare.value_normalizer import get_normalizer
from src.login.auto_tax_login import CHANJET_TASK_URL
from src.login.browser_manager import BrowserManager
from src.login.task_login_flow import TaskLoginFlow
from src.models.field_mapping import DataType, FieldMapping
from src.navigation.navigation_engine import NavigationEngine
from src.registry.tax_type_registry import TaxTypeRegistry


LOGGER = logging.getLogger("compare_tax_forms")

WORKBOOK_ROOT = (
    Path.home()
    / "xwechat_files"
    / "wxid_ok3uvjq21ydu22_098f"
    / "msg"
    / "file"
    / "2026-04"
)
WECHAT_FILE_ROOT = Path.home() / "xwechat_files" / "wxid_ok3uvjq21ydu22_098f" / "msg" / "file"
LOCAL_WORKBOOK_ROOT = Path("mappings") / "id_workbooks"
VAT_WORKBOOK = "增值税小规模纳税人ID定义-5.16增加附列资料二ok-8.27增加发票归集统计表ok-12.29ok-1.16ok-2.10ok菁菁开发提供主表附加明细ID修改.xlsx"
VAT_GENERAL_WORKBOOK = "增值税一般纳税人.xlsx"
CIT_WORKBOOK = "企业所得税主表.xlsx"
CULTURE_FEE_WORKBOOK = "文化事业建设费ID-9.16ok-9.30ok-1.7ok.xlsx"
CONSUMPTION_TAX_MAIN_WORKBOOK = "消费税及附加税费申报表-2026-05-28 09_50_52.xlsx(1).xlsx"
CONSUMPTION_TAX_SURCHARGE_WORKBOOK = "消费税附加税费计算表-2026-05-28 09_50_58.xlsx(1).xlsx"
WORKBOOK_CACHE_ROOT = Path("runtime") / "workbook_cache"
UNSUPPORTED_SHEET_PROTECTION_ATTRS = ("allowResizeRows", "allowResizeColumns")

FIELD_ID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
TEXT_FIELDS = {"nsrsbh", "nsrmc", "nsrxm", "nsssq", "djxh", "sbpm"}
CONSUMPTION_TAX_TEXT_PREFIXES = (
    "nsrsbh",
    "nsrmc",
    "nsrxm",
    "skssq",
    "ysxfpmc",
    "jldw",
    "jmxzdm",
    "bqsfsy",
    "jmzc",
    "syjmzc",
    "jbr",
    "jbrsfzh",
    "dlr",
    "dljg",
    "slswjg",
    "blrysfz",
    "lxfs",
)
CONSUMPTION_TAX_RATE_PREFIXES = ("desl", "blsl", "sfl", "jzbl")
CONSUMPTION_TAX_MAIN_RENDER_SENTINEL_FIELDS = (
    "bqybtse_xfsjfjsfsbb",
    "bnljybtse_xfsjfjsfsbb",
    "bqcswhjssybtse_xfsjfjsfsbb",
    "bnljcswhjssybtse_xfsjfjsfsbb",
    "bqjyffjybtse_xfsjfjsfsbb",
    "bnljjyffjybtse_xfsjfjsfsbb",
    "bqdfjyfjybtse_xfsjfjsfsbb",
    "bnljdfjyfjybtse_xfsjfjsfsbb",
)
CONSUMPTION_TAX_NON_API_PREFIXES = (
    "nsrsbh",
    "nsrmc",
    "nsrxm",
    "skssq",
    "bqsfsy",
    "jmzc",
    "syjmzc",
    "jbr",
    "jbrsfzh",
    "dlr",
    "dljg",
    "slswjg",
    "blrysfz",
    "lxfs",
)
DASH_ZERO_VALUES = {"——", "—", "-", "/", "－"}
PROBLEM_STATUSES = {"mismatch", "api_missing", "web_missing", "parse_error", "mapping_error"}
QUALITY_WARNING_STATUSES = PROBLEM_STATUSES | {"both_missing"}
WEB_EXTRACTION_DEFAULT_MIN_COVERAGE = 0.6
WEB_EXTRACTION_RECOVERY_ATTEMPTS = 3
WEB_EXTRACTION_RECOVERY_MAX_SECONDS = 60
UNDECLARED_APPENDIX_MENU_PRECHECK_SECONDS = 5
UNDECLARED_JUST_CLICKED_MENU_CONFIRM_SECONDS = 90
WEB_EXTRACTION_MIN_COVERAGE_BY_TARGET = {
    "vat_general_main": 0.6,
    "vat_general_appendix1": 0.6,
    "vat_general_appendix2": 0.6,
    "vat_general_appendix3": 0.6,
    "vat_general_appendix4": 0.6,
    "vat_general_appendix5": 0.6,
    "vat_small_main": 0.75,
    "vat_small_appendix1": 0.75,
    "vat_small_appendix2": 0.6,
    "cit_a_main": 0.6,
    "culture_fee_main": 0.6,
    "culture_fee_deduction": 0.6,
    "consumption_tax_main": 0.75,
    "consumption_tax_surcharge": 0.75,
}
WEB_EXTRACTION_SCROLL_RETRY_TARGETS = {
    "vat_general_main",
    "vat_general_appendix1",
    "vat_general_appendix2",
    "vat_general_appendix3",
    "vat_general_appendix4",
    "vat_general_appendix5",
    "vat_small_main",
    "vat_small_appendix1",
    "vat_small_appendix2",
    "cit_a_main",
    "culture_fee_main",
    "culture_fee_deduction",
    "consumption_tax_main",
    "consumption_tax_surcharge",
}
WEB_EXTRACTION_RECOVERY_SCROLL_POSITIONS = (
    (0.0, 0.0),
    (0.15, 0.0),
    (0.35, 0.0),
    (0.5, 0.0),
    (0.67, 0.0),
    (0.84, 0.0),
    (1.0, 0.0),
    (1.0, 0.5),
    (1.0, 1.0),
    (0.5, 1.0),
    (0.0, 1.0),
)
FORM_VALUE_TOKEN_PATTERN = r"(?:——|—|－|-?\d[\d,]*(?:\.\d+)?)"
FORM_VALUE_TOKEN_PATTERN = FORM_VALUE_TOKEN_PATTERN.replace(r"(?:\.\d+)?)", r"(?:\.\d+)?%?)")
FORM_VALUE_TOKEN_RE = re.compile(FORM_VALUE_TOKEN_PATTERN)
QUERY_URL_HINTS = ("sbxxcx", "sbxx/sbxxcx", "zhcx/sbxx")
UNDECLARED_CONSUMPTION_TAX_PATH = "/sbzx/view/lzsfjssb/#/declare/xfssb?jyjkId=30"
TAX_DIGITAL_ACCOUNT_HINTS = (
    "\u7a0e\u52a1\u6570\u5b57\u8d26\u6237",
    "\u8d26\u6237\u67e5\u8be2",
    "\u53d1\u7968\u4e1a\u52a1",
    "\u5408\u89c4\u7406\u7a0e",
)
TAX_PORTAL_HINTS = (
    "\u6211\u8981\u67e5\u8be2",
    "\u6211\u8981\u529e\u7a0e",
    "\u7533\u62a5\u4fe1\u606f\u67e5\u8be2",
    "\u672c\u671f\u5e94\u7533\u62a5",
    "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801",
    "\u7eb3\u7a0e\u4eba\u8bc6\u522b\u53f7",
)
API_EXCEL_VALUE_FILL = PatternFill("solid", fgColor="E2F0D9")
API_EXCEL_MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")
UNDECLARED_VAT_GENERAL_PATH = "/sbzx/view/lzsfjssb/#/declare/zzsybnsrsb?jyjkId=10"
UNDECLARED_VAT_SMALL_PATH = "/sbzx/view/lzsfjssb/#/declare/zzsxgmnsrsb?jyjkId=10"
UNDECLARED_CULTURE_FEE_SHANDONG_PATH_TEMPLATE = (
    "/sbzx/view/sdsfsgjssb/#/yyzx/whsyjsf/tb/yssb"
    "?SssqQ={period_start}&SssqZ={period_end}&ZspmDmList=302170200"
)
UNDECLARED_SUPPORTED_TAX_TYPES = {
    "VAT_GENERAL",
    "VAT_SMALL_SCALE",
    "CIT_A_PREPAY",
    "CULTURE_FEE",
    "CONSUMPTION_TAX",
}
VAT_APPENDIX5_ROW_LABELS = {
    "_cjs": "城市维护建设税",
    "_jyfj": "教育费附加",
    "_jyf": "教育费附加",
    "_dfjyfj": "地方教育附加",
    "_hj": "合计",
}
VAT_APPENDIX5_ROW_NUMBERS = {
    "城市维护建设税": "1",
    "教育费附加": "2",
    "地方教育附加": "3",
    "合计": "4",
}
VAT_APPENDIX5_GRID_BASE_COL = 3
VAT_APPENDIX5_GRID_WIDTH = 14
VAT_APPENDIX5_SINGLE_VALUE_LABELS = {
    "dqxztze": "当期新增投资额",
    "sqldkdmje": "上期留抵可抵免金额",
    "jzxqkdmje": "结转下期可抵免金额",
    "dqxzkyykcdldtse": "当期新增可用于扣除的留抵退税额",
    "sqjckyykcdldtse": "上期结存可用于扣除的留抵退税额",
    "jzxqkyykcdldtse": "结转下期可用于扣除的留抵退税额",
}
VAT_APPENDIX5_TEXT_FIELDS = {
    "bqsfsyxwqylslfjmzc",
    "jmzcsyzt",
    "syjmzcqsj",
    "syjmzczsj",
    "bqsfsysdjspycjrhxqydmzc",
    "jbr",
    "jbrsfzhm",
    "lxfs",
}
UNDECLARED_VAT_MENU_KEYWORDS = {
    "vat_small_main": ("主表", "增值税及附加税费申报表"),
    "vat_small_appendix1": ("附列资料一", "扣除项目明细"),
    "vat_small_appendix2": ("附列资料二", "附加税费情况表"),
    "cit_a_main": ("A200000", "企业所得税月（季）度预缴纳税申报表"),
    "vat_general_main": ("主表", "增值税及附加税费申报表"),
    "vat_general_appendix1": ("附列资料一", "本期销售情况明细"),
    "vat_general_appendix2": ("附列资料二", "本期进项税额明细"),
    "vat_general_appendix3": ("附列资料三", "扣除项目明细"),
    "vat_general_appendix4": ("附列资料四", "税额抵减情况表"),
    "vat_general_appendix5": ("附列资料五", "附加税费情况表"),
    "culture_fee_main": ("文化事业建设费申报表",),
    "culture_fee_deduction": ("应税服务减除项目清单",),
    "consumption_tax_main": ("消费税及附加税费申报表", "主表"),
    "consumption_tax_surcharge": ("消费税附加税费计算表",),
}


UNDECLARED_TARGET_BODY_KEYWORDS = {
    "vat_general_appendix1": (
        "\u589e\u503c\u7a0e\u7eb3\u7a0e\u7533\u62a5\u8868\u9644\u5217\u8d44\u6599\uff08\u4e00\uff09",
        "\u672c\u671f\u9500\u552e\u60c5\u51b5\u660e\u7ec6",
        "\u9879\u76ee\u53ca\u680f\u6b21",
    ),
}

RECENT_UNDECLARED_MENU_TARGETS: dict[int, tuple[str, float]] = {}


@dataclass(frozen=True)
class CompareTarget:
    target_id: str
    tax_type: str
    tax_code: str
    api_table: str
    workbook_name: str
    sheet_name: str
    form_code: str
    form_name: str
    query_keywords: tuple[str, ...]
    detail_form_keywords: tuple[str, ...] = ()
    loader: str = "layout_scan"
    extractor: str = "layout_scan"


TARGETS: dict[str, CompareTarget] = {
    "vat_small_main": CompareTarget(
        target_id="vat_small_main",
        tax_type="VAT_SMALL_SCALE",
        tax_code="sz_zzs",
        api_table="zzszb_qc",
        workbook_name=VAT_WORKBOOK,
        sheet_name="主表",
        form_code="VAT_SMALL_SCALE_MAIN",
        form_name="增值税及附加税费申报表（小规模纳税人适用）",
        query_keywords=("增值税及附加税费申报表", "小规模纳税人适用"),
        loader="vat_main",
        extractor="line_table",
    ),
    "vat_small_appendix1": CompareTarget(
        target_id="vat_small_appendix1",
        tax_type="VAT_SMALL_SCALE",
        tax_code="sz_zzs",
        api_table="zzsflzl_qc",
        workbook_name=VAT_WORKBOOK,
        sheet_name="附列资料一",
        form_code="VAT_SMALL_SCALE_APPENDIX1",
        form_name="增值税及附加税费申报表（小规模纳税人适用）附列资料（一）（服务、不动产和无形资产扣除项目明细）",
        query_keywords=("增值税及附加税费申报表", "小规模纳税人适用"),
        detail_form_keywords=("附列资料（一）", "服务、不动产和无形资产扣除项目明细"),
    ),
    "vat_small_appendix2": CompareTarget(
        target_id="vat_small_appendix2",
        tax_type="VAT_SMALL_SCALE",
        tax_code="sz_zzs",
        api_table="zzsflzl2_qc",
        workbook_name=VAT_WORKBOOK,
        sheet_name="附列资料二",
        form_code="VAT_SMALL_SCALE_APPENDIX2",
        form_name="增值税及附加税费申报表（小规模纳税人适用）附列资料（二）（附加税费情况表）",
        query_keywords=("增值税及附加税费申报表", "小规模纳税人适用"),
        detail_form_keywords=("附列资料（二）", "附加税费情况表"),
    ),
    "cit_a_main": CompareTarget(
        target_id="cit_a_main",
        tax_type="CIT_A_PREPAY",
        tax_code="sz_qysds",
        api_table="qysds_zb_qc",
        workbook_name=CIT_WORKBOOK,
        sheet_name="主表",
        form_code="CIT_A_MAIN",
        form_name="A200000 中华人民共和国企业所得税月（季）度预缴纳税申报表（A类）",
        query_keywords=("企业所得税", "月", "季"),
        detail_form_keywords=("A200000", "企业所得税月（季）度预缴纳税申报表（A类）"),
    ),
    "vat_general_main": CompareTarget(
        target_id="vat_general_main",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzszb_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="主表",
        form_code="VAT_GENERAL_MAIN",
        form_name="增值税纳税申报表（一般纳税人适用）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("增值税及附加税费申报表", "一般纳税人适用"),
    ),
    "vat_general_appendix1": CompareTarget(
        target_id="vat_general_appendix1",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzsfb1_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="附表1",
        form_code="VAT_GENERAL_APPENDIX1",
        form_name="增值税纳税申报表附列资料（一）（本期销售情况明细）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("附列资料（一）", "本期销售情况明细"),
    ),
    "vat_general_appendix2": CompareTarget(
        target_id="vat_general_appendix2",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzsfb2_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="附表2",
        form_code="VAT_GENERAL_APPENDIX2",
        form_name="增值税纳税申报表附列资料（二）（本期进项税额明细）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("附列资料（二）", "本期进项税额明细"),
    ),
    "vat_general_appendix3": CompareTarget(
        target_id="vat_general_appendix3",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzsfb3_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="附表3",
        form_code="VAT_GENERAL_APPENDIX3",
        form_name="增值税纳税申报表附列资料（三）（服务、不动产和无形资产扣除项目明细）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("附列资料（三）", "扣除项目明细"),
    ),
    "vat_general_appendix4": CompareTarget(
        target_id="vat_general_appendix4",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzsfb4_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="附表4",
        form_code="VAT_GENERAL_APPENDIX4",
        form_name="增值税纳税申报表附列资料（四）（税额抵减情况表）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("附列资料（四）", "税额抵减情况表"),
    ),
    "vat_general_appendix5": CompareTarget(
        target_id="vat_general_appendix5",
        tax_type="VAT_GENERAL",
        tax_code="sz_zzs",
        api_table="zzsfb5_qc",
        workbook_name=VAT_GENERAL_WORKBOOK,
        sheet_name="附表5",
        form_code="VAT_GENERAL_APPENDIX5",
        form_name="增值税及附加税费申报表附列资料（五）（附加税费情况表）",
        query_keywords=("增值税", "一般纳税人适用"),
        detail_form_keywords=("附列资料（五）", "附加税费情况表"),
    ),
    "culture_fee_main": CompareTarget(
        target_id="culture_fee_main",
        tax_type="CULTURE_FEE",
        tax_code="sz_whsyjsf",
        api_table="",
        workbook_name=CULTURE_FEE_WORKBOOK,
        sheet_name="文化事业建设费申报表",
        form_code="CULTURE_FEE_MAIN",
        form_name="文化事业建设费申报表",
        query_keywords=("文化事业建设费",),
        detail_form_keywords=("文化事业建设费申报表",),
    ),
    "culture_fee_deduction": CompareTarget(
        target_id="culture_fee_deduction",
        tax_type="CULTURE_FEE",
        tax_code="sz_whsyjsf",
        api_table="",
        workbook_name=CULTURE_FEE_WORKBOOK,
        sheet_name="应税服务减除项目清单",
        form_code="CULTURE_FEE_DEDUCTION",
        form_name="应税服务减除项目清单",
        query_keywords=("文化事业建设费",),
        detail_form_keywords=("应税服务减除项目清单",),
    ),
    "consumption_tax_main": CompareTarget(
        target_id="consumption_tax_main",
        tax_type="CONSUMPTION_TAX",
        tax_code="sz_xfs",
        api_table="xfszb_qc",
        workbook_name=CONSUMPTION_TAX_MAIN_WORKBOOK,
        sheet_name="Sheet1",
        form_code="CONSUMPTION_TAX_MAIN",
        form_name="消费税及附加税费申报表",
        query_keywords=("消费税及附加税费申报表",),
        detail_form_keywords=("消费税及附加税费申报表",),
    ),
    "consumption_tax_surcharge": CompareTarget(
        target_id="consumption_tax_surcharge",
        tax_type="CONSUMPTION_TAX",
        tax_code="sz_xfs",
        api_table="xfsfb1_qc",
        workbook_name=CONSUMPTION_TAX_SURCHARGE_WORKBOOK,
        sheet_name="Sheet1",
        form_code="CONSUMPTION_TAX_SURCHARGE",
        form_name="消费税附加税费计算表",
        query_keywords=("消费税及附加税费申报表",),
        detail_form_keywords=("消费税附加税费计算表",),
    ),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare configured tax forms by taskId.")
    parser.add_argument("--task-id", required=True, help="Outer Chanjet taskId.")
    parser.add_argument(
        "--targets",
        default="auto",
        help=f"Comma-separated target ids, 'auto', or 'all'. Available: {', '.join(TARGETS)}",
    )
    parser.add_argument("--config-root", default="config")
    parser.add_argument("--cdp-port", type=int, default=9222)
    parser.add_argument("--mode", choices=["auto", "connect", "launch"], default="auto")
    parser.add_argument("--chrome-path", default=r"C:\Program Files\Google\Chrome\Application\chrome.exe")
    parser.add_argument("--user-data-dir", default="./browser_profile/etax_compare_forms")
    parser.add_argument("--plugin-path", default=r"C:\Users\Administrator\Downloads\EtaxPlugin")
    parser.add_argument("--chanjet-timeout", type=int, default=300)
    parser.add_argument("--tax-timeout", type=int, default=600)
    parser.add_argument("--tax-login-strategy", choices=["direct_first", "plugin_first"], default="plugin_first")
    parser.add_argument(
        "--declaration-status-override",
        choices=["", "filed", "unfiled"],
        default="",
        help="Use this declaration status only when task execution logs do not expose one.",
    )
    parser.add_argument("--skip-api", action="store_true", help="Only load workbook mappings; do not call Chanjet API.")
    parser.add_argument("--skip-browser", action="store_true", help="Only load API/workbook mapping; do not open tax pages.")
    parser.add_argument("--skip-pdf", action="store_true", help="Do not save a PDF copy of the compared tax page.")
    parser.add_argument("--log-level", default="INFO")
    return parser.parse_args()


def setup_logging(level: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )


def resolve_targets(value: str) -> list[CompareTarget]:
    if value.strip().lower() == "auto":
        raise ValueError("'auto' targets require API data; call resolve_auto_targets after fetch_api")
    if value.strip().lower() == "all":
        return list(TARGETS.values())
    resolved = []
    for raw_id in value.split(","):
        target_id = raw_id.strip()
        if not target_id:
            continue
        if target_id not in TARGETS:
            raise KeyError(f"Unknown target '{target_id}'. Available: {', '.join(TARGETS)}")
        resolved.append(TARGETS[target_id])
    return resolved


def is_auto_targets(value: str) -> bool:
    return value.strip().lower() == "auto"


def resolve_auto_targets(api_by_tax: dict[str, Any], mappings_by_target: dict[str, list[FieldMapping]]) -> list[CompareTarget]:
    selected_ids: list[str] = []

    # CIT is independent and can be selected directly from its configured API table.
    if target_api_coverage(api_by_tax, TARGETS["cit_a_main"], mappings_by_target["cit_a_main"]) > 0:
        selected_ids.append("cit_a_main")

    culture_fee_ids = ("culture_fee_main", "culture_fee_deduction")
    if any(target_api_coverage(api_by_tax, TARGETS[target_id], mappings_by_target[target_id]) > 0 for target_id in culture_fee_ids):
        selected_ids.extend(culture_fee_ids)

    consumption_tax_ids = ("consumption_tax_main", "consumption_tax_surcharge")
    if any(target_api_coverage(api_by_tax, TARGETS[target_id], mappings_by_target[target_id]) > 0 for target_id in consumption_tax_ids):
        selected_ids.extend(consumption_tax_ids)

    vat_ids = resolve_auto_vat_targets(api_by_tax, mappings_by_target)
    selected_ids.extend(vat_ids)

    selected = [TARGETS[target_id] for target_id in TARGETS if target_id in selected_ids]
    LOGGER.info("Auto-selected targets: %s", ", ".join(target.target_id for target in selected) or "(none)")
    return selected


def resolve_auto_vat_targets(api_by_tax: dict[str, Any], mappings_by_target: dict[str, list[FieldMapping]]) -> list[str]:
    general_ids = [
        "vat_general_main",
        "vat_general_appendix1",
        "vat_general_appendix2",
        "vat_general_appendix3",
        "vat_general_appendix4",
        "vat_general_appendix5",
    ]
    small_ids = ["vat_small_main", "vat_small_appendix1", "vat_small_appendix2"]

    general_appendices = [
        target_id
        for target_id in general_ids[1:]
        if target_has_prefixed_table(api_by_tax, TARGETS[target_id])
        and target_api_coverage(api_by_tax, TARGETS[target_id], mappings_by_target[target_id]) > 0
    ]
    small_appendices = [
        target_id
        for target_id in small_ids[1:]
        if target_has_prefixed_table(api_by_tax, TARGETS[target_id])
        and target_api_coverage(api_by_tax, TARGETS[target_id], mappings_by_target[target_id]) > 0
    ]

    if general_appendices:
        return general_ids
    if small_appendices:
        return small_ids

    general_main_coverage = target_api_coverage(api_by_tax, TARGETS["vat_general_main"], mappings_by_target["vat_general_main"])
    small_main_coverage = target_api_coverage(api_by_tax, TARGETS["vat_small_main"], mappings_by_target["vat_small_main"])
    if general_main_coverage <= 0 and small_main_coverage <= 0:
        return []
    if general_main_coverage >= small_main_coverage:
        return ["vat_general_main"]
    return ["vat_small_main"]


def target_has_prefixed_table(api_by_tax: dict[str, Any], target: CompareTarget) -> bool:
    tax_data = api_by_tax.get(target.tax_code, {})
    if not isinstance(tax_data, dict):
        return False
    if target.api_table in tax_data and tax_data.get(target.api_table) not in (None, "", {}, []):
        return True
    prefix = f"{target.api_table}."
    return any(str(key).startswith(prefix) for key in tax_data)


def target_api_coverage(api_by_tax: dict[str, Any], target: CompareTarget, mappings: list[FieldMapping]) -> int:
    return sum(1 for mapping in mappings if is_effective_api_value(api_value(api_by_tax, target, mapping.field_id)))


def filter_targets_with_api_data(
    api_by_tax: dict[str, Any],
    targets: list[CompareTarget],
    mappings_by_target: dict[str, list[FieldMapping]],
) -> list[CompareTarget]:
    for target in targets:
        mappings = mappings_by_target.get(target.target_id, [])
        coverage = target_api_coverage(api_by_tax, target, mappings)
        if coverage <= 0:
            LOGGER.info(
                "Keeping target %s although API field coverage is 0/%s; tax_code=%s. "
                "The selected tax type must keep all registered forms visible in the verification output.",
                target.target_id,
                len(mappings),
                target.tax_code,
            )
    return targets


def is_effective_api_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def find_workbook(name: str) -> Path:
    search_roots = [LOCAL_WORKBOOK_ROOT, WORKBOOK_ROOT]
    matches = [
        p
        for root in search_roots
        if root.exists()
        for p in root.rglob(name)
        if not p.name.startswith("~$")
    ]
    if not matches and WECHAT_FILE_ROOT.exists():
        direct_candidates = [WECHAT_FILE_ROOT / name]
        direct_candidates.extend(
            child / name for child in WECHAT_FILE_ROOT.iterdir() if child.is_dir()
        )
        matches = [p for p in direct_candidates if p.exists() and not p.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Workbook not found: {name}; searched={search_roots + [WECHAT_FILE_ROOT]}")
    return matches[0]


def workbook_has_unsupported_protection_attrs(workbook: Path) -> bool:
    try:
        with ZipFile(workbook) as archive:
            for name in archive.namelist():
                if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                    continue
                data = archive.read(name)
                if any(attr.encode("utf-8") in data for attr in UNSUPPORTED_SHEET_PROTECTION_ATTRS):
                    return True
    except Exception:
        return False
    return False


def sanitized_workbook_path(workbook: Path) -> Path:
    stat = workbook.stat()
    key = hashlib.sha1(
        f"{workbook.resolve()}|{stat.st_mtime_ns}|{stat.st_size}".encode("utf-8")
    ).hexdigest()[:16]
    WORKBOOK_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
    cache_path = WORKBOOK_CACHE_ROOT / f"{workbook.stem}_{key}.xlsx"
    if cache_path.exists():
        return cache_path

    with ZipFile(workbook, "r") as source, ZipFile(cache_path, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                text = data.decode("utf-8")
                for attr in UNSUPPORTED_SHEET_PROTECTION_ATTRS:
                    text = re.sub(rf'\s{re.escape(attr)}="[^"]*"', "", text)
                data = text.encode("utf-8")
            target.writestr(info, data)
    LOGGER.info("Sanitized workbook protection attributes for openpyxl: %s -> %s", workbook, cache_path)
    return cache_path


def load_workbook_compat(workbook: Path, *, data_only: bool, read_only: bool = False):
    source = sanitized_workbook_path(workbook) if workbook_has_unsupported_protection_attrs(workbook) else workbook
    return load_workbook(source, data_only=data_only, read_only=read_only)


def load_mappings(target: CompareTarget) -> list[FieldMapping]:
    workbook = find_workbook(target.workbook_name)
    if target.loader == "vat_main":
        mappings = load_vat_main_mappings(target, workbook)
    else:
        mappings = load_layout_scan_mappings(target, workbook)
    LOGGER.info("Loaded %s field IDs for %s from %s/%s", len(mappings), target.target_id, workbook, target.sheet_name)
    return mappings


def load_auto_mappings() -> dict[str, list[FieldMapping]]:
    mappings_by_target: dict[str, list[FieldMapping]] = {}
    for target in TARGETS.values():
        try:
            mappings_by_target[target.target_id] = load_mappings(target)
        except FileNotFoundError as exc:
            mappings_by_target[target.target_id] = []
            LOGGER.warning(
                "Skipping auto candidate %s because its mapping workbook is unavailable: %s",
                target.target_id,
                exc,
            )
    return mappings_by_target


def load_vat_main_mappings(target: CompareTarget, workbook: Path) -> list[FieldMapping]:
    wb = load_workbook_compat(workbook, data_only=True, read_only=True)
    ws = wb[target.sheet_name]
    mappings: list[FieldMapping] = []
    seen: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx in (2, 3, 4):
            for value in row:
                field_id = str(value).strip() if value is not None else ""
                if field_id in TEXT_FIELDS and field_id not in seen:
                    seen.add(field_id)
                    mappings.append(make_mapping(target, field_id, DataType.TEXT, row_idx=row_idx))

        line_no = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not line_no:
            continue
        for ordinal, value in enumerate(row[3:7]):
            field_id = str(value).strip() if value is not None else ""
            if FIELD_ID_RE.match(field_id) and field_id not in seen:
                seen.add(field_id)
                mappings.append(
                    make_mapping(
                        target,
                        field_id,
                        infer_data_type(field_id),
                        line_no=line_no,
                        row_idx=row_idx,
                        col_idx=ordinal,
                    )
                )
    wb.close()
    return mappings


def load_layout_scan_mappings(target: CompareTarget, workbook: Path) -> list[FieldMapping]:
    wb = load_workbook_compat(workbook, data_only=True, read_only=True)
    if target.sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {target.sheet_name}; available={wb.sheetnames}")
    ws = wb[target.sheet_name]
    mappings: list[FieldMapping] = []
    seen: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            field_id = str(value).strip() if value is not None else ""
            if not FIELD_ID_RE.match(field_id) or field_id in seen or not should_include_mapping_field(target, field_id):
                continue
            seen.add(field_id)
            mappings.append(
                make_mapping(
                    target,
                    field_id,
                    infer_data_type(field_id),
                    line_no=find_line_no(row, col_idx),
                    row_name=find_row_name(row, col_idx),
                    row_idx=row_idx,
                    col_idx=col_idx,
                )
            )
    wb.close()
    return mappings


def find_line_no(row: tuple[Any, ...], col_idx: int | None = None) -> str:
    search_values = row[: max(col_idx - 1, 0)] if col_idx is not None else row[:4]
    for value in reversed(search_values):
        text = str(value).strip() if value is not None else ""
        if re.fullmatch(r"(FZ)?\d+[A-Za-z]?(\.\d+)?([=＝（(].*)?", text):
            return text
    return ""


def find_row_name(row: tuple[Any, ...], col_idx: int) -> str:
    candidates = []
    for value in row[: max(col_idx - 1, 0)]:
        text = str(value).strip() if value is not None else ""
        if not text or FIELD_ID_RE.match(text) or re.fullmatch(r"\d+(\.\d+)?", text):
            continue
        candidates.append(text)
    return candidates[-1] if candidates else ""


def should_include_mapping_field(target: CompareTarget, field_id: str) -> bool:
    if target.tax_type != "CONSUMPTION_TAX":
        return True
    lower = field_id.lower()
    if lower in {"a", "b", "c"}:
        return False
    return not lower.startswith(CONSUMPTION_TAX_NON_API_PREFIXES)


def infer_data_type(field_id: str) -> DataType:
    lower = field_id.lower()
    if lower.startswith(CONSUMPTION_TAX_TEXT_PREFIXES):
        return DataType.TEXT
    if lower.startswith(CONSUMPTION_TAX_RATE_PREFIXES):
        return DataType.RATE
    if lower in TEXT_FIELDS or re.fullmatch(r"(kpfnsrsbh|kpfdwmc|fwxmmc|pzzl|pzhm)(_\d+)?", lower):
        return DataType.TEXT
    if lower.startswith("fl") or lower.endswith("sl") or "taxrate" in lower or lower.endswith("_rate"):
        return DataType.RATE
    if "cyrs" in lower or lower.endswith("rs") or lower.endswith("count"):
        return DataType.INTEGER
    return DataType.AMOUNT


def data_type_for_target_field(target: CompareTarget, field_id: str, data_type: DataType) -> DataType:
    if target.target_id == "vat_general_appendix5" and field_id in VAT_APPENDIX5_TEXT_FIELDS:
        return DataType.TEXT
    return data_type


def make_mapping(
    target: CompareTarget,
    field_id: str,
    data_type: DataType,
    line_no: str = "",
    row_name: str = "",
    row_idx: int | None = None,
    col_idx: int | None = None,
) -> FieldMapping:
    return FieldMapping(
        tax_type=target.tax_type,
        form_code=target.form_code,
        form_name=target.form_name,
        sheet_name=target.sheet_name,
        field_id=field_id,
        display_name=field_id,
        row_name=row_name,
        line_no=line_no,
        data_type=data_type_for_target_field(target, field_id, data_type),
        web_cell_id=field_id,
        web_row_index=row_idx,
        web_col_index=col_idx,
        api_json_path=f"$.data.{target.tax_code}.{target.api_table}.{field_id}",
    )


def fetch_api_payload(task_id: str) -> dict[str, Any]:
    response = APIClient().fetch_by_task_id(task_id)
    if response.get("error"):
        raise RuntimeError(f"API fetch failed: {response.get('error')}")
    return response


def fetch_api(task_id: str) -> tuple[dict[str, Any], str]:
    response = fetch_api_payload(task_id)
    return response.get("data", {}), response.get("province", "")


def extract_expected_tax_no(api_response: dict[str, Any]) -> str:
    param_json = api_response.get("paramJson") or {}
    if isinstance(param_json, str):
        try:
            param_json = json.loads(param_json)
        except (json.JSONDecodeError, TypeError):
            param_json = {}
    if not isinstance(param_json, dict):
        return ""
    direct_tax_no = str(param_json.get("taxNo") or "").strip()
    if direct_tax_no:
        return direct_tax_no
    cookies = param_json.get("cookies") or {}
    if isinstance(cookies, dict):
        user_info = cookies.get("user_info") or {}
        if isinstance(user_info, dict):
            for key in ("tax_no", "taxNo", "creditCode", "taxpayer_id"):
                value = str(user_info.get(key) or "").strip()
                if value:
                    return value
    return ""


def target_period_range(api_response: dict[str, Any], target: CompareTarget) -> tuple[str, str]:
    """Resolve target tax period as ISO dates, e.g. 2026-05-01/2026-05-31."""

    param_json = api_response.get("paramJson") or {}
    if isinstance(param_json, str):
        try:
            param_json = json.loads(param_json)
        except (TypeError, json.JSONDecodeError):
            param_json = {}
    if isinstance(param_json, dict):
        range_from_number_data = period_range_from_number_data(param_json.get("numberData"), target)
        if range_from_number_data:
            return range_from_number_data
        range_from_period = period_range_from_yyyymm(param_json.get("period"))
        if range_from_period:
            return range_from_period
    range_from_api_fields = period_range_from_api_fields(api_response.get("data", {}).get(target.tax_code, {}))
    if range_from_api_fields:
        return range_from_api_fields
    now = datetime.now()
    last_day = calendar.monthrange(now.year, now.month)[1]
    return f"{now.year:04d}-{now.month:02d}-01", f"{now.year:04d}-{now.month:02d}-{last_day:02d}"


def period_range_from_number_data(number_data: Any, target: CompareTarget) -> tuple[str, str] | None:
    if not isinstance(number_data, list):
        return None
    target_hints = [target.form_name, *target.query_keywords, *target.detail_form_keywords]
    for item in number_data:
        if not isinstance(item, dict):
            continue
        text = " ".join(str(item.get(key) or "") for key in ("name", "pzzlDm", "zsxmMc", "zspmMc"))
        if not any(hint and hint in text for hint in target_hints):
            continue
        start = normalize_iso_date(item.get("skssqq") or item.get("sssqQ") or item.get("SssqQ"))
        end = normalize_iso_date(item.get("skssqz") or item.get("sssqZ") or item.get("SssqZ"))
        if start and end:
            return start, end
    return None


def period_range_from_yyyymm(value: Any) -> tuple[str, str] | None:
    text = re.sub(r"\D", "", str(value or ""))
    if len(text) < 6:
        return None
    year = int(text[:4])
    month = int(text[4:6])
    if not 1 <= month <= 12:
        return None
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def period_range_from_api_fields(tax_data: Any) -> tuple[str, str] | None:
    if not isinstance(tax_data, dict):
        return None
    for key, value in tax_data.items():
        if not any(hint in str(key).lower() for hint in ("nsssq", "sssq", "period")):
            continue
        text = str(value or "")
        dates = re.findall(r"\d{4}[-年]\d{1,2}[-月]\d{1,2}", text)
        normalized = [normalize_iso_date(item) for item in dates]
        normalized = [item for item in normalized if item]
        if len(normalized) >= 2:
            return normalized[0], normalized[1]
    return None


def normalize_iso_date(value: Any) -> str:
    text = str(value or "").strip()
    match = re.search(r"(\d{4})[-年](\d{1,2})[-月](\d{1,2})", text)
    if not match:
        return ""
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


def fetch_current_period_flag(task_id: str, tax_code: str = "") -> bool | None:
    """Return latest '成功保存数据-是否是当期' logInfo as bool.

    None means the task log does not contain this marker, so callers should keep
    the normal declared-query flow.
    """
    try:
        flag = fetch_task_current_period_flag(task_id, tax_code=tax_code, timeout=20)
    except Exception as exc:
        LOGGER.warning("Could not read current-period marker from task logs; continuing normal flow: %s", exc)
        return None
    if flag is None:
        LOGGER.info("No current-period marker found in task execution logs")
    else:
        LOGGER.info("Task current-period marker parsed=%s", flag)
    return flag


def fetch_current_period_flags_for_targets(
    task_id: str,
    targets: list[CompareTarget],
) -> dict[str, bool | None]:
    status_targets = [target for target in targets if target_requires_declaration_status(target)]
    if not status_targets:
        return {}
    try:
        logs = fetch_task_execution_logs(task_id, timeout=20)
    except Exception as exc:
        LOGGER.warning("Could not read current-period markers from task logs; continuing normal flow: %s", exc)
        return {target.target_id: None for target in status_targets}
    flags: dict[str, bool | None] = {}
    for target in status_targets:
        flag = current_period_flag_from_logs(logs, tax_code=target.tax_code)
        flags[target.target_id] = flag
        if flag is None:
            LOGGER.info("No current-period marker found in task execution logs for %s/%s", target.target_id, target.tax_code)
        else:
            LOGGER.info("Task current-period marker parsed for %s/%s: %s", target.target_id, target.tax_code, flag)
    return flags


def target_requires_declaration_status(target: CompareTarget) -> bool:
    return target.tax_type not in {"CBJ_PERSONAL", "CBJ_ANNUAL"}


def supports_undeclared_tax_page(target: CompareTarget) -> bool:
    return target.tax_type in UNDECLARED_SUPPORTED_TAX_TYPES


def ensure_supported_declaration_flow(target: CompareTarget, current_period_flag: bool | None) -> None:
    if current_period_flag is not False:
        return
    if supports_undeclared_tax_page(target):
        return
    raise RuntimeError(
        f"Unsupported undeclared tax-bureau flow for target={target.target_id}, tax_type={target.tax_type}. "
        "This tax type is known to be unfiled, but its undeclared page strategy has not been implemented."
    )


def effective_current_period_flag_for_target(
    target: CompareTarget,
    current_period_flag: bool | None,
    declaration_status_override: str = "",
) -> bool | None:
    """Use the conservative unfiled path when the backend does not expose declaration status."""
    if current_period_flag is not None:
        return current_period_flag
    override_flag = declaration_status_override_flag(declaration_status_override)
    if override_flag is not None and target_requires_declaration_status(target):
        return override_flag
    if current_period_flag is None and target_requires_declaration_status(target):
        return False
    return current_period_flag


def declaration_status_override_flag(value: str) -> bool | None:
    text = str(value or "").strip().lower()
    if text == "filed":
        return True
    if text == "unfiled":
        return False
    return None


def fallback_status_flag(
    raw_current_period_flag: bool | None,
    target_current_period_flag: bool | None,
    declaration_status_override: str = "",
) -> bool | None:
    if raw_current_period_flag is not None:
        return raw_current_period_flag
    if declaration_status_override_flag(declaration_status_override) is not None:
        return target_current_period_flag
    return None


def current_period_flag_source(raw_current_period_flag: bool | None, declaration_status_override: str = "") -> str:
    if raw_current_period_flag is not None:
        return "execution_log"
    if declaration_status_override_flag(declaration_status_override) is not None:
        return "override"
    return "default"


def current_period_flag_message(flag: bool | None) -> str:
    if flag is True:
        return "filed"
    if flag is False:
        return "unfiled"
    return "unknown"


def can_fallback_to_declared_query_after_undeclared_unavailable(current_period_flag: bool | None) -> bool:
    return current_period_flag is None


def can_fallback_to_declared_query_after_already_declared_conflict(current_period_flag: bool | None) -> bool:
    return current_period_flag is None


def api_value(api_by_tax: dict[str, Any], target: CompareTarget, field_id: str) -> Any:
    tax_data = api_by_tax.get(target.tax_code, {})
    if not isinstance(tax_data, dict):
        return None
    direct_value = raw_api_value(tax_data, target, field_id)
    fallback_value = fallback_api_value(tax_data, target, field_id, direct_value)
    return fallback_value if fallback_value is not None else direct_value


def raw_api_value(tax_data: dict[str, Any], target: CompareTarget, field_id: str) -> Any:
    table_key = f"{target.api_table}.{field_id}" if target.api_table else field_id
    if table_key in tax_data:
        return tax_data[table_key]
    return tax_data.get(field_id)


VAT_GENERAL_APPENDIX1_API_VALUE_ALIASES = {
    "hjxse_6": ("asysljsxse_jzjtxm_bys", "yshwxse_jzjtxm_bys"),
}


def fallback_api_value(tax_data: dict[str, Any], target: CompareTarget, field_id: str, direct_value: Any) -> Any:
    if target.target_id != "vat_general_appendix1":
        return None
    if field_id not in VAT_GENERAL_APPENDIX1_API_VALUE_ALIASES:
        return None
    if is_effective_api_value(direct_value) and not api_amount_value_is_zero(direct_value):
        return None
    for source_field_id in VAT_GENERAL_APPENDIX1_API_VALUE_ALIASES[field_id]:
        source_value = tax_data.get(source_field_id)
        if source_value is None:
            source_value = tax_data.get(f"{TARGETS['vat_general_main'].api_table}.{source_field_id}")
        if is_effective_api_value(source_value) and not api_amount_value_is_zero(source_value):
            return source_value
    return None


def api_amount_value_is_zero(value: Any) -> bool:
    parsed = parse_amount(value)
    return parsed == Decimal("0") if parsed is not None else False


def api_has_comparable_value(api_by_tax: dict[str, Any], target: CompareTarget, field_id: str) -> bool:
    value = api_value(api_by_tax, target, field_id)
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def mappings_for_comparison(
    target: CompareTarget,
    mappings: list[FieldMapping],
    api_by_tax: dict[str, Any],
) -> list[FieldMapping]:
    return [
        mapping
        for mapping in mappings
        if not should_exclude_from_comparison(target, mapping.field_id)
        and api_has_comparable_value(api_by_tax, target, mapping.field_id)
    ]


def should_exclude_from_comparison(target: CompareTarget, field_id: str) -> bool:
    if target.target_id == "culture_fee_main" and field_id == "fl_bys":
        return True
    return False


def connect_browser(args: argparse.Namespace) -> BrowserManager:
    bm = BrowserManager()
    if args.mode in {"auto", "connect"}:
        try:
            bm.connect_cdp(args.cdp_port)
            return bm
        except Exception as exc:
            if args.mode == "connect":
                raise
            LOGGER.info("CDP connect failed, launching Chrome instead: %s", exc)
            bm.close()
            bm = BrowserManager()
    bm.launch_with_extension(
        {
            "user_data_dir": args.user_data_dir,
            "cdp_port": args.cdp_port,
            "plugin_path": args.plugin_path,
            "chrome_path": getattr(args, "chrome_path", r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        }
    )
    return bm


def wait_for_chanjet_login(bm: BrowserManager, timeout: int) -> Any:
    start = time.time()
    while time.time() - start < timeout:
        page = bm.find_page_by_url("chanjet.com")
        if page:
            title = page.title()
            if title and "登录" not in title and "Login" not in title:
                return page
        time.sleep(3)
    raise TimeoutError("Chanjet login was not detected before timeout")


def get_web_config(config_root: str, tax_type: str):
    registry = TaxTypeRegistry()
    registry.load_all_from_dir(config_root)
    if tax_type in registry.list_all():
        return registry.get(tax_type).forms[0].web_config
    return registry.get("VAT_SMALL_SCALE").forms[0].web_config


def find_existing_tax_page(bm: BrowserManager, province: str, expected_tax_no: str = ""):
    if not province:
        return None
    expected_tax_no = (expected_tax_no or "").strip()
    if not expected_tax_no:
        LOGGER.info("Skipping existing tax bureau page reuse: expected taxpayer id is unavailable")
        return None
    host = f"etax.{province}.chinatax.gov.cn"
    scored_candidates = []
    for page in bm.get_all_pages():
        try:
            if urllib.parse.urlparse(page.url or "").hostname != host:
                continue
            text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
            url = page.url or ""
            lower_url = url.lower()
            if is_loading_page_url(url) or is_tpass_login_page_url(url):
                LOGGER.info(
                    "Skipping existing tax bureau page because it is not usable; task login is required: %s",
                    url,
                )
                continue
            if not text.strip():
                LOGGER.info(
                    "Skipping existing tax bureau page because page text is empty; task login is required: %s",
                    url,
                )
                continue
            if "loginb" in lower_url:
                LOGGER.info(
                    "Skipping existing tax bureau portal page; task login is required to refresh taxpayer session: %s",
                    url,
                )
                continue
            if expected_tax_no not in text:
                LOGGER.info(
                    "Skipping existing tax bureau page because taxpayer id does not match: expected=%s url=%s",
                    expected_tax_no,
                    page.url,
                )
                continue
            detail_context = (
                "sbxxcx/detail" in lower_url
                or "sbxxcxxq" in lower_url
                or ("申报信息查询详情" in text and "主附表表单" in text)
            )
            if any(hint in lower_url for hint in QUERY_URL_HINTS) and not detail_context:
                LOGGER.info(
                    "Skipping existing declaration query list page; task login is required to confirm taxpayer context: %s",
                    url,
                )
                continue
            logged_in_hints = (
                "\u6211\u8981\u67e5\u8be2",
                "\u6211\u8981\u529e\u7a0e",
                "\u672c\u671f\u5e94\u7533\u62a5",
                "\u6211\u7684\u5f85\u529e",
                "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801",
                "\u7eb3\u7a0e\u4eba\u8bc6\u522b\u53f7",
                "\u8ddd\u672c\u6708\u5f81\u671f\u7ed3\u675f",
                "\u7533\u62a5\u4fe1\u606f\u67e5\u8be2",
                "\u7533\u62a5\u4fe1\u606f\u67e5\u8be2\u8be6\u60c5",
                "\u8fd4\u56de",
                "\u589e\u503c\u7a0e",
                "\u63d0\u4ea4\u7533\u62a5",
                "\u7a0e\u52a1\u6570\u5b57\u8d26\u6237",
                "\u8d26\u6237\u67e5\u8be2",
            )
            if not any(hint in text for hint in logged_in_hints):
                continue
            score = 0
            if detail_context:
                score += 100
            if "\u7533\u62a5\u4fe1\u606f\u67e5\u8be2" in text:
                score += 80
            if "\u6211\u8981\u67e5\u8be2" in text:
                score += 40
            if "\u672c\u671f\u5e94\u7533\u62a5" in text:
                score += 30
            if "\u7a0e\u52a1\u6570\u5b57\u8d26\u6237" in text or "\u8d26\u6237\u67e5\u8be2" in text:
                score += 25
            if "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801" in text or "\u7eb3\u7a0e\u4eba\u8bc6\u522b\u53f7" in text:
                score += 20
            if "loginb" in url.lower():
                score -= 5
            scored_candidates.append((score, page))
        except Exception:
            continue
    if scored_candidates:
        scored_candidates.sort(key=lambda item: item[0], reverse=True)
        return scored_candidates[0][1]
    candidates = []
    for page in bm.get_all_pages():
        try:
            url = page.url or ""
            if urllib.parse.urlparse(url).hostname != host:
                continue
            if "loginb" in url.lower() or is_loading_page_url(url) or is_tpass_login_page_url(url):
                continue
            text = page.evaluate("document.body ? document.body.innerText.slice(0, 3000) : ''")
            if not text.strip():
                continue
            if expected_tax_no not in text:
                continue
            lower_url = url.lower()
            detail_context = (
                "sbxxcx/detail" in lower_url
                or "sbxxcxxq" in lower_url
                or ("申报信息查询详情" in text and "主附表表单" in text)
            )
            if any(hint in lower_url for hint in QUERY_URL_HINTS) and not detail_context:
                continue
            if any(
                hint in text
                for hint in (
                    "我要查询",
                    "我要办税",
                    "纳税人识别号",
                    "统一社会信用代码",
                    "申报信息查询",
                    "申报信息查询详情",
                    "首页",
                    "返回",
                    "增值税",
                    "提交申报",
                    "税务数字账户",
                    "账户查询",
                )
            ):
                candidates.append(page)
        except Exception:
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            continue
    for page in candidates:
        if "sbxxcxxq" in (page.url or ""):
            return page
    if candidates:
        return candidates[0]
    return None


def login_tax_page_for_task(
    bm: BrowserManager,
    chanjet_page,
    args: argparse.Namespace,
    province: str,
    expected_tax_no: str = "",
):
    flow = TaskLoginFlow(
        bm,
        timeout=args.tax_timeout,
        login_strategy=getattr(args, "tax_login_strategy", "plugin_first"),
    )
    tax_page, info = flow.login(chanjet_page, args.task_id)
    LOGGER.info(
        "Logged into tax bureau: province=%s inner_task_id=%s url=%s",
        info.province,
        info.inner_task_id,
        tax_page.url,
    )
    if info.province and info.province != province:
        LOGGER.info(
            "Using task-login province override for tax bureau navigation: %s -> %s",
            province,
            info.province,
        )
        province = info.province
    if expected_tax_no and info.tax_no and expected_tax_no != info.tax_no:
        LOGGER.warning(
            "Task login taxpayer id differs from API paramJson: api=%s login=%s",
            expected_tax_no,
            info.tax_no,
        )
    return tax_page, province, info


def is_query_page(page) -> bool:
    url = (page.url or "").lower()
    if is_declaration_detail_page(page):
        return False
    return any(hint in url for hint in QUERY_URL_HINTS)


def is_declaration_detail_page(page) -> bool:
    url = (page.url or "").lower()
    if "sbxxcx/detail" in url or "sbxxcxxq" in url:
        return True
    try:
        text = page.evaluate("document.body ? document.body.innerText.slice(0, 1000) : ''")
    except Exception:
        return False
    return "申报信息查询详情" in text and "主附表表单" in text


def declaration_detail_matches_target(page, target: CompareTarget) -> bool:
    try:
        text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
    except Exception:
        return False
    normalized = re.sub(r"\s+", "", text)
    keyword_sets = [target.query_keywords, target.detail_form_keywords, (target.form_name,)]
    for keywords in keyword_sets:
        clean_keywords = [re.sub(r"\s+", "", str(keyword)) for keyword in keywords if str(keyword).strip()]
        if clean_keywords and all(keyword in normalized for keyword in clean_keywords):
            return True
    return False


def ensure_target_page(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
    web_config,
    prefer_detail_form_switch: bool = False,
    api_response: dict[str, Any] | None = None,
):
    target_query_period = target_period_range(api_response, target) if api_response else None
    tried_detail_form_switch = False
    if not is_target_detail_page(page, target, mappings):
        if is_declaration_detail_page(page):
            if declaration_detail_matches_target(page, target):
                LOGGER.info("Reusing current declaration detail page for %s", target.target_id)
            elif prefer_detail_form_switch and target.detail_form_keywords:
                tried_detail_form_switch = True
                LOGGER.info("Trying in-detail form switch for %s before returning to declaration query", target.target_id)
            else:
                LOGGER.info("Current declaration detail does not match %s; returning to declaration query", target.target_id)
                page = navigate_to_query_page_robust(page, web_config)
        elif not is_query_page(page):
            LOGGER.info("Navigating to declaration information query page")
            page = navigate_to_query_page_robust(page, web_config)
        if not is_declaration_detail_page(page) and not is_target_detail_page(page, target, mappings):
            if not is_ready_declaration_query_page(page):
                host = extract_etax_host(page.url or "")
                recovered_page = wait_for_declaration_query_page(page, host, timeout=45) if host else None
                if recovered_page:
                    page = recovered_page
                else:
                    raise RuntimeError(f"Could not navigate to declaration query page before opening target={target.target_id}; url={page.url}")
            LOGGER.info("Opening declaration row for %s", target.target_id)
            opened_page = click_declaration_row(
                page,
                target.query_keywords,
                target_query_period,
                allow_period_fallback=target.target_id == "cit_a_main",
            )
            if opened_page:
                page = opened_page

    page.wait_for_load_state("domcontentloaded", timeout=15000)
    time.sleep(2)
    if target.detail_form_keywords:
        select_result = select_detail_form(page, target.detail_form_keywords)
        if select_result == "not_found":
            if tried_detail_form_switch:
                LOGGER.info("In-detail form switch did not find %s; reopening from declaration query", target.target_id)
                page = navigate_to_query_page_robust(page, web_config)
                if not is_ready_declaration_query_page(page):
                    raise RuntimeError(f"Could not recover declaration query page for target={target.target_id}; url={page.url}")
                opened_page = click_declaration_row(
                    page,
                    target.query_keywords,
                    target_query_period,
                    allow_period_fallback=target.target_id == "cit_a_main",
                )
                if opened_page:
                    page = opened_page
                page.wait_for_load_state("domcontentloaded", timeout=15000)
                time.sleep(2)
                select_result = select_detail_form(page, target.detail_form_keywords)
                if select_result != "not_found":
                    deadline = time.time() + 30
                    while time.time() < deadline:
                        if is_target_detail_page(page, target, mappings):
                            return page
                        time.sleep(1)
            if is_target_detail_page(page, target, mappings):
                LOGGER.info(
                    "Detail form selector was not found for %s, but current page already matches target",
                    target.target_id,
                )
                return page
            raise RuntimeError(
                f"Detail form was not found for target={target.target_id}, "
                f"keywords={target.detail_form_keywords}"
            )
        deadline = time.time() + 30
        while time.time() < deadline:
            if is_target_detail_page(page, target, mappings):
                return page
            time.sleep(1)
        raise RuntimeError(
            f"Detail form was selected but target page was not confirmed for "
            f"target={target.target_id}, keywords={target.detail_form_keywords}, result={select_result}"
        )
    if not is_target_detail_page(page, target, mappings):
        raise RuntimeError(
            f"Target page was not confirmed before extraction for target={target.target_id}; "
            f"url={page.url}"
        )
    return page


def can_switch_detail_form_between(previous: CompareTarget | None, current: CompareTarget) -> bool:
    if previous is None:
        return False
    if not current.detail_form_keywords:
        return False
    return (
        previous.tax_code == current.tax_code
        and previous.tax_type == current.tax_type
        and previous.query_keywords == current.query_keywords
    )


def find_context_tax_page(page, required_hints: tuple[str, ...]):
    try:
        current_host = urllib.parse.urlparse(page.url or "").hostname
    except Exception:
        current_host = ""
    if not current_host:
        return None
    for candidate in list(page.context.pages):
        try:
            url = candidate.url or ""
            if urllib.parse.urlparse(url).hostname != current_host:
                continue
            if "loading" in url.lower():
                continue
            text = candidate.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
            if all(hint in text for hint in required_hints):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def wait_for_context_tax_page(page, required_hints: tuple[str, ...], timeout: int = 20):
    deadline = time.time() + timeout
    while time.time() < deadline:
        candidate = find_context_tax_page(page, required_hints)
        if candidate:
            return candidate
        time.sleep(1)
    return None


def find_context_query_page(page):
    try:
        current_host = urllib.parse.urlparse(page.url or "").hostname
    except Exception:
        current_host = ""
    if not current_host:
        return None
    for candidate in list(page.context.pages):
        try:
            if urllib.parse.urlparse(candidate.url or "").hostname != current_host:
                continue
            if is_query_page(candidate):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def find_context_digital_account_page(page):
    try:
        current_host = urllib.parse.urlparse(page.url or "").hostname
    except Exception:
        current_host = ""
    if not current_host:
        return None
    for candidate in list(page.context.pages):
        try:
            url = candidate.url or ""
            lower_url = url.lower()
            if urllib.parse.urlparse(url).hostname != current_host:
                continue
            if "loading" in lower_url:
                continue
            if "/szzh/" not in lower_url and "/szzh/?" not in lower_url:
                continue
            if "/szzh/?" in lower_url:
                candidate.bring_to_front()
                return candidate
            text = candidate.evaluate("document.body ? document.body.innerText.slice(0, 2000) : ''")
            if "账户查询" in text or "税务数字账户" in text:
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def wait_for_context_digital_account_page(page, timeout: int = 30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        candidate = find_context_digital_account_page(page)
        if candidate:
            return candidate
        time.sleep(1)
    return None


def find_context_page_by_url_fragment(page, fragment: str):
    for candidate in list(page.context.pages):
        try:
            if fragment in (candidate.url or ""):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def wait_for_context_page_by_url_fragment(page, fragment: str, timeout: int = 30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        candidate = find_context_page_by_url_fragment(page, fragment)
        if candidate:
            return candidate
        time.sleep(1)
    return None


def wait_for_context_query_page(page, timeout: int = 30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if is_query_page(page):
            return page
        candidate = find_context_query_page(page)
        if candidate:
            return candidate
        time.sleep(1)
    return None


def page_contains_hints(page, hints: tuple[str, ...]) -> bool:
    try:
        text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
    except Exception:
        return False
    return all(hint in text for hint in hints)


def open_declaration_query_direct(page, host: str, timeout: int = 30):
    target_url = f"https://{host}/szzh/zhcx/sbxx/sbxxcx"
    LOGGER.info("Opening declaration query page directly: %s", target_url)
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        if "interrupted by another navigation" not in str(exc):
            LOGGER.info("Direct declaration query navigation failed: %s", exc)
        else:
            LOGGER.info("Direct declaration query navigation was interrupted; waiting for redirected page")
    try:
        page.wait_for_url("**/szzh/zhcx/sbxx/sbxxcx**", timeout=timeout * 1000)
    except Exception:
        pass
    query_page = wait_for_context_query_page(page, timeout=5)
    if query_page:
        return query_page
    if "loading" in (page.url or "").lower():
        LOGGER.info("Declaration query is still on loading page; waiting for redirect: %s", page.url)
        return wait_for_context_query_page(page, timeout=timeout)
    return None


def navigate_to_query_page(page, web_config):
    url = page.url or ""
    host_match = re.search(r"https://(etax\.[^/]+)", url)
    if host_match:
        host = host_match.group(1)
        province_match = re.search(r"etax\.([^.]+)\.chinatax\.gov\.cn", host)
        province = province_match.group(1) if province_match else ""
        if province == "sichuan":
            query_page = open_declaration_query_direct(page, host, timeout=45)
            if query_page:
                return query_page
            LOGGER.info("Sichuan direct declaration query did not become usable; falling back to tax digital account")
        if "/szzh/" not in (page.url or ""):
            szzh_url = f"https://{host}/szc/szzh/sjswszzh/spHandler?cdlj=/szzh/szzh/"
            for attempt in range(1, 3):
                szzh_page = wait_for_context_digital_account_page(page, timeout=1)
                if szzh_page:
                    page = szzh_page
                    break
                LOGGER.info("Opening tax digital account before declaration query: %s (attempt %s/2)", szzh_url, attempt)
                try:
                    page.goto(szzh_url, wait_until="domcontentloaded", timeout=30000)
                except Exception as exc:
                    if "interrupted by another navigation" not in str(exc) and attempt == 2:
                        raise
                    LOGGER.info("Tax digital account navigation was interrupted; waiting for a usable account page")
                szzh_page = wait_for_context_digital_account_page(page, timeout=12)
                if szzh_page:
                    page = szzh_page
                    break
                LOGGER.warning("Tax digital account page did not become usable after attempt %s", attempt)
                continue
                portal_page = wait_for_context_tax_page(page, ("税务数字账户",), timeout=15)
                if portal_page:
                    page = portal_page
                LOGGER.info("Opening tax digital account before declaration query: %s (attempt %s/3)", szzh_url, attempt)
                try:
                    page.goto(szzh_url, wait_until="domcontentloaded", timeout=30000)
                except Exception as exc:
                    if "interrupted by another navigation" not in str(exc) and attempt == 3:
                        raise
                    LOGGER.info("Tax digital account navigation was interrupted; waiting for a usable account page")
                szzh_page = wait_for_context_digital_account_page(page, timeout=12)
                if szzh_page:
                    page = szzh_page
                    break
                LOGGER.warning("Tax digital account page did not become usable after attempt %s", attempt)
            else:
                szzh_page = wait_for_context_digital_account_page(page, timeout=15)
                if szzh_page:
                    page = szzh_page
                else:
                    raise RuntimeError(f"Could not open tax digital account page; current url={page.url}")
        target_url = f"https://{host}/szzh/zhcx/sbxx/sbxxcx"
        LOGGER.info("Opening declaration query page directly: %s", target_url)
        last_nav_error = None
        for attempt in range(1, 4):
            try:
                page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
                last_nav_error = None
            except Exception as exc:
                last_nav_error = exc
                if "interrupted by another navigation" not in str(exc) or attempt == 3:
                    raise
                LOGGER.info("Declaration query navigation was interrupted; retrying after root page settles")
                query_page = wait_for_context_query_page(page, timeout=8)
                if query_page:
                    return query_page
                settled_page = wait_for_context_digital_account_page(page, timeout=10)
                if settled_page:
                    page = settled_page
                continue
            query_page = wait_for_context_query_page(page, timeout=30)
            if query_page:
                return query_page
            settled_page = wait_for_context_digital_account_page(page, timeout=5)
            if settled_page:
                page = settled_page
        if last_nav_error:
            raise last_nav_error
        if "loading" in (page.url or "").lower():
            query_page = open_declaration_query_direct(page, host, timeout=45)
            if query_page:
                return query_page
            raise RuntimeError(f"Declaration query page did not open; still on loading page: {page.url}")
    portal_page = find_context_tax_page(page, ("我要查询",))
    if portal_page:
        page = portal_page
    NavigationEngine(page).navigate_to_form(web_config)
    time.sleep(3)
    if is_query_page(page):
        return page
    query_page = find_context_query_page(page)
    if query_page:
        return query_page
    return page


def extract_etax_host(url: str) -> str:
    match = re.search(r"https://(etax\.[^/]+)", url or "")
    return match.group(1) if match else ""


def is_undeclared_vat_page_url(url: str) -> bool:
    return UNDECLARED_VAT_GENERAL_PATH.split("#", 1)[0] in (url or "")


def is_loading_page_url(url: str) -> bool:
    return "/loading" in (url or "").lower()


def is_tpass_login_page_url(url: str) -> bool:
    lower_url = (url or "").lower()
    return "tpass." in lower_url and "#/login" in lower_url


class DeclarationQueryAuthError(RuntimeError):
    """Raised when declaration-query recovery lands on the unified login page."""


def declaration_query_auth_error_for_url(url: str, host: str = "") -> DeclarationQueryAuthError:
    detail = f"host={host}; url={url}" if host else f"url={url}"
    return DeclarationQueryAuthError(
        "Tax bureau login state or digital account authentication is not ready: "
        f"declaration query did not become usable before timeout; {detail}. "
        "Please retry through EtaxPlugin login."
    )


class UndeclaredTaxAlreadyDeclaredError(RuntimeError):
    """Raised when an undeclared entry reports that the period was already declared."""


class UndeclaredTaxTargetUnavailableError(RuntimeError):
    """Raised when the inferred undeclared target is absent from the tax home."""


UndeclaredVatAlreadyDeclaredError = UndeclaredTaxAlreadyDeclaredError


def etax_hostname(host_or_url: str) -> str:
    if not host_or_url:
        return ""
    value = host_or_url
    if "://" not in value:
        value = f"https://{value}"
    try:
        return urllib.parse.urlparse(value).hostname or ""
    except Exception:
        return ""


def context_pages_for_etax_host(page, host: str):
    expected_hostname = etax_hostname(host)
    if not expected_hostname:
        return []
    pages = []
    for candidate in list(page.context.pages):
        try:
            candidate_url = candidate.url or ""
            if etax_hostname(candidate_url) == expected_hostname:
                pages.append(candidate)
        except Exception:
            continue
    return pages


def find_context_query_page_for_host(page, host: str):
    for candidate in context_pages_for_etax_host(page, host):
        try:
            if is_ready_declaration_query_page(candidate):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def find_context_digital_account_page_for_host(page, host: str):
    for candidate in context_pages_for_etax_host(page, host):
        try:
            url = candidate.url or ""
            lower_url = url.lower()
            if "loading" in lower_url:
                continue
            if "/szzh/" not in lower_url and "/szzh/?" not in lower_url:
                continue
            if "/szc/szzh/" in lower_url and not page_has_visible_loading(candidate):
                candidate.bring_to_front()
                return candidate
            if lower_url.rstrip("/").endswith("/szzh") or "/szzh/?" in lower_url:
                candidate.bring_to_front()
                return candidate
            text = candidate.evaluate("document.body ? document.body.innerText.slice(0, 3000) : ''")
            if any(hint in text for hint in TAX_DIGITAL_ACCOUNT_HINTS):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def page_has_visible_loading(page) -> bool:
    try:
        return bool(
            page.evaluate(
                """() => {
                    const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                    const loadingSelectors = [
                        '.el-loading-mask',
                        '.ant-spin-spinning',
                        '.t-loading',
                        '.vxe-loading',
                        '[class*="loading"]',
                        '[class*="Loading"]'
                    ];
                    for (const selector of loadingSelectors) {
                        if (Array.from(document.querySelectorAll(selector)).some(visible)) return true;
                    }
                    const text = String(document.body && document.body.innerText || '').replace(/\\s+/g, '');
                    return text === '加载中' || text.includes('正在加载') || text.includes('请稍候');
                }"""
            )
        )
    except Exception:
        return False


def wait_for_declaration_query_page(page, host: str, timeout: int = 60):
    deadline = time.time() + timeout
    last_logged_url = ""
    last_log_at = 0.0
    loading_started_at = None
    while time.time() < deadline:
        try:
            if is_ready_declaration_query_page(page):
                return page
        except Exception:
            pass
        candidate = find_context_query_page_for_host(page, host)
        if candidate:
            return candidate
        try:
            current_url = page.url or ""
        except Exception:
            current_url = ""
        if is_tpass_login_page_url(current_url):
            LOGGER.info("Declaration query wait stopped on tpass login page: %s", current_url)
            return None
        now = time.time()
        if is_loading_page_url(current_url):
            if loading_started_at is None:
                loading_started_at = now
            elif now - loading_started_at >= min(25, timeout):
                LOGGER.info("Declaration query stayed on loading page for %.1fs; trying recovery path", now - loading_started_at)
                return None
        else:
            loading_started_at = None
        if current_url and ("loading" in current_url.lower() or current_url != last_logged_url) and now - last_log_at >= 5:
            LOGGER.info("Waiting for declaration query page; current url=%s", current_url)
            last_logged_url = current_url
            last_log_at = now
        time.sleep(1)
    return None


def is_ready_declaration_query_page(page) -> bool:
    if not is_query_page(page):
        return False
    try:
        url = page.url or ""
    except Exception:
        return False
    if "loading" in url.lower():
        return False
    try:
        text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
    except Exception:
        return False
    if not text.strip():
        return False
    ready_hints = (
        "\u7533\u62a5\u4fe1\u606f\u67e5\u8be2",
        "\u7a0e\u6b3e\u6240\u5c5e\u671f",
        "\u5f81\u6536\u9879\u76ee",
        "\u7533\u62a5\u8868",
        "\u67e5\u8be2",
    )
    if not any(hint in text for hint in ready_hints):
        return False
    if "/szzh/zhcx/sbxx/sbxxcx" in url:
        try:
            return bool(
                page.evaluate(
                    """() => {
                        function findComponents(el, out = []) {
                            if (el.__vue__) out.push(el.__vue__);
                            if (el.__vueParentComponent) out.push(el.__vueParentComponent);
                            for (const child of el.children || []) findComponents(child, out);
                            return out;
                        }
                        const comp = findComponents(document.body).find((v) => {
                            const name = (v.$options && v.$options.name) || (v.type && v.type.name) || v.name || '';
                            return name === 'sbxxcx';
                        });
                        const dataRoot = comp && (comp.$data || comp.ctx || comp.setupState || {});
                        return !!(dataRoot && dataRoot.formData);
                    }"""
                )
            )
        except Exception:
            return False
    return True


def wait_for_digital_account_page_for_host(page, host: str, timeout: int = 30, soft_timeout: int = 12):
    start = time.time()
    deadline = time.time() + timeout
    while time.time() < deadline:
        query_page = find_context_query_page_for_host(page, host)
        if query_page:
            LOGGER.info("Declaration query became ready while waiting for tax digital account: %.1fs", time.time() - start)
            return query_page
        candidate = find_context_digital_account_page_for_host(page, host)
        if candidate:
            LOGGER.info("Tax digital account became usable: %.1fs", time.time() - start)
            return candidate
        elapsed = time.time() - start
        if elapsed >= soft_timeout:
            LOGGER.info(
                "Tax digital account readiness not detected after %.1fs; continuing with direct declaration query",
                elapsed,
            )
            return None
        time.sleep(0.35)
    LOGGER.info("Tax digital account wait timed out after %.1fs", time.time() - start)
    return None


def open_declaration_query_with_wait(page, host: str, timeout: int = 60):
    target_url = f"https://{host}/szzh/zhcx/sbxx/sbxxcx"
    LOGGER.info("Opening declaration query page directly: %s", target_url)
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        message = str(exc)
        if "interrupted by another navigation" in message:
            LOGGER.info("Declaration query navigation was interrupted; waiting for redirected page")
        else:
            LOGGER.info("Declaration query navigation did not settle immediately: %s", exc)
    return wait_for_declaration_query_page(page, host, timeout=timeout)


def open_digital_account_with_wait(page, host: str, timeout: int = 30):
    existing_query_page = find_context_query_page_for_host(page, host)
    if existing_query_page:
        return existing_query_page
    existing_page = find_context_digital_account_page_for_host(page, host)
    if existing_page:
        return existing_page
    szzh_url = f"https://{host}/szc/szzh/sjswszzh/spHandler?cdlj=/szzh/szzh/"
    LOGGER.info("Opening tax digital account before declaration query: %s", szzh_url)
    try:
        page.goto(szzh_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        message = str(exc)
        if "interrupted by another navigation" in message:
            LOGGER.info("Tax digital account navigation was interrupted; waiting for redirected page")
        else:
            LOGGER.info("Tax digital account navigation did not settle immediately: %s", exc)
    return wait_for_digital_account_page_for_host(page, host, timeout=timeout)


def is_tax_portal_page(page) -> bool:
    try:
        url = page.url or ""
    except Exception:
        return False
    if is_loading_page_url(url) or is_tpass_login_page_url(url):
        return False
    try:
        text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
    except Exception:
        return False
    return any(hint in text for hint in TAX_PORTAL_HINTS)


def wait_for_tax_portal_page(page, host: str, timeout: int = 30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            current_url = page.url or ""
        except Exception:
            current_url = ""
        if is_tpass_login_page_url(current_url):
            LOGGER.info("Tax portal wait stopped on tpass login page: %s", current_url)
            return None
        if extract_etax_host(current_url) == host and is_tax_portal_page(page):
            return page
        time.sleep(1)
    return page if is_tax_portal_page(page) else None


def find_context_tax_portal_page_for_host(page, host: str):
    for candidate in context_pages_for_etax_host(page, host):
        try:
            if is_tax_portal_page(candidate):
                candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


def reset_tax_bureau_tab_to_loginb(page, host: str, timeout: int = 30):
    login_url = f"https://{host}/loginb/"
    LOGGER.info("Resetting current tax-bureau tab before declaration query recovery: %s", login_url)
    try:
        page.goto(login_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        message = str(exc)
        if "interrupted by another navigation" in message:
            LOGGER.info("Tax-bureau reset navigation was interrupted; waiting for redirected page")
        else:
            LOGGER.info("Tax-bureau reset navigation did not settle immediately: %s", exc)
    return wait_for_tax_portal_page(page, host, timeout=timeout)


def navigate_to_query_from_tax_portal(page, host: str, timeout: int = 60):
    if not is_tax_portal_page(page):
        return None
    LOGGER.info("Navigating to declaration query from tax-bureau portal menu")
    try:
        result = page.evaluate(
            """async () => {
                const normalize = (value) => String(value || '').replace(/\\s+/g, '');
                const wait = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
                const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                const dispatchPointer = (el) => {
                    for (const type of ['mouseover', 'mouseenter', 'mousemove']) {
                        el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
                    }
                };
                const clickByLabels = (labels) => {
                    const wanted = labels.map(normalize);
                    const nodes = Array.from(document.querySelectorAll(
                        'a, button, li, span, div, [role=button], [role=menuitem], .el-menu-item, .ant-menu-item'
                    )).filter((el) => {
                        if (!visible(el)) return false;
                        const text = normalize(el.innerText || el.textContent || el.getAttribute('title'));
                        return text && text.length <= 120;
                    });
                    const exact = nodes.find((el) => wanted.includes(normalize(el.innerText || el.textContent || el.getAttribute('title'))));
                    const partial = exact || nodes.find((el) => {
                        const text = normalize(el.innerText || el.textContent || el.getAttribute('title'));
                        return wanted.some((label) => text.includes(label));
                    });
                    if (!partial) return '';
                    partial.scrollIntoView({ block: 'center', inline: 'center' });
                    dispatchPointer(partial);
                    partial.click();
                    return normalize(partial.innerText || partial.textContent || partial.getAttribute('title'));
                };

                const clicked = [];
                const first = clickByLabels(['我要查询']);
                if (first) clicked.push(first);
                await wait(1000);
                const second = clickByLabels(['一户式查询', '申报信息查询']);
                if (second) clicked.push(second);
                await wait(1000);
                if (!clicked.some((item) => item.includes('申报信息查询'))) {
                    const third = clickByLabels(['申报信息查询']);
                    if (third) clicked.push(third);
                }
                await wait(5000);
                return clicked.join('>');
            }"""
        )
        LOGGER.info("Tax portal declaration query menu result: %s", result or "not_clicked")
    except Exception as exc:
        LOGGER.info("Tax portal declaration query menu navigation failed: %s", exc)
        return None
    return wait_for_declaration_query_page(page, host, timeout=timeout)


def open_declaration_query_via_sp_handler(
    page,
    host: str,
    timeout: int = 60,
    *,
    fail_on_tpass: bool = False,
):
    target_url = f"https://{host}/szc/szzh/sjswszzh/spHandler?cdlj=/szzh/zhcx/sbxx/sbxxcx"
    LOGGER.info("Opening declaration query through tax digital account handler: %s", target_url)
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        message = str(exc)
        if "interrupted by another navigation" in message:
            LOGGER.info("Declaration query handler navigation was interrupted; waiting for redirected page")
        else:
            LOGGER.info("Declaration query handler navigation did not settle immediately: %s", exc)
    if is_tpass_login_page_url(page.url or ""):
        message = (
            "税局登录态或数字账户认证已失效：申报查询入口跳回统一登录页，"
            "无法进入申报信息查询页。请人工重新进入对应税局/数字账户后重试。"
        )
        LOGGER.info("Declaration query handler reached tpass login page; recovery path is not authenticated")
        if fail_on_tpass:
            raise DeclarationQueryAuthError(message)
        return None
    return wait_for_declaration_query_page(page, host, timeout=timeout)


def recover_declaration_query_in_current_tab(page, host: str, timeout: int = 60):
    portal_page = find_context_tax_portal_page_for_host(page, host)
    if not portal_page:
        portal_page = reset_tax_bureau_tab_to_loginb(page, host, timeout=30)
    if not portal_page:
        return None

    query_page = navigate_to_query_from_tax_portal(portal_page, host, timeout=timeout)
    if query_page:
        return query_page

    query_page = open_declaration_query_with_wait(portal_page, host, timeout=timeout)
    if query_page:
        return query_page

    query_page = open_declaration_query_via_sp_handler(portal_page, host, timeout=timeout, fail_on_tpass=True)
    if query_page:
        return query_page
    return None


def open_fresh_tax_bureau_tab(page, host: str):
    try:
        fresh_page = page.context.new_page()
    except Exception as exc:
        LOGGER.info("Could not open a fresh tax-bureau tab for recovery: %s", exc)
        return None

    login_url = f"https://{host}/loginb/"
    LOGGER.info("Opening fresh tax-bureau tab for declaration query recovery: %s", login_url)
    try:
        fresh_page.goto(login_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        message = str(exc)
        if "interrupted by another navigation" in message:
            LOGGER.info("Fresh tax-bureau tab navigation was interrupted; continuing with redirected page")
        else:
            LOGGER.info("Fresh tax-bureau tab navigation did not settle immediately: %s", exc)
    return fresh_page


def recover_declaration_query_in_fresh_tab(page, host: str, timeout: int = 60):
    fresh_page = open_fresh_tax_bureau_tab(page, host)
    if not fresh_page:
        return None

    try:
        query_page = open_declaration_query_with_wait(fresh_page, host, timeout=timeout)
        if query_page:
            return query_page

        query_page = open_declaration_query_via_sp_handler(fresh_page, host, timeout=timeout, fail_on_tpass=True)
        if query_page:
            return query_page

        szzh_page = open_digital_account_with_wait(fresh_page, host, timeout=30)
        if szzh_page:
            if is_ready_declaration_query_page(szzh_page):
                return szzh_page
            fresh_page = szzh_page
        query_page = open_declaration_query_with_wait(fresh_page, host, timeout=timeout)
        if query_page:
            return query_page
    except DeclarationQueryAuthError:
        raise
    except Exception as exc:
        LOGGER.info("Fresh-tab declaration query recovery failed: %s", exc)
    return None


def find_any_context_etax_host(page) -> str:
    for candidate in list(page.context.pages):
        try:
            host = extract_etax_host(candidate.url or "")
        except Exception:
            host = ""
        if host:
            return host
    return ""


def navigate_to_query_page_robust(page, web_config):
    host = extract_etax_host(page.url or "") or find_any_context_etax_host(page)
    if host:
        query_page = wait_for_declaration_query_page(page, host, timeout=3)
        if query_page:
            return query_page

        current_url = page.url or ""
        if is_undeclared_vat_page_url(current_url) or is_loading_page_url(current_url):
            auth_error = None
            try:
                query_page = recover_declaration_query_in_current_tab(page, host, timeout=60)
                if query_page:
                    return query_page
            except DeclarationQueryAuthError as exc:
                auth_error = exc
                LOGGER.info("Current-tab declaration query recovery hit auth page; trying fresh tab")
            try:
                query_page = recover_declaration_query_in_fresh_tab(page, host, timeout=60)
                if query_page:
                    return query_page
            except DeclarationQueryAuthError as exc:
                auth_error = auth_error or exc
            if auth_error:
                raise auth_error

        szzh_page = open_digital_account_with_wait(page, host, timeout=30)
        if szzh_page:
            if is_ready_declaration_query_page(szzh_page):
                return szzh_page
            page = szzh_page
            query_page = open_declaration_query_with_wait(page, host, timeout=60)
            if query_page:
                return query_page

        query_page = open_declaration_query_with_wait(page, host, timeout=60)
        if query_page:
            return query_page

        if is_loading_page_url(page.url or ""):
            auth_error = None
            try:
                query_page = recover_declaration_query_in_current_tab(page, host, timeout=60)
                if query_page:
                    return query_page
            except DeclarationQueryAuthError as exc:
                auth_error = exc
                LOGGER.info("Current-tab declaration query recovery hit auth page; trying fresh tab")
            try:
                query_page = recover_declaration_query_in_fresh_tab(page, host, timeout=60)
                if query_page:
                    return query_page
            except DeclarationQueryAuthError as exc:
                auth_error = auth_error or exc
            if auth_error:
                raise auth_error

        for attempt in range(1, 3):
            LOGGER.info("Declaration query was not ready; retrying through tax digital account (attempt %s/2)", attempt)
            szzh_page = open_digital_account_with_wait(page, host, timeout=30)
            if szzh_page:
                if is_ready_declaration_query_page(szzh_page):
                    return szzh_page
                page = szzh_page
            query_page = open_declaration_query_with_wait(page, host, timeout=60)
            if query_page:
                return query_page

        query_page = wait_for_declaration_query_page(page, host, timeout=45)
        if query_page:
            return query_page
        if is_loading_page_url(page.url or "") or is_tpass_login_page_url(page.url or ""):
            raise declaration_query_auth_error_for_url(page.url or "", host)
        return page

    portal_page = find_context_tax_page(page, ("鎴戣鏌ヨ",))
    if portal_page:
        page = portal_page
    NavigationEngine(page).navigate_to_form(web_config)
    time.sleep(3)
    if is_query_page(page):
        return page
    query_page = find_context_query_page(page)
    if query_page:
        return query_page
    return page


def navigate_to_undeclared_vat_page(page, province: str):
    """Backward-compatible wrapper for the VAT undeclared declaration page."""
    return navigate_to_undeclared_tax_page(page, province, TARGETS["vat_general_main"], {})


def navigate_to_undeclared_tax_page(page, province: str, target: CompareTarget, api_response: dict[str, Any]):
    """Open the province-specific undeclared declaration entry for the target tax type."""
    current_url = page.url or ""
    if is_expected_undeclared_entry_url(current_url, target):
        LOGGER.info("Already on undeclared declaration page for %s: %s", target.tax_type, current_url)
        time.sleep(5)
        return page
    origin_match = re.match(r"(https://etax\.[^/]+)", current_url)
    fallback_origin = origin_match.group(1) if origin_match else ""
    target_urls = undeclared_entry_urls(page, province, target, api_response, fallback_origin)

    last_error = None
    for target_url in target_urls:
        LOGGER.info("Current-period marker is false; opening undeclared declaration URL for %s: %s", target.target_id, target_url)
        try:
            page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(5)
            if is_tpass_login_page_url(page.url or ""):
                raise declaration_query_auth_error_for_url(page.url or "", extract_etax_host(target_url))
            target_page = wait_for_undeclared_entry_page(page, target, timeout=12)
            if target_page:
                return target_page
            LOGGER.info(
                "Undeclared direct URL did not land on target entry for %s; url=%s",
                target.target_id,
                page.url,
            )
        except DeclarationQueryAuthError:
            raise
        except Exception as exc:
            last_error = exc
            LOGGER.warning("Undeclared URL failed for %s: %s", target.target_id, exc)
            time.sleep(1)

    if last_error:
        LOGGER.info("Falling back to tax home undeclared navigation for %s after URL failure", target.target_id)
    return navigate_to_undeclared_tax_page_from_home(page, province, target)


def undeclared_entry_urls(
    page,
    province: str,
    target: CompareTarget,
    api_response: dict[str, Any],
    fallback_origin: str = "",
) -> list[str]:
    target_urls: list[str] = []
    path = undeclared_entry_path(target, api_response, province)
    if not path:
        return []
    if fallback_origin:
        target_urls.append(f"{fallback_origin}{path}")
    if province:
        target_urls.append(f"https://etax.{province}.chinatax.gov.cn:8443{path}")
    return list(dict.fromkeys(target_urls))


def undeclared_entry_path(target: CompareTarget, api_response: dict[str, Any], province: str = "") -> str:
    if target.tax_type == "VAT_GENERAL":
        return UNDECLARED_VAT_GENERAL_PATH
    if target.tax_type == "VAT_SMALL_SCALE":
        return UNDECLARED_VAT_SMALL_PATH
    if target.tax_type == "CONSUMPTION_TAX":
        return UNDECLARED_CONSUMPTION_TAX_PATH
    if target.tax_type == "CULTURE_FEE" and (province or "").lower() == "shandong":
        period_start, period_end = target_period_range(api_response, target)
        return UNDECLARED_CULTURE_FEE_SHANDONG_PATH_TEMPLATE.format(
            period_start=urllib.parse.quote(period_start),
            period_end=urllib.parse.quote(period_end),
        )
    return ""


def is_expected_undeclared_entry_url(url: str, target: CompareTarget) -> bool:
    lower_url = (url or "").lower()
    if target.tax_type == "VAT_GENERAL":
        return "/declare/zzsybnsrsb" in lower_url
    if target.tax_type == "VAT_SMALL_SCALE":
        return "/declare/zzsxgmnsrsb" in lower_url
    if target.tax_type == "CONSUMPTION_TAX":
        return "/declare/xfssb" in lower_url or "jyjkid=30" in lower_url
    if target.tax_type == "CULTURE_FEE":
        return "sdsfsgjssb" in lower_url and "whsyjsf" in lower_url
    return False


def navigate_to_undeclared_tax_page_from_home(page, province: str, target: CompareTarget):
    origin = current_etax_origin(page, province)
    home_url = f"{origin}/loginb/"
    LOGGER.info("Opening tax bureau home before undeclared navigation for %s: %s", target.target_id, home_url)
    try:
        page.goto(home_url, wait_until="domcontentloaded", timeout=30000)
    except Exception as exc:
        LOGGER.info("Tax home navigation did not settle immediately: %s", exc)
    time.sleep(3)
    raise_tax_auth_error_if_needed(page, target)
    if (
        target.tax_type == "CIT_A_PREPAY"
        and is_undeclared_target_unavailable_page(page, target)
        and not is_undeclared_home_redirect_page(page, target)
    ):
        snippet = page_text_snippet(page, limit=300)
        raise UndeclaredTaxTargetUnavailableError(
            f"Undeclared tax target is unavailable for target={target.target_id}; "
            f"url={page.url}; page_text={snippet}"
        )
    click_result = ""
    max_click_attempts = 3
    for attempt in range(1, max_click_attempts + 1):
        try:
            click_result = click_tax_home_declare_entry_scoped(page, target)
            break
        except Exception as exc:
            if not is_navigation_context_destroyed_error(exc):
                raise
            LOGGER.warning(
                "Tax home declare entry click for %s was interrupted by navigation; attempt=%s/%s: %s",
                target.target_id,
                attempt,
                max_click_attempts,
                exc,
            )
            wait_for_navigation_after_context_destroyed(page)
            raise_tax_auth_error_if_needed(page, target)
            target_page = wait_for_undeclared_entry_page(page, target, timeout=5)
            if target_page:
                LOGGER.info(
                    "Tax home declare entry click for %s reached target page after interrupted navigation: %s",
                    target.target_id,
                    target_page.url,
                )
                return target_page
            if attempt >= max_click_attempts:
                click_result = "navigation_interrupted"
                break
    LOGGER.info("Tax home declare entry click result for %s: %s", target.target_id, click_result)
    raise_tax_auth_error_if_needed(page, target)
    time.sleep(5)
    raise_tax_auth_error_if_needed(page, target)
    target_page = wait_for_undeclared_entry_page(page, target, timeout=10)
    if target_page:
        return target_page
    target_page = find_context_undeclared_entry_page(page, target)
    if target_page is not page:
        LOGGER.info(
            "Switched to newly opened undeclared tax page for %s: %s",
            target.target_id,
            target_page.url,
        )
        return target_page
    return page


def is_navigation_context_destroyed_error(exc: Exception) -> bool:
    message = str(exc)
    return "Execution context was destroyed" in message or "most likely because of a navigation" in message


def wait_for_navigation_after_context_destroyed(page) -> None:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception:
        pass
    time.sleep(1)


def wait_for_undeclared_entry_page(page, target: CompareTarget, timeout: int = 10):
    deadline = time.time() + timeout
    while time.time() < deadline:
        raise_tax_auth_error_if_needed(page, target)
        target_page = find_context_undeclared_entry_page(page, target)
        if target_page is not page:
            LOGGER.info(
                "Found undeclared target page for %s while waiting: %s",
                target.target_id,
                target_page.url,
            )
            return target_page
        try:
            if is_expected_undeclared_entry_url(page.url or "", target):
                return page
        except Exception:
            pass
        time.sleep(1)
    return None


def find_context_undeclared_entry_page(page, target: CompareTarget):
    """Return a same-context tab that contains the target undeclared declaration shell."""
    try:
        pages = list(getattr(page.context, "pages", []) or [])
    except Exception:
        return page
    if not pages:
        return page
    for candidate in reversed(pages):
        try:
            url = candidate.url or ""
        except Exception:
            continue
        if not is_expected_undeclared_entry_url(url, target):
            continue
        if target.tax_type == "CONSUMPTION_TAX":
            try:
                body_text = candidate.evaluate("document.body ? document.body.innerText.slice(0, 3000) : ''")
            except Exception:
                body_text = ""
            if "消费税及附加税费申报" not in str(body_text or ""):
                continue
        return candidate
    return page


def undeclared_home_target_keyword_groups(target: CompareTarget) -> tuple[tuple[str, ...], ...]:
    if target.tax_type == "VAT_GENERAL":
        return (
            ("\u589e\u503c\u7a0e\u53ca\u9644\u52a0\u7a0e\u8d39\u7533\u62a5",),
            ("\u4e00\u822c\u7eb3\u7a0e\u4eba", "\u4e00\u822c"),
        )
    if target.tax_type == "VAT_SMALL_SCALE":
        return (
            ("\u589e\u503c\u7a0e\u53ca\u9644\u52a0\u7a0e\u8d39\u7533\u62a5",),
            ("\u5c0f\u89c4\u6a21\u7eb3\u7a0e\u4eba", "\u5c0f\u89c4\u6a21"),
        )
    if target.tax_type == "CIT_A_PREPAY":
        return (
            ("\u4f01\u4e1a\u6240\u5f97\u7a0e", "\u5c45\u6c11\u4f01\u4e1a", "\u4f01\u4e1a\u6240", "A200000"),
            ("A\u7c7b", "A200000", "\u67e5\u8d26\u5f81\u6536"),
        )
    if target.tax_type == "CULTURE_FEE":
        return (("\u6587\u5316\u4e8b\u4e1a\u5efa\u8bbe\u8d39",),)
    if target.tax_type == "CONSUMPTION_TAX":
        return (("\u6d88\u8d39\u7a0e\u53ca\u9644\u52a0\u7a0e\u8d39\u7533\u62a5", "\u6d88\u8d39\u7a0e"),)
    return tuple((keyword,) for keyword in (target.detail_form_keywords or target.query_keywords or (target.form_name,)))


def compact_page_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def text_matches_keyword_groups(text: str, groups: tuple[tuple[str, ...], ...]) -> bool:
    compact = compact_page_text(text)
    if not groups:
        return False
    return all(any(compact_page_text(keyword) in compact for keyword in group if keyword) for group in groups)


def home_declare_scope_text(text: str, limit: int = 1800) -> str:
    compact = compact_page_text(text)
    starts = [
        compact.find(compact_page_text(keyword))
        for keyword in ("\u672c\u671f\u5e94\u7533\u62a5", "\u5e94\u7533\u62a5", "\u7533\u62a5\u6e05\u518c")
        if compact.find(compact_page_text(keyword)) >= 0
    ]
    if not starts:
        return ""
    start = min(starts)
    end = len(compact)
    for marker in (
        "\u70ed\u95e8\u670d\u52a1",
        "\u5730\u65b9\u7279\u8272",
        "\u6211\u7684\u5f85\u529e",
        "\u529e\u7a0e\u65e5\u5386",
        "\u901a\u77e5\u516c\u544a",
    ):
        pos = compact.find(compact_page_text(marker), start + 1)
        if pos >= 0:
            end = min(end, pos)
    return compact[start : min(end, start + limit)]


def cit_a_scope_is_rejected(text: str) -> bool:
    compact = compact_page_text(text)
    return "\u6838\u5b9a\u5f81\u6536" in compact or "B\u7c7b" in compact or "\uff22\u7c7b" in compact


def undeclared_home_target_in_declare_scope(text: str, target: CompareTarget) -> bool:
    if target.tax_type != "CIT_A_PREPAY":
        return text_matches_keyword_groups(text, undeclared_home_target_keyword_groups(target))
    scope_text = home_declare_scope_text(text)
    if not scope_text or cit_a_scope_is_rejected(scope_text):
        return False
    return text_matches_keyword_groups(scope_text, undeclared_home_target_keyword_groups(target))


def target_home_status_segment(text: str, target: CompareTarget, limit: int = 360) -> str:
    compact = compact_page_text(text)
    groups = undeclared_home_target_keyword_groups(target)
    if not groups:
        return ""
    positions: list[int] = []
    for group in groups:
        group_positions = [
            compact.find(compact_page_text(keyword))
            for keyword in group
            if keyword and compact.find(compact_page_text(keyword)) >= 0
        ]
        if not group_positions:
            return ""
        positions.append(min(group_positions))
    start = min(positions)
    return compact[start : start + limit]


def undeclared_home_target_declaration_status(page, target: CompareTarget) -> str:
    snippet = page_text_snippet(page, limit=5000)
    if not snippet:
        return ""
    search_text = home_declare_scope_text(snippet) if target.tax_type == "CIT_A_PREPAY" else snippet
    if target.tax_type == "CIT_A_PREPAY" and (not search_text or cit_a_scope_is_rejected(search_text)):
        return ""
    segment = target_home_status_segment(search_text, target)
    if not segment:
        return ""
    if "\u5df2\u7533\u62a5" in segment and (
        "\u66f4\u6b63" in segment
        or "\u4f5c\u5e9f" in segment
        or "\u7533\u62a5\u6210\u529f" in segment
    ):
        return "filed"
    if "\u672a\u7533\u62a5" in segment and (
        "\u586b\u5199\u7533\u62a5\u8868" in segment
        or "\u529e\u7406" in segment
        or "\u53bb\u7533\u62a5" in segment
        or "\u7acb\u5373\u7533\u62a5" in segment
    ):
        return "unfiled"
    return ""


def is_undeclared_home_redirect_page(page, target: CompareTarget) -> bool:
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""
    snippet = page_text_snippet(page, limit=1500)
    if not snippet:
        return False
    home_groups = (
        ("\u672c\u671f\u5e94\u7533\u62a5", "\u586b\u5199\u7533\u62a5\u8868", "\u6211\u8981\u529e\u7a0e"),
        ("\u586b\u5199\u7533\u62a5\u8868", "\u6211\u8981\u586b\u8868", "\u7ee7\u7eed\u7533\u62a5", "\u529e\u7406"),
    )
    if not text_matches_keyword_groups(snippet, home_groups):
        return False
    if not undeclared_home_target_in_declare_scope(snippet, target):
        return False
    return "/loginb" in url or "/mhzx/" in url or not is_expected_undeclared_entry_url(url, target)


def is_undeclared_target_unavailable_page(page, target: CompareTarget) -> bool:
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""
    if is_expected_undeclared_entry_url(url, target) or is_tpass_login_page_url(url):
        return False
    if is_tax_auth_error_page(page):
        return False
    if "/loginb" in url or "/mhzx/" in url:
        return True
    return False


def is_tax_auth_error_page(page) -> bool:
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""
    if "/mhzx/api/mh/tpass/code" in url:
        return True
    snippet = page_text_snippet(page, limit=500)
    auth_error_markers = (
        "\u6388\u6743\u7801\u4e0d\u80fd\u4e3a\u7a7a",
        "authorization code",
        "auth code",
    )
    return any(marker.lower() in snippet.lower() for marker in auth_error_markers)


def raise_tax_auth_error_if_needed(page, target: CompareTarget) -> None:
    try:
        url = page.url or ""
    except Exception:
        url = ""
    if not (is_tpass_login_page_url(url) or is_tax_auth_error_page(page)):
        return
    snippet = page_text_snippet(page, limit=300)
    raise DeclarationQueryAuthError(
        "Tax bureau login state or digital account authentication expired: "
        f"undeclared entry returned to unified login page before target={target.target_id}; "
        f"url={url}; page_text={snippet}"
    )


def recover_undeclared_entry_from_home_redirect(page, target: CompareTarget):
    if not is_undeclared_home_redirect_page(page, target):
        return None
    home_status = undeclared_home_target_declaration_status(page, target)
    if home_status == "filed":
        snippet = page_text_snippet(page, limit=300)
        raise UndeclaredTaxAlreadyDeclaredError(
            f"Tax bureau home shows the target already declared before undeclared entry recovery; "
            f"target={target.target_id}; url={page.url}; page_text={snippet}"
        )
    LOGGER.info(
        "Undeclared page for %s appears to be tax home/portal after redirect; retrying target entry click. url=%s",
        target.target_id,
        page.url,
    )
    last_click_result = ""
    for attempt in range(1, 4):
        try:
            click_result = click_tax_home_declare_entry_scoped(page, target)
        except Exception as exc:
            LOGGER.info("Tax home undeclared recovery click failed for %s: %s", target.target_id, exc)
            return None
        LOGGER.info(
            "Tax home undeclared recovery click result for %s attempt %s/3: %s",
            target.target_id,
            attempt,
            click_result,
        )
        time.sleep(3)
        target_page = wait_for_undeclared_entry_page(page, target, timeout=8)
        if target_page:
            return target_page
        target_page = find_context_undeclared_entry_page(page, target)
        if target_page is not page:
            return target_page
        if not is_undeclared_home_redirect_page(page, target):
            return None
        if click_result == last_click_result:
            LOGGER.info(
                "Tax home undeclared recovery repeated the same click result for %s; stopping. result=%s",
                target.target_id,
                click_result,
            )
            return None
        last_click_result = click_result
    return None


def current_etax_origin(page, province: str) -> str:
    current_url = page.url or ""
    origin_match = re.match(r"(https://etax\.[^/]+)", current_url)
    if origin_match:
        return origin_match.group(1)
    return f"https://etax.{province}.chinatax.gov.cn:8443"


def click_tax_home_declare_entry(page, target: CompareTarget) -> str:
    return str(
        page.evaluate(
            """async (keywords) => {
                const normalize = (value) => String(value || '').replace(/\\s+/g, '');
                const sleep = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const clickFirst = (labels, scopeText) => {
                    const wanted = labels.map(normalize).filter(Boolean);
                    const nodes = Array.from(document.querySelectorAll('button, a, [role=button], .t-button, .el-button, .ant-btn'))
                        .filter(visible)
                        .map((el) => ({
                            el,
                            text: normalize(el.innerText || el.textContent || el.getAttribute('title') || el.getAttribute('aria-label')),
                            scope: normalize(el.closest('li, tr, .card, .item, .t-card, .el-card, [class*=card], [class*=item]')?.innerText || '')
                        }))
                        .filter((item) => item.text && item.text.length <= 80);
                    const scoped = nodes.find((item) =>
                        scopeText && item.scope.includes(scopeText) && wanted.some((label) => item.text.includes(label))
                    );
                    const node = scoped || nodes.find((item) => wanted.some((label) => item.text.includes(label)));
                    if (!node) return '';
                    node.el.scrollIntoView({ block: 'center' });
                    node.el.click();
                    return node.text;
                };
                const targetText = keywords.map(normalize).find(Boolean) || '';
                const steps = [
                    ['我要办税', '办税'],
                    ['税费申报及缴纳', '纳税申报', '申报缴纳'],
                    ['填写申报表', '申报表填写', '我要填表', '继续申报', '办理']
                ];
                const clicked = [];
                for (const labels of steps) {
                    let text = '';
                    for (let attempt = 0; attempt < 16; attempt += 1) {
                        text = clickFirst(labels, targetText);
                        if (text) break;
                        await sleep(500);
                    }
                    if (text) {
                        clicked.push(text);
                        await sleep(1200);
                    }
                }
                return clicked.length ? `clicked:${clicked.join('>')}` : 'entry_not_found';
            }""",
            list(target.detail_form_keywords or target.query_keywords or (target.form_name,)),
        )
    )


def click_tax_home_declare_entry_scoped(page, target: CompareTarget) -> str:
    return str(
        page.evaluate(
            """async (payload) => {
                const normalize = (value) => String(value || '').replace(/\\s+/g, '');
                const sleep = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const elementText = (el) => normalize(
                    el?.innerText || el?.textContent || el?.getAttribute?.('title') || el?.getAttribute?.('aria-label') || ''
                );
                const taxType = String(payload?.taxType || '');
                const rawKeywords = Array.isArray(payload?.keywords) ? payload.keywords : [];
                const keywordGroupsByTaxType = {
                    VAT_GENERAL: [
                        ['\\u589e\\u503c\\u7a0e\\u53ca\\u9644\\u52a0\\u7a0e\\u8d39\\u7533\\u62a5'],
                        ['\\u4e00\\u822c\\u7eb3\\u7a0e\\u4eba', '\\u4e00\\u822c']
                    ],
                    VAT_SMALL_SCALE: [
                        ['\\u589e\\u503c\\u7a0e\\u53ca\\u9644\\u52a0\\u7a0e\\u8d39\\u7533\\u62a5'],
                        ['\\u5c0f\\u89c4\\u6a21\\u7eb3\\u7a0e\\u4eba', '\\u5c0f\\u89c4\\u6a21']
                    ],
                    CIT_A_PREPAY: [
                        ['\\u4f01\\u4e1a\\u6240\\u5f97\\u7a0e', '\\u5c45\\u6c11\\u4f01\\u4e1a', '\\u4f01\\u4e1a\\u6240', 'A200000'],
                        ['A\\u7c7b', 'A200000', '\\u67e5\\u8d26\\u5f81\\u6536']
                    ],
                    CULTURE_FEE: [['\\u6587\\u5316\\u4e8b\\u4e1a\\u5efa\\u8bbe\\u8d39']],
                    CONSUMPTION_TAX: [['\\u6d88\\u8d39\\u7a0e\\u53ca\\u9644\\u52a0\\u7a0e\\u8d39\\u7533\\u62a5', '\\u6d88\\u8d39\\u7a0e']]
                };
                const keywordGroups = keywordGroupsByTaxType[taxType] || rawKeywords.map((keyword) => [keyword]);
                const textMatchesTarget = (text) => {
                    const normalized = normalize(text);
                    return keywordGroups.length > 0 && keywordGroups.every((group) =>
                        group.map(normalize).filter(Boolean).some((keyword) => normalized.includes(keyword))
                    );
                };
                const clickElement = (el) => {
                    el.scrollIntoView({ block: 'center' });
                    el.click();
                    return elementText(el);
                };
                const actionLabels = [
                    '\\u586b\\u5199\\u7533\\u62a5\\u8868',
                    '\\u6211\\u8981\\u586b\\u8868',
                    '\\u7ee7\\u7eed\\u7533\\u62a5',
                    '\\u7ee7\\u7eed\\u529e\\u7406',
                    '\\u7acb\\u5373\\u7533\\u62a5',
                    '\\u53bb\\u7533\\u62a5',
                    '\\u586b\\u62a5',
                    '\\u586b\\u8868',
                    '\\u529e\\u7406',
                    '\\u8fdb\\u5165'
                ].map(normalize);
                const forbiddenScopeLabels = [
                    '\\u529e\\u7a0e\\u8fdb\\u5ea6\\u53ca\\u7ed3\\u679c\\u4fe1\\u606f\\u67e5\\u8be2',
                    '\\u8fdb\\u5ea6\\u53ca\\u7ed3\\u679c',
                    '\\u529e\\u7a0e\\u8fdb\\u5ea6',
                    '\\u7533\\u62a5\\u67e5\\u8be2',
                    '\\u7a0e\\u8d39\\u7f34\\u7eb3',
                    '\\u53d1\\u7968\\u4e1a\\u52a1',
                    '\\u793e\\u4fdd\\u8d39\\u4e1a\\u52a1',
                    '\\u7a0e\\u52a1\\u6570\\u5b57\\u8d26\\u6237',
                    '\\u6211\\u7684\\u5f85\\u529e',
                    '\\u901a\\u77e5\\u516c\\u544a',
                    '\\u70ed\\u95e8\\u670d\\u52a1',
                    '\\u5730\\u65b9\\u7279\\u8272'
                ].map(normalize);
                const declareScopeLabels = [
                    '\\u672c\\u671f\\u5e94\\u7533\\u62a5',
                    '\\u5e94\\u7533\\u62a5',
                    '\\u7533\\u62a5\\u6e05\\u518c',
                    '\\u7a0e\\u8d39\\u7533\\u62a5\\u53ca\\u7f34\\u7eb3',
                    '\\u7eb3\\u7a0e\\u7533\\u62a5',
                    '\\u7533\\u62a5\\u7f34\\u7eb3'
                ].map(normalize);
                const containerSelector = [
                    'tr',
                    'li',
                    '[role=row]',
                    '.card',
                    '.item',
                    '.t-card',
                    '.el-card',
                    '.ant-card',
                    '[class*=card]',
                    '[class*=item]',
                    '[class*=todo]',
                    '[class*=declare]',
                    '[class*=tax]'
                ].join(', ');
                const hasAnyLabel = (text, labels) => {
                    const normalized = normalize(text);
                    return labels.some((label) => normalized.includes(label));
                };
                const isRejectedCitScope = (text) => {
                    const normalized = normalize(text);
                    return taxType === 'CIT_A_PREPAY' && (
                        normalized.includes('\\u6838\\u5b9a\\u5f81\\u6536') ||
                        normalized.includes('B\\u7c7b') ||
                        normalized.includes('\\uff22\\u7c7b')
                    );
                };
                const closestScopeText = (el, fallbackRoot) => normalize(
                    el?.closest?.(containerSelector)?.innerText ||
                    el?.closest?.(containerSelector)?.textContent ||
                    fallbackRoot?.innerText ||
                    fallbackRoot?.textContent ||
                    ''
                );
                const isForbiddenScope = (text) => hasAnyLabel(text, forbiddenScopeLabels) || isRejectedCitScope(text);
                const isDeclareScope = (text) => hasAnyLabel(text, declareScopeLabels);
                const findActionIn = (root) => {
                    const rootScope = elementText(root);
                    const buttons = Array.from(root.querySelectorAll?.('button, a, [role=button], .t-button, .el-button, .ant-btn') || [])
                        .filter(visible)
                        .map((el) => ({ el, text: elementText(el), scope: closestScopeText(el, root), rootScope }))
                        .filter((item) => item.text && item.text.length <= 80)
                        .filter((item) => !isForbiddenScope(item.text) && !isForbiddenScope(item.scope))
                        .filter((item) =>
                            isDeclareScope(item.scope) ||
                            textMatchesTarget(item.scope) ||
                            isDeclareScope(item.rootScope) ||
                            textMatchesTarget(item.rootScope)
                        )
                        .filter((item) => actionLabels.some((label) => item.text.includes(label)))
                        .sort((a, b) => a.text.length - b.text.length);
                    return buttons[0]?.el || null;
                };
                const clickFirst = (labels) => {
                    const wanted = labels.map(normalize).filter(Boolean);
                    const node = Array.from(document.querySelectorAll('button, a, [role=button], .t-button, .el-button, .ant-btn'))
                        .filter(visible)
                        .map((el) => ({ el, text: elementText(el), scope: closestScopeText(el, document.body) }))
                        .filter((item) => item.text && item.text.length <= 80)
                        .filter((item) => !isForbiddenScope(item.text) && !isForbiddenScope(item.scope))
                        .filter((item) => wanted.some((label) => item.text.includes(label)))
                        .sort((a, b) => {
                            const exactA = wanted.some((label) => a.text === label) ? 0 : 1;
                            const exactB = wanted.some((label) => b.text === label) ? 0 : 1;
                            return exactA - exactB || a.text.length - b.text.length;
                        })[0];
                    return node ? clickElement(node.el) : '';
                };
                const clickScopedTargetAction = (allowTargetTitle) => {
                    let targetTitle = null;
                    const containers = Array.from(document.querySelectorAll(containerSelector))
                        .filter(visible)
                        .filter((el) => textMatchesTarget(el.innerText || el.textContent || ''))
                        .map((el) => {
                            const rect = el.getBoundingClientRect();
                            return { el, area: rect.width * rect.height, text: elementText(el) };
                        })
                        .filter((item) => item.area > 0)
                        .filter((item) => !isForbiddenScope(item.text))
                        .filter((item) => (
                            taxType === 'CIT_A_PREPAY'
                                ? isDeclareScope(item.text)
                                : (isDeclareScope(item.text) || textMatchesTarget(item.text))
                        ))
                        .sort((a, b) => a.area - b.area || a.text.length - b.text.length);
                    for (const item of containers) {
                        const action = findActionIn(item.el);
                        if (action) return `target_action:${clickElement(action)}`;
                    }
                    const targetNode = Array.from(document.querySelectorAll('button, a, [role=button], .t-button, .el-button, .ant-btn'))
                        .filter(visible)
                        .map((el) => ({ el, text: elementText(el), scope: closestScopeText(el, document.body) }))
                        .filter((item) => textMatchesTarget(item.text))
                        .filter((item) => !isForbiddenScope(item.text) && !isForbiddenScope(item.scope))
                        .filter((item) => taxType !== 'CIT_A_PREPAY' || isDeclareScope(item.scope))
                        .map((item) => item.el)[0] || null;
                    if (targetNode) {
                        targetTitle = targetNode;
                        for (let ancestor = targetNode.parentElement, depth = 0; ancestor && depth < 8; ancestor = ancestor.parentElement, depth += 1) {
                            const ancestorText = elementText(ancestor);
                            if (!textMatchesTarget(ancestorText) || isForbiddenScope(ancestorText)) continue;
                            const action = findActionIn(ancestor);
                            if (action) return `target_action:${clickElement(action)}`;
                        }
                    }
                    if (targetTitle && allowTargetTitle) return `target_title:${clickElement(targetTitle)}`;
                    return '';
                };
                const clicked = [];
                const expectedEntryByTaxType = {
                    VAT_GENERAL: '/declare/zzsybnsrsb',
                    VAT_SMALL_SCALE: '/declare/zzsxgmnsrsb',
                    CONSUMPTION_TAX: '/declare/xfssb',
                    CULTURE_FEE: 'whsyjsf'
                };
                const onExpectedEntry = () => {
                    const expected = expectedEntryByTaxType[taxType] || '';
                    return expected && String(location.href || '').toLowerCase().includes(expected.toLowerCase());
                };
                const firstTargetText = clickScopedTargetAction(true);
                if (firstTargetText) {
                    clicked.push(firstTargetText);
                    await sleep(1800);
                    return clicked.length ? `clicked:${clicked.join('>')}` : 'entry_not_found';
                }
                const navSteps = [
                    ['\\u6211\\u8981\\u529e\\u7a0e', '\\u529e\\u7a0e'],
                    ['\\u7a0e\\u8d39\\u7533\\u62a5\\u53ca\\u7f34\\u7eb3', '\\u7eb3\\u7a0e\\u7533\\u62a5', '\\u7533\\u62a5\\u7f34\\u7eb3']
                ];
                for (const labels of navSteps) {
                    let text = '';
                    for (let attempt = 0; attempt < 12; attempt += 1) {
                        text = clickFirst(labels);
                        if (text) break;
                        await sleep(500);
                    }
                    if (text) {
                        clicked.push(text);
                        await sleep(1200);
                    }
                }
                for (let attempt = 0; attempt < 20; attempt += 1) {
                    const text = clickScopedTargetAction(true);
                    if (text) {
                        clicked.push(text);
                        await sleep(1800);
                        break;
                    }
                    await sleep(500);
                }
                return clicked.length ? `clicked:${clicked.join('>')}` : 'entry_not_found';
            }""",
            {
                "keywords": list(target.detail_form_keywords or target.query_keywords or (target.form_name,)),
                "taxType": target.tax_type,
            },
        )
    )


def prepare_undeclared_page_for_target(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping] | None = None,
    allow_home_recovery: bool = True,
):
    """On undeclared pages, open the editable declaration form before extraction."""
    raise_tax_auth_error_if_needed(page, target)
    home_status = undeclared_home_target_declaration_status(page, target)
    if home_status == "filed":
        snippet = page_text_snippet(page, limit=300)
        raise UndeclaredTaxAlreadyDeclaredError(
            f"Tax bureau home shows the target already declared before undeclared form preparation; "
            f"target={target.target_id}; url={page.url}; page_text={snippet}"
        )
    if allow_home_recovery and is_undeclared_home_redirect_page(page, target):
        recovered_page = recover_undeclared_entry_from_home_redirect(page, target)
        if recovered_page:
            return prepare_undeclared_page_for_target(
                recovered_page,
                target,
                mappings,
                allow_home_recovery=False,
            )
        if is_undeclared_target_unavailable_page(page, target):
            snippet = page_text_snippet(page, limit=300)
            raise UndeclaredTaxTargetUnavailableError(
                f"Undeclared tax target is unavailable for target={target.target_id}; "
                f"url={page.url}; page_text={snippet}"
            )
    if target.tax_type == "CIT_A_PREPAY" and is_undeclared_target_unavailable_page(page, target):
        snippet = page_text_snippet(page, limit=300)
        raise UndeclaredTaxTargetUnavailableError(
            f"Undeclared tax target is unavailable for target={target.target_id}; "
            f"url={page.url}; page_text={snippet}"
        )
    result = "fill_button_not_found"
    menu_keywords = UNDECLARED_VAT_MENU_KEYWORDS.get(target.target_id)
    deadline = time.time() + 45
    while time.time() < deadline:
        raise_tax_auth_error_if_needed(page, target)
        try:
            result = page.evaluate(
                """(targetId) => {
                    const normalize = (value) => String(value || '').replace(/\\s+/g, '');
                    const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                    const body = normalize(document.body ? document.body.innerText : '');
                    const formHints = [
                        '\\u62a5\\u8868\\u5217\\u8868',
                        '\\u7a0e\\u6b3e\\u8ba1\\u7b97',
                        '\\u9500\\u552e\\u989d',
                        '\\u9500\\u9879\\u7a0e\\u989d',
                        '\\u8fdb\\u9879\\u7a0e\\u989d',
                    ];
                    if (
                        body.includes('\\u672c\\u5c5e\\u671f\\u5df2\\u7533\\u62a5') ||
                        (body.includes('\\u5df2\\u7533\\u62a5') && body.includes('\\u7533\\u62a5\\u66f4\\u6b63\\u4e0e\\u4f5c\\u5e9f'))
                    ) {
                        return 'already_declared';
                    }
                    const entryLabels = [
                        '\\u7ee7\\u7eed\\u7533\\u62a5',
                        '\\u7ee7\\u7eed\\u529e\\u7406',
                        '\\u6211\\u8981\\u586b\\u8868',
                        '\\u586b\\u5199\\u7533\\u62a5\\u8868',
                    ].map(normalize);
                    const entryButton = Array.from(document.querySelectorAll(
                        'button, a, [role=button], .t-button, .el-button, .ant-btn'
                    )).filter((el) => visible(el)).find((el) => {
                        const text = normalize(el.innerText || el.textContent || el.getAttribute('title') || el.getAttribute('aria-label'));
                        return entryLabels.some((label) => text === label || text.includes(label));
                    });
                    if (entryButton) {
                        const entryText = normalize(entryButton.innerText || entryButton.textContent || entryButton.getAttribute('title') || entryButton.getAttribute('aria-label'));
                        entryButton.scrollIntoView({ block: 'center' });
                        entryButton.click();
                        if (entryText.includes('\\u7ee7\\u7eed\\u7533\\u62a5') || entryText.includes('\\u7ee7\\u7eed\\u529e\\u7406')) {
                            return `clicked_continue_${targetId}`;
                        }
                        return `clicked_${targetId}`;
                    }
                    if (formHints.some((hint) => body.includes(hint))) return 'already_form_view';

                    const preferredLabels = [
                        '\\u6211\\u8981\\u586b\\u8868',
                        '\\u586b\\u5199\\u7533\\u62a5\\u8868',
                        '\\u586b\\u62a5',
                        '\\u586b\\u8868',
                        '\\u529e\\u7406',
                        '\\u8fdb\\u5165',
                    ].map(normalize);
                    const blockedLabels = [
                        '\\u63d0\\u4ea4',
                        '\\u4fdd\\u5b58',
                        '\\u5220\\u9664',
                        '\\u4f5c\\u5e9f',
                        '\\u8fd4\\u56de',
                        '\\u53d6\\u6d88',
                        '\\u6253\\u5370',
                        '\\u5bfc\\u51fa',
                    ].map(normalize);
                    const nodes = Array.from(document.querySelectorAll(
                        'button, a, [role=button], .t-button, .el-button, .ant-btn'
                    )).filter((el) => visible(el));
                    const candidates = nodes
                        .map((el) => ({ el, text: normalize(el.innerText || el.textContent || el.getAttribute('title') || el.getAttribute('aria-label')) }))
                        .filter((item) => item.text && item.text.length <= 40)
                        .filter((item) => !blockedLabels.some((label) => item.text.includes(label)))
                        .filter((item) => preferredLabels.some((label) => item.text === label || item.text.includes(label)))
                        .sort((a, b) => a.text.length - b.text.length);
                    const button = candidates[0] && candidates[0].el;
                    if (!button) return 'fill_button_not_found';
                    button.scrollIntoView({ block: 'center' });
                    button.click();
                    return `clicked_${targetId}`;
                }""",
                target.target_id,
            )
        except Exception as exc:
            LOGGER.info("Waiting for undeclared page to settle for %s: %s", target.target_id, exc)
            time.sleep(2)
            continue
        if str(result).startswith("clicked_continue_"):
            LOGGER.info("Clicked undeclared continue button for %s; waiting for fill-table entry", target.target_id)
            time.sleep(3)
            continue
        if result != "fill_button_not_found":
            break
        time.sleep(2)
    LOGGER.info("Undeclared form prepare result for %s: %s", target.target_id, result)
    if result == "already_declared":
        snippet = page_text_snippet(page, limit=300)
        raise UndeclaredTaxAlreadyDeclaredError(
            f"Tax bureau reports the task already declared while task log says current-period=false; "
            f"target={target.target_id}; url={page.url}; page_text={snippet}"
        )
    if result == "fill_button_not_found":
        if allow_home_recovery:
            recovered_page = recover_undeclared_entry_from_home_redirect(page, target)
            if recovered_page:
                return prepare_undeclared_page_for_target(
                    recovered_page,
                    target,
                    mappings,
                    allow_home_recovery=False,
                )
        raise_tax_auth_error_if_needed(page, target)
        expected = target_title_keywords(target, menu_keywords or ())
        target_page = find_context_undeclared_entry_page(page, target)
        if target_page is not page:
            LOGGER.info(
                "Using newly opened undeclared tax page for %s while preparing form: %s",
                target.target_id,
                target_page.url,
            )
            page = target_page
        if (
            expected
            and target_content_visible_in_any_scope(page, target, mappings, expected)
        ):
            LOGGER.info("Undeclared tax form appears already open for %s; continuing without fill button", target.target_id)
            result = "already_form_view"
        elif is_undeclared_target_unavailable_page(page, target):
            snippet = page_text_snippet(page, limit=300)
            raise UndeclaredTaxTargetUnavailableError(
                f"Undeclared tax target is unavailable for target={target.target_id}; "
                f"url={page.url}; page_text={snippet}"
            )
        else:
            snippet = page_text_snippet(page, limit=300)
            raise RuntimeError(
                f"Could not open undeclared tax form before target={target.target_id}; "
                f"url={page.url}; page_text={snippet}"
            )
    time.sleep(3)

    target_page = find_context_undeclared_entry_page(page, target)
    if target_page is not page:
        LOGGER.info(
            "Using newly opened undeclared tax page for %s before menu selection: %s",
            target.target_id,
            target_page.url,
        )
        page = target_page

    if menu_keywords:
        expected = target_title_keywords(target, menu_keywords)
        if target.target_id in UNDECLARED_TARGET_BODY_KEYWORDS:
            for _ in range(UNDECLARED_APPENDIX_MENU_PRECHECK_SECONDS):
                if undeclared_target_body_keywords_visible(page, target):
                    LOGGER.info(
                        "Undeclared target content is already visible for %s before menu selection: %s",
                        target.target_id,
                        expected,
                    )
                    dismiss_known_undeclared_info_dialogs(page)
                    return page
                time.sleep(1)
        elif target_content_visible_in_any_scope(page, target, mappings, expected):
            LOGGER.info(
                "Undeclared target content is already visible for %s before menu selection: %s",
                target.target_id,
                expected,
            )
            dismiss_known_undeclared_info_dialogs(page)
            return page
        if allow_home_recovery and is_undeclared_home_redirect_page(page, target):
            recovered_page = recover_undeclared_entry_from_home_redirect(page, target)
            if recovered_page:
                return prepare_undeclared_page_for_target(
                    recovered_page,
                    target,
                    mappings,
                    allow_home_recovery=False,
                )
        menu_result = select_undeclared_vat_menu_item(page, menu_keywords)
        LOGGER.info("Undeclared menu select result for %s: %s", target.target_id, menu_result)
        if menu_result == "clicked_menu_item" and is_fast_confirmable_undeclared_menu_target(target):
            mark_recent_undeclared_menu_target(page, target)
            dismiss_known_undeclared_info_dialogs(page)
            LOGGER.info(
                "Accepted just-clicked undeclared menu target for %s; deferring heavy confirmation to extraction",
                target.target_id,
            )
            return page
        if menu_result == "menu_item_not_found":
            if target_content_visible_in_any_scope(page, target, mappings, expected):
                LOGGER.info("Undeclared target content became visible while selecting menu for %s: %s", target.target_id, expected)
                dismiss_known_undeclared_info_dialogs(page)
                return page
            if allow_home_recovery:
                recovered_page = recover_undeclared_entry_from_home_redirect(page, target)
                if recovered_page:
                    return prepare_undeclared_page_for_target(
                        recovered_page,
                        target,
                        mappings,
                        allow_home_recovery=False,
                    )
            if is_undeclared_target_unavailable_page(page, target):
                menu_items = visible_undeclared_menu_items(page)
                raise UndeclaredTaxTargetUnavailableError(
                    f"Undeclared tax menu item was not available for target={target.target_id}; "
                    f"keywords={menu_keywords}; visible_menu={menu_items}; url={page.url}"
                )
            menu_items = visible_undeclared_menu_items(page)
            raise RuntimeError(
                f"Undeclared tax menu item was not found for target={target.target_id}; "
                f"keywords={menu_keywords}; visible_menu={menu_items}; url={page.url}"
            )
        if not wait_for_undeclared_target_visible(page, target, menu_keywords, mappings=mappings):
            if allow_home_recovery:
                recovered_page = recover_undeclared_entry_from_home_redirect(page, target)
                if recovered_page:
                    return prepare_undeclared_page_for_target(
                        recovered_page,
                        target,
                        mappings,
                        allow_home_recovery=False,
                    )
            raise RuntimeError(
                f"Undeclared tax target was not confirmed after menu selection for "
                f"target={target.target_id}; keywords={menu_keywords}; url={page.url}"
            )
        dismiss_known_undeclared_info_dialogs(page)
        return page

    if target.detail_form_keywords:
        select_result = select_detail_form(page, target.detail_form_keywords)
        if select_result == "not_found":
            raise RuntimeError(
                f"Detail form selector was not found on undeclared page for target={target.target_id}; "
                f"keywords={target.detail_form_keywords}; url={page.url}"
            )
        dismiss_known_undeclared_info_dialogs(page)
    return page


def open_undeclared_target_with_auth_retry(
    *,
    bm: BrowserManager,
    chanjet_page,
    args: argparse.Namespace,
    tax_page,
    province: str,
    expected_tax_no: str,
    target: CompareTarget,
    api_response: dict[str, Any],
    mappings: list[FieldMapping] | None = None,
):
    retried_auth = False
    while True:
        try:
            tax_page = navigate_to_undeclared_tax_page(tax_page, province, target, api_response)
            tax_page = prepare_undeclared_page_for_target(tax_page, target, mappings)
            return tax_page, province
        except DeclarationQueryAuthError as exc:
            if retried_auth:
                raise
            retried_auth = True
            LOGGER.warning(
                "Undeclared entry auth failed for target=%s; refreshing task tax login once and retrying: %s",
                target.target_id,
                exc,
            )
            tax_page, province, _ = login_tax_page_for_task(
                bm,
                chanjet_page,
                args,
                province,
                expected_tax_no,
            )


def select_undeclared_vat_menu_item(page, keywords: tuple[str, ...], timeout: int = 30) -> str:
    """Click a left-side undeclared report menu item after the menu has finished rendering."""
    deadline = time.time() + timeout
    last_result = "menu_item_not_found"
    while time.time() < deadline:
        try:
            result = page.evaluate(
                """(keywords) => {
                    const normalize = (value) => String(value || '')
                        .replace(/\\s+/g, '')
                        .replace(/[（）()]/g, '');
                    const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                    const wanted = keywords.map(normalize).filter(Boolean);
                    const menuSelector = [
                        '.gt-collapse-menu-sidebar-content li.t-menu__item',
                        '.gt-collapse-menu-sidebar-content [class*=menu__item]',
                        'li[role=menuitem]',
                        '[role=treeitem]',
                        '.el-menu-item',
                        '.ant-menu-item',
                        '.t-menu__item',
                        '.t-tabs__nav-item',
                    ].join(', ');
                    const items = Array.from(document.querySelectorAll(menuSelector)).filter((el) => {
                        if (!visible(el)) return false;
                        const cls = String(el.className || '');
                        if (cls.includes('t-submenu ') || cls === 't-submenu' || cls.includes('t-submenu t-is-opened')) {
                            return false;
                        }
                        const text = normalize(el.textContent || el.innerText || '');
                        if (!text || text.length > 160) return false;
                        return true;
                    });
                    const item = items
                        .filter((el) => {
                            const text = normalize(el.textContent || el.innerText || '');
                            return wanted.every((kw) => text.includes(kw));
                        })
                        .sort((a, b) => {
                            const ar = a.getBoundingClientRect();
                            const br = b.getBoundingClientRect();
                            const areaA = ar.width * ar.height;
                            const areaB = br.width * br.height;
                            const textA = normalize(a.textContent || a.innerText || '');
                            const textB = normalize(b.textContent || b.innerText || '');
                            return textA.length - textB.length || areaA - areaB;
                        })[0];
                    if (!item) {
                        const visibleMenuText = items
                            .map((el) => normalize(el.textContent || el.innerText || ''))
                            .filter((text) => (
                                text.includes('主表') ||
                                text.includes('附列资料') ||
                                text.includes('增值税') ||
                                text.includes('企业所得税') ||
                                text.includes('文化事业') ||
                                text.includes('消费税') ||
                                text.includes('A200000') ||
                                text.includes('申报表') ||
                                text.includes('计算表') ||
                                text.includes('清单')
                            ));
                        if (!visibleMenuText.length) return 'menu_not_ready';
                        return 'menu_item_not_found';
                    }
                    if (String(item.className || '').includes('t-is-active')) return 'already_selected';
                    item.scrollIntoView({ block: 'center' });
                    item.click();
                    return 'clicked_menu_item';
                }""",
                list(keywords),
            )
        except Exception as exc:
            LOGGER.info("Waiting for undeclared tax menu to settle for keywords=%s: %s", keywords, exc)
            result = "menu_not_ready"
        last_result = str(result)
        if last_result in {"already_selected", "clicked_menu_item"}:
            return last_result
        time.sleep(1)
    return "menu_item_not_found" if last_result == "menu_not_ready" else last_result


def visible_undeclared_menu_items(page, limit: int = 20) -> list[str]:
    try:
        items = page.evaluate(
            """(limit) => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const selector = [
                    '.gt-collapse-menu-sidebar-content li.t-menu__item',
                    '.gt-collapse-menu-sidebar-content [class*=menu__item]',
                    'li[role=menuitem]',
                    '[role=treeitem]',
                    '.el-menu-item',
                    '.ant-menu-item',
                    '.t-menu__item',
                    '.t-tabs__nav-item',
                    'a',
                    'button',
                ].join(', ');
                const seen = new Set();
                return Array.from(document.querySelectorAll(selector))
                    .filter(visible)
                    .map((el) => String(el.innerText || el.textContent || '').trim().replace(/\\s+/g, ' '))
                    .filter((text) => text && text.length <= 160)
                    .filter((text) => {
                        if (seen.has(text)) return false;
                        seen.add(text);
                        return true;
                    })
                    .slice(0, limit);
            }""",
            limit,
        )
    except Exception:
        return []
    return [str(item) for item in (items or [])]


def dismiss_known_undeclared_info_dialogs(page) -> str:
    """Close read-only undeclared tax tips that cover the form evidence."""
    try:
        result = page.evaluate(
            """() => {
                const unsafeSubmitHints = [
                    '\u786e\u8ba4\u7ee7\u7eed\u63d0\u4ea4',
                    '\u7ee7\u7eed\u63d0\u4ea4',
                    '\u63d0\u4ea4\u7533\u62a5'
                ];
                const knownTips = [
                    '本表由发生应税交易允许从含税销售额中扣除相关价款后计算销项税额或应纳税额的纳税人'
                ];
                const visible = (el) => {
                    if (!el) return false;
                    const rect = el.getBoundingClientRect();
                    return rect.width > 0 && rect.height > 0;
                };
                const visibleDialogs = Array.from(document.querySelectorAll(
                    '.t-dialog__position, .t-dialog, [role=dialog]'
                )).filter(visible);
                const unsafeDialog = visibleDialogs.find((el) => {
                    const text = el.innerText || el.textContent || '';
                    return unsafeSubmitHints.some((hint) => text.includes(hint));
                });
                if (unsafeDialog) {
                    const rect = unsafeDialog.getBoundingClientRect();
                    const cancelButtons = Array.from(document.querySelectorAll('button, .t-button'))
                        .filter((button) => {
                            if (!visible(button)) return false;
                            const text = (button.innerText || button.textContent || '').trim();
                            if (!['\u53d6\u6d88', '\u5173\u95ed'].includes(text)) return false;
                            const b = button.getBoundingClientRect();
                            const cx = b.left + b.width / 2;
                            const cy = b.top + b.height / 2;
                            return cx >= rect.left && cx <= rect.right && cy >= rect.top && cy <= rect.bottom;
                        });
                    const button = cancelButtons[cancelButtons.length - 1];
                    if (button) {
                        button.click();
                        return 'clicked_cancel_unsafe_submit';
                    }
                }
                const dialogs = Array.from(document.querySelectorAll(
                    '.t-dialog__position, .t-dialog, [role=dialog]'
                )).filter((el) => {
                    if (!visible(el)) return false;
                    const text = el.innerText || el.textContent || '';
                    return knownTips.some((tip) => text.includes(tip));
                }).sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return (ar.width * ar.height) - (br.width * br.height);
                });
                const dialog = dialogs[0];
                if (!dialog) return 'not_found';
                const rect = dialog.getBoundingClientRect();
                const confirmButtons = Array.from(document.querySelectorAll('button, .t-button'))
                    .filter((button) => {
                        if (!visible(button)) return false;
                        const text = (button.innerText || button.textContent || '').trim();
                        if (!['确定', '确认'].includes(text)) return false;
                        const b = button.getBoundingClientRect();
                        const cx = b.left + b.width / 2;
                        const cy = b.top + b.height / 2;
                        return cx >= rect.left && cx <= rect.right && cy >= rect.top && cy <= rect.bottom;
                    });
                const button = confirmButtons[confirmButtons.length - 1];
                if (!button) return 'confirm_not_found';
                button.click();
                return 'clicked';
            }"""
        )
    except Exception as exc:
        LOGGER.debug("Failed to dismiss undeclared tax info dialog: %s", exc)
        return "error"
    if str(result).startswith("clicked"):
        LOGGER.info("Dismissed undeclared tax info dialog: %s", result)
        time.sleep(1)
    return str(result)


def page_text_snippet(page, limit: int = 300) -> str:
    try:
        text = page.evaluate("(limit) => document.body ? document.body.innerText.slice(0, limit) : ''", limit)
    except Exception:
        return ""
    return re.sub(r"\s+", " ", str(text or "")).strip()


def wait_for_undeclared_target_visible(
    page,
    target: CompareTarget,
    menu_keywords: tuple[str, ...],
    timeout: int = 30,
    settle_seconds: int = 4,
    mappings: list[FieldMapping] | None = None,
) -> bool:
    expected = target_title_keywords(target, menu_keywords)
    selected_expected = target_selection_keywords(target, menu_keywords)
    deadline = time.time() + timeout
    while time.time() < deadline:
        selected = is_undeclared_target_selected(page, selected_expected)
        visible = selected and target_content_visible_in_any_scope(page, target, mappings, expected)
        if not visible and selected and mappings:
            try:
                scope = select_target_content_scope(page, target, mappings, expected)
                field_count = count_business_fields(scope, mappings, max_fields=30)
            except Exception:
                field_count = 0
            visible = field_count >= max(3, min(8, len(mappings) // 8))
        if visible:
            LOGGER.info("Undeclared target visible for %s: %s", target.target_id, expected)
            if settle_seconds > 0:
                time.sleep(settle_seconds)
            return True
        time.sleep(1)
    LOGGER.warning("Timed out waiting for undeclared target content: %s keywords=%s", target.target_id, expected)
    return False


def is_undeclared_target_selected(page, expected_keywords: tuple[str, ...]) -> bool:
    try:
        return bool(
            page.evaluate(
                """(keywords) => {
                    const normalize = (value) => String(value || '')
                        .replace(/\\s+/g, '')
                        .replace(/[（）()]/g, '');
                    const wanted = keywords.map(normalize).filter(Boolean);
                    if (!wanted.length) return false;
                    const activeItems = Array.from(document.querySelectorAll(
                        '.gt-collapse-menu-sidebar-content .t-is-active, ' +
                        '.gt-collapse-menu-sidebar-content [class*=active], ' +
                        '.t-is-active, .is-active, .active, [aria-selected=true], ' +
                        '.el-menu-item.is-active, .ant-menu-item-selected'
                    )).filter((el) => el.offsetParent !== null);
                    return activeItems.some((el) => {
                        const text = normalize(el.innerText || el.textContent || '');
                        return wanted.every((kw) => text.includes(kw));
                    });
                }""",
                list(expected_keywords),
            )
        )
    except Exception:
        return False


def is_undeclared_target_content_visible(page, expected_keywords: tuple[str, ...]) -> bool:
    try:
        body_text = page.evaluate(
            """() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const excluded = [
                    '.gt-collapse-menu-sidebar',
                    '.gt-collapse-menu-sidebar-content',
                    '.t-menu',
                    '.t-menu__item',
                    '.t-submenu',
                    'aside',
                    'nav',
                ].join(', ');
                const nodes = Array.from(document.querySelectorAll('body *')).filter((el) => {
                    if (!visible(el) || el.closest(excluded)) return false;
                    const rect = el.getBoundingClientRect();
                    if (rect.width < 120 || rect.height < 8) return false;
                    if (rect.left < 220) return false;
                    const text = String(el.innerText || el.textContent || '').trim();
                    if (!text) return false;
                    return true;
                });
                const text = nodes
                    .map((el) => String(el.innerText || el.textContent || '').trim())
                    .join('\\n');
                return text || (document.body ? document.body.innerText.slice(0, 8000) : '');
            }"""
        )
    except Exception:
        return False
    normalized = re.sub(r"[\s（）()]+", "", str(body_text or ""))
    wanted = [re.sub(r"[\s（）()]+", "", str(keyword)) for keyword in expected_keywords if str(keyword).strip()]
    return bool(wanted) and all(keyword in normalized for keyword in wanted)


def page_scopes_for_content(page) -> list[Any]:
    """Return the main page plus accessible child frames for form detection/extraction."""
    scopes: list[Any] = [page]
    for frame in getattr(page, "frames", []) or []:
        try:
            main_frame = getattr(page, "main_frame", None)
            if main_frame is not None and frame == main_frame:
                continue
        except Exception:
            pass
        scopes.append(frame)
    return scopes


def scope_url(scope) -> str:
    try:
        return str(getattr(scope, "url", "") or "")
    except Exception:
        return ""


def scope_text(scope, limit: int = 12000) -> str:
    try:
        text = scope.evaluate(
            """() => document.body ? String(document.body.innerText || document.body.textContent || '').slice(0, 12000) : ''"""
        )
    except Exception:
        return ""
    return str(text or "")[:limit]


def normalized_form_text(value: Any) -> str:
    return re.sub(r"[\s锛堬級()]+", "", str(value or ""))


def scope_has_expected_keywords(scope, expected_keywords: tuple[str, ...]) -> bool:
    wanted = [normalized_form_text(keyword) for keyword in expected_keywords if str(keyword).strip()]
    if not wanted:
        return False
    text = normalized_form_text(scope_text(scope))
    return all(keyword in text for keyword in wanted)


def target_scope_url_score(scope, target: CompareTarget) -> int:
    url = scope_url(scope).lower()
    score = 0
    if "/form/index.html" in url:
        score -= 30
    if target.target_id == "culture_fee_main" and "bda0610334" in url:
        score += 80
    if target.target_id == "culture_fee_deduction" and "whsyjsf" in url and "index.html" not in url:
        score += 30
    return score


def select_target_content_scope(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping] | None = None,
    expected_keywords: tuple[str, ...] | None = None,
):
    """Pick the page/frame that actually contains the target form content."""
    expected = expected_keywords or target.detail_form_keywords or target.query_keywords or (target.form_name,)
    best_scope = page
    best_score = -1
    best_field_count = 0
    for scope in page_scopes_for_content(page):
        text = scope_text(scope)
        normalized = normalized_form_text(text)
        wanted = [normalized_form_text(keyword) for keyword in expected if str(keyword).strip()]
        keyword_hits = sum(1 for keyword in wanted if keyword and keyword in normalized)
        field_count = 0
        if mappings and (keyword_hits or scope is page):
            try:
                field_count = count_business_fields(scope, mappings, max_fields=30)
            except Exception:
                field_count = 0
        text_score = min(20, len(text) // 500) if keyword_hits else 0
        score = keyword_hits * 100 + field_count + text_score + target_scope_url_score(scope, target)
        if score > best_score:
            best_scope = scope
            best_score = score
            best_field_count = field_count
    if best_scope is not page and best_score > 0:
        LOGGER.info(
            "Selected embedded tax form scope for %s: score=%s fields=%s url=%s",
            target.target_id,
            best_score,
            best_field_count,
            scope_url(best_scope),
        )
        return best_scope
    return page


def target_content_visible_in_any_scope(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping] | None = None,
    expected_keywords: tuple[str, ...] | None = None,
) -> bool:
    expected = expected_keywords or target.detail_form_keywords or target.query_keywords or (target.form_name,)
    if undeclared_target_body_keywords_visible(page, target):
        return True
    if is_undeclared_target_content_visible(page, expected):
        return True
    scope = select_target_content_scope(page, target, mappings, expected)
    if scope is page:
        return False
    if scope_has_expected_keywords(scope, expected):
        return True
    if mappings:
        try:
            return count_business_fields(scope, mappings, max_fields=30) >= max(3, min(8, len(mappings) // 8))
        except Exception:
            return False
    return False


def undeclared_target_body_keywords_visible(page, target: CompareTarget) -> bool:
    body_keywords = UNDECLARED_TARGET_BODY_KEYWORDS.get(target.target_id)
    if not body_keywords:
        return False
    return scope_has_expected_keywords(page, body_keywords)


def target_title_keywords(target: CompareTarget, menu_keywords: tuple[str, ...]) -> tuple[str, ...]:
    if target.target_id == "vat_general_main":
        return ("增值税及附加税费申报表",)
    if "appendix" in target.target_id:
        return menu_keywords
    return target.detail_form_keywords or menu_keywords


def target_selection_keywords(target: CompareTarget, menu_keywords: tuple[str, ...]) -> tuple[str, ...]:
    """Keywords for active left-menu selection; can be stricter than form title text."""
    return menu_keywords or target_title_keywords(target, menu_keywords)


def is_undeclared_target_confirmed(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
) -> bool:
    if undeclared_target_body_keywords_visible(page, target):
        return True
    menu_keywords = UNDECLARED_VAT_MENU_KEYWORDS.get(target.target_id)
    expected = target_title_keywords(target, menu_keywords or ())
    selected_expected = target_selection_keywords(target, menu_keywords or ())
    if menu_keywords:
        selected = is_undeclared_target_selected(page, selected_expected)
        content_visible = target_content_visible_in_any_scope(page, target, mappings, expected)
        if selected and content_visible:
            return True
        embedded_scope = select_target_content_scope(page, target, mappings, expected)
        if embedded_scope is not page and content_visible:
            return True
        if not selected:
            if content_visible and has_visible_form_business_evidence(embedded_scope, target, mappings):
                return True
            if target.tax_type == "CONSUMPTION_TAX" and content_visible:
                try:
                    field_count = count_business_fields(page, mappings, max_fields=30)
                except Exception:
                    field_count = 0
                if field_count >= max(3, min(8, len(mappings) // 8)):
                    LOGGER.info(
                        "Accepting visible consumption-tax form content despite unconfirmed active menu: "
                        "target=%s fields=%s",
                        target.target_id,
                        field_count,
                    )
                    return True
            return False
        try:
            field_count = count_business_fields(embedded_scope, mappings, max_fields=30)
        except Exception:
            field_count = 0
        return bool(mappings) and field_count >= max(3, min(8, len(mappings) // 8))
    return bool(expected) and target_content_visible_in_any_scope(page, target, mappings, expected)


def has_visible_form_business_evidence(scope, target: CompareTarget, mappings: list[FieldMapping]) -> bool:
    """Accept visible undeclared form content without relying solely on active-menu state."""
    if not mappings:
        return False
    required = max(3, min(8, len(mappings) // 8))
    try:
        field_count = count_business_fields(scope, mappings, max_fields=30)
    except Exception:
        return False
    if field_count < required:
        return False
    LOGGER.info(
        "Accepting visible undeclared form content despite unconfirmed active menu: target=%s fields=%s required=%s",
        target.target_id,
        field_count,
        required,
    )
    return True


def is_fast_confirmable_undeclared_menu_target(target: CompareTarget) -> bool:
    """VAT appendices have explicit left-menu entries; a successful click is a strong navigation signal."""
    return target.tax_type == "VAT_GENERAL" and target.target_id.startswith("vat_general_appendix")


def mark_recent_undeclared_menu_target(page, target: CompareTarget) -> None:
    RECENT_UNDECLARED_MENU_TARGETS[id(page)] = (target.target_id, time.time())


def recently_clicked_undeclared_menu_target(page, target: CompareTarget) -> bool:
    if not is_fast_confirmable_undeclared_menu_target(target):
        return False
    recorded = RECENT_UNDECLARED_MENU_TARGETS.get(id(page))
    if not recorded:
        return False
    target_id, clicked_at = recorded
    if target_id != target.target_id:
        return False
    return (time.time() - clicked_at) <= UNDECLARED_JUST_CLICKED_MENU_CONFIRM_SECONDS


def confirm_target_page_for_evidence(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
    current_period_flag: bool | None,
) -> None:
    """Refuse extraction/PDF evidence when the browser is not on the intended form."""
    if current_period_flag is False and supports_undeclared_tax_page(target):
        if recently_clicked_undeclared_menu_target(page, target):
            LOGGER.info(
                "Target page accepted by recent undeclared menu click marker: %s",
                target.target_id,
            )
            return
        if is_undeclared_target_confirmed(page, target, mappings):
            return
        raise RuntimeError(
            f"Target page was not confirmed before extraction/PDF for target={target.target_id}; "
            f"expected={target_title_keywords(target, UNDECLARED_VAT_MENU_KEYWORDS.get(target.target_id) or ())}; "
            f"url={page.url}"
        )

    if is_target_detail_page(page, target, mappings):
        return
    raise RuntimeError(
        f"Target page was not confirmed before extraction/PDF for target={target.target_id}; "
        f"url={page.url}"
    )


def is_target_detail_page(page, target: CompareTarget, mappings: list[FieldMapping]) -> bool:
    try:
        body_text = page.evaluate("document.body ? document.body.innerText.slice(0, 5000) : ''")
    except Exception:
        body_text = ""
    url = (page.url or "").lower()
    detail_context = (
        "sbxxcx/detail" in url
        or "sbxxcxxq" in url
        or ("申报信息查询详情" in body_text and "主附表表单" in body_text)
        or "报表列表" in body_text
    )
    if any(hint in url for hint in QUERY_URL_HINTS) and not detail_context:
        return False
    if target.detail_form_keywords:
        if all(keyword in body_text for keyword in target.detail_form_keywords):
            return detail_context
        if target.tax_type in {"CULTURE_FEE", "CIT_A_PREPAY"}:
            return False
    if not detail_context:
        return False
    return count_business_fields(page, mappings, max_fields=30) >= max(5, min(12, len(mappings) // 4))


def has_any_field(page, mappings: list[FieldMapping]) -> bool:
    for mapping in mappings[:30]:
        if extract_field_value(page, mapping) not in (None, ""):
            return True
    return False


def count_business_fields(page, mappings: list[FieldMapping], max_fields: int | None = None) -> int:
    count = 0
    checked = 0
    for mapping in mappings:
        if mapping.field_id in TEXT_FIELDS:
            continue
        if max_fields is not None and checked >= max_fields:
            break
        checked += 1
        if extract_field_value(page, mapping) not in (None, ""):
            count += 1
    return count


def refresh_declaration_query_results(page, period_range: tuple[str, str] | None = None) -> str:
    """Reset stale query filters and run the declaration query again."""

    period_start, period_end = period_range or ("", "")
    try:
        result = page.evaluate(
            """async (payload) => {
                const periodStart = String(payload?.periodStart || '');
                const periodEnd = String(payload?.periodEnd || '');
                const normalize = (value) => String(value || '').replace(/\\s+/g, '');
                const wait = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
                const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                function findComponents(el, out = []) {
                    if (el.__vue__) out.push(el.__vue__);
                    if (el.__vueParentComponent) out.push(el.__vueParentComponent);
                    for (const child of el.children || []) findComponents(child, out);
                    return out;
                }
                const clickByLabels = (labels) => {
                    const wanted = labels.map(normalize);
                    const nodes = Array.from(document.querySelectorAll(
                        'button, a, [role=button], .el-button, .ant-btn, .t-button'
                    )).filter(visible);
                    const exact = nodes.find((el) => {
                        const text = normalize(el.innerText || el.textContent || el.getAttribute('aria-label'));
                        return wanted.includes(text);
                    });
                    const partial = exact || nodes.find((el) => {
                        const text = normalize(el.innerText || el.textContent || el.getAttribute('aria-label'));
                        return wanted.some((label) => text.includes(label));
                    });
                    if (!partial) return '';
                    partial.scrollIntoView({ block: 'center', inline: 'center' });
                    partial.click();
                    return normalize(partial.innerText || partial.textContent || partial.getAttribute('aria-label'));
                };

                const reset = clickByLabels(['重置', '清空']);
                if (reset) await wait(1000);
                let periodSet = '';
                let comp = null;
                let formData = null;
                if (periodStart && periodEnd) {
                    let dataRoot = null;
                    for (let attempt = 0; attempt < 10 && !formData; attempt += 1) {
                        comp = findComponents(document.body).find((v) => {
                            const name = (v.$options && v.$options.name) || (v.type && v.type.name) || v.name || '';
                            return name === 'sbxxcx';
                        });
                        dataRoot = comp && (comp.$data || comp.ctx || comp.setupState || {});
                        formData = dataRoot && dataRoot.formData;
                        if (!formData) await wait(500);
                    }
                    if (formData) {
                        formData.skssqq = periodStart;
                        formData.skssqz = periodEnd;
                        formData.sbrqq = '';
                        formData.sbrqz = '';
                        formData.pageNum = 1;
                        for (const item of dataRoot.querySearchConfig || []) {
                            if (item && item.key === 'skssqq') item.value = periodStart;
                            if (item && item.key === 'skssqz') item.value = periodEnd;
                            if (item && item.key === 'sbrqq') item.value = '';
                            if (item && item.key === 'sbrqz') item.value = '';
                        }
                        periodSet = 'vue_formData';
                        if (typeof comp.$forceUpdate === 'function') comp.$forceUpdate();
                        if (comp.proxy && typeof comp.proxy.$forceUpdate === 'function') comp.proxy.$forceUpdate();
                        if (typeof comp.$nextTick === 'function') await new Promise((resolve) => comp.$nextTick(resolve));
                        if (comp.proxy && typeof comp.proxy.$nextTick === 'function') await new Promise((resolve) => comp.proxy.$nextTick(resolve));
                    }
                }
                if (periodSet && comp && formData && typeof comp.handlerSearch === 'function') {
                    const maybe = comp.handlerSearch(formData);
                    if (maybe && typeof maybe.then === 'function') await maybe;
                    await wait(5000);
                    return `reset=${reset || 'none'};period=${periodSet || 'none'};query=handlerSearch`;
                }
                const query = clickByLabels(['查询', '搜索']);
                if (periodSet && comp && formData && typeof comp.handlerSearch === 'function') {
                    const maybe = comp.handlerSearch(formData);
                    if (maybe && typeof maybe.then === 'function') await maybe;
                    await wait(5000);
                    return `reset=${reset || 'none'};period=${periodSet || 'none'};query=handlerSearch`;
                } else if (query) {
                    await wait(5000);
                }
                return `reset=${reset || 'none'};period=${periodSet || 'none'};query=${query || 'none'}`;
            }""",
            {
                "periodStart": period_start,
                "periodEnd": period_end,
            },
        )
        LOGGER.info("Declaration query refresh result: %s", result)
        return str(result)
    except Exception as exc:
        LOGGER.warning("Declaration query refresh failed: %s", exc)
        return f"error:{exc}"


def wait_for_declaration_detail_page(page, before_pages: set[Any], timeout: int = 30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        candidates = list(page.context.pages)
        for candidate in candidates:
            try:
                if candidate in before_pages:
                    continue
                if is_declaration_detail_page(candidate):
                    candidate.bring_to_front()
                    return candidate
            except Exception:
                continue
        for candidate in candidates:
            try:
                if is_declaration_detail_page(candidate):
                    candidate.bring_to_front()
                    return candidate
            except Exception:
                continue
        time.sleep(1)
    return page


def legacy_click_declaration_row_once_without_period(page, keywords: tuple[str, ...]) -> str:
    result = page.evaluate(
        """(keywords) => {
            const normalize = (value) => String(value || '')
                .replace(/\\s+/g, '')
                .replace(/[（）()《》]/g, '');
            const wanted = keywords.map(normalize);
            const includesWanted = (value) => {
                const text = normalize(value);
                return wanted.every((kw) => text.includes(kw));
            };

            function findComponents(el, out = []) {
                if (el.__vue__) out.push(el.__vue__);
                for (const child of el.children || []) findComponents(child, out);
                return out;
            }
            const comp = findComponents(document.body).find((v) => v.$options && v.$options.name === 'sbxxcx');
            if (comp && comp.$data && Array.isArray(comp.$data.data)) {
                const rowIndex = comp.$data.data.findIndex((row) => {
                    const text = Object.values(row || {}).map((v) => String(v ?? '')).join('');
                    return includesWanted(text);
                });
                if (rowIndex >= 0 && typeof comp.rehandleClickOp === 'function') {
                    comp.rehandleClickOp(comp.$data.data[rowIndex], rowIndex);
                    return 'clicked_vue';
                }
            }
            const nodes = Array.from(document.querySelectorAll('tr, .el-table__row, li, div'));
            const hit = nodes.find((el) => {
                const text = el.innerText || el.textContent || '';
                return text && text.length < 1000 && includesWanted(text);
            });
            if (!hit) return 'not_found';
            const row = hit.closest('tr, .el-table__row, li, [class*=row]') || hit;
            const buttons = Array.from(row.querySelectorAll('a, button, [role=button], .ant-btn, .el-button'))
                .filter((el) => el.offsetParent !== null);
            const detail = buttons.find((el) => /(查看|详情|申报|查询|打开)/.test(el.innerText || el.textContent || ''));
            (detail || buttons[buttons.length - 1] || hit).click();
            return 'clicked_dom';
        }""",
        list(keywords),
    )
    return str(result)


def legacy_click_declaration_row_without_period(page, keywords: tuple[str, ...]):
    before_pages = set(page.context.pages)
    result = click_declaration_row_once(page, keywords)
    if result == "not_found" and is_query_page(page):
        refresh_declaration_query_results(page, period_range)
        result = click_declaration_row_once(page, keywords)
    if not str(result).startswith("clicked"):
        raise RuntimeError(f"Declaration row was not found for keywords={keywords}; result={result}")
    time.sleep(5)
    detail_page = wait_for_declaration_detail_page(page, before_pages)
    if not is_declaration_detail_page(detail_page):
        raise RuntimeError(f"Declaration row was clicked but detail page did not open; keywords={keywords}; result={result}")
    return detail_page


def click_declaration_row_once(
    page,
    keywords: tuple[str, ...],
    period_range: tuple[str, str] | None = None,
    require_unique: bool = False,
) -> str:
    period_start, period_end = period_range or ("", "")
    payload = {
        "keywords": list(keywords),
        "periodStart": period_start,
        "periodEnd": period_end,
    }
    if require_unique:
        payload["requireUnique"] = True
    result = page.evaluate(
        """(payload) => {
            const keywords = Array.isArray(payload?.keywords) ? payload.keywords : [];
            const periodStart = String(payload?.periodStart || '');
            const periodEnd = String(payload?.periodEnd || '');
            const requireUnique = !!payload?.requireUnique;
            const normalize = (value) => String(value || '')
                .replace(/\\s+/g, '')
                .replace(/[\u3000\u0028\u0029\uff08\uff09\u300a\u300b\u3001\u3002\uff0c\uff1a\uff1b\u3010\u3011\[\]{}]/g, '');
            const digitsOnly = (value) => String(value || '').replace(/\\D/g, '');
            const periodStartDigits = digitsOnly(periodStart);
            const periodEndDigits = digitsOnly(periodEnd);
            const periodMonthDigits = periodStartDigits.slice(0, 6);
            const wanted = keywords.map(normalize);
            const includesWanted = (value) => {
                const text = normalize(value);
                return wanted.every((kw) => text.includes(kw));
            };
            const includesPeriod = (value) => {
                if (!periodStartDigits || !periodEndDigits) return true;
                const digits = digitsOnly(value);
                if (digits.includes(periodStartDigits) && digits.includes(periodEndDigits)) return true;
                return !!periodMonthDigits && digits.includes(periodMonthDigits);
            };
            const matchesTarget = (value) => includesWanted(value) && includesPeriod(value);

            function findComponents(el, out = []) {
                if (el.__vue__) out.push(el.__vue__);
                for (const child of el.children || []) findComponents(child, out);
                return out;
            }
            const comp = findComponents(document.body).find((v) => v.$options && v.$options.name === 'sbxxcx');
            if (comp && comp.$data && Array.isArray(comp.$data.data)) {
                const rowIndexes = comp.$data.data
                    .map((row, index) => {
                        const text = Object.values(row || {}).map((v) => String(v ?? '')).join('');
                        return matchesTarget(text) ? index : -1;
                    })
                    .filter((index) => index >= 0);
                if (requireUnique && rowIndexes.length > 1) return 'multiple_matches';
                const rowIndex = rowIndexes[0] ?? -1;
                if (rowIndex >= 0 && typeof comp.rehandleClickOp === 'function') {
                    comp.rehandleClickOp(comp.$data.data[rowIndex], rowIndex);
                    return 'clicked_vue';
                }
            }
            const nodes = Array.from(document.querySelectorAll('tr, .el-table__row, li, div'));
            const hits = nodes.filter((el) => {
                const text = el.innerText || el.textContent || '';
                return text && text.length < 1000 && matchesTarget(text);
            });
            if (requireUnique && hits.length > 1) return 'multiple_matches';
            const hit = hits[0];
            if (!hit) return periodStartDigits ? 'not_found_period' : 'not_found';
            const row = hit.closest('tr, .el-table__row, li, [class*=row]') || hit;
            const buttons = Array.from(row.querySelectorAll('a, button, [role=button], .ant-btn, .el-button'))
                .filter((el) => el.offsetParent !== null);
            const detailKeywords = ['\u67e5\u770b', '\u8be6\u60c5', '\u7533\u62a5', '\u67e5\u8be2', '\u6253\u5f00'];
            const detail = buttons.find((el) => {
                const text = String(el.innerText || el.textContent || '');
                return detailKeywords.some((keyword) => text.includes(keyword));
            });
            (detail || buttons[buttons.length - 1] || hit).click();
            return 'clicked_dom';
        }""",
        payload,
    )
    return str(result)


def click_declaration_row(
    page,
    keywords: tuple[str, ...],
    period_range: tuple[str, str] | None = None,
    allow_period_fallback: bool = False,
):
    before_pages = set(page.context.pages)
    result = click_declaration_row_once(page, keywords, period_range)
    if result in {"not_found", "not_found_period"} and is_query_page(page):
        refresh_result = refresh_declaration_query_results(page, period_range)
        if period_range and "period=none" in refresh_result:
            LOGGER.info("Declaration query period controls were not ready; retrying refresh after a short wait")
            time.sleep(5)
            refresh_declaration_query_results(page, period_range)
        result = click_declaration_row_once(page, keywords, period_range)
    if result == "not_found_period" and allow_period_fallback:
        LOGGER.info(
            "Declaration row was not found by period; retrying with unique keyword-only match for keywords=%s",
            keywords,
        )
        result = click_declaration_row_once(page, keywords, None, require_unique=True)
    if not str(result).startswith("clicked"):
        period_detail = f"; period={period_range[0]}~{period_range[1]}" if period_range else ""
        raise RuntimeError(f"Declaration row was not found for keywords={keywords}{period_detail}; result={result}")
    time.sleep(5)
    detail_page = wait_for_declaration_detail_page(page, before_pages)
    if not is_declaration_detail_page(detail_page):
        raise RuntimeError(f"Declaration row was clicked but detail page did not open; keywords={keywords}; result={result}")
    return detail_page


def select_detail_form(page, keywords: tuple[str, ...]) -> str:
    result = page.evaluate(
        """async (keywords) => {
            const normalize = (value) => String(value || '')
                .replace(/\\s+/g, '')
                .replace(/[（）()《》]/g, '');
            const wanted = keywords.map(normalize);
            const includesWanted = (value) => {
                const text = normalize(value);
                return wanted.every((kw) => text.includes(kw));
            };

            const body = document.body ? document.body.innerText : '';
            const detailContext = location.href.includes('sbxxcx/detail')
                || location.href.includes('sbxxcxxq')
                || (body.includes('申报信息查询详情') && body.includes('主附表表单'))
                || body.includes('报表列表');
            if (detailContext && includesWanted(body)) return 'already_visible';

            function findVueComponents(el, out = []) {
                if (el.__vue__) out.push(el.__vue__);
                if (el.__vueParentComponent) out.push(el.__vueParentComponent);
                for (const child of el.children || []) findVueComponents(child, out);
                return out;
            }
            const components = findVueComponents(document.body);
            const detailComp = components.find((comp) => {
                const tree = comp && comp.sbmxxqzfbTree;
                return Array.isArray(tree) && typeof comp.selectChange === 'function';
            });
            if (detailComp) {
                const option = detailComp.sbmxxqzfbTree.find((item) => includesWanted(item && item.label));
                if (option) {
                    const beforeHtml = String(detailComp.sbxxcxDetail || '');
                    const ret = detailComp.selectChange(option.value, { option });
                    if (ret && typeof ret.then === 'function') await ret;
                    const deadline = Date.now() + 30000;
                    while (Date.now() < deadline) {
                        await new Promise((resolve) => setTimeout(resolve, 500));
                        const bodyNow = document.body ? document.body.innerText : '';
                        const htmlNow = String(detailComp.sbxxcxDetail || '');
                        if (includesWanted(bodyNow) || (htmlNow !== beforeHtml && includesWanted(htmlNow))) {
                            return `tdesign_vue_select:${option.value}`;
                        }
                    }
                    return `tdesign_vue_select_unconfirmed:${option.value}`;
                }
            }

            const all = Array.from(document.querySelectorAll('*')).filter((el) => el.offsetParent !== null);
            const exact = all.find((el) => {
                if (['HTML', 'BODY'].includes(el.tagName)) return false;
                const text = el.innerText || el.textContent || '';
                if (!text || text.length > 300) return false;
                return includesWanted(text);
            });
            if (exact) {
                exact.click();
                return 'clicked_visible_text';
            }

            const labels = all.filter((el) => /主附表表单|附表|附列资料/.test(el.innerText || el.textContent || ''));
            for (const label of labels) {
                const container = label.closest('.el-form-item, .ant-form-item, tr, div') || label.parentElement;
                const trigger = container && Array.from(container.querySelectorAll('input, .el-select, .ant-select, .t-select, .t-select-input, button, [role=combobox]'))
                    .find((el) => el.offsetParent !== null);
                if (trigger) {
                    trigger.click();
                    break;
                }
            }

            const options = Array.from(document.querySelectorAll('.el-select-dropdown__item, .ant-select-item, .t-select-option, .t-option, li, [role=option]'))
                .filter((el) => el.offsetParent !== null);
            const option = options.find((el) => {
                return includesWanted(el.innerText || el.textContent || '');
            });
            if (option) {
                option.click();
                return 'selected_option';
            }
            return 'not_found';
        }""",
        list(keywords),
    )
    LOGGER.info("Detail form select result for %s: %s", keywords, result)
    time.sleep(2)
    return str(result)


def extract_field_value(page, mapping: FieldMapping) -> Any:
    return page.evaluate(
        """(mapping) => {
            const fieldId = mapping.field_id;
            const cssEscape = window.CSS && CSS.escape ? CSS.escape(fieldId) : fieldId.replace(/([:.\\[\\],=])/g, '\\\\$1');
            const selectors = [
                `#${cssEscape}`,
                `[name="${fieldId}"]`,
                `[data-cell-id="${fieldId}"]`,
                `[data-field-id="${fieldId}"]`,
                `[data-field="${fieldId}"]`,
                `[id$="${fieldId}"]`
            ];
            for (const selector of selectors) {
                const el = document.querySelector(selector);
                if (!el) continue;
                const tag = el.tagName.toLowerCase();
                if (tag === 'input' || tag === 'textarea' || tag === 'select') return el.value ?? '';
                const value = el.getAttribute('value');
                if (value !== null && value !== '') return value;
                const text = (el.innerText || el.textContent || '').trim();
                if (text !== '') return text;
            }

            function findComponents(el, out = []) {
                if (!el) return out;
                if (el.__vue__) out.push(el.__vue__);
                if (el.__vueParentComponent) out.push(el.__vueParentComponent);
                for (const child of el.children || []) findComponents(child, out);
                return out;
            }
            function isScalarValue(value) {
                return ['string', 'number', 'boolean'].includes(typeof value);
            }
            function findKey(obj, key, depth = 0, seen = new Set()) {
                try {
                    if (!obj || typeof obj !== 'object' || depth > 6 || seen.has(obj)) return undefined;
                    if (obj === window || obj === document || obj.nodeType) return undefined;
                    seen.add(obj);
                    if (Object.prototype.hasOwnProperty.call(obj, key)) return obj[key];
                    const keys = Object.keys(obj).slice(0, 160);
                    for (const k of keys) {
                        let value;
                        try { value = obj[k]; } catch (_) { continue; }
                        if (isScalarValue(value)) continue;
                        const found = findKey(value, key, depth + 1, seen);
                        if (found !== undefined) return found;
                    }
                } catch (_) {
                    return undefined;
                }
                return undefined;
            }
            const vueExactFieldValue = (() => {
                if (mapping.form_code !== 'VAT_GENERAL_APPENDIX1') return null;
                for (const comp of findComponents(document.body)) {
                    const roots = [
                        comp && comp.$data,
                        comp && comp._data,
                        comp && comp.$props,
                        comp && comp.props,
                        comp && comp.setupState,
                        comp && comp.ctx,
                    ];
                    for (const root of roots) {
                        const value = findKey(root, fieldId);
                        if (value === undefined || value === null) continue;
                        if (isScalarValue(value)) return String(value).trim();
                    }
                }
                return null;
            })();
            if (vueExactFieldValue !== null) return vueExactFieldValue;

            const vueVatValue = (() => {
                const match = fieldId.match(/^(.*)_(ybxm_bys|ybxm_bnlj|jzjtxm_bys|jzjtxm_bnlj)$/);
                if (!match) return null;
                const suffixToLine = { ybxm_bys: '1', ybxm_bnlj: '2', jzjtxm_bys: '3', jzjtxm_bnlj: '4' };
                const alias = {
                    cswhjssbqybtse: 'bqybtsecjs',
                    jyffjbqybtse: 'bqybtsejyfj',
                    dfjyfjbqybtse: 'bqybtsedfjyfj',
                };
                const property = alias[match[1]] || match[1];
                const line = suffixToLine[match[2]];

                for (const comp of findComponents(document.body)) {
                    const rows = findKey(comp.$data, 'sbZzsYbnsr') || findKey(comp, 'sbZzsYbnsr');
                    if (!Array.isArray(rows)) continue;
                    const row = rows.find((item) => String(item && item.ewblxh) === line);
                    if (!row || row[property] === undefined || row[property] === null) continue;
                    return String(row[property]);
                }
                return null;
            })();
            if (vueVatValue !== null) return vueVatValue;

            if (mapping.line_no !== '' && mapping.web_col_index !== null && mapping.web_col_index !== undefined) {
                const wantedLine = String(mapping.line_no).match(/^\\d+/)?.[0];
                const normalizeLabel = (value) => String(value || '')
                    .replace(/\\s+/g, '')
                    .replace(/^其中[:：]?/, '')
                    .replace(/[（]/g, '(')
                    .replace(/[）]/g, ')')
                    .replace(/[，]/g, ',')
                    .replace(/[＝]/g, '=');
                const wantedName = normalizeLabel(mapping.row_name);
                const rows = Array.from(document.querySelectorAll('table tr'));
                const readCellValue = (cell) => {
                    const control = cell.querySelector('input, textarea, select');
                    if (control) return String(control.value ?? '').trim();
                    const attrValue = cell.getAttribute('value');
                    if (attrValue !== null && attrValue !== '') return String(attrValue).trim();
                    return String(cell.innerText || cell.textContent || '').trim();
                };
                const hasControlValue = (cell) => !!cell.querySelector('input, textarea, select');
                for (const tr of rows) {
                    const cells = Array.from(tr.querySelectorAll('th,td'));
                    const texts = cells.map((td) => (td.innerText || td.textContent || '').trim());
                    const rowText = normalizeLabel(texts.join(''));
                    const lineIndex = texts.findIndex((text, idx) => {
                        if (idx > 4) return false;
                        const m = text.match(/^(FZ\\d+|\\d+(?:\\.\\d+)?)(.*)$/);
                        if (!m || m[1] !== wantedLine) return false;
                        const rest = (m[2] || '').trim();
                        return rest === '' || /^[()（）=+\\-×\\\\\\s]/.test(rest);
                    });
                    if (lineIndex < 0) continue;
                    if (wantedName && !rowText.includes(wantedName) && mapping.form_code !== 'CULTURE_FEE_MAIN') continue;
                    const offset = Number(mapping.web_col_index);
                    const looksAmount = (value) => {
                        const text = String(value || '').trim().replace(/[,\\s]/g, '');
                        return text === ''
                            || text === '——'
                            || text === '-'
                            || /^-?\d+(\.\d+)?%?$/.test(text);
                    };
                    const mainSuffix = fieldId.match(/_(ybxm_bys|ybxm_bnlj|jzjtxm_bys|jzjtxm_bnlj)$/);
                    if (mapping.form_code === 'VAT_GENERAL_MAIN' && mainSuffix) {
                        const relativeOffset = {
                            ybxm_bys: 1,
                            ybxm_bnlj: 2,
                            jzjtxm_bys: 3,
                            jzjtxm_bnlj: 4,
                        }[mainSuffix[1]];
                        const mainIndex = lineIndex + relativeOffset;
                        if (mainIndex >= 0 && mainIndex < cells.length) {
                            const value = (cells[mainIndex].innerText || cells[mainIndex].textContent || '').trim();
                            if (looksAmount(value)) return value;
                        }
                    }
                    const appendix2Field = fieldId.match(/^(fs|je|se)_/);
                    if (appendix2Field) {
                        const relativeOffset = { fs: 1, je: 2, se: 3 }[appendix2Field[1]];
                        const appendix2Index = lineIndex + relativeOffset;
                        if (appendix2Index >= 0 && appendix2Index < cells.length) {
                            const value = (cells[appendix2Index].innerText || cells[appendix2Index].textContent || '').trim();
                            if (looksAmount(value)) return value;
                        }
                    }
                    if (mapping.form_code === 'CULTURE_FEE_MAIN') {
                        const cultureFeeSuffix = fieldId.match(/_(bys|bnlj)$/);
                        if (cultureFeeSuffix) {
                            const valueCells = cells.slice(lineIndex + 1).filter((cell) => {
                                const value = readCellValue(cell);
                                if (hasControlValue(cell)) return true;
                                return looksAmount(value);
                            });
                            const valueIndex = { bys: 0, bnlj: 1 }[cultureFeeSuffix[1]];
                            if (valueIndex >= 0 && valueIndex < valueCells.length) {
                                const value = readCellValue(valueCells[valueIndex]);
                                if (mapping.data_type !== 'amount' || looksAmount(value)) return value;
                            }
                        }
                    }
                    if (mapping.form_code === 'VAT_GENERAL_APPENDIX1') {
                        const appendix1Index = lineIndex + offset - 5;
                        if (appendix1Index >= 0 && appendix1Index < cells.length) {
                            const value = (cells[appendix1Index].innerText || cells[appendix1Index].textContent || '').trim();
                            if (looksAmount(value)) return value;
                        }
                    }
                    const candidateIndexes = [
                        offset - 1,
                        lineIndex + offset - 1,
                        lineIndex + 1 + Math.max(0, offset - 4),
                        offset - 2,
                        cells.length - 1,
                    ];
                    const seenIndexes = new Set();
                    for (const index of candidateIndexes) {
                        if (index < 0 || index >= cells.length || seenIndexes.has(index)) continue;
                        seenIndexes.add(index);
                        const cell = cells[index];
                        const value = readCellValue(cell);
                        if (mapping.data_type !== 'amount' || looksAmount(value)) return value;
                    }
                }
                const mainSuffix = fieldId.match(/_(ybxm_bys|ybxm_bnlj|jzjtxm_bys|jzjtxm_bnlj)$/);
                if (mapping.form_code === 'VAT_GENERAL_MAIN' && mainSuffix && wantedLine) {
                    const looksAmount = (value) => {
                        const text = String(value || '').trim().replace(/[,\\s]/g, '');
                        return text === ''
                            || text === '——'
                            || text === '-'
                            || /^-?\d+(\.\d+)?%?$/.test(text);
                    };
                    const relativeOffset = {
                        ybxm_bys: 1,
                        ybxm_bnlj: 2,
                        jzjtxm_bys: 3,
                        jzjtxm_bnlj: 4,
                    }[mainSuffix[1]];
                    const linePattern = new RegExp(`^${wantedLine}(?!\\\\d)`);
                    for (const tr of rows) {
                        const cells = Array.from(tr.querySelectorAll('th,td'));
                        const texts = cells.map((td) => (td.innerText || td.textContent || '').trim());
                        const lineIndex = texts.findIndex((text, idx) => idx <= 2 && linePattern.test(normalizeLabel(text)));
                        if (lineIndex < 0) continue;
                        const mainIndex = lineIndex + relativeOffset;
                        if (mainIndex >= 0 && mainIndex < cells.length) {
                            const value = (cells[mainIndex].innerText || cells[mainIndex].textContent || '').trim();
                            if (looksAmount(value)) return value;
                        }
                    }
                }
            }
            return null;
        }""",
        mapping.model_dump(mode="json"),
    )


def extract_web_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    data = {mapping.field_id: extract_field_value(page, mapping) for mapping in mappings}
    found = sum(1 for value in data.values() if value not in (None, ""))
    LOGGER.info("Extracted %s/%s web fields", found, len(mappings))
    return data


def extract_web_data_with_scroll_retry(page, target: CompareTarget, mappings: list[FieldMapping]) -> dict[str, Any]:
    data = extract_web_data(page, mappings)
    if target.target_id not in WEB_EXTRACTION_SCROLL_RETRY_TARGETS:
        return data

    found, total, ratio = web_extraction_coverage(data, mappings)
    retry_threshold = WEB_EXTRACTION_MIN_COVERAGE_BY_TARGET.get(
        target.target_id,
        WEB_EXTRACTION_DEFAULT_MIN_COVERAGE,
    )
    if total <= 0 or ratio >= retry_threshold:
        return data
    if found > 0:
        LOGGER.info(
            "%s web extraction coverage is low after first read: %s/%s (%.2f%%); "
            "deferring targeted recovery to comparable web_missing fields",
            target.target_id,
            found,
            total,
            ratio * 100,
        )
        return data

    missing = [mapping for mapping in mappings if data.get(mapping.field_id) in (None, "")]
    LOGGER.warning(
        "%s web extraction coverage is low before scroll retry: %s/%s (%.2f%%); retrying %s missing field(s)",
        target.target_id,
        found,
        total,
        ratio * 100,
        len(missing),
    )
    for vertical_ratio in (0, 0.25, 0.5, 0.75, 1):
        for horizontal_ratio in (0, 0.5, 1):
            scroll_page_for_extraction(page, vertical_ratio, horizontal_ratio)
            if vertical_ratio or horizontal_ratio:
                time.sleep(0.25)
            retry_mappings = [mapping for mapping in missing if data.get(mapping.field_id) in (None, "")]
            if not retry_mappings:
                restore_page_scroll_for_extraction(page)
                return data
            partial = extract_web_data(page, retry_mappings)
            for field_id, value in partial.items():
                if value not in (None, "") and data.get(field_id) in (None, ""):
                    data[field_id] = value

    restore_page_scroll_for_extraction(page)
    found_after, _, ratio_after = web_extraction_coverage(data, mappings)
    LOGGER.info(
        "%s web extraction coverage after scroll retry: %s/%s (%.2f%%)",
        target.target_id,
        found_after,
        total,
        ratio_after * 100,
    )
    return data


def scroll_page_for_extraction(page, vertical_ratio: float, horizontal_ratio: float) -> None:
    page.evaluate(
        """({verticalRatio, horizontalRatio}) => {
            const clamp = (value) => Math.max(0, Math.min(1, Number(value) || 0));
            const yRatio = clamp(verticalRatio);
            const xRatio = clamp(horizontalRatio);
            const scrollables = Array.from(document.querySelectorAll('*')).filter((el) => {
                const style = getComputedStyle(el);
                const canScroll = /(auto|scroll)/.test(style.overflow + style.overflowY + style.overflowX);
                return canScroll && (el.scrollHeight > el.clientHeight + 8 || el.scrollWidth > el.clientWidth + 8);
            });
            for (const el of scrollables) {
                try {
                    el.scrollTop = Math.max(0, el.scrollHeight - el.clientHeight) * yRatio;
                    el.scrollLeft = Math.max(0, el.scrollWidth - el.clientWidth) * xRatio;
                } catch (_) {}
            }
            window.scrollTo({
                top: Math.max(0, document.documentElement.scrollHeight - window.innerHeight) * yRatio,
                left: Math.max(0, document.documentElement.scrollWidth - window.innerWidth) * xRatio,
                behavior: 'instant',
            });
        }""",
        {"verticalRatio": vertical_ratio, "horizontalRatio": horizontal_ratio},
    )


def restore_page_scroll_for_extraction(page) -> None:
    try:
        scroll_page_for_extraction(page, 0, 0)
    except Exception:
        pass


def web_value_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def merge_web_data_fill_missing(base: dict[str, Any], supplement: dict[str, Any]) -> dict[str, Any]:
    merged = dict(base)
    for field_id, value in supplement.items():
        if web_value_present(value) and not web_value_present(merged.get(field_id)):
            merged[field_id] = value
    return merged


def web_missing_field_ids_for_compare(
    target: CompareTarget,
    mappings: list[FieldMapping],
    api_by_tax: dict[str, Any],
    web_raw: dict[str, Any],
) -> list[str]:
    result = compare_target(target, mappings, api_by_tax, web_raw)
    return [
        field.field_id
        for field in result.field_results
        if getattr(field.status, "value", field.status) == "web_missing"
    ]


def extraction_scopes_for_target(page, target: CompareTarget, mappings: list[FieldMapping]) -> list[Any]:
    selected = select_target_content_scope(page, target, mappings)
    scopes: list[Any] = []
    seen: set[int] = set()
    for scope in [selected, *page_scopes_for_content(page)]:
        key = id(scope)
        if key in seen:
            continue
        seen.add(key)
        scopes.append(scope)
    return scopes


def wait_for_render_cycle(scope) -> None:
    try:
        scope.evaluate(
            """() => new Promise((resolve) => {
                let done = false;
                const finish = () => {
                    if (done) return;
                    done = true;
                    resolve();
                };
                setTimeout(finish, 800);
                requestAnimationFrame(() => requestAnimationFrame(finish));
            })"""
        )
    except Exception:
        pass


def scope_has_visible_loading_for_extraction(scope) -> bool:
    try:
        return bool(
            scope.evaluate(
                """() => {
                    const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                    const selectors = [
                        '.el-loading-mask',
                        '.ant-spin-spinning',
                        '.t-loading',
                        '.vxe-loading',
                        '[class*="loading"]',
                        '[class*="Loading"]'
                    ];
                    return selectors.some((selector) => Array.from(document.querySelectorAll(selector)).some(visible));
                }"""
            )
        )
    except Exception:
        return False


def wait_for_web_extraction_settle(page, target: CompareTarget, mappings: list[FieldMapping], timeout: float = 8) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        scopes = extraction_scopes_for_target(page, target, mappings)
        if not any(scope_has_visible_loading_for_extraction(scope) for scope in scopes):
            for scope in scopes:
                wait_for_render_cycle(scope)
            return
        time.sleep(0.5)


def force_target_render_for_extraction(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
    attempt: int,
) -> None:
    if supports_undeclared_tax_page(target):
        dismiss_known_undeclared_info_dialogs(page)
    started = time.time()
    deadline = started + min(WEB_EXTRACTION_RECOVERY_MAX_SECONDS, 18 + attempt * 6)
    wait_for_web_extraction_settle(page, target, mappings, timeout=4 + attempt)
    scopes = extraction_scopes_for_target(page, target, mappings)
    for vertical_ratio, horizontal_ratio in WEB_EXTRACTION_RECOVERY_SCROLL_POSITIONS:
        if time.time() >= deadline:
            LOGGER.warning(
                "%s web extraction render recovery attempt %s reached %.1fs budget; stopping scroll sweep",
                target.target_id,
                attempt,
                time.time() - started,
            )
            return
        for scope in scopes:
            try:
                scroll_page_for_extraction(scope, vertical_ratio, horizontal_ratio)
                wait_for_render_cycle(scope)
            except Exception as exc:
                LOGGER.warning(
                    "%s web extraction render recovery attempt %s stopped at scroll=(%.2f,%.2f): %s",
                    target.target_id,
                    attempt,
                    vertical_ratio,
                    horizontal_ratio,
                    exc,
                )
                return
        time.sleep(0.12)
    wait_for_web_extraction_settle(page, target, mappings, timeout=3)


def extract_missing_web_fields_after_scroll_sweep(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
) -> dict[str, Any]:
    if not mappings:
        return {}
    scopes = extraction_scopes_for_target(page, target, mappings)
    data = {mapping.field_id: None for mapping in mappings}
    started = time.time()
    deadline = started + WEB_EXTRACTION_RECOVERY_MAX_SECONDS
    try:
        for vertical_ratio, horizontal_ratio in WEB_EXTRACTION_RECOVERY_SCROLL_POSITIONS:
            if time.time() >= deadline:
                LOGGER.warning(
                    "%s missing-field extraction scroll sweep reached %.1fs budget; remaining=%s",
                    target.target_id,
                    time.time() - started,
                    ",".join(
                        mapping.field_id
                        for mapping in mappings
                        if not web_value_present(data.get(mapping.field_id))
                    )[:300],
                )
                return data
            for scope in scopes:
                try:
                    scroll_page_for_extraction(scope, vertical_ratio, horizontal_ratio)
                    wait_for_render_cycle(scope)
                except Exception as exc:
                    LOGGER.warning(
                        "%s missing-field extraction scroll sweep stopped at scroll=(%.2f,%.2f): %s",
                        target.target_id,
                        vertical_ratio,
                        horizontal_ratio,
                        exc,
                    )
                    return data
            time.sleep(0.12)
            retry_mappings = [mapping for mapping in mappings if not web_value_present(data.get(mapping.field_id))]
            if not retry_mappings:
                return data
            for scope in scopes:
                try:
                    partial = extract_web_data_for_target(scope, target, retry_mappings)
                except Exception as exc:
                    LOGGER.warning("%s missing-field extraction read stopped: %s", target.target_id, exc)
                    return data
                data = merge_web_data_fill_missing(data, partial)
                retry_mappings = [
                    mapping for mapping in mappings if not web_value_present(data.get(mapping.field_id))
                ]
                if not retry_mappings:
                    return data
        return data
    finally:
        for scope in scopes:
            restore_page_scroll_for_extraction(scope)


def extract_web_data_reliably(
    page,
    target: CompareTarget,
    mappings: list[FieldMapping],
    comparison_mappings: list[FieldMapping],
    api_by_tax: dict[str, Any],
    current_period_flag: bool | None = None,
) -> dict[str, Any]:
    extraction_mappings = web_extraction_mappings_for_compare(target, mappings, comparison_mappings)
    data = extract_web_data_for_target(page, target, extraction_mappings)
    recovery_started = time.time()
    recovery_deadline = recovery_started + WEB_EXTRACTION_RECOVERY_MAX_SECONDS
    for attempt in range(1, WEB_EXTRACTION_RECOVERY_ATTEMPTS + 1):
        if time.time() >= recovery_deadline:
            LOGGER.warning(
                "%s web extraction recovery reached %.1fs budget before attempt %s; continuing with current data",
                target.target_id,
                time.time() - recovery_started,
                attempt,
            )
            break
        low_coverage = is_low_web_extraction_coverage(target, data, comparison_mappings)
        missing_field_ids = web_missing_field_ids_for_compare(target, comparison_mappings, api_by_tax, data)
        if not low_coverage and not missing_field_ids:
            return data
        if low_coverage and not missing_field_ids:
            found, total, ratio = web_extraction_coverage(data, comparison_mappings)
            LOGGER.info(
                "%s web extraction coverage is below threshold but has no comparable web_missing; "
                "skipping expensive recovery: coverage=%s/%s (%.2f%%)",
                target.target_id,
                found,
                total,
                ratio * 100,
            )
            return data

        found, total, ratio = web_extraction_coverage(data, comparison_mappings)
        LOGGER.warning(
            "%s web extraction incomplete before recovery attempt %s/%s: coverage=%s/%s (%.2f%%), web_missing=%s%s",
            target.target_id,
            attempt,
            WEB_EXTRACTION_RECOVERY_ATTEMPTS,
            found,
            total,
            ratio * 100,
            len(missing_field_ids),
            f" fields={','.join(missing_field_ids[:12])}" if missing_field_ids else "",
        )
        missing_field_id_set = set(missing_field_ids)
        missing_mappings = [
            mapping for mapping in comparison_mappings if mapping.field_id in missing_field_id_set
        ]
        render_scope_mappings = missing_mappings or extraction_mappings[:30]
        try:
            force_target_render_for_extraction(page, target, render_scope_mappings, attempt)
            retry_data = extract_web_data_for_target(page, target, extraction_mappings)
        except Exception as exc:
            LOGGER.warning(
                "%s web extraction recovery attempt %s stopped after %.1fs: %s",
                target.target_id,
                attempt,
                time.time() - recovery_started,
                exc,
            )
            break
        data = merge_web_data_fill_missing(data, retry_data)

        missing_field_ids = web_missing_field_ids_for_compare(target, comparison_mappings, api_by_tax, data)
        if missing_field_ids:
            if time.time() >= recovery_deadline:
                LOGGER.warning(
                    "%s web extraction recovery budget exhausted before missing-field sweep; remaining=%s",
                    target.target_id,
                    ",".join(missing_field_ids[:30]),
                )
                break
            missing_field_id_set = set(missing_field_ids)
            missing_mappings = [
                mapping for mapping in comparison_mappings if mapping.field_id in missing_field_id_set
            ]
            try:
                supplement = extract_missing_web_fields_after_scroll_sweep(page, target, missing_mappings)
            except Exception as exc:
                LOGGER.warning("%s web extraction missing-field sweep stopped: %s", target.target_id, exc)
                break
            data = merge_web_data_fill_missing(data, supplement)

    missing_field_ids = web_missing_field_ids_for_compare(target, comparison_mappings, api_by_tax, data)
    if missing_field_ids:
        found, total, ratio = web_extraction_coverage(data, comparison_mappings)
        LOGGER.error(
            "%s web extraction still has missing fields after recovery: coverage=%s/%s (%.2f%%), fields=%s",
            target.target_id,
            found,
            total,
            ratio * 100,
            ",".join(missing_field_ids[:30]),
        )
    return data


def web_extraction_mappings_for_compare(
    target: CompareTarget,
    mappings: list[FieldMapping],
    comparison_mappings: list[FieldMapping],
) -> list[FieldMapping]:
    field_ids = {mapping.field_id for mapping in comparison_mappings}
    mapping_by_id = {mapping.field_id: mapping for mapping in mappings}
    if target.target_id == "vat_general_main":
        for mapping in comparison_mappings:
            alias_source = VAT_GENERAL_MAIN_WEB_VALUE_ALIASES.get(mapping.field_id)
            if alias_source:
                field_ids.add(alias_source)
            if should_force_subtract_current_month_for_vat(target, mapping) or should_subtract_current_month_for_vat(
                target,
                mapping,
            ):
                field_ids.add(mapping.field_id.removesuffix("_ybxm_bnlj") + "_ybxm_bys")
    selected = [mapping_by_id[field_id] for field_id in field_ids if field_id in mapping_by_id]
    selected.sort(key=lambda item: mappings.index(item))
    return selected


VAT_GENERAL_MAIN_WEB_VALUE_ALIASES = {
    "qmwjse_ybxm_bys": "qcwjse_ybxm_bys",
    "qmldse_ybxm_bys": "sqldse_ybxm_bys",
    "qmldse_ybxm_bnlj": "sqldse_ybxm_bnlj",
    "qmldse_jzjtxm_bys": "sqldse_jzjtxm_bys",
    "qmldse_jzjtxm_bnlj": "sqldse_jzjtxm_bnlj",
}

VAT_GENERAL_MAIN_FORCE_SUBTRACT_CURRENT_MONTH_FIELDS = {
    "qmwjse_ybxm_bnlj",
}

VAT_GENERAL_MAIN_SUBTRACT_CURRENT_MONTH_SUFFIXES = (
    ("_ybxm_bnlj", "_ybxm_bys"),
    ("_jzjtxm_bnlj", "_jzjtxm_bys"),
)

VAT_GENERAL_MAIN_CUMULATIVE_DIRECT_FIELDS = {
    "qcwjse_ybxm_bnlj",
    "qmldse_ybxm_bnlj",
    "qmldse_jzjtxm_bnlj",
    "sjdkse_jzjtxm_bnlj",
}


def apply_target_web_value_rules(target: CompareTarget, web_raw: dict[str, Any]) -> dict[str, Any]:
    if target.target_id != "vat_general_main":
        return web_raw
    for target_field, source_field in VAT_GENERAL_MAIN_WEB_VALUE_ALIASES.items():
        if target_field in web_raw:
            web_raw[target_field] = web_raw.get(source_field)
    if "sjdkse_jzjtxm_bnlj" in web_raw and web_value_looks_zero(web_raw.get("sjdkse_jzjtxm_bnlj")):
        web_raw["sjdkse_jzjtxm_bnlj"] = web_raw.get("sjdkse_jzjtxm_bys")
    derived_qcwjse = derive_vat_general_jizhengjitui_opening_unpaid(web_raw)
    if derived_qcwjse is not None:
        web_raw["qcwjse_jzjtxm_bnlj"] = derived_qcwjse
    return web_raw


def web_value_looks_zero(value: Any) -> bool:
    parsed = parse_amount(value)
    return parsed == Decimal("0") if parsed is not None else False


def derive_vat_general_jizhengjitui_opening_unpaid(web_raw: dict[str, Any]) -> str | None:
    ending_unpaid = parse_amount(web_raw.get("qmwjse_jzjtxm_bys"))
    tax_due_total = parse_amount(web_raw.get("ynsehj_jzjtxm_bys"))
    paid_previous = parse_amount(web_raw.get("bqjnsqynse_jzjtxm_bnlj"))
    if ending_unpaid is None or tax_due_total is None or paid_previous is None:
        return None
    value = ending_unpaid + paid_previous - tax_due_total
    return f"{value.quantize(Decimal('0.01'))}"


VAT_SMALL_APPENDIX1_BLOCK_FIELDS = (
    ("qcye", "bqfse", "bqkce", "qmye"),
    ("ysfwxsqbhssr", "ysfwxsbqkce", "ysfwxshsxse", "ysfwxsbhsxse"),
    ("qcye5", "bqfse5", "bqkce5", "qmye5"),
    ("ysfwxsqbhssr5", "ysfwxsbqkce5", "ysfwxshsxse5", "ysfwxsbhsxse5"),
)

VAT_SMALL_APPENDIX2_PREFIX_TO_VALUE_INDEX = {
    "jsyjzzsybzzs": 0,
    "zzsxe": 1,
    "slzsl": 2,
    "bqynse": 3,
    "bqjmse": 4,
    "xwqylslfjmzcjzbl": 5,
    "xwqylslfjmzcjze": 6,
    "bqyjse": 7,
    "bqybtse": 8,
}

VAT_SMALL_MAIN_SUFFIX_TO_VALUE_INDEX = {
    "_hwjlw_bqs": 0,
    "_fwjbdc_bqs": 1,
    "_hwjlw_bnlj": 2,
    "_fwjbdc_bnlj": 3,
}


def extract_body_text(page) -> str:
    return str(page.evaluate("document.body ? document.body.innerText : ''") or "")


VAT_GENERAL_APPENDIX1_GRID_BASE_COL = 6


def extract_vat_general_appendix1_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    parsed = parse_vat_general_appendix1_rows(extract_table_text_rows(page), mappings)
    return {mapping.field_id: parsed.get(mapping.field_id) for mapping in mappings}


def parse_vat_general_appendix1_text(text: str, mappings: list[FieldMapping]) -> dict[str, Any]:
    lines = normalize_page_text_lines(text)
    row_cache: dict[tuple[str, str], list[str]] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        row_name = str(mapping.row_name or "").strip()
        if not line_no or not row_name or mapping.web_col_index is None:
            continue
        key = (line_no, row_name)
        if key in row_cache:
            continue
        row_cache[key] = vat_general_appendix1_values_for_row(lines, line_no, row_name)

    data: dict[str, Any] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        row_name = str(mapping.row_name or "").strip()
        if not line_no or not row_name or mapping.web_col_index is None:
            continue
        values = row_cache.get((line_no, row_name), [])
        value_index = int(mapping.web_col_index) - VAT_GENERAL_APPENDIX1_GRID_BASE_COL
        if 0 <= value_index < len(values):
            data[mapping.field_id] = values[value_index]
    return data


def vat_general_appendix1_values_for_row(lines: list[str], line_no: str, row_name: str) -> list[str]:
    wanted_row = compact_label(row_name)
    for line in lines:
        if wanted_row not in compact_label(line):
            continue
        tokens = amount_tokens(line)
        if not tokens:
            continue
        start_index = 0
        rate_match = re.match(r"^(\d+(?:\.\d+)?)%", row_name)
        if rate_match and tokens and tokens[0].replace(",", "") == rate_match.group(1):
            start_index = 1
        line_index = next(
            (
                index
                for index in range(start_index, len(tokens))
                if tokens[index].replace(",", "") == line_no
            ),
            -1,
        )
        if line_index < 0:
            continue
        return tokens[line_index + 1 : line_index + 15]
    return []


def extract_table_text_rows(page) -> list[list[str]]:
    try:
        rows = page.evaluate(
            """() => Array.from(document.querySelectorAll('tr'))
                .map((tr) => Array.from(tr.querySelectorAll('th,td'))
                    .map((cell) => String(cell.textContent || '').trim())
                    .filter((text) => text !== ''))
                .filter((cells) => cells.length > 0)"""
        )
    except Exception:
        return []
    return [[str(cell) for cell in row] for row in rows if isinstance(row, list)]


def parse_vat_general_appendix1_rows(rows: list[list[str]], mappings: list[FieldMapping]) -> dict[str, Any]:
    row_cache: dict[tuple[str, str], list[str]] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        row_name = str(mapping.row_name or "").strip()
        if not line_no or not row_name or mapping.web_col_index is None:
            continue
        key = (line_no, row_name)
        if key in row_cache:
            continue
        row_cache[key] = vat_general_appendix1_values_for_cells(rows, line_no, row_name)

    data: dict[str, Any] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        row_name = str(mapping.row_name or "").strip()
        if not line_no or not row_name or mapping.web_col_index is None:
            continue
        values = row_cache.get((line_no, row_name), [])
        value_index = int(mapping.web_col_index) - VAT_GENERAL_APPENDIX1_GRID_BASE_COL
        if 0 <= value_index < len(values):
            data[mapping.field_id] = values[value_index]
    return data


def vat_general_appendix1_values_for_cells(rows: list[list[str]], line_no: str, row_name: str) -> list[str]:
    wanted_row = compact_label(row_name)
    for cells in rows:
        compact_cells = [compact_label(cell) for cell in cells]
        if not any(wanted_row and wanted_row in cell for cell in compact_cells):
            continue
        line_index = next(
            (
                index
                for index, cell in enumerate(cells)
                if re.fullmatch(r"\d+(?:\.\d+)?", str(cell).replace(",", "").strip())
                and str(cell).replace(",", "").strip() == line_no
            ),
            -1,
        )
        if line_index < 0:
            continue
        values = [normalize_appendix3_amount_cell(cell) for cell in cells[line_index + 1 : line_index + 15]]
        return values
    return []


def extract_vat_small_appendix1_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    parsed = parse_vat_small_appendix1_text(extract_body_text(page), mappings)
    fallback = extract_web_data_with_scroll_retry(
        page,
        TARGETS["vat_small_appendix1"],
        [mapping for mapping in mappings if mapping.field_id not in parsed],
    )
    fallback.update({field_id: value for field_id, value in parsed.items() if value not in (None, "")})
    return {mapping.field_id: fallback.get(mapping.field_id) for mapping in mappings}


def parse_vat_small_appendix1_text(text: str, mappings: list[FieldMapping]) -> dict[str, Any]:
    lines = normalize_page_text_lines(text)
    value_blocks: list[list[str]] = []
    for line in lines:
        tokens = amount_tokens(line)
        if len(tokens) == 4 and any("." in token for token in tokens):
            value_blocks.append(tokens)
    data: dict[str, Any] = {}
    mapping_ids = {mapping.field_id for mapping in mappings}
    for block_index, field_ids in enumerate(VAT_SMALL_APPENDIX1_BLOCK_FIELDS):
        if block_index >= len(value_blocks):
            break
        for field_id, value in zip(field_ids, value_blocks[block_index]):
            if field_id in mapping_ids:
                data[field_id] = value
    return data


def extract_vat_small_appendix2_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    parsed = parse_vat_small_appendix2_text(extract_body_text(page), mappings)
    fallback = extract_web_data_with_scroll_retry(
        page,
        TARGETS["vat_small_appendix2"],
        [mapping for mapping in mappings if mapping.field_id not in parsed],
    )
    fallback.update({field_id: value for field_id, value in parsed.items() if value not in (None, "")})
    return {mapping.field_id: fallback.get(mapping.field_id) for mapping in mappings}


def parse_vat_small_appendix2_text(text: str, mappings: list[FieldMapping]) -> dict[str, Any]:
    lines = normalize_page_text_lines(text)
    row_values: dict[str, list[str]] = {}
    for label in ("城市维护建设税", "教育费附加", "地方教育附加", "合计"):
        values = values_after_row_label(lines, label)
        if values:
            row_values[label] = values

    data: dict[str, Any] = {}
    for mapping in mappings:
        row_name = str(mapping.row_name or "").strip()
        if first_line_number(mapping.line_no) == "4" or mapping.field_id.endswith("_hj"):
            row_name = "合计"
        values = row_values.get(row_name)
        if not values:
            continue
        prefix = appendix2_field_prefix(mapping.field_id)
        value_index = VAT_SMALL_APPENDIX2_PREFIX_TO_VALUE_INDEX.get(prefix)
        if value_index is None or value_index >= len(values):
            if prefix in {"bqjmsejmxzdm", "bqyjse", "bqybtse"}:
                data[mapping.field_id] = "0.00"
            continue
        data[mapping.field_id] = normalize_percent_ratio(values[value_index]) if prefix == "slzsl" else values[value_index]
    return data


def extract_vat_small_main_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    parsed = parse_vat_small_main_text(extract_body_text(page), mappings)
    return {mapping.field_id: parsed.get(mapping.field_id) for mapping in mappings}


def parse_vat_small_main_text(text: str, mappings: list[FieldMapping]) -> dict[str, Any]:
    lines = normalize_page_text_lines(text)
    values_by_line: dict[str, list[str]] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        row_name = str(mapping.row_name or "").strip()
        if not line_no or not row_name or line_no in values_by_line:
            continue
        values = small_main_values_for_row(lines, line_no, row_name)
        if values:
            values_by_line[line_no] = values
    line_numbers = sorted(
        {first_line_number(mapping.line_no) for mapping in mappings if first_line_number(mapping.line_no)},
        key=lambda item: int(Decimal(item)),
    )
    for index in range(len(lines)):
        window = " ".join(lines[index : index + 12])
        for line_no in line_numbers:
            if line_no in values_by_line:
                continue
            values = small_main_values_after_line_no(window, line_no)
            if values:
                values_by_line[line_no] = values

    data: dict[str, Any] = {}
    for mapping in mappings:
        line_no = first_line_number(mapping.line_no)
        if not line_no:
            continue
        values = values_by_line.get(line_no)
        if not values:
            continue
        value_index = small_main_value_index(mapping.field_id)
        if value_index is None or value_index >= len(values):
            continue
        data[mapping.field_id] = small_main_value_for_field(mapping.field_id, values, value_index)
    fill_vat_small_main_derived_totals(data)
    return data


def small_main_values_for_row(lines: list[str], line_no: str, row_name: str) -> list[str]:
    wanted_row = compact_label(row_name)
    if not wanted_row:
        return []
    for index, line in enumerate(lines):
        if wanted_row not in compact_label(line):
            continue
        window = " ".join(lines[index : index + 12])
        values = small_main_values_after_line_no(window, line_no)
        if values:
            return values
    return []


def amount_tokens(text: Any) -> list[str]:
    return FORM_VALUE_TOKEN_RE.findall(str(text or ""))


def values_after_row_label(lines: list[str], label: str) -> list[str]:
    compact_label_text = compact_label(label)
    for index, line in enumerate(lines):
        compact_line = compact_label(line)
        if compact_label_text not in compact_line:
            continue
        suffix = line.split(label, 1)[1] if label in line else line
        window = " ".join([suffix, *lines[index + 1 : index + 12]])
        values = amount_tokens(window)
        if values:
            return values
    return []


def appendix2_field_prefix(field_id: str) -> str:
    text = str(field_id or "")
    for suffix in ("_cjs", "_jyfj", "_dfjyfj", "_hj"):
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def small_main_value_index(field_id: str) -> int | None:
    text = str(field_id or "")
    for suffix, index in VAT_SMALL_MAIN_SUFFIX_TO_VALUE_INDEX.items():
        if text.endswith(suffix):
            return index
    return None


def small_main_value_for_field(field_id: str, values: list[str], value_index: int) -> str:
    if field_id.endswith("_hwjlw_bnlj") and len(values) > 2:
        return subtract_amount_text(values[2], values[0])
    if field_id.endswith("_fwjbdc_bnlj") and len(values) > 3:
        return subtract_amount_text(values[3], values[1])
    return values[value_index]


def subtract_amount_text(total: str, current: str) -> str:
    total_amount = parse_amount(total)
    current_amount = parse_amount(current)
    if total_amount is None or current_amount is None:
        return total
    amount = total_amount - current_amount
    return f"{amount.quantize(Decimal('0.01'))}"


def fill_vat_small_main_derived_totals(data: dict[str, Any]) -> None:
    for suffix in ("_hwjlw_bqs", "_fwjbdc_bqs", "_hwjlw_bnlj", "_fwjbdc_bnlj"):
        target_field = f"ynsehj{suffix}"
        if data.get(target_field) not in (None, ""):
            continue
        tax_due = data.get(f"bqynse{suffix}")
        reduction = data.get(f"bqynsejze{suffix}")
        if tax_due in (None, "") or reduction in (None, ""):
            continue
        data[target_field] = subtract_amount_text(str(tax_due), str(reduction))


def small_main_values_after_line_no(line: str, line_no: str) -> list[str]:
    pattern = re.compile(
        rf"(?<![\d.+=＝+＋\-－≥≤]){re.escape(line_no)}(?:\s*(?:[=＝（(][^\s]*[\)）]?)?)?\s+"
        rf"(?P<values>(?:{FORM_VALUE_TOKEN_PATTERN}\s*){{4,8}})"
    )
    match = pattern.search(line)
    if not match:
        return []
    values = collapse_repeated_dash_tokens(amount_tokens(match.group("values")))
    if len(values) < 4:
        return []
    if len(values) >= 4 and not any(small_main_value_token_looks_like_amount(value) for value in values[:4]):
        return []
    return values


def small_main_value_token_looks_like_amount(value: Any) -> bool:
    text = str(value or "")
    return "." in text or "," in text or text in DASH_ZERO_VALUES


def collapse_repeated_dash_tokens(tokens: list[str]) -> list[str]:
    collapsed: list[str] = []
    index = 0
    while index < len(tokens):
        token = tokens[index]
        if token in DASH_ZERO_VALUES and index + 1 < len(tokens) and tokens[index + 1] in DASH_ZERO_VALUES:
            collapsed.append(token)
            index += 2
            continue
        collapsed.append(token)
        index += 1
    return collapsed[:4] if len(collapsed) >= 4 else collapsed


def extract_web_data_for_target(page, target: CompareTarget, mappings: list[FieldMapping]) -> dict[str, Any]:
    extraction_scope = page if target.target_id == "vat_general_appendix1" else select_target_content_scope(page, target, mappings)
    if extraction_scope is not page:
        LOGGER.info("Using embedded tax form scope for web extraction: target=%s url=%s", target.target_id, scope_url(extraction_scope))
    page = extraction_scope
    if target.target_id == "vat_general_appendix1":
        data = extract_vat_general_appendix1_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with VAT general appendix1 parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("VAT general appendix1 parser found no fields; continuing without generic fallback")
        return apply_target_web_value_rules(target, data)
    if target.target_id == "vat_small_main":
        data = extract_vat_small_main_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with small VAT main parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Small VAT main parser found no fields; falling back to generic extraction")
    if target.target_id == "vat_small_appendix1":
        data = extract_vat_small_appendix1_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with small VAT appendix1 parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Small VAT appendix1 parser found no fields; falling back to generic extraction")
    if target.target_id == "vat_small_appendix2":
        data = extract_vat_small_appendix2_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with small VAT appendix2 parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Small VAT appendix2 parser found no fields; falling back to generic extraction")
    if target.target_id in {"consumption_tax_main", "consumption_tax_surcharge"}:
        data = extract_consumption_tax_data(page, target, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with consumption-tax parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Consumption-tax parser found no fields; falling back to generic extraction")
    if target.target_id == "vat_general_appendix3":
        data = extract_vat_general_appendix3_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with appendix3 parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Appendix3 parser found no fields; falling back to generic extraction")
    if target.target_id == "vat_general_appendix5":
        data = extract_vat_general_appendix5_data(page, mappings)
        found = sum(1 for value in data.values() if value not in (None, ""))
        if found:
            LOGGER.info("Extracted %s/%s web fields with appendix5 parser", found, len(mappings))
            return apply_target_web_value_rules(target, data)
        LOGGER.warning("Appendix5 parser found no fields; falling back to generic extraction")
    return apply_target_web_value_rules(target, extract_web_data_with_scroll_retry(page, target, mappings))


def extract_consumption_tax_data(page, target: CompareTarget, mappings: list[FieldMapping]) -> dict[str, Any]:
    rows = collect_consumption_tax_rows(page)
    if target.target_id == "consumption_tax_main":
        parsed = parse_consumption_tax_main_rows(rows)
        missing_sentinel_fields = [
            field_id
            for field_id in CONSUMPTION_TAX_MAIN_RENDER_SENTINEL_FIELDS
            if field_id in {mapping.field_id for mapping in mappings} and parsed.get(field_id) in (None, "")
        ]
        if missing_sentinel_fields:
            LOGGER.info(
                "Consumption-tax main summary fields missing after first render read; retrying full scroll render: %s",
                ", ".join(missing_sentinel_fields),
            )
            rows = collect_consumption_tax_rows(page, retry=True)
            parsed = parse_consumption_tax_main_rows(rows)
    else:
        parsed = parse_consumption_tax_surcharge_rows(rows)
    return {mapping.field_id: parsed.get(mapping.field_id) for mapping in mappings}


def collect_consumption_tax_rows(page, retry: bool = False) -> list[list[Any]]:
    positions = (
        ("current", "top", "middle", "lower_middle", "near_bottom", "bottom", "near_bottom", "bottom")
        if retry
        else ("current", "bottom", "top")
    )
    collected: list[list[Any]] = []
    for position in positions:
        if position != "current":
            scroll_consumption_tax_render_area(page, position)
            time.sleep(0.35)
        collected.extend(read_consumption_tax_table_rows(page))
    return dedupe_table_rows(collected)


def scroll_consumption_tax_render_area(page, position: str) -> None:
    try:
        page.evaluate(
            """(position) => {
                const ratios = {
                    top: 0,
                    middle: 0.5,
                    lower_middle: 0.72,
                    near_bottom: 0.9,
                    bottom: 1
                };
                const ratio = ratios[position] ?? 0;
                const scrollables = Array.from(document.querySelectorAll('*')).filter((el) => {
                    const style = getComputedStyle(el);
                    return /(auto|scroll)/.test(style.overflow + style.overflowY + style.overflowX)
                        && (el.scrollHeight > el.clientHeight || el.scrollWidth > el.clientWidth);
                });
                const toBottom = position === 'bottom';
                for (const el of scrollables) {
                    el.scrollTop = Math.max(0, el.scrollHeight - el.clientHeight) * ratio;
                    if (position === 'top') el.scrollLeft = 0;
                    if (toBottom) el.scrollLeft = el.scrollWidth;
                    el.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
                window.scrollTo(
                    toBottom ? document.body.scrollWidth : 0,
                    Math.max(0, document.body.scrollHeight - window.innerHeight) * ratio
                );
                window.dispatchEvent(new Event('scroll'));
            }""",
            position,
        )
    except Exception as exc:  # pragma: no cover - defensive browser fallback
        LOGGER.debug("Could not force consumption-tax scroll render: %s", exc)


def read_consumption_tax_table_rows(page) -> list[list[Any]]:
    return page.evaluate(
        """() => {
            const valueOf = (cell) => {
                const controls = Array.from(cell.querySelectorAll('input, textarea, select'));
                const values = controls
                    .map((item) => String(item.value || item.getAttribute('value') || '').trim())
                    .filter(Boolean);
                if (values.length) return values.join(' ');
                const title = cell.getAttribute('title');
                if (title) return String(title).trim();
                return String(cell.innerText || cell.textContent || '').trim();
            };
            return Array.from(document.querySelectorAll('table tr')).map((tr) =>
                Array.from(tr.querySelectorAll('th,td')).map(valueOf)
            );
        }"""
    )


def dedupe_table_rows(rows: list[list[Any]]) -> list[list[Any]]:
    deduped: list[list[Any]] = []
    seen: set[str] = set()
    for row in rows:
        normalized = [str(cell or "").strip() for cell in row]
        if not any(normalized):
            continue
        key = "\x1f".join(normalized)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def parse_consumption_tax_main_rows(rows: list[list[Any]]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    normalized_rows = [[str(cell or "").strip() for cell in row] for row in rows]
    product_start = -1
    for idx, row in enumerate(normalized_rows):
        joined = "".join(row)
        if "6=1" in joined and "2" in joined and "5" in joined:
            product_start = idx + 1
            break
    if product_start >= 0:
        product_no = 1
        for row in normalized_rows[product_start:]:
            if not row:
                continue
            first_cell = normalize_form_text(row[0])
            if "合计" in first_cell or "鍚堣" in first_cell:
                if len(row) >= 10:
                    data["bqynse6_xfsjfjsfsbb"] = row[8]
                    data["bnynse6_xfsjfjsfsbb"] = row[9]
                break
            if product_no > 5:
                break
            if len(row) >= 10:
                prefix_to_index = {
                    "ysxfpmc": 0,
                    "desl": 1,
                    "blsl": 2,
                    "jldw": 3,
                    "bqxssl": 4,
                    "bnxssl": 5,
                    "bqxse": 6,
                    "bnxse": 7,
                    "bqynse": 8,
                    "bnynse": 9,
                }
                for prefix, cell_index in prefix_to_index.items():
                    value = normalize_form_text(row[cell_index]) if prefix in {"ysxfpmc", "jldw"} else row[cell_index]
                    data[f"{prefix}{product_no}_xfsjfjsfsbb"] = value
                product_no += 1

    summary_fields = [
        (("本期减免税额",), ("7",), "bqsjmse_xfsjfjsfsbb", "bnljjmse_xfsjfjsfsbb"),
        (("期初留抵税额",), ("8",), "bqqcldse_xfsjfjsfsbb", None),
        (("本期准予扣除税额",), ("9",), "bqzydkse_xfsjfjsfsbb", "bnljzydkse_xfsjfjsfsbb"),
        (("本期应扣除税额",), ("10=8+9", "10"), "bqykcse_xfsjfjsfsbb", None),
        (("本期实际扣除税额",), ("11",), "bqsjkcse_xfsjfjsfsbb", "bnljsjkcse_xfsjfjsfsbb"),
        (("期末留抵税额",), ("12=10-11", "12"), "bqqmlqse_xfsjfjsfsbb", None),
        (("本期预缴税额",), ("13",), "bqyjse_xfsjfjsfsbb", None),
        (("城市维护建设税本期应补退税额", "城市维护建设税本期应补（退）税额"), ("15",), "bqcswhjssybtse_xfsjfjsfsbb", "bnljcswhjssybtse_xfsjfjsfsbb"),
        (("教育费附加本期应补退费额", "教育费附加本期应补（退）费额"), ("16",), "bqjyffjybtse_xfsjfjsfsbb", "bnljjyffjybtse_xfsjfjsfsbb"),
        (("地方教育附加本期应补退费额", "地方教育附加本期应补（退）费额"), ("17",), "bqdfjyfjybtse_xfsjfjsfsbb", "bnljdfjyfjybtse_xfsjfjsfsbb"),
        (("本期应补退税额", "本期应补（退）税额"), ("14=6-7-11-13", "14"), "bqybtse_xfsjfjsfsbb", "bnljybtse_xfsjfjsfsbb"),
    ]
    for row in normalized_rows:
        if len(row) < 3:
            continue
        row_label = normalize_form_text(row[0])
        for labels, codes, current_field, cumulative_field in summary_fields:
            if consumption_summary_row_matches(row, row_label, labels, codes):
                current_value, cumulative_value = consumption_summary_values(row, has_cumulative=bool(cumulative_field))
                data[current_field] = current_value
                if cumulative_field:
                    data[cumulative_field] = cumulative_value
                break
    return data


def consumption_summary_row_matches(
    row: list[str],
    row_label: str,
    labels: tuple[str, ...],
    codes: tuple[str, ...],
) -> bool:
    normalized_labels = [normalize_form_text(label) for label in labels]
    if any(label and label in row_label for label in normalized_labels):
        return True
    for cell in row[:2]:
        cell_text = normalize_form_text(cell).replace("＝", "=")
        for code in codes:
            normalized_code = normalize_form_text(code).replace("＝", "=")
            if not normalized_code:
                continue
            if cell_text == normalized_code or cell_text.startswith(normalized_code):
                return True
    return False


def consumption_summary_values(row: list[str], has_cumulative: bool) -> tuple[Any, Any]:
    amount_cells = [cell for cell in row[1:] if parse_amount(cell) is not None]
    if has_cumulative:
        if len(amount_cells) >= 2:
            return amount_cells[-2], amount_cells[-1]
        if amount_cells:
            return amount_cells[-1], None
        return None, None
    if amount_cells:
        return amount_cells[-1], None
    return None, None


def parse_consumption_tax_surcharge_rows(rows: list[list[Any]]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    row_labels = {
        "城市维护建设税": 1,
        "教育费附加": 2,
        "地方教育附加": 3,
        "合计": 4,
    }
    normal_rows = [[str(cell or "").strip() for cell in row] for row in rows]
    for row in normal_rows:
        if len(row) < 5:
            continue
        normalized_label = normalize_form_text(row[0])
        row_no = next((value for label, value in row_labels.items() if label in normalized_label), None)
        if row_no is None:
            continue
        if row_no in (1, 2, 3) and len(row) >= 12:
            prefix_to_index = {
                "jsyjsfsse": 2,
                "sfl": 3,
                "bqynsfe": 4,
                "jmxzdm": 5,
                "jmsfe": 6,
                "jzbl": 8,
                "jze": 9,
                "bqyjsfe": 10,
                "bqybtsfe": 11,
            }
            for prefix, cell_index in prefix_to_index.items():
                value = row[cell_index]
                if prefix == "jzbl":
                    value = normalize_percent_ratio(value)
                data[f"{prefix}{row_no}_xfsfjsfjsb"] = value
        elif row_no == 4 and len(row) >= 12:
            data["bqynsfe4_xfsfjsfjsb"] = row[4]
            data["jmsfe4_xfsfjsfjsb"] = row[6]
            data["jze4_xfsfjsfjsb"] = row[9]
            data["bqyjsfe4_xfsfjsfjsb"] = row[10]
            data["bqybtsfe4_xfsfjsfjsb"] = row[11]
    return data


def normalize_percent_ratio(value: Any) -> str:
    text_value = str(value or "").strip()
    has_percent_suffix = text_value.endswith(("%", "％"))
    if has_percent_suffix:
        text_value = text_value[:-1]
    amount = parse_amount(text_value)
    if amount is None:
        return str(value or "")
    if has_percent_suffix or abs(amount) > Decimal("1"):
        amount = amount / Decimal("100")
    text = format(amount, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def normalize_form_text(value: Any) -> str:
    return (
        str(value or "")
        .replace("\u2002", "")
        .replace("\xa0", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace(" ", "")
        .replace("（", "")
        .replace("）", "")
        .replace("(", "")
        .replace(")", "")
    )


def extract_vat_general_appendix3_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    rows = page.evaluate(
        """() => {
            const clean = (value) => String(value ?? '')
                .replace(/\\u00a0/g, ' ')
                .replace(/\\u3000/g, ' ')
                .trim();
            const valueOf = (cell) => {
                const input = cell.querySelector('input, textarea, select');
                if (input) return clean(input.value);
                const value = cell.getAttribute('value');
                if (value !== null && value !== '') return clean(value);
                return clean(cell.innerText || cell.textContent || '');
            };
            const tables = Array.from(document.querySelectorAll('table'));
            const table = tables.find((item) => {
                const text = item.innerText || item.textContent || '';
                return text.includes('附列资料（三）') || text.includes('扣除项目明细');
            }) || tables.find((item) => {
                const text = item.innerText || item.textContent || '';
                return text.includes('13%税率的项目') && text.includes('期末余额');
            });
            if (!table) return [];
            return Array.from(table.querySelectorAll('tr')).map((tr) =>
                Array.from(tr.querySelectorAll('th,td')).map(valueOf)
            );
        }"""
    )
    return parse_vat_general_appendix3_rows(rows or [], mappings)


def parse_vat_general_appendix3_rows(rows: list[list[Any]], mappings: list[FieldMapping]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for mapping in mappings:
        if mapping.data_type != DataType.AMOUNT:
            continue
        if not mapping.line_no or mapping.web_col_index is None:
            continue
        value = appendix3_value_from_rows(rows, mapping)
        if value is not None:
            data[mapping.field_id] = value
    return {mapping.field_id: data.get(mapping.field_id) for mapping in mappings}


def appendix3_value_from_rows(rows: list[list[Any]], mapping: FieldMapping) -> str | None:
    wanted_line = first_line_number(mapping.line_no)
    wanted_name = compact_appendix3_label(mapping.row_name)
    if not wanted_line or not wanted_name:
        return None

    for row in rows:
        cells = [str(value or "").strip() for value in row]
        if not cells:
            continue
        compact_cells = [compact_appendix3_label(cell) for cell in cells]
        line_index = appendix3_line_index(compact_cells, wanted_line)
        if line_index < 0:
            continue
        row_text = compact_appendix3_label("".join(cells))
        if wanted_name and wanted_name not in row_text:
            continue
        label_index = appendix3_label_index(compact_cells, wanted_name)
        value_start = max(line_index, label_index) + 1
        value_index = value_start + int(mapping.web_col_index) - 3
        if value_index < 0 or value_index >= len(cells):
            return None
        return normalize_appendix3_amount_cell(cells[value_index])
    return None


def compact_appendix3_label(value: Any) -> str:
    return (
        compact_label(value)
        .replace("（", "(")
        .replace("）", ")")
        .replace("，", ",")
        .replace("：", ":")
    )


def first_line_number(value: Any) -> str:
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    return match.group(0) if match else ""


def appendix3_line_index(cells: list[str], wanted_line: str) -> int:
    for index, cell in enumerate(cells[:4]):
        if cell == wanted_line:
            return index
    line_pattern = re.compile(rf"^{re.escape(wanted_line)}(?!\d)")
    for index, cell in enumerate(cells[:4]):
        if line_pattern.match(cell):
            return index
    return -1


def appendix3_label_index(cells: list[str], wanted_name: str) -> int:
    for index, cell in enumerate(cells[:4]):
        if wanted_name and wanted_name in cell:
            return index
    return -1


def normalize_appendix3_amount_cell(value: Any) -> str:
    text = str(value or "").strip()
    text = (
        text.replace("\u00a0", "")
        .replace("\u3000", "")
        .replace(",", "")
        .replace("￥", "")
        .replace("¥", "")
    )
    text = re.sub(r"\s+", "", text)
    if not text or text in DASH_ZERO_VALUES:
        return "0.00"
    return text


def extract_vat_general_appendix5_data(page, mappings: list[FieldMapping]) -> dict[str, Any]:
    text = page.evaluate("document.body ? document.body.innerText : ''")
    parsed = parse_vat_general_appendix5_text(str(text or ""), mappings)
    fallback = extract_web_data(page, [mapping for mapping in mappings if mapping.field_id not in parsed])
    fallback.update({field_id: value for field_id, value in parsed.items() if value not in (None, "")})
    return {mapping.field_id: fallback.get(mapping.field_id) for mapping in mappings}


def parse_vat_general_appendix5_text(text: str, mappings: list[FieldMapping]) -> dict[str, Any]:
    lines = normalize_page_text_lines(text)
    data: dict[str, Any] = {}
    rows = extract_vat_appendix5_grid_rows(lines)
    partial_rows = extract_vat_appendix5_partial_rows(lines)
    has_grid_rows = bool(rows)

    for mapping in mappings:
        field_id = mapping.field_id
        row_label = vat_appendix5_row_label_for_field(field_id)
        if row_label and row_label in rows and mapping.web_col_index is not None:
            index = int(mapping.web_col_index) - VAT_APPENDIX5_GRID_BASE_COL
            values = rows[row_label]
            if 0 <= index < len(values):
                data[field_id] = values[index]
            continue
        if (
            row_label
            and field_id.startswith("bqyjse_")
            and mapping.web_col_index == 15
            and appendix5_partial_row_supports_zero_default(partial_rows.get(row_label, []))
        ):
            data[field_id] = "0.00"
            continue
        if (
            row_label
            and mapping.data_type == DataType.AMOUNT
            and has_grid_rows
            and not appendix5_text_contains_label(lines, row_label)
            and row_label != "合计"
        ):
            data[field_id] = "0.00"
            continue

        if field_id == "nsrmc":
            data[field_id] = value_after_label(lines, "纳税人名称")
        elif field_id == "nsssq":
            start = value_after_label(lines, "税（费）款所属期起")
            end = value_after_label(lines, "税（费）款所属期止")
            data[field_id] = f"{start}至{end}" if start and end else start or end
        elif field_id == "bqsfsyxwqylslfjmzc":
            data[field_id] = checked_option_after_label(lines, "本期是否适用小微企业")
        elif field_id == "jmzcsyzt":
            data[field_id] = checked_option_after_label(lines, "减征政策适用主体")
        elif field_id == "syjmzcqsj":
            data[field_id] = value_after_label(lines, "适用减免政策日期起")
        elif field_id == "syjmzczsj":
            data[field_id] = value_after_label(lines, "适用减免政策日期止")
        elif field_id == "bqsfsysdjspycjrhxqydmzc":
            data[field_id] = checked_option_after_label(lines, "本期是否适用试点建设培育产教融合型企业抵免政策")
        elif field_id in VAT_APPENDIX5_SINGLE_VALUE_LABELS:
            label = VAT_APPENDIX5_SINGLE_VALUE_LABELS[field_id]
            value = numeric_value_after_label(lines, label, mapping.line_no)
            if value:
                data[field_id] = value
            elif has_grid_rows and not appendix5_text_contains_label(lines, label):
                data[field_id] = "0.00"
    return data


def normalize_page_text_lines(text: str) -> list[str]:
    normalized = (
        str(text or "")
        .replace("\u2002", " ")
        .replace("\xa0", " ")
        .replace("\u3000", " ")
    )
    return [line.strip() for line in normalized.splitlines() if line.strip()]


def appendix5_text_contains_label(lines: list[str], label: str) -> bool:
    wanted = compact_label(label)
    return any(wanted in compact_label(line) for line in lines)


def extract_vat_appendix5_grid_rows(lines: list[str]) -> dict[str, list[str]]:
    rows: dict[str, list[str]] = {}
    row_labels = set(VAT_APPENDIX5_ROW_LABELS.values())
    stop_labels = row_labels | {
        "本期是否适用试点建设培育产教融合型企业抵免政策",
        "可用于扣除的增值税留抵退税额使用情况",
        "当期新增投资额",
    }
    for index, line in enumerate(lines):
        label, value_start, inline_values = find_appendix5_row_label(lines, index, row_labels)
        if not label or label in rows:
            continue
        values = list(inline_values)
        if len(values) < VAT_APPENDIX5_GRID_WIDTH:
            values.extend(collect_appendix5_row_values(lines, value_start, stop_labels, label))
        values = strip_leading_row_number(values, label)
        values = align_appendix5_grid_values(values)
        if len(values) >= VAT_APPENDIX5_GRID_WIDTH:
            rows[label] = values[:VAT_APPENDIX5_GRID_WIDTH]
    return rows


def extract_vat_appendix5_partial_rows(lines: list[str]) -> dict[str, list[str]]:
    rows: dict[str, list[str]] = {}
    row_labels = set(VAT_APPENDIX5_ROW_LABELS.values())
    stop_labels = row_labels | {
        "本期是否适用试点建设培育产教融合型企业抵免政策",
        "可用于扣除的增值税留抵退税额使用情况",
        "当期新增投资额",
    }
    for index, line in enumerate(lines):
        label, value_start, inline_values = find_appendix5_row_label(lines, index, row_labels)
        if not label or label in rows:
            continue
        values = list(inline_values)
        values.extend(collect_appendix5_partial_row_values(lines, value_start, stop_labels))
        rows[label] = strip_leading_row_number(values, label)
    return rows


def find_appendix5_row_label(lines: list[str], index: int, row_labels: set[str]) -> tuple[str, int, list[str]]:
    line = lines[index]
    cells = split_appendix5_cells(line, preserve_empty=True)
    for label in row_labels:
        wanted = compact_label(label)
        for cell_index, cell in enumerate(cells):
            current_cell = compact_label(cell)
            if is_appendix5_label_cell(current_cell, wanted):
                return label, index + 1, appendix5_value_tokens(cells[cell_index + 1 :], preserve_empty=True)

    current = compact_label(line)
    next_cells = split_appendix5_cells(lines[index + 1], preserve_empty=True) if index + 1 < len(lines) else []
    next_first = compact_label(next_cells[0]) if next_cells else ""
    for label in row_labels:
        wanted = compact_label(label)
        if is_appendix5_label_cell(current, wanted):
            return label, index + 1, []
        if current and (current + next_first).startswith(wanted):
            return label, index + 2, appendix5_value_tokens(next_cells[1:], preserve_empty=True)
    return "", index + 1, []


def is_appendix5_label_cell(current: str, wanted: str) -> bool:
    if current == wanted:
        return True
    return current.startswith(wanted) and len(current) <= len(wanted) + 12


def split_appendix5_cells(line: Any, preserve_empty: bool = False) -> list[str]:
    text = (
        str(line or "")
        .replace("\u2002", " ")
        .replace("\xa0", " ")
        .replace("\u3000", " ")
        .strip()
    )
    if not text:
        return []
    if "\t" not in text:
        return [text]
    cells = [cell.strip() for cell in text.split("\t")]
    if preserve_empty:
        while cells and not cells[0]:
            cells.pop(0)
        while cells and not cells[-1]:
            cells.pop()
        return cells
    return [cell for cell in cells if cell]


def compact_label(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def align_appendix5_grid_values(values: list[str]) -> list[str]:
    aligned = list(values)
    while len(aligned) > VAT_APPENDIX5_GRID_WIDTH and len(aligned) > 7 and aligned[7] == "":
        aligned.pop(7)
    if len(aligned) == VAT_APPENDIX5_GRID_WIDTH - 1:
        aligned.insert(7, "")
        return aligned
    return aligned


def collect_appendix5_row_values(lines: list[str], start: int, stop_labels: set[str], label: str = "") -> list[str]:
    values: list[str] = []
    for line in lines[start : start + 120]:
        compact_line = compact_label(line)
        if values and any(is_appendix5_label_cell(compact_line, compact_label(label)) for label in stop_labels):
            break
        values.extend(appendix5_value_tokens(split_appendix5_cells(line, preserve_empty=True), preserve_empty="\t" in line))
    return values


def collect_appendix5_partial_row_values(lines: list[str], start: int, stop_labels: set[str]) -> list[str]:
    values: list[str] = []
    for line in lines[start : start + 120]:
        compact_line = compact_label(line)
        if any(is_appendix5_label_cell(compact_line, compact_label(label)) for label in stop_labels):
            break
        values.extend(appendix5_value_tokens(split_appendix5_cells(line, preserve_empty=True), preserve_empty="\t" in line))
    return values


def strip_leading_row_number(values: list[str], label: str = "") -> list[str]:
    if not values:
        return values
    expected = VAT_APPENDIX5_ROW_NUMBERS.get(label)
    if expected and values[0] == expected:
        return values[1:]
    if len(values) > VAT_APPENDIX5_GRID_WIDTH and values[0] in {"1", "2", "3", "4"}:
        return values[1:]
    return values


def normalize_appendix5_value_token(value: Any) -> str | None:
    text = str(value or "").strip()
    text = text.replace("\u2002", " ").replace("\xa0", " ")
    text = re.sub(r"\s+", "", text)
    if not text:
        return None
    if text in {"|", "｜", "请选择"}:
        return ""
    if "|" in text or "｜" in text:
        return ""
    if text in {"—", "——", "--", "-", "－"}:
        return "0.00"
    if re.fullmatch(r"-?\d+(?:,\d{3})*(?:\.\d+)?%?", text) or re.fullmatch(r"-?\d+(?:\.\d+)?%?", text):
        return text
    return None


def appendix5_value_tokens(values: list[Any], preserve_empty: bool = False) -> list[str]:
    tokens: list[str] = []
    for value in values:
        if preserve_empty and str(value or "").strip() == "":
            tokens.append("")
            continue
        token = normalize_appendix5_value_token(value)
        if token is not None:
            tokens.append(token)
    return tokens


def appendix5_partial_row_supports_zero_default(values: list[str]) -> bool:
    if not values:
        return False
    amount_values = []
    for value in values:
        text = str(value or "").strip()
        if not text or text.endswith("%"):
            continue
        amount = parse_amount(text)
        if amount is None:
            continue
        amount_values.append(amount)
    return bool(amount_values) and all(amount == 0 for amount in amount_values)


def vat_appendix5_row_label_for_field(field_id: str) -> str:
    for suffix, label in VAT_APPENDIX5_ROW_LABELS.items():
        if field_id.endswith(suffix):
            return label
    return ""


def value_after_label(lines: list[str], label: str) -> str:
    for index, line in enumerate(lines):
        cells = split_appendix5_cells(line)
        for cell_index, cell in enumerate(cells):
            if label not in cell:
                continue
            inline = cell.split(label, 1)[1].strip(" ：:")
            if inline:
                return inline
            if cell_index + 1 < len(cells):
                return cells[cell_index + 1]
            for candidate in lines[index + 1 : index + 5]:
                if candidate and candidate not in {":", "："}:
                    return candidate
    return ""


def numeric_value_after_label(lines: list[str], label: str, line_no: Any = "") -> str:
    skip_line_no = str(line_no or "").strip()
    for index, line in enumerate(lines):
        cells = split_appendix5_cells(line, preserve_empty=True)
        inline_candidates: list[Any] = []
        for cell_index, cell in enumerate(cells):
            if label in cell:
                inline = cell.split(label, 1)[1].strip(" ：:")
                if inline:
                    inline_candidates.append(inline)
                inline_candidates.extend(cells[cell_index + 1 :])
                break
        if inline_candidates:
            for candidate in inline_candidates:
                token = normalize_appendix5_value_token(candidate)
                if token is None:
                    continue
                if skip_line_no and token == skip_line_no:
                    continue
                return token
            continue
        if label not in line:
            continue
        for candidate in lines[index + 1 : index + 8]:
            token = normalize_appendix5_value_token(candidate)
            if token is None:
                continue
            if skip_line_no and token == skip_line_no:
                continue
            return token
    return ""


def checked_option_after_label(lines: list[str], label: str) -> str:
    for index, line in enumerate(lines):
        cells = split_appendix5_cells(line)
        search_windows: list[str] = []
        for cell_index, cell in enumerate(cells):
            if label not in cell:
                continue
            inline = cell.split(label, 1)[1].strip(" ：:")
            if inline and contains_checked_option_marker(inline):
                search_windows.append(inline)
            search_windows.extend(cells[cell_index + 1 : cell_index + 3])
            break
        if label in line and not any(contains_checked_option_marker(window) for window in search_windows):
            search_windows.extend(lines[index + 1 : index + 4])
        for window in search_windows:
            options = re.findall(r"([■☑√●])\s*([^□■☑√●]+)", window)
            for marker, option in options:
                if marker:
                    return option.strip()
    return ""


def contains_checked_option_marker(value: Any) -> bool:
    return any(marker in str(value or "") for marker in ("■", "☑", "√", "●"))


def web_extraction_coverage(web_raw: dict[str, Any], mappings: list[FieldMapping]) -> tuple[int, int, float]:
    comparable = [mapping for mapping in mappings if mapping.compare and mapping.field_id not in TEXT_FIELDS]
    total = len(comparable)
    if total <= 0:
        return 0, 0, 1.0
    found = sum(1 for mapping in comparable if web_value_present(web_raw.get(mapping.field_id)))
    return found, total, found / total


def is_low_web_extraction_coverage(target: CompareTarget, web_raw: dict[str, Any], mappings: list[FieldMapping]) -> bool:
    found, total, ratio = web_extraction_coverage(web_raw, mappings)
    if total <= 0:
        return False
    if found == 0:
        return True
    threshold = WEB_EXTRACTION_MIN_COVERAGE_BY_TARGET.get(
        target.target_id,
        WEB_EXTRACTION_DEFAULT_MIN_COVERAGE,
    )
    return ratio < threshold


def comparison_quality_issues(
    target: CompareTarget,
    result,
    low_web_coverage: bool,
    current_period_flag: bool | None = None,
) -> list[str]:
    summary = result.summary
    issues: list[str] = []
    if summary.mismatch_count:
        issues.append(f"mismatch={summary.mismatch_count}")
    if summary.api_missing_count:
        issues.append(f"api_missing={summary.api_missing_count}")
    if summary.web_missing_count:
        issues.append(f"web_missing={summary.web_missing_count}")
    if summary.parse_error_count:
        issues.append(f"parse_error={summary.parse_error_count}")
    if summary.mapping_error_count:
        issues.append(f"mapping_error={summary.mapping_error_count}")
    if low_web_coverage and summary.web_missing_count:
        issues.append("low_web_extraction_coverage")
    both_missing_ratio = summary.both_missing_count / summary.total_fields if summary.total_fields else 0
    if target.target_id == "vat_general_appendix5":
        both_missing_threshold = 0.3
    elif target.tax_type == "CONSUMPTION_TAX":
        both_missing_threshold = 0.8
    else:
        both_missing_threshold = 0.5
    if both_missing_ratio >= both_missing_threshold and summary.both_missing_count:
        issues.append(f"both_missing_ratio={both_missing_ratio:.2%}")
    return issues


def compare_target(
    target: CompareTarget,
    mappings: list[FieldMapping],
    api_by_tax: dict[str, Any],
    web_raw: dict[str, Any],
    current_period_flag: bool | None = None,
):
    mappings = mappings_for_comparison(target, mappings, api_by_tax)
    web_raw = apply_target_web_value_rules(target, dict(web_raw))
    api_norm = {}
    web_norm = {}
    for mapping in mappings:
        normalizer = get_normalizer(mapping.data_type)
        api_norm[mapping.field_id] = normalizer.normalize(clean_value(api_value(api_by_tax, target, mapping.field_id), mapping))
        web_value = adjusted_web_value_for_compare(target, mapping, web_raw, current_period_flag=current_period_flag)
        web_norm[mapping.field_id] = normalizer.normalize(clean_value(web_value, mapping))
    result = Comparator().compare_all(
        mappings=mappings,
        api_data=api_norm,
        web_data=web_norm,
        batch_id=target.target_id,
        company_name="",
        taxpayer_id="",
        period="",
    )
    if not result.tax_type:
        result.tax_type = target.tax_type
    if not result.form_code:
        result.form_code = target.form_code
    if not result.form_name:
        result.form_name = target.form_name
    return result


def adjusted_web_value_for_compare(
    target: CompareTarget,
    mapping: FieldMapping,
    web_raw: dict[str, Any],
    current_period_flag: bool | None = None,
) -> Any:
    value = web_raw.get(mapping.field_id)
    if target.target_id == "vat_general_main" and mapping.field_id == "sjdkse_jzjtxm_bnlj" and web_value_looks_zero(value):
        return web_raw.get("sjdkse_jzjtxm_bys")
    if target.target_id == "vat_general_main" and mapping.field_id == "qcwjse_jzjtxm_bnlj":
        derived = derive_vat_general_jizhengjitui_opening_unpaid(web_raw)
        if derived is not None:
            return derived
    force_subtract = should_force_subtract_current_month_for_vat(
        target,
        mapping,
    ) or should_force_subtract_current_month_for_undeclared_vat(target, mapping, current_period_flag)
    if not force_subtract and not should_subtract_current_month_for_vat(target, mapping):
        return value
    if isinstance(value, str) and value.strip() in DASH_ZERO_VALUES | {""}:
        return value

    monthly_field_id = vat_general_current_month_field_id(mapping.field_id)
    if not monthly_field_id:
        return value
    cumulative = parse_amount(value)
    current_month = parse_amount(web_raw.get(monthly_field_id))
    if cumulative is None or current_month is None:
        return value
    if not force_subtract and cumulative < current_month:
        return value
    adjusted = cumulative - current_month
    return f"{adjusted.quantize(Decimal('0.01'))}"


def should_force_subtract_current_month_for_vat(target: CompareTarget, mapping: FieldMapping) -> bool:
    return (
        target.target_id == "vat_general_main"
        and mapping.field_id in VAT_GENERAL_MAIN_FORCE_SUBTRACT_CURRENT_MONTH_FIELDS
    )


VAT_GENERAL_MAIN_UNDECLARED_FORCE_SUBTRACT_CURRENT_MONTH_FIELDS = {
    "sqldse_ybxm_bnlj",
    "qmldse_ybxm_bnlj",
}


def should_force_subtract_current_month_for_undeclared_vat(
    target: CompareTarget,
    mapping: FieldMapping,
    current_period_flag: bool | None,
) -> bool:
    return (
        current_period_flag is False
        and target.target_id == "vat_general_main"
        and mapping.field_id in VAT_GENERAL_MAIN_UNDECLARED_FORCE_SUBTRACT_CURRENT_MONTH_FIELDS
    )


def should_subtract_current_month_for_vat(target: CompareTarget, mapping: FieldMapping) -> bool:
    return (
        target.tax_code == "sz_zzs"
        and mapping.data_type == DataType.AMOUNT
        and any(mapping.field_id.endswith(cumulative_suffix) for cumulative_suffix, _ in VAT_GENERAL_MAIN_SUBTRACT_CURRENT_MONTH_SUFFIXES)
        and mapping.field_id not in VAT_GENERAL_MAIN_CUMULATIVE_DIRECT_FIELDS
    )


def vat_general_current_month_field_id(field_id: str) -> str:
    for cumulative_suffix, current_suffix in VAT_GENERAL_MAIN_SUBTRACT_CURRENT_MONTH_SUFFIXES:
        if field_id.endswith(cumulative_suffix):
            return field_id.removesuffix(cumulative_suffix) + current_suffix
    return ""


def parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in DASH_ZERO_VALUES:
        return Decimal("0")
    text = re.sub(r"\s+", "", text)
    text = text.replace(",", "").replace("￥", "").replace("¥", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def clean_value(value: Any, mapping: FieldMapping) -> Any:
    if mapping.data_type == DataType.AMOUNT and isinstance(value, str) and value.strip() in DASH_ZERO_VALUES:
        return "0.00"
    return value


def task_output_dir(task_id: str) -> Path:
    return Path("./output/reports") / str(task_id)


def save_result(
    task_id: str,
    target: CompareTarget,
    result,
    province: str = "",
    current_period_flag: bool | None = None,
    quality_issues: list[str] | None = None,
    not_comparable_reason: str = "",
) -> Path:
    output_dir = task_output_dir(task_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"{target.target_id}_compare_{task_id}_{ts}.json"
    payload = json.loads(result.model_dump_json())
    payload["province"] = province
    payload["current_period_flag"] = current_period_flag
    payload["declaration_status"] = declaration_status_for_target(target, current_period_flag)
    payload["quality_issues"] = quality_issues or []
    if not_comparable_reason:
        payload["not_comparable"] = True
        payload["not_comparable_reason"] = not_comparable_reason
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def declaration_status_for_target(target: CompareTarget, current_period_flag: bool | None) -> str:
    if current_period_flag is False:
        return "未申报"
    if current_period_flag is True:
        return "已申报"
    return "未申报"


def save_api_filled_workbook(
    task_id: str,
    target: CompareTarget,
    mappings: list[FieldMapping],
    api_by_tax: dict[str, Any],
    output_dir: Path,
) -> Path | None:
    """Write API values back into the ID workbook layout for human review."""

    workbook = find_workbook(target.workbook_name)
    wb = load_workbook_compat(workbook, data_only=False)
    try:
        if target.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {target.sheet_name}; available={wb.sheetnames}")
        ws = wb[target.sheet_name]
        for sheet in list(wb.worksheets):
            if sheet.title != target.sheet_name:
                wb.remove(sheet)
        ws.title = safe_excel_sheet_title(target.sheet_name)

        filled = 0
        missing = 0
        for mapping in mappings:
            cell = find_mapping_cell(ws, mapping)
            if cell is None:
                continue
            raw_value = api_value(api_by_tax, target, mapping.field_id)
            value = api_excel_value(raw_value, mapping)
            cell.value = value
            cell.fill = API_EXCEL_MISSING_FILL if value == "" else API_EXCEL_VALUE_FILL
            cell.comment = Comment(
                f"field_id: {mapping.field_id}\napi_path: {mapping.api_json_path}\napi_raw_value: {raw_value if raw_value is not None else ''}",
                "Codex",
            )
            if value == "":
                missing += 1
            else:
                filled += 1

        excel_dir = output_dir / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = excel_dir / f"{safe_filename(target.form_name)}_{target.target_id}_{task_id}_{ts}_api_filled.xlsx"
        wb.save(excel_path)
        LOGGER.info(
            "Saved API-filled Excel for %s: %s (filled=%s missing=%s)",
            target.target_id,
            excel_path,
            filled,
            missing,
        )
        return excel_path
    finally:
        wb.close()


def safe_excel_sheet_title(value: str) -> str:
    text = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(value or "Sheet")).strip()
    return (text or "Sheet")[:31]


def api_excel_value(value: Any, mapping: FieldMapping) -> Any:
    if value is None:
        return ""
    value = clean_value(value, mapping)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def find_mapping_cell(ws, mapping: FieldMapping):
    row_idx = mapping.web_row_index
    col_idx = mapping.web_col_index
    if row_idx and col_idx:
        cell = writable_cell(ws, row_idx, col_idx)
        if cell is not None and str(cell.value or "").strip() == mapping.field_id:
            return cell

    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip() == mapping.field_id:
                return writable_cell(ws, cell.row, cell.column)
    return None


def writable_cell(ws, row_idx: int, col_idx: int):
    cell = ws.cell(row=row_idx, column=col_idx)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return None


def save_web_pdf(page, task_id: str, target: CompareTarget, output_dir: Path) -> Path | None:
    """Save a PDF copy of the currently compared tax page for later verification."""
    pdf_dir = output_dir / "pdf"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_path = pdf_dir / f"{safe_filename(target.form_name)}_{target.target_id}_{task_id}_{ts}.pdf"

    prepared = False
    try:
        try:
            prepare_page_for_pdf_snapshot(page)
            prepared = True
            page.emulate_media(media="print")
            page.pdf(
                path=str(pdf_path),
                format="A4",
                landscape=True,
                scale=0.72,
                print_background=True,
                prefer_css_page_size=False,
                margin={"top": "8mm", "right": "8mm", "bottom": "8mm", "left": "8mm"},
            )
            LOGGER.info("Saved tax page PDF snapshot for %s: %s", target.target_id, pdf_path)
            cleanup_page_pdf_snapshot(page)
            prepared = False
            return pdf_path
        except Exception as exc:
            LOGGER.warning("page.pdf failed for %s: %s", target.target_id, exc)

        try:
            session = page.context.new_cdp_session(page)
            result = session.send(
                "Page.printToPDF",
                {
                    "printBackground": True,
                    "preferCSSPageSize": False,
                    "paperWidth": 11.69,
                    "paperHeight": 8.27,
                    "scale": 0.72,
                    "marginTop": 0.31,
                    "marginRight": 0.31,
                    "marginBottom": 0.31,
                    "marginLeft": 0.31,
                },
            )
            pdf_path.write_bytes(base64.b64decode(result["data"]))
            try:
                session.detach()
            except Exception:
                pass
            LOGGER.info("Saved tax page PDF via CDP for %s: %s", target.target_id, pdf_path)
            return pdf_path
        except Exception as exc:
            LOGGER.warning("Could not save PDF for %s: %s", target.target_id, exc)

        downloaded = try_download_pdf_from_page(page, pdf_path)
        if downloaded:
            LOGGER.info("Saved tax PDF by page download for %s: %s", target.target_id, downloaded)
            return downloaded
        return None
    finally:
        if prepared:
            cleanup_page_pdf_snapshot(page)


def prepare_page_for_pdf_snapshot(page) -> None:
    page.evaluate(
        """() => {
            const root = document.documentElement;
            root.classList.add('codex-pdf-export');
            let style = document.getElementById('codex-pdf-export-style');
            if (!style) {
                style = document.createElement('style');
                style.id = 'codex-pdf-export-style';
                style.textContent = `
html.codex-pdf-export,
html.codex-pdf-export body,
html.codex-pdf-export #app,
html.codex-pdf-export [id="app"],
html.codex-pdf-export .page-container,
html.codex-pdf-export .g-layout,
html.codex-pdf-export .g-layout--main,
html.codex-pdf-export .g-main,
html.codex-pdf-export .t-layout,
html.codex-pdf-export .t-layout__content,
html.codex-pdf-export main,
html.codex-pdf-export .page-content,
html.codex-pdf-export .wrapper,
html.codex-pdf-export .report-template,
html.codex-pdf-export .is-scroll-panel,
html.codex-pdf-export .gov-tax-report-container,
html.codex-pdf-export .tax-report,
html.codex-pdf-export .gt-collapse-menu,
html.codex-pdf-export .gt-collapse-menu-content,
html.codex-pdf-export .gt-collapse-menu-content_box,
html.codex-pdf-export .gt-collapse-menu-main,
html.codex-pdf-export .gt-collapse-menu-panel,
html.codex-pdf-export .t-table,
html.codex-pdf-export .t-table__content,
html.codex-pdf-export .t-table__body,
html.codex-pdf-export .t-table__body-wrapper,
html.codex-pdf-export .el-table,
html.codex-pdf-export .el-table__body-wrapper,
html.codex-pdf-export .ant-table,
html.codex-pdf-export .ant-table-body {
  height: auto !important;
  max-height: none !important;
  min-height: 0 !important;
  overflow: visible !important;
}
html.codex-pdf-export .page-content,
html.codex-pdf-export .wrapper,
html.codex-pdf-export .report-template,
html.codex-pdf-export .gt-collapse-menu-content,
html.codex-pdf-export .gt-collapse-menu-main {
  width: auto !important;
  max-width: none !important;
}
html.codex-pdf-export .gt-collapse-menu-sidebar,
html.codex-pdf-export .gt-layout-footerbar,
html.codex-pdf-export .gt-layout-footerbar_fix,
html.codex-pdf-export .t-layout__sider,
html.codex-pdf-export .t-affix,
html.codex-pdf-export [class*="footerbar"],
html.codex-pdf-export [class*="countdown"],
html.codex-pdf-export [class*="operation"] {
  display: none !important;
}
html.codex-pdf-export table {
  page-break-inside: auto;
}
html.codex-pdf-export thead {
  display: table-header-group;
}
html.codex-pdf-export tr,
html.codex-pdf-export img,
html.codex-pdf-export svg {
  page-break-inside: avoid;
}
html.codex-pdf-export * {
  -webkit-print-color-adjust: exact !important;
  print-color-adjust: exact !important;
}
`;
                document.head.appendChild(style);
            }
            const scrollables = Array.from(document.querySelectorAll('*')).filter((el) => {
                const style = getComputedStyle(el);
                return /(auto|scroll)/.test(style.overflow + style.overflowY + style.overflowX)
                    && (el.scrollHeight > el.clientHeight || el.scrollWidth > el.clientWidth);
            });
            for (const el of scrollables) {
                el.scrollTop = el.scrollHeight;
                el.scrollLeft = 0;
            }
            window.scrollTo(0, document.body.scrollHeight);
            for (const el of scrollables) {
                el.scrollTop = 0;
                el.scrollLeft = 0;
            }
            window.scrollTo(0, 0);
        }"""
    )
    time.sleep(1)


def cleanup_page_pdf_snapshot(page) -> None:
    try:
        page.evaluate("document.documentElement.classList.remove('codex-pdf-export')")
    except Exception:
        pass
    try:
        page.emulate_media(media="screen")
    except Exception:
        pass


def safe_filename(value: str, max_length: int = 120) -> str:
    invalid = set('<>:"/\\|?*')
    text = "".join("_" if char in invalid or ord(char) < 32 else char for char in str(value or "").strip())
    text = re.sub(r"\s+", "", text).strip("._ ")
    return (text or "税表")[:max_length]


def escape_html(value: Any) -> str:
    if value is None or value == "":
        return '<span class="empty">空</span>'
    return html.escape(str(value))


def relative_report_link(base_path: Path, target_path: str | Path | None) -> str:
    if not target_path:
        return ""
    path = Path(target_path)
    try:
        return path.resolve().relative_to(base_path.parent.resolve()).as_posix()
    except ValueError:
        try:
            return Path(urllib.parse.quote(str(path))).as_posix()
        except Exception:
            return str(path)


def status_label(status: str) -> str:
    labels = {
        "mismatch": ("不一致", "bad"),
        "api_missing": ("接口缺失", "warn"),
        "web_missing": ("网页缺失", "warn"),
        "parse_error": ("解析失败", "bad"),
        "mapping_error": ("映射错误", "bad"),
        "both_missing": ("双方为空", "muted"),
        "tolerance_match": ("容差通过", "ok"),
        "match": ("一致", "ok"),
        "skip": ("跳过", "muted"),
    }
    label, cls = labels.get(status, (status, "muted"))
    return f'<span class="pill {cls}">{html.escape(label)}</span>'


def problem_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    status_order = {
        "mismatch": 0,
        "parse_error": 1,
        "mapping_error": 2,
        "api_missing": 3,
        "web_missing": 4,
    }
    return [
        item
        for item in sorted(
            fields,
            key=lambda x: (
                status_order.get(str(x.get("status", "")), 99),
                str(x.get("line_no") or ""),
                str(x.get("field_id") or ""),
            ),
        )
        if str(item.get("status")) in PROBLEM_STATUSES
    ]


def effective_pass_rate(summary: dict[str, Any], fields: list[dict[str, Any]]) -> tuple[int, int, float]:
    total = int(summary.get("total_fields", len(fields)) or 0)
    both_missing = int(summary.get("both_missing_count", 0) or 0)
    skip = int(summary.get("skip_count", 0) or 0)
    denominator = max(0, total - both_missing - skip)
    passed = int(summary.get("match_count", 0) or 0) + int(summary.get("tolerance_match_count", 0) or 0)
    rate = round(passed / denominator * 100, 2) if denominator else 100.0
    return passed, denominator, rate


def render_problem_table(fields: list[dict[str, Any]]) -> str:
    problems = problem_fields(fields)
    if not problems:
        return '<p class="quiet">没有发现需要优先处理的问题字段。</p>'
    rows = []
    for item in problems:
        business_name = item.get("business_name") or item.get("row_name") or item.get("column_name") or item.get("display_name")
        rows.append(
            f"""
            <tr>
              <td>{status_label(str(item.get("status", "")))}</td>
              <td>{escape_html(item.get("line_no"))}</td>
              <td>{escape_html(business_name)}</td>
              <td><code>{escape_html(item.get("field_id"))}</code></td>
              <td>{escape_html(item.get("api_raw_value"))}</td>
              <td>{escape_html(item.get("web_raw_value"))}</td>
              <td>{escape_html(item.get("diff_value"))}</td>
              <td>{escape_html(item.get("detail"))}</td>
            </tr>
            """
        )
    return f"""
    <table>
      <thead>
        <tr>
          <th>状态</th><th>行次</th><th>项目</th><th>字段</th><th>接口值</th><th>网页值</th><th>差异</th><th>说明</th>
        </tr>
      </thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    """


def render_task_summary_report(task_id: str, outputs: list[dict[str, Any]]) -> Path | None:
    reports = []
    output_dir = task_output_dir(task_id)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"compare_summary_{task_id}_{ts}.html"

    for item in outputs:
        report_path = Path(str(item.get("report_path", "")))
        if not report_path.exists():
            continue
        report = json.loads(report_path.read_text(encoding="utf-8"))
        reports.append({**item, "report": report})

    if not reports:
        return None

    summary_rows = []
    detail_sections = []
    total_problems = 0
    for index, item in enumerate(reports, start=1):
        report = item["report"]
        summary = report.get("summary", {})
        fields = report.get("field_results", [])
        problems = problem_fields(fields)
        total_problems += len(problems)
        passed, denominator, effective_rate = effective_pass_rate(summary, fields)
        quality_issues = item.get("quality_issues") or report.get("quality_issues") or []
        quality_html = "<br>".join(escape_html(issue) for issue in quality_issues) if quality_issues else '<span class="quiet">-</span>'
        form_name = str(report.get("form_name") or report.get("form_code") or item.get("target_id") or "")
        anchor = f"form-{index}"
        report_link = relative_report_link(output_path, item.get("report_path"))
        pdf_link = relative_report_link(output_path, item.get("pdf_path"))
        api_excel_link = relative_report_link(output_path, item.get("api_excel_path"))
        pdf_html = f'<a href="{html.escape(pdf_link)}">PDF</a>' if pdf_link else '<span class="quiet">未生成</span>'
        api_excel_html = (
            f'<a href="{html.escape(api_excel_link)}">接口Excel</a>'
            if api_excel_link
            else '<span class="quiet">未生成</span>'
        )
        summary_rows.append(
            f"""
            <tr>
              <td><a href="#{anchor}">{html.escape(form_name)}</a></td>
              <td>{len(problems)}</td>
              <td>{effective_rate}% <span class="quiet">({passed}/{denominator})</span></td>
              <td>{escape_html(summary.get("match_rate"))}%</td>
              <td>{escape_html(summary.get("mismatch_count"))}</td>
              <td>{escape_html(summary.get("api_missing_count"))}</td>
              <td>{escape_html(summary.get("web_missing_count"))}</td>
              <td>{escape_html(summary.get("both_missing_count"))}</td>
              <td>{quality_html}</td>
              <td><a href="{html.escape(report_link)}">JSON</a> / {api_excel_html} / {pdf_html}</td>
            </tr>
            """
        )
        detail_sections.append(
            f"""
            <section id="{anchor}" class="form-block">
              <div class="section-title">
                <h2>{html.escape(form_name)}</h2>
                <div class="quiet">有效通过率 {effective_rate}% ({passed}/{denominator})，原始通过率 {escape_html(summary.get("match_rate"))}%</div>
              </div>
              {f'<div class="quiet">质量风险：{quality_html}</div>' if quality_issues else ''}
              {render_problem_table(fields)}
            </section>
            """
        )

    title = f"taskId {task_id} 对比汇总"
    output_path.write_text(
        f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      color-scheme: light;
      --text: #172033;
      --muted: #667085;
      --line: #d8dee8;
      --bg: #f5f7fb;
      --surface: #ffffff;
      --bad: #b42318;
      --warn: #b54708;
      --ok: #067647;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif; color: var(--text); background: var(--bg); }}
    main {{ max-width: 1280px; margin: 0 auto; padding: 28px 24px 48px; }}
    h1 {{ margin: 0 0 8px; font-size: 24px; }}
    h2 {{ margin: 0; font-size: 18px; }}
    a {{ color: #175cd3; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .quiet {{ color: var(--muted); font-size: 13px; }}
    .topline {{ display: flex; justify-content: space-between; gap: 16px; align-items: flex-end; margin-bottom: 20px; }}
    .metric-strip {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin: 18px 0 20px; }}
    .metric {{ background: var(--surface); border: 1px solid var(--line); border-radius: 8px; padding: 14px 16px; }}
    .metric-label {{ color: var(--muted); font-size: 13px; }}
    .metric-value {{ font-size: 24px; font-weight: 700; margin-top: 4px; }}
    table {{ width: 100%; border-collapse: collapse; background: var(--surface); border: 1px solid var(--line); border-radius: 8px; overflow: hidden; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 10px 12px; text-align: left; vertical-align: top; font-size: 13px; }}
    th {{ background: #eef2f7; color: #344054; font-weight: 700; white-space: nowrap; }}
    tr:last-child td {{ border-bottom: 0; }}
    code {{ font-family: Consolas, monospace; font-size: 12px; }}
    .form-block {{ margin-top: 22px; }}
    .section-title {{ display: flex; justify-content: space-between; gap: 12px; align-items: baseline; margin: 0 0 10px; }}
    .pill {{ display: inline-block; min-width: 56px; text-align: center; border-radius: 999px; padding: 2px 8px; font-size: 12px; font-weight: 700; }}
    .pill.bad {{ color: var(--bad); background: #fee4e2; }}
    .pill.warn {{ color: var(--warn); background: #fef0c7; }}
    .pill.ok {{ color: var(--ok); background: #dcfae6; }}
    .pill.muted {{ color: #475467; background: #eaecf0; }}
    .empty {{ color: var(--muted); }}
    @media (max-width: 820px) {{
      main {{ padding: 20px 14px 36px; }}
      .topline, .section-title {{ display: block; }}
      .metric-strip {{ grid-template-columns: 1fr; }}
      table {{ display: block; overflow-x: auto; }}
    }}
  </style>
</head>
<body>
  <main>
    <div class="topline">
      <div>
        <h1>{html.escape(title)}</h1>
        <div class="quiet">生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}；所有主表和附表结果集中在本页。</div>
      </div>
    </div>
    <section class="metric-strip">
      <div class="metric"><div class="metric-label">表单数量</div><div class="metric-value">{len(reports)}</div></div>
      <div class="metric"><div class="metric-label">需要优先处理的问题字段</div><div class="metric-value">{total_problems}</div></div>
      <div class="metric"><div class="metric-label">结果目录</div><div class="metric-value" style="font-size:16px;">{html.escape(str(output_dir))}</div></div>
    </section>
    <table>
      <thead>
        <tr>
          <th>表单</th><th>问题字段</th><th>有效通过率</th><th>原始通过率</th><th>不一致</th><th>接口缺失</th><th>网页缺失</th><th>双方为空</th><th>质量风险</th><th>文件</th>
        </tr>
      </thead>
      <tbody>{''.join(summary_rows)}</tbody>
    </table>
    {''.join(detail_sections)}
  </main>
</body>
</html>
""",
        encoding="utf-8",
    )
    LOGGER.info("Saved combined task report: %s", output_path)
    return output_path


def try_download_pdf_from_page(page, pdf_path: Path) -> Path | None:
    labels = ("下载PDF", "PDF下载", "下载", "导出PDF", "导出")
    for label in labels:
        try:
            exists = page.evaluate(
                """(label) => {
                    const nodes = Array.from(document.querySelectorAll('a, button, [role=button]'))
                        .filter((el) => el.offsetParent !== null);
                    return nodes.some((el) => (el.innerText || el.textContent || '').replace(/\\s+/g, '').includes(label));
                }""",
                label,
            )
            if not exists:
                continue
            with page.expect_download(timeout=8000) as download_info:
                page.evaluate(
                    """(label) => {
                        const nodes = Array.from(document.querySelectorAll('a, button, [role=button]'))
                            .filter((el) => el.offsetParent !== null);
                        const node = nodes.find((el) => (el.innerText || el.textContent || '').replace(/\\s+/g, '').includes(label));
                        if (!node) return false;
                        node.scrollIntoView({ block: 'center' });
                        node.click();
                        return true;
                    }""",
                    label,
                )
            download = download_info.value
            download.save_as(str(pdf_path))
            return pdf_path
        except Exception:
            continue
    return None


def main() -> int:
    args = parse_args()
    setup_logging(args.log_level)
    return run_compare(args)


def run_compare(args: argparse.Namespace) -> int:
    """Run the taskId-driven comparison flow.

    This function is intentionally importable so the repository has one
    real comparison implementation while multiple CLI wrappers can call it.
    Prefer invoking it through ``main.py`` for day-to-day use.
    """

    if args.skip_api:
        if is_auto_targets(args.targets):
            raise ValueError("--targets auto cannot be used with --skip-api")
        targets = resolve_targets(args.targets)
        mappings_by_target = {target.target_id: load_mappings(target) for target in targets}
        return 0

    api_response = fetch_api_payload(args.task_id)
    api_by_tax = api_response.get("data", {})
    province = api_response.get("province", "")
    expected_tax_no = extract_expected_tax_no(api_response)
    LOGGER.info("Fetched API data: province=%s tax_codes=%s", province, ", ".join(api_by_tax.keys()))
    if is_auto_targets(args.targets):
        mappings_by_target = load_auto_mappings()
        targets = resolve_auto_targets(api_by_tax, mappings_by_target)
    else:
        targets = resolve_targets(args.targets)
        mappings_by_target = {target.target_id: load_mappings(target) for target in targets}
    if not targets:
        LOGGER.warning("No compare targets selected for task_id=%s", args.task_id)
        return 2
    targets = filter_targets_with_api_data(api_by_tax, targets, mappings_by_target)
    if not targets:
        LOGGER.warning("No compare targets have API field coverage for task_id=%s", args.task_id)
        return 2
    current_period_flags_by_target = fetch_current_period_flags_for_targets(args.task_id, targets)

    if args.skip_browser:
        for target in targets:
            present = sum(1 for mapping in mappings_by_target[target.target_id] if api_value(api_by_tax, target, mapping.field_id) is not None)
            LOGGER.info("%s API field coverage: %s/%s", target.target_id, present, len(mappings_by_target[target.target_id]))
        logged_flags = {key: value for key, value in current_period_flags_by_target.items() if value is not None}
        if logged_flags:
            LOGGER.info("Task current-period flags: %s", logged_flags)
        return 0

    bm = connect_browser(args)
    try:
        page = bm.find_page_by_url("chanjet.com") or bm.get_page()
        page.goto(CHANJET_TASK_URL, wait_until="domcontentloaded", timeout=30000)
        chanjet_page = wait_for_chanjet_login(bm, args.chanjet_timeout)

        tax_page = find_existing_tax_page(bm, province, expected_tax_no)
        if tax_page:
            LOGGER.info(
                "Reusing existing tax bureau page: province=%s tax_no=%s url=%s",
                province,
                expected_tax_no,
                tax_page.url,
            )
        else:
            tax_page, province, _ = login_tax_page_for_task(
                bm,
                chanjet_page,
                args,
                province,
                expected_tax_no,
            )

        exit_code = 0
        run_outputs: list[dict[str, Any]] = []
        previous_declared_target: CompareTarget | None = None
        for target in targets:
            mappings = mappings_by_target[target.target_id]
            web_config = get_web_config(args.config_root, target.tax_type)
            raw_current_period_flag = current_period_flags_by_target.get(target.target_id)
            declaration_status_override = str(getattr(args, "declaration_status_override", "") or "")
            target_current_period_flag = effective_current_period_flag_for_target(
                target,
                raw_current_period_flag,
                declaration_status_override=declaration_status_override,
            )
            comparison_mappings = mappings_for_comparison(target, mappings, api_by_tax)
            not_comparable_reason = ""
            if len(comparison_mappings) != len(mappings):
                LOGGER.info(
                    "%s compare policy: ignoring %s field(s) absent from API response",
                    target.target_id,
                    len(mappings) - len(comparison_mappings),
                )
            if not comparison_mappings:
                not_comparable_reason = "api_no_comparable_fields"
                LOGGER.warning(
                    "%s has no comparable API fields after target selection; "
                    "writing an evidence-only report without forcing tax-page form navigation.",
                    target.target_id,
                )
                result = compare_target(target, comparison_mappings, api_by_tax, {})
                output_path = save_result(
                    args.task_id,
                    target,
                    result,
                    province=province,
                    current_period_flag=target_current_period_flag,
                    quality_issues=[],
                    not_comparable_reason=not_comparable_reason,
                )
                api_excel_path = save_api_filled_workbook(args.task_id, target, mappings, api_by_tax, output_path.parent)
                run_outputs.append(
                    {
                        "target_id": target.target_id,
                        "form_name": target.form_name,
                        "report_path": str(output_path),
                        "api_excel_path": str(api_excel_path) if api_excel_path else "",
                        "pdf_path": "",
                        "quality_issues": [],
                    }
                )
                LOGGER.info(
                    "%s complete: total=0 match=0 mismatch=0 api_missing=0 web_missing=0 match_rate=100.0%% report=%s api_excel=%s pdf=",
                    target.target_id,
                    output_path,
                    api_excel_path or "",
                )
                continue
            if raw_current_period_flag is None and target_current_period_flag is False:
                LOGGER.info(
                    "No current-period marker found; treating target=%s as undeclared for tax-bureau navigation (source=%s)",
                    target.target_id,
                    current_period_flag_source(raw_current_period_flag, declaration_status_override),
                )
            if raw_current_period_flag is None and target_current_period_flag is True:
                LOGGER.info(
                    "No current-period marker found; treating target=%s as declared for tax-bureau navigation (source=%s)",
                    target.target_id,
                    current_period_flag_source(raw_current_period_flag, declaration_status_override),
                )
            ensure_supported_declaration_flow(target, target_current_period_flag)
            fallback_current_period_flag = fallback_status_flag(
                raw_current_period_flag,
                target_current_period_flag,
                declaration_status_override=declaration_status_override,
            )
            if target_current_period_flag is False and supports_undeclared_tax_page(target):
                try:
                    tax_page, province = open_undeclared_target_with_auth_retry(
                        bm=bm,
                        chanjet_page=chanjet_page,
                        args=args,
                        tax_page=tax_page,
                        province=province,
                        expected_tax_no=expected_tax_no,
                        target=target,
                        api_response=api_response,
                        mappings=mappings,
                    )
                    previous_declared_target = None
                except UndeclaredTaxAlreadyDeclaredError as exc:
                    if not can_fallback_to_declared_query_after_already_declared_conflict(fallback_current_period_flag):
                        raise
                    LOGGER.warning(
                        "Tax bureau reports already declared while declaration status is unknown; "
                        "falling back to declared-query flow: %s",
                        exc,
                    )
                    raw_current_period_flag = True
                    current_period_flags_by_target[target.target_id] = True
                    target_current_period_flag = True
                    tax_page = ensure_target_page(
                        tax_page,
                        target,
                        mappings,
                        web_config,
                        prefer_detail_form_switch=False,
                        api_response=api_response,
                    )
                    previous_declared_target = target if is_declaration_detail_page(tax_page) else None
                except UndeclaredTaxTargetUnavailableError as exc:
                    if not can_fallback_to_declared_query_after_undeclared_unavailable(fallback_current_period_flag):
                        raise
                    LOGGER.warning(
                        "Undeclared target is unavailable while declaration status is unknown; "
                        "falling back to declared-query flow: %s",
                        exc,
                    )
                    raw_current_period_flag = True
                    current_period_flags_by_target[target.target_id] = True
                    target_current_period_flag = True
                    tax_page = ensure_target_page(
                        tax_page,
                        target,
                        mappings,
                        web_config,
                        prefer_detail_form_switch=False,
                        api_response=api_response,
                    )
                    previous_declared_target = target if is_declaration_detail_page(tax_page) else None
            else:
                tax_page = ensure_target_page(
                    tax_page,
                    target,
                    mappings,
                    web_config,
                    prefer_detail_form_switch=can_switch_detail_form_between(previous_declared_target, target),
                    api_response=api_response,
                )
                previous_declared_target = target if is_declaration_detail_page(tax_page) else None
            LOGGER.info("Confirming target page before extraction: %s", target.target_id)
            confirm_target_page_for_evidence(tax_page, target, mappings, target_current_period_flag)
            LOGGER.info("Target page confirmed; extracting web data: %s", target.target_id)
            web_data = extract_web_data_reliably(
                tax_page,
                target,
                mappings,
                comparison_mappings,
                api_by_tax,
                current_period_flag=target_current_period_flag,
            )
            low_web_coverage = is_low_web_extraction_coverage(target, web_data, comparison_mappings)
            if low_web_coverage:
                found, total, ratio = web_extraction_coverage(web_data, comparison_mappings)
                LOGGER.warning(
                    "%s low web extraction coverage: %s/%s (%.2f%%)",
                    target.target_id,
                    found,
                    total,
                    ratio * 100,
                )
            result = compare_target(
                target,
                comparison_mappings,
                api_by_tax,
                web_data,
                current_period_flag=target_current_period_flag,
            )
            quality_issues = comparison_quality_issues(
                target,
                result,
                low_web_coverage,
                current_period_flag=target_current_period_flag,
            )
            if quality_issues:
                LOGGER.warning("%s quality issues: %s", target.target_id, "; ".join(quality_issues))
            output_path = save_result(
                args.task_id,
                target,
                result,
                province=province,
                current_period_flag=target_current_period_flag,
                quality_issues=quality_issues,
                not_comparable_reason=not_comparable_reason,
            )
            api_excel_path = save_api_filled_workbook(args.task_id, target, mappings, api_by_tax, output_path.parent)
            pdf_path = None
            if not args.skip_pdf:
                pdf_path = save_web_pdf(tax_page, args.task_id, target, output_path.parent)
            run_outputs.append(
                {
                    "target_id": target.target_id,
                    "form_name": target.form_name,
                    "report_path": str(output_path),
                    "api_excel_path": str(api_excel_path) if api_excel_path else "",
                    "pdf_path": str(pdf_path) if pdf_path else "",
                    "quality_issues": quality_issues,
                }
            )
            summary = result.summary
            LOGGER.info(
                "%s complete: total=%s match=%s mismatch=%s api_missing=%s web_missing=%s match_rate=%s%% report=%s api_excel=%s pdf=%s",
                target.target_id,
                summary.total_fields,
                summary.match_count + summary.tolerance_match_count,
                summary.mismatch_count,
                summary.api_missing_count,
                summary.web_missing_count,
                summary.match_rate,
                output_path,
                api_excel_path or "",
                pdf_path or "",
            )
            if quality_issues:
                exit_code = 1
        combined_path = render_task_summary_report(args.task_id, run_outputs)
        if combined_path:
            LOGGER.info("Combined compare report: %s", combined_path)
        return exit_code
    finally:
        bm.close()


if __name__ == "__main__":
    raise SystemExit(main())
