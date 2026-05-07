"""Generate a mock VAT mapping Excel file for testing."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill

MOCK_FIELDS = [
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "应征增值税不含税销售额（3%征收率）",
        "栏次": "1",
        "列名称": "本期数-货物及劳务",
        "接口字段ID": "sales_3_goods",
        "JSONPath": "$.data.salesGoods3Percent",
        "数据类型": "amount",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable tr[data-line='1'] td[data-col='goods_current']",
        "容差": 0.01,
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "应征增值税不含税销售额（3%征收率）",
        "栏次": "1",
        "列名称": "本期数-服务不动产无形资产",
        "接口字段ID": "sales_3_service",
        "JSONPath": "$.data.salesService3Percent",
        "数据类型": "amount",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable tr[data-line='1'] td[data-col='service_current']",
        "容差": 0.01,
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "应征增值税不含税销售额（5%征收率）",
        "栏次": "2",
        "列名称": "本期数-货物及劳务",
        "接口字段ID": "sales_5_goods",
        "JSONPath": "$.data.salesGoods5Percent",
        "数据类型": "amount",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable tr[data-line='2'] td[data-col='goods_current']",
        "容差": 0.01,
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "税款所属期",
        "栏次": "—",
        "列名称": "税款所属期",
        "接口字段ID": "tax_period",
        "JSONPath": "$.data.taxPeriod",
        "数据类型": "text",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable .tax-period-value",
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "本期应纳税额",
        "栏次": "15",
        "列名称": "本期应纳税额",
        "接口字段ID": "tax_due_current",
        "JSONPath": "$.data.taxDueCurrent",
        "数据类型": "amount",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "是",
        "网页selector": "#declareTable tr[data-line='15'] td[data-col='current']",
        "备注": "计算项，暂不比对",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "税率",
        "栏次": "—",
        "列名称": "适用税率",
        "接口字段ID": "tax_rate",
        "JSONPath": "$.data.taxRate",
        "数据类型": "rate",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable .tax-rate-value",
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "申报日期",
        "栏次": "—",
        "列名称": "申报日期",
        "接口字段ID": "declare_date",
        "JSONPath": "$.data.declareDate",
        "数据类型": "date",
        "是否比对": "是",
        "是否必填": "是",
        "是否计算项": "否",
        "网页selector": "#declareTable .declare-date-value",
        "备注": "",
    },
    {
        "税种": "增值税",
        "表单编码": "VAT_SMALL_SCALE_MAIN",
        "表单名称": "增值税及附加税费申报表（小规模纳税人适用）",
        "表单版本": "2026",
        "页码": 1,
        "行名称": "免税销售额",
        "栏次": "9",
        "列名称": "免税销售额",
        "接口字段ID": "tax_exempt_sales",
        "JSONPath": "$.data.taxExemptSales",
        "数据类型": "empty_or_dash",
        "是否比对": "是",
        "是否必填": "否",
        "是否计算项": "否",
        "网页selector": "#declareTable tr[data-line='9'] td[data-col='current']",
        "备注": "可能为横线或空值",
    },
]


def create_mock_excel(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "增值税主表"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=1, column=1, value="增值税小规模纳税人申报表 - 字段映射关系")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Header row
    headers = list(MOCK_FIELDS[0].keys())
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows
    for row_idx, field in enumerate(MOCK_FIELDS, 3):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=field.get(header))

    # Create a second sheet for supplemental form
    ws2 = wb.create_sheet("附列资料")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws2.cell(row=1, column=1, value="增值税附列资料 - 字段映射关系")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

    sup_headers = ["税种", "表单编码", "行名称", "栏次", "列名称",
                   "接口字段ID", "JSONPath", "数据类型", "是否比对", "备注"]
    for col_idx, header in enumerate(sup_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sup_data = [
        ["增值税", "VAT_SMALL_SCALE_SUPPLEMENT", "服务不动产无形资产(3%)", "1",
         "本期数", "service_3_current", "$.data.service3Current", "amount", "是", ""],
        ["增值税", "VAT_SMALL_SCALE_SUPPLEMENT", "服务不动产无形资产(5%)", "2",
         "本期数", "service_5_current", "$.data.service5Current", "amount", "是", ""],
    ]
    for row_idx, row_data in enumerate(sup_data, 3):
        for col_idx, val in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_path)
    print(f"Mock Excel created: {output_path}")


if __name__ == "__main__":
    output = str(Path(__file__).parent.parent.parent / "mappings" / "vat_small_scale_mapping.xlsx")
    create_mock_excel(output)