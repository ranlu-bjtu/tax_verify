import shutil
import sys
import tempfile
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from scripts.compare_tax_forms import (  # noqa: E402
    TARGETS,
    can_fallback_to_declared_query_after_undeclared_unavailable,
    can_fallback_to_declared_query_after_already_declared_conflict,
    compare_target,
    comparison_quality_issues,
    declaration_status_for_target,
    fallback_status_flag,
    effective_current_period_flag_for_target,
    ensure_supported_declaration_flow,
    extract_consumption_tax_data,
    load_layout_scan_mappings,
    load_workbook_compat,
    mappings_for_comparison,
    parse_consumption_tax_main_rows,
    parse_consumption_tax_surcharge_rows,
    resolve_auto_targets,
)
from src.models.field_mapping import DataType, FieldMapping  # noqa: E402


def inject_unsupported_sheet_protection_attr(workbook: Path) -> None:
    temp = workbook.with_suffix(".tmp.xlsx")
    with ZipFile(workbook, "r") as source, ZipFile(temp, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                if "<sheetProtection" not in text:
                    text = text.replace(
                        "<sheetData>",
                        '<sheetProtection sheet="1" allowResizeRows="1"/><sheetData>',
                    )
                data = text.encode("utf-8")
            target.writestr(info, data)
    workbook.unlink()
    shutil.move(temp, workbook)


def test_load_workbook_compat_handles_unsupported_sheet_protection_attr():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "protected.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "field_id"
        wb.save(path)
        wb.close()
        inject_unsupported_sheet_protection_attr(path)

        loaded = load_workbook_compat(path, data_only=True, read_only=True)
        try:
            values = [row[0] for row in loaded.active.iter_rows(values_only=True)]
        finally:
            loaded.close()

        assert values == ["field_id"]


def test_consumption_tax_layout_loader_filters_non_api_labels():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "consumption.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["纳税人识别号", "nsrsbh_xfsjfjsfsbb", "a"])
        ws.append(["项目", "ysxfpmc1_xfsjfjsfsbb", "bqxse1_xfsjfjsfsbb"])
        wb.save(path)
        wb.close()

        mappings = load_layout_scan_mappings(TARGETS["consumption_tax_main"], path)

        assert [mapping.field_id for mapping in mappings] == [
            "ysxfpmc1_xfsjfjsfsbb",
            "bqxse1_xfsjfjsfsbb",
        ]


def test_auto_targets_include_consumption_tax_forms_when_api_has_fields():
    mappings_by_target = {target_id: [] for target_id in TARGETS}
    mappings_by_target["consumption_tax_main"] = load_layout_scan_mappings(
        TARGETS["consumption_tax_main"],
        _workbook_with_fields("xfs-main.xlsx", ["bqxse1_xfsjfjsfsbb"]),
    )
    mappings_by_target["consumption_tax_surcharge"] = load_layout_scan_mappings(
        TARGETS["consumption_tax_surcharge"],
        _workbook_with_fields("xfs-surcharge.xlsx", ["jze1_xfsfjsfjsb"]),
    )
    api_by_tax = {
        "sz_xfs": {
            "xfszb_qc.bqxse1_xfsjfjsfsbb": "1",
            "xfsfb1_qc.jze1_xfsfjsfjsb": "2",
        }
    }

    selected = resolve_auto_targets(api_by_tax, mappings_by_target)

    assert [target.target_id for target in selected] == [
        "consumption_tax_main",
        "consumption_tax_surcharge",
    ]


def test_auto_targets_keep_all_consumption_tax_forms_when_one_form_has_api_fields():
    mappings_by_target = {target_id: [] for target_id in TARGETS}
    mappings_by_target["consumption_tax_main"] = load_layout_scan_mappings(
        TARGETS["consumption_tax_main"],
        _workbook_with_fields("xfs-main.xlsx", ["bqxse1_xfsjfjsfsbb"]),
    )
    mappings_by_target["consumption_tax_surcharge"] = load_layout_scan_mappings(
        TARGETS["consumption_tax_surcharge"],
        _workbook_with_fields("xfs-surcharge.xlsx", ["jze1_xfsfjsfjsb"]),
    )
    api_by_tax = {
        "sz_xfs": {
            "xfszb_qc.bqxse1_xfsjfjsfsbb": "1",
        }
    }

    selected = resolve_auto_targets(api_by_tax, mappings_by_target)

    assert [target.target_id for target in selected] == [
        "consumption_tax_main",
        "consumption_tax_surcharge",
    ]


def test_zero_api_field_report_keeps_target_metadata():
    target = TARGETS["consumption_tax_surcharge"]
    mappings = [
        FieldMapping(
            tax_type=target.tax_type,
            form_code=target.form_code,
            form_name=target.form_name,
            field_id="jze1_xfsfjsfjsb",
            display_name="jze1",
            data_type=DataType.AMOUNT,
        )
    ]

    result = compare_target(target, mappings, {"sz_xfs": {}}, {})

    assert result.summary.total_fields == 0
    assert result.tax_type == target.tax_type
    assert result.form_code == target.form_code
    assert result.form_name == target.form_name


def test_consumption_tax_unknown_declaration_status_defaults_to_unfiled():
    target = TARGETS["consumption_tax_main"]

    assert declaration_status_for_target(target, None) == "未申报"
    assert declaration_status_for_target(target, True) == "已申报"
    assert declaration_status_for_target(target, False) == "未申报"
    assert effective_current_period_flag_for_target(target, None) is False
    assert effective_current_period_flag_for_target(target, None, declaration_status_override="filed") is True
    assert effective_current_period_flag_for_target(target, None, declaration_status_override="unfiled") is False
    assert effective_current_period_flag_for_target(target, True) is True
    assert effective_current_period_flag_for_target(target, True, declaration_status_override="unfiled") is True
    assert effective_current_period_flag_for_target(target, False) is False
    assert fallback_status_flag(None, False) is None
    assert fallback_status_flag(None, False, declaration_status_override="unfiled") is False
    assert "declaration_status_unknown" not in comparison_quality_issues(
        target,
        _comparison_result_stub(),
        low_web_coverage=False,
        current_period_flag=None,
    )


def test_unknown_status_only_case_can_fallback_when_unfiled_entry_is_absent():
    assert can_fallback_to_declared_query_after_undeclared_unavailable(None) is True
    assert can_fallback_to_declared_query_after_undeclared_unavailable(False) is False
    assert can_fallback_to_declared_query_after_undeclared_unavailable(True) is False


def test_known_unfiled_status_cannot_fallback_when_tax_bureau_reports_already_declared():
    assert can_fallback_to_declared_query_after_already_declared_conflict(None) is True
    assert can_fallback_to_declared_query_after_already_declared_conflict(False) is False
    assert can_fallback_to_declared_query_after_already_declared_conflict(True) is False


def test_consumption_tax_unfiled_flow_is_supported():
    ensure_supported_declaration_flow(TARGETS["consumption_tax_main"], False)


def test_consumption_tax_main_parser_uses_current_period_columns():
    rows = [
        ["1", "2", "3", "4", "a", "5", "b", "6=1×4+2×5", "c"],
        ["金银首饰", "0.00", "0.05", "", "0.00", "0.00", "206,256.64", "1,089,600.87", "10,312.83", "54,480.04"],
        ["合计", "——", "——", "——", "——", "——", "——", "——", "10,312.83", "54,480.04"],
        ["本期应补（退）税额", "14=6-7-11-13", "10,312.83", "54,480.04"],
        ["城市维护建设税本期应补（退）税额", "15", "360.95", "1,906.80"],
    ]

    parsed = parse_consumption_tax_main_rows(rows)

    assert parsed["ysxfpmc1_xfsjfjsfsbb"] == "金银首饰"
    assert parsed["bqxse1_xfsjfjsfsbb"] == "206,256.64"
    assert parsed["bqynse1_xfsjfjsfsbb"] == "10,312.83"
    assert parsed["bqybtse_xfsjfjsfsbb"] == "10,312.83"
    assert parsed["bnljybtse_xfsjfjsfsbb"] == "54,480.04"
    assert parsed["bqcswhjssybtse_xfsjfjsfsbb"] == "360.95"


def test_consumption_tax_main_parser_finds_existing_surcharge_summary_rows():
    rows = [
        ["1", "2", "3", "4", "a", "5", "b", "6=1×4+2×5", "c"],
        ["金银首饰", "0.00", "0.05", "", "0.00", "0.00", "142,344.40", "6,403,029.40", "7,117.22", "320,151.47"],
        ["合计", "—", "—", "—", "—", "—", "—", "—", "7,117.22", "320,151.47"],
        ["本期应补（退）税额", "14=6-7-11-13", "7,117.22", "320,151.47"],
        ["城市维护建设税本期应补（退）税额", "15", "249.10", "11,205.29"],
        ["教育费附加本期应补（退）费额", "16", "106.76", "4,802.25"],
        ["地方教育附加本期应补（退）费额", "17", "71.17", "3,201.51"],
    ]

    parsed = parse_consumption_tax_main_rows(rows)

    assert parsed["bqybtse_xfsjfjsfsbb"] == "7,117.22"
    assert parsed["bnljybtse_xfsjfjsfsbb"] == "320,151.47"
    assert parsed["bqcswhjssybtse_xfsjfjsfsbb"] == "249.10"
    assert parsed["bnljcswhjssybtse_xfsjfjsfsbb"] == "11,205.29"
    assert parsed["bqjyffjybtse_xfsjfjsfsbb"] == "106.76"
    assert parsed["bnljjyffjybtse_xfsjfjsfsbb"] == "4,802.25"
    assert parsed["bqdfjyfjybtse_xfsjfjsfsbb"] == "71.17"
    assert parsed["bnljdfjyfjybtse_xfsjfjsfsbb"] == "3,201.51"


def test_consumption_tax_extraction_collects_rows_after_scroll_render():
    target = TARGETS["consumption_tax_main"]
    mappings = [
        FieldMapping(
            tax_type=target.tax_type,
            form_code=target.form_code,
            form_name=target.form_name,
            field_id=field_id,
            display_name=field_id,
            data_type=DataType.AMOUNT,
        )
        for field_id in [
            "bqybtse_xfsjfjsfsbb",
            "bnljybtse_xfsjfjsfsbb",
            "bqcswhjssybtse_xfsjfjsfsbb",
            "bnljcswhjssybtse_xfsjfjsfsbb",
            "bqjyffjybtse_xfsjfjsfsbb",
            "bnljjyffjybtse_xfsjfjsfsbb",
            "bqdfjyfjybtse_xfsjfjsfsbb",
            "bnljdfjyfjybtse_xfsjfjsfsbb",
        ]
    ]

    class FakePage:
        def __init__(self):
            self.position = "current"

        def evaluate(self, script, *args):
            if args:
                self.position = args[0]
                return None
            if "document.querySelectorAll('table tr')" not in script:
                return None
            if self.position == "bottom":
                return [
                    ["本期应补（退）税额", "14=6-7-11-13", "7,117.22", "320,151.47"],
                    ["城市维护建设税本期应补（退）税额", "15", "249.10", "11,205.29"],
                    ["教育费附加本期应补（退）费额", "16", "106.76", "4,802.25"],
                    ["地方教育附加本期应补（退）费额", "17", "71.17", "3,201.51"],
                ]
            return [
                ["1", "2", "3", "4", "a", "5", "b", "6=1×4+2×5", "c"],
                ["金银首饰", "0.00", "0.05", "", "0.00", "0.00", "142,344.40", "6,403,029.40", "7,117.22", "320,151.47"],
                ["合计", "—", "—", "—", "—", "—", "—", "—", "7,117.22", "320,151.47"],
            ]

    data = extract_consumption_tax_data(FakePage(), target, mappings)

    assert data["bqybtse_xfsjfjsfsbb"] == "7,117.22"
    assert data["bnljybtse_xfsjfjsfsbb"] == "320,151.47"
    assert data["bqcswhjssybtse_xfsjfjsfsbb"] == "249.10"
    assert data["bnljcswhjssybtse_xfsjfjsfsbb"] == "11,205.29"
    assert data["bqjyffjybtse_xfsjfjsfsbb"] == "106.76"
    assert data["bnljjyffjybtse_xfsjfjsfsbb"] == "4,802.25"
    assert data["bqdfjyfjybtse_xfsjfjsfsbb"] == "71.17"
    assert data["bnljdfjyfjybtse_xfsjfjsfsbb"] == "3,201.51"


def test_consumption_tax_retry_scans_near_bottom_virtual_rows():
    target = TARGETS["consumption_tax_main"]
    mappings = [
        FieldMapping(
            tax_type=target.tax_type,
            form_code=target.form_code,
            form_name=target.form_name,
            field_id="bqybtse_xfsjfjsfsbb",
            display_name="bqybtse_xfsjfjsfsbb",
            data_type=DataType.AMOUNT,
        )
    ]

    class FakePage:
        def __init__(self):
            self.position = "current"
            self.positions = []

        def evaluate(self, script, *args):
            if args:
                self.position = args[0]
                self.positions.append(self.position)
                return None
            if "document.querySelectorAll('table tr')" not in script:
                return None
            if self.position == "near_bottom":
                return [["14=6-7-11-13", "", "7,117.22", "320,151.47"]]
            return [
                ["1", "2", "3", "4", "a", "5", "b", "6=1*4+2*5", "c"],
                ["summary", "", "", "", "", "", "", "", "7,117.22", "320,151.47"],
            ]

    page = FakePage()
    data = extract_consumption_tax_data(page, target, mappings)

    assert data["bqybtse_xfsjfjsfsbb"] == "7,117.22"
    assert "near_bottom" in page.positions


def test_culture_fee_rate_field_is_excluded_from_comparison():
    target = TARGETS["culture_fee_main"]
    mappings = [
        FieldMapping(
            tax_type=target.tax_type,
            form_code=target.form_code,
            form_name=target.form_name,
            field_id="fl_bys",
            display_name="fl_bys",
            data_type=DataType.RATE,
        ),
        FieldMapping(
            tax_type=target.tax_type,
            form_code=target.form_code,
            form_name=target.form_name,
            field_id="yjfe_bys",
            display_name="yjfe_bys",
            data_type=DataType.AMOUNT,
        ),
    ]
    api_by_tax = {
        target.tax_code: {
            "fl_bys": "0.00",
            "yjfe_bys": "0.00",
        }
    }

    comparison_mappings = mappings_for_comparison(target, mappings, api_by_tax)

    assert [mapping.field_id for mapping in comparison_mappings] == ["yjfe_bys"]


def test_consumption_tax_surcharge_parser_maps_rows_and_totals():
    rows = [
        ["城市维护建设税", "市区", "10,312.83", "0.07", "721.90", "", "0.00", "减免性质", "50.00", "360.95", "0.00", "360.95"],
        ["教育费附加", "教育费附加", "10,312.83", "0.03", "309.38", "", "0.00", "减免性质", "50.00", "154.69", "0.00", "154.69"],
        ["合计", "——", "——", "——", "1,237.54", "——", "0.00", "——", "——", "618.77", "0.00", "618.77"],
    ]

    parsed = parse_consumption_tax_surcharge_rows(rows)

    assert parsed["jsyjsfsse1_xfsfjsfjsb"] == "10,312.83"
    assert parsed["sfl2_xfsfjsfjsb"] == "0.03"
    assert parsed["jzbl1_xfsfjsfjsb"] == "0.5"
    assert parsed["jze1_xfsfjsfjsb"] == "360.95"
    assert parsed["bqynsfe4_xfsfjsfjsb"] == "1,237.54"
    assert parsed["bqybtsfe4_xfsfjsfjsb"] == "618.77"


def _workbook_with_fields(name: str, fields: list[str]) -> Path:
    temp_root = Path(tempfile.gettempdir()) / "tax_verify_consumption_tests"
    temp_root.mkdir(parents=True, exist_ok=True)
    path = temp_root / name
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(fields)
    wb.save(path)
    wb.close()
    return path


class _SummaryStub:
    mismatch_count = 0
    api_missing_count = 0
    web_missing_count = 0
    parse_error_count = 0
    mapping_error_count = 0
    both_missing_count = 0
    total_fields = 1


class _ComparisonResultStub:
    summary = _SummaryStub()


def _comparison_result_stub():
    return _ComparisonResultStub()


if __name__ == "__main__":
    test_load_workbook_compat_handles_unsupported_sheet_protection_attr()
    test_consumption_tax_layout_loader_filters_non_api_labels()
    test_auto_targets_include_consumption_tax_forms_when_api_has_fields()
    test_auto_targets_keep_all_consumption_tax_forms_when_one_form_has_api_fields()
    test_zero_api_field_report_keeps_target_metadata()
    test_consumption_tax_unknown_declaration_status_defaults_to_unfiled()
    test_unknown_status_only_case_can_fallback_when_unfiled_entry_is_absent()
    test_consumption_tax_unfiled_flow_is_supported()
    test_consumption_tax_main_parser_uses_current_period_columns()
    test_consumption_tax_main_parser_finds_existing_surcharge_summary_rows()
    test_consumption_tax_extraction_collects_rows_after_scroll_render()
    test_consumption_tax_retry_scans_near_bottom_virtual_rows()
    test_culture_fee_rate_field_is_excluded_from_comparison()
    test_consumption_tax_surcharge_parser_maps_rows_and_totals()
    print("All consumption tax support tests passed!")
